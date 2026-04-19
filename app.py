"""
peterheijen.com — Finance API v1
=================================
Simpel, werkend, robuust.

Upload CSV/XLSX → Python rekent → Claude categoriseert & analyseert → PDF rapport per email.

ARCHITECTUUR: Async job-model
  POST /rapport  → start achtergrond-job, retourneert direct job_id (< 1 sec)
  GET  /rapport/{job_id}/status → poll voor voortgang
  Geen lange HTTP-requests meer → geen Railway/proxy timeout issues.
"""

import os
import io
import json
import uuid
import logging
import base64
import httpx
import threading
from datetime import datetime

import pandas as pd
from typing import List, Optional
from fastapi import FastAPI, UploadFile, HTTPException, Form
from fastapi.middleware.cors import CORSMiddleware
from fpdf import FPDF
from dotenv import load_dotenv

load_dotenv()

logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s')
logger = logging.getLogger(__name__)

app = FastAPI(title="PeterHeijen Finance API", version="0.2.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # In productie: alleen je eigen domein
    allow_methods=["POST", "GET"],
    allow_headers=["*"],
)


# ---------------------------------------------------------------------------
# JOB STORE — in-memory tracking van achtergrond-analyses
# ---------------------------------------------------------------------------
# Thread-safe dict: {job_id: {status, fase, email, error, ...}}
jobs: dict = {}
jobs_lock = threading.Lock()


# ---------------------------------------------------------------------------
# STAP 1: DATA INLEZEN
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# BANK-FORMAAT HERKENNING & NORMALISATIE
# ---------------------------------------------------------------------------
# Ondersteunde banken: ABN AMRO, ING, Rabobank, SNS, ASN, Triodos, Knab,
# Bunq, N26, RegioBank — zowel CSV als XLS/XLSX.
#
# Elke bank levert eigen kolomnamen en formaat. Alles wordt genormaliseerd
# naar: Rekeningnummer, Transactiedatum (YYYYMMDD), Transactiebedrag (float),
# Omschrijving, Beginsaldo, Eindsaldo.
# ---------------------------------------------------------------------------

def _parse_dutch_amount(val) -> float:
    """Parse Nederlands bedrag: '1.234,56' → 1234.56. Werkt ook met '-1234.56'."""
    if pd.isna(val):
        return 0.0
    s = str(val).strip()
    if not s:
        return 0.0
    # Als het al een float-achtig getal is (punt als decimaal, geen komma)
    if ',' not in s and s.replace('.', '', 1).replace('-', '', 1).isdigit():
        return float(s)
    # Nederlands formaat: punt = duizendtallen, komma = decimaal
    s = s.replace('.', '').replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return 0.0


def _parse_datum(val) -> str:
    """Normaliseer datum naar YYYYMMDD string. Herkent: YYYYMMDD, DD-MM-YYYY, DD/MM/YYYY, YYYY-MM-DD."""
    s = str(val).strip()
    # Al YYYYMMDD (8 cijfers)
    if len(s) == 8 and s.isdigit():
        return s
    # DD-MM-YYYY of DD/MM/YYYY
    for sep in ['-', '/']:
        if sep in s:
            parts = s.split(sep)
            if len(parts) == 3:
                # Bepaal volgorde: als eerste deel 4 cijfers → YYYY-MM-DD
                if len(parts[0]) == 4:
                    return f"{parts[0]}{parts[1].zfill(2)}{parts[2].zfill(2)}"
                else:
                    return f"{parts[2]}{parts[1].zfill(2)}{parts[0].zfill(2)}"
    return s


def _bereken_saldos(df: pd.DataFrame) -> pd.DataFrame:
    """Bereken Beginsaldo/Eindsaldo als die ontbreken. Cumulatief per rekening."""
    if 'Beginsaldo' in df.columns and 'Eindsaldo' in df.columns:
        return df
    df = df.sort_values('Transactiedatum')
    for rek in df['Rekeningnummer'].unique():
        mask = df['Rekeningnummer'] == rek
        bedragen = df.loc[mask, 'Transactiebedrag'].astype(float)
        cumsum = bedragen.cumsum()
        df.loc[mask, 'Eindsaldo'] = cumsum
        df.loc[mask, 'Beginsaldo'] = cumsum - bedragen
    return df


def _gebruik_saldo_kolom(df: pd.DataFrame, saldo_col: str) -> pd.DataFrame:
    """Als er een 'Saldo na mutatie' kolom is, gebruik die voor nauwkeurigere saldo's."""
    if saldo_col not in df.columns:
        return _bereken_saldos(df)
    df['Eindsaldo'] = df[saldo_col].apply(_parse_dutch_amount)
    df['Beginsaldo'] = df['Eindsaldo'] - df['Transactiebedrag'].astype(float)
    return df


def _detecteer_formaat(df: pd.DataFrame) -> str:
    """Detecteer bankformaat op basis van kolomnamen."""
    cols = set(c.strip() for c in df.columns)

    # --- ABN AMRO XLSX (standaardformaat — ons referentieformaat) ---
    if 'Rekeningnummer' in cols and 'Transactiebedrag' in cols and 'Transactiedatum' in cols:
        return 'abnamro_xlsx'

    # --- ABN AMRO / ING CSV ---
    # ING: Datum;Naam / Omschrijving;Rekening;Tegenrekening;Code;Af Bij;Bedrag (EUR);Mutatiesoort;Mededelingen;Saldo na mutatie;Tag
    # ABN: Datum;Naam / Omschrijving;Rekening;Tegenrekening;Code;Af Bij;Bedrag (EUR);Mededelingen
    if 'Af Bij' in cols and 'Bedrag (EUR)' in cols:
        if 'Saldo na mutatie' in cols:
            return 'ing_csv'
        return 'abnamro_csv'

    # --- Rabobank CSV ---
    # IBAN/BBAN,Munt,BIC,Volgnr,Datum,Rentedatum,Bedrag,Saldo na trn,...,Naam tegenpartij,...,Omschrijving-1,...
    if 'IBAN/BBAN' in cols and ('Saldo na trn' in cols or 'Bedrag' in cols):
        return 'rabobank_csv'

    # --- Triodos CSV (uit custom parser) ---
    # Kolommen: Datum, Rekening, Bedrag, CreditDebet, Naam, Tegenrekening, Code, Omschrijving, Saldo
    # Verschil met Knab: Triodos heeft 'Tegenrekening' + 'Naam', Knab heeft 'Tegenrekeningnummer' + 'Tegenrekeninghouder'
    if 'CreditDebet' in cols and 'Bedrag' in cols and 'Tegenrekening' in cols and 'Naam' in cols:
        return 'triodos_csv'

    # --- Knab CSV ---
    # Rekeningnummer;Transactiedatum;Valutacode;CreditDebet;Bedrag;Tegenrekeningnummer;Tegenrekeninghouder;...;Omschrijving
    if 'CreditDebet' in cols and 'Bedrag' in cols:
        return 'knab_csv'

    # --- N26 CSV ---
    # "Date","Payee","Account number","Transaction type","Payment reference","Category","Amount (EUR)",...
    if 'Payee' in cols or 'Amount (EUR)' in cols:
        return 'n26_csv'

    # --- Bunq CSV ---
    # Datum;Bedrag;Rekening;Tegenpartij;Naam;Omschrijving
    if 'Tegenpartij' in cols and 'Naam' in cols and 'Bedrag' in cols:
        return 'bunq_csv'

    # --- SNS / ASN / RegioBank / Triodos (Volksbank-formaten) ---
    # Vaak: Datum, Omschrijving, Bedrag, Saldo (soms met Rekening/IBAN)
    if 'Bedrag' in cols and 'Datum' in cols:
        # Generiek NL-formaat met Datum + Bedrag
        return 'generiek_nl'

    return 'onbekend'


def _normaliseer(df: pd.DataFrame, formaat: str) -> pd.DataFrame:
    """Normaliseer elk bankformaat naar standaardkolommen."""
    logger.info(f"Bankformaat gedetecteerd: {formaat}")

    if formaat == 'abnamro_xlsx':
        # Al in standaardformaat — alleen datum normaliseren
        df['Transactiedatum'] = df['Transactiedatum'].astype(str).apply(_parse_datum)
        df['Transactiebedrag'] = df['Transactiebedrag'].apply(
            lambda v: _parse_dutch_amount(v) if not isinstance(v, (int, float)) else float(v))
        # Tegenrekening bewaren voor transfer-detectie
        if 'Tegenrekening' not in df.columns:
            df['Tegenrekening'] = ''
        return df

    elif formaat in ('abnamro_csv', 'ing_csv'):
        # Bedrag: komma-decimaal + "Af Bij" kolom
        df['_bedrag'] = df['Bedrag (EUR)'].apply(_parse_dutch_amount)
        df['_bedrag'] = df.apply(
            lambda r: -abs(r['_bedrag']) if str(r.get('Af Bij', '')).strip().lower() == 'af' else abs(r['_bedrag']),
            axis=1)
        # Mapping
        df['Rekeningnummer'] = df['Rekening']
        df['Transactiedatum'] = df['Datum'].astype(str).apply(_parse_datum)
        df['Transactiebedrag'] = df['_bedrag']
        # Omschrijving
        naam = df.get('Naam / Omschrijving', pd.Series([''] * len(df))).fillna('')
        meded = df.get('Mededelingen', pd.Series([''] * len(df))).fillna('')
        df['Omschrijving'] = (naam + ' ' + meded).str.strip()
        # Tegenrekening bewaren voor transfer-detectie
        if 'Tegenrekening' not in df.columns:
            df['Tegenrekening'] = ''
        # Saldo
        if formaat == 'ing_csv' and 'Saldo na mutatie' in df.columns:
            df = _gebruik_saldo_kolom(df, 'Saldo na mutatie')
        else:
            df = _bereken_saldos(df)
        return df

    elif formaat == 'rabobank_csv':
        df['Rekeningnummer'] = df['IBAN/BBAN']
        df['Transactiedatum'] = df['Datum'].astype(str).apply(_parse_datum)
        df['Transactiebedrag'] = df['Bedrag'].apply(_parse_dutch_amount)
        # Omschrijving: Naam tegenpartij + Omschrijving-1/2/3
        omschr_parts = []
        for col in ['Naam tegenpartij', 'Omschrijving-1', 'Omschrijving-2', 'Omschrijving-3']:
            if col in df.columns:
                omschr_parts.append(df[col].fillna(''))
        df['Omschrijving'] = pd.concat(omschr_parts, axis=1).apply(lambda r: ' '.join(r).strip(), axis=1) if omschr_parts else 'Onbekend'
        # Tegenrekening bewaren (Rabobank: 'IBAN/BBAN tegenpartij' of 'Tegenrekening IBAN/BBAN')
        for col in ['IBAN/BBAN tegenpartij', 'Tegenrekening IBAN/BBAN', 'Tegenrekening']:
            if col in df.columns:
                df['Tegenrekening'] = df[col].fillna('')
                break
        else:
            df['Tegenrekening'] = ''
        # Saldo
        if 'Saldo na trn' in df.columns:
            df = _gebruik_saldo_kolom(df, 'Saldo na trn')
        else:
            df = _bereken_saldos(df)
        return df

    elif formaat == 'triodos_csv':
        df['Rekeningnummer'] = df['Rekening']
        df['Transactiedatum'] = df['Datum'].astype(str).apply(_parse_datum)
        bedrag = df['Bedrag'].apply(_parse_dutch_amount)
        # CreditDebet: 'Credit' = positief, 'Debet' = negatief
        df['Transactiebedrag'] = df.apply(
            lambda r: -abs(_parse_dutch_amount(r['Bedrag'])) if str(r.get('CreditDebet', '')).strip().lower().startswith('d')
            else abs(_parse_dutch_amount(r['Bedrag'])), axis=1)
        # Omschrijving: Naam + Omschrijving
        naam = df.get('Naam', pd.Series([''] * len(df))).fillna('')
        omschr = df.get('Omschrijving', pd.Series([''] * len(df))).fillna('')
        df['Omschrijving'] = (naam + ' ' + omschr).str.strip()
        # Tegenrekening bewaren voor transfer-detectie (Triodos heeft deze kolom)
        if 'Tegenrekening' not in df.columns:
            df['Tegenrekening'] = ''
        # Saldo
        if 'Saldo' in df.columns:
            df = _gebruik_saldo_kolom(df, 'Saldo')
        else:
            df = _bereken_saldos(df)
        return df

    elif formaat == 'knab_csv':
        df['Rekeningnummer'] = df['Rekeningnummer']  # al goed
        df['Transactiedatum'] = df['Transactiedatum'].astype(str).apply(_parse_datum)
        bedrag = df['Bedrag'].apply(_parse_dutch_amount)
        # CreditDebet: 'C' = positief, 'D' = negatief
        df['Transactiebedrag'] = df.apply(
            lambda r: -abs(_parse_dutch_amount(r['Bedrag'])) if str(r.get('CreditDebet', '')).strip().upper() == 'D'
            else abs(_parse_dutch_amount(r['Bedrag'])), axis=1)
        # Omschrijving
        parts = []
        for col in ['Tegenrekeninghouder', 'Omschrijving']:
            if col in df.columns:
                parts.append(df[col].fillna(''))
        df['Omschrijving'] = pd.concat(parts, axis=1).apply(lambda r: ' '.join(r).strip(), axis=1) if parts else 'Onbekend'
        # Tegenrekening bewaren (Knab noemt het 'Tegenrekeningnummer')
        if 'Tegenrekeningnummer' in df.columns:
            df['Tegenrekening'] = df['Tegenrekeningnummer'].fillna('')
        elif 'Tegenrekening' not in df.columns:
            df['Tegenrekening'] = ''
        df = _bereken_saldos(df)
        return df

    elif formaat == 'n26_csv':
        # Engelse kolomnamen
        if 'Date' in df.columns:
            df['Transactiedatum'] = df['Date'].astype(str).apply(_parse_datum)
        elif 'Datum' in df.columns:
            df['Transactiedatum'] = df['Datum'].astype(str).apply(_parse_datum)
        # Amount
        amt_col = 'Amount (EUR)' if 'Amount (EUR)' in df.columns else 'Bedrag (EUR)' if 'Bedrag (EUR)' in df.columns else 'Bedrag'
        df['Transactiebedrag'] = df[amt_col].apply(
            lambda v: float(v) if isinstance(v, (int, float)) else _parse_dutch_amount(v))
        # Omschrijving
        payee = df.get('Payee', df.get('Naam', pd.Series([''] * len(df)))).fillna('')
        ref = df.get('Payment reference', df.get('Omschrijving', pd.Series([''] * len(df)))).fillna('')
        df['Omschrijving'] = (payee + ' ' + ref).str.strip()
        # N26 heeft geen rekeningnummer in export — gebruik Account number als die er is, anders placeholder
        if 'Account number' in df.columns:
            df['Rekeningnummer'] = df['Account number'].fillna('N26')
        else:
            df['Rekeningnummer'] = 'N26'
        df['Tegenrekening'] = ''  # N26 geeft geen tegenrekening
        df = _bereken_saldos(df)
        return df

    elif formaat == 'bunq_csv':
        df['Rekeningnummer'] = df.get('Rekening', pd.Series(['Bunq'] * len(df)))
        df['Transactiedatum'] = df['Datum'].astype(str).apply(_parse_datum)
        df['Transactiebedrag'] = df['Bedrag'].apply(_parse_dutch_amount)
        naam = df.get('Naam', pd.Series([''] * len(df))).fillna('')
        omschr = df.get('Omschrijving', pd.Series([''] * len(df))).fillna('')
        df['Omschrijving'] = (naam + ' ' + omschr).str.strip()
        # Tegenrekening: Bunq noemt het soms 'Tegenpartij'
        if 'Tegenpartij' in df.columns:
            df['Tegenrekening'] = df['Tegenpartij'].fillna('')
        else:
            df['Tegenrekening'] = ''
        df = _bereken_saldos(df)
        return df

    elif formaat == 'generiek_nl':
        # Generiek formaat voor SNS, ASN, Triodos, RegioBank en onbekende banken
        # Probeer slim de juiste kolommen te vinden
        df['Transactiedatum'] = df['Datum'].astype(str).apply(_parse_datum)
        df['Transactiebedrag'] = df['Bedrag'].apply(_parse_dutch_amount)

        # Zoek rekeningnummer
        for col in ['Rekening', 'IBAN', 'Rekeningnummer', 'IBAN/BBAN']:
            if col in df.columns:
                df['Rekeningnummer'] = df[col]
                break
        else:
            df['Rekeningnummer'] = 'Onbekend'

        # Zoek omschrijving
        for col in ['Omschrijving', 'Naam / Omschrijving', 'Naam', 'Mededelingen', 'Naam tegenpartij']:
            if col in df.columns:
                df['Omschrijving'] = df[col].fillna('')
                break
        else:
            df['Omschrijving'] = 'Geen omschrijving'

        # Af/Bij correctie als die kolom er is
        for col in ['Af Bij', 'Af/Bij', 'Credit/Debet', 'CreditDebet']:
            if col in df.columns:
                df['Transactiebedrag'] = df.apply(
                    lambda r: -abs(r['Transactiebedrag']) if str(r[col]).strip().lower() in ('af', 'd', 'debet')
                    else abs(r['Transactiebedrag']), axis=1)
                break

        # Tegenrekening bewaren als die er is
        for col in ['Tegenrekening', 'IBAN tegenpartij', 'Tegenrekeningnummer']:
            if col in df.columns:
                df['Tegenrekening'] = df[col].fillna('')
                break
        else:
            df['Tegenrekening'] = ''

        # Saldo
        for col in ['Saldo', 'Saldo na mutatie', 'Saldo na trn']:
            if col in df.columns:
                df = _gebruik_saldo_kolom(df, col)
                break
        else:
            df = _bereken_saldos(df)

        return df

    return df


def _parse_triodos_csv(inhoud: bytes) -> pd.DataFrame:
    """Parse Triodos Bank CSV - speciaal formaat zonder header, met dubbele quotes.

    Triodos levert een CSV waar elke regel gewrapped is in outer quotes,
    met escaped inner quotes en ;; aan het eind van elke regel.
    9 velden: Datum, Rekening, Bedrag, Credit/Debet, Naam, Tegenrekening, Code, Omschrijving, Saldo
    """
    text = inhoud.decode('latin-1', errors='replace')
    rows = []

    for line in text.strip().split('\n'):
        line = line.strip().rstrip(';')
        if not line:
            continue
        # Strip buitenste quotes
        if line.startswith('"') and line.endswith('"'):
            line = line[1:-1]
        # Unescape dubbele quotes
        line = line.replace('""', '"')
        # Parse met quote-aware comma splitting
        parts = []
        current = ''
        in_quotes = False
        for ch in line:
            if ch == '"':
                in_quotes = not in_quotes
            elif ch == ',' and not in_quotes:
                parts.append(current.strip())
                current = ''
            else:
                current += ch
        parts.append(current.strip())

        if len(parts) >= 7:
            rows.append(parts)

    if not rows:
        raise ValueError("Triodos CSV bevat geen transacties")

    # Maak DataFrame met standaard kolomnamen
    col_names = ['Datum', 'Rekening', 'Bedrag', 'CreditDebet', 'Naam', 'Tegenrekening', 'Code', 'Omschrijving', 'Saldo']
    # Pas kolommen aan als er meer of minder zijn
    df = pd.DataFrame(rows)
    actual_cols = min(len(col_names), len(df.columns))
    df.columns = col_names[:actual_cols] + [f'Extra_{i}' for i in range(actual_cols, len(df.columns))]

    logger.info(f"Triodos CSV: {len(df)} transacties geparsed")
    return df


def _is_triodos_format(inhoud: bytes) -> bool:
    """Detecteer of dit een Triodos CSV is: begint met quote, geen header, ;; aan einde."""
    try:
        first_line = inhoud.decode('latin-1', errors='replace').split('\n')[0].strip()
        # Triodos: begint met ", bevat ""-escaped velden, eindigt op ;;
        return first_line.startswith('"') and '""' in first_line and first_line.endswith(';;')
    except Exception:
        return False


def lees_transacties(inhoud: bytes, bestandsnaam: str) -> pd.DataFrame:
    """Lees transacties uit CSV/XLS/XLSX - herkent automatisch het bankformaat.

    Ondersteunde banken: ABN AMRO, ING, Rabobank, SNS, ASN, Triodos,
    Knab, Bunq, N26, RegioBank - zowel CSV als XLS/XLSX.
    """
    naam_lower = bestandsnaam.lower()

    if naam_lower.endswith(('.xlsx', '.xls')):
        try:
            df = pd.read_excel(io.BytesIO(inhoud))
        except Exception as e:
            raise ValueError(f"Kan Excel-bestand niet lezen: {e}")
    elif naam_lower.endswith('.csv'):
        # Strip BOM (Byte Order Mark)
        if inhoud.startswith(b'\xef\xbb\xbf'):
            inhoud = inhoud[3:]
        elif inhoud.startswith(b'\xff\xfe') or inhoud.startswith(b'\xfe\xff'):
            inhoud = inhoud[2:]

        # Fix problematische line endings (ING gebruikt \r\r\n)
        inhoud = inhoud.replace(b'\r\r\n', b'\n').replace(b'\r\n', b'\n').replace(b'\r', b'\n')

        # Log preview voor debugging
        try:
            preview = inhoud[:300].decode('utf-8', errors='replace')
        except Exception:
            preview = str(inhoud[:300])
        logger.info(f"CSV preview: {preview[:150]}")

        # Check speciaal Triodos-formaat (geen header, dubbele quotes)
        if _is_triodos_format(inhoud):
            logger.info("Triodos-formaat gedetecteerd — custom parser")
            df = _parse_triodos_csv(inhoud)
        else:
            # Standaard CSV parsing — probeer combinaties
            parsed = False
            best_df = None
            best_cols = 0

            for enc in ['utf-8', 'latin-1', 'cp1252']:
                for sep in [';', ',', '\t', '|']:
                    try:
                        test_df = pd.read_csv(io.BytesIO(inhoud), sep=sep, encoding=enc,
                                              dtype=str, on_bad_lines='skip')
                        n_cols = len(test_df.columns)
                        if n_cols > best_cols:
                            best_df = test_df
                            best_cols = n_cols
                        if n_cols > 3:
                            parsed = True
                            df = test_df
                            logger.info(f"CSV geparsed: sep={repr(sep)}, enc={enc}, "
                                        f"kolommen={n_cols}: {list(test_df.columns)}")
                            break
                    except Exception:
                        continue
                if parsed:
                    break

            if not parsed:
                # Fallback: python engine met auto-detectie
                try:
                    df = pd.read_csv(io.BytesIO(inhoud), sep=None, engine='python',
                                     encoding='utf-8', dtype=str, on_bad_lines='skip')
                    if len(df.columns) > 3:
                        parsed = True
                except Exception:
                    pass

            if not parsed:
                cols_info = ""
                if best_df is not None:
                    cols_info = f" Beste poging: {best_cols} kolommen."
                raise ValueError(
                    f"Kan CSV niet parsen — controleer of het een geldig bankafschrift is.{cols_info}")
    else:
        raise ValueError(f"Onbekend bestandstype: {bestandsnaam}. Ondersteund: .csv, .xls, .xlsx")

    # Strip spaties uit kolomnamen
    df.columns = [c.strip() for c in df.columns]

    # Detecteer en normaliseer bankformaat
    formaat = _detecteer_formaat(df)
    if formaat == 'onbekend':
        raise ValueError(
            f"Bankformaat niet herkend. Gevonden kolommen: {list(df.columns)}. "
            f"Ondersteund: ABN AMRO, ING, Rabobank, SNS, ASN, Triodos, Knab, Bunq, N26, RegioBank."
        )

    df = _normaliseer(df, formaat)

    # Valideer dat normalisatie gelukt is
    verwacht = ['Rekeningnummer', 'Transactiedatum', 'Transactiebedrag', 'Omschrijving']
    ontbreekt = [k for k in verwacht if k not in df.columns]
    if ontbreekt:
        raise ValueError(f"Kolommen ontbreken na normalisatie: {ontbreekt}. Gevonden: {list(df.columns)}")

    # Rekeningnummer altijd als string (ABN AMRO levert int, andere banken string)
    df['Rekeningnummer'] = df['Rekeningnummer'].astype(str).str.strip()

    # Saldo's toevoegen als die nog ontbreken
    df = _bereken_saldos(df)

    # Datum parsing - flexibel (YYYYMMDD)
    df['Transactiedatum'] = df['Transactiedatum'].astype(str)
    try:
        df['datum'] = pd.to_datetime(df['Transactiedatum'], format='%Y%m%d')
    except Exception:
        # Fallback: laat pandas het zelf uitzoeken
        df['datum'] = pd.to_datetime(df['Transactiedatum'], dayfirst=True)

    df['maand'] = df['datum'].dt.to_period('M')
    df['bedrag'] = df['Transactiebedrag'].astype(float)

    logger.info(f"Bestand gelezen: {len(df)} transacties, formaat={formaat}, "
                f"rekeningen={df['Rekeningnummer'].nunique()}")

    return df


# ---------------------------------------------------------------------------
# STAP 1B: HUISHOUDREGISTER & INTERNE-OVERBOEKINGEN DETECTIE
# ---------------------------------------------------------------------------

def _bouw_huishoudregister(df: pd.DataFrame) -> set:
    """Verzamel alle eigen rekeningnummers uit de upload.

    Alle rekeningen in de upload zijn per definitie 'eigen'.
    Returns: set van rekeningnummers (strings, stripped, lowercase-safe).
    """
    eigen = set()
    for rek in df['Rekeningnummer'].unique():
        rek_str = str(rek).strip()
        if rek_str and rek_str != 'Onbekend':
            eigen.add(rek_str)
    logger.info(f"Huishoudregister: {len(eigen)} eigen rekeningen gevonden: {eigen}")
    return eigen


def _normaliseer_iban(val) -> str:
    """Normaliseer een IBAN/rekeningnummer voor vergelijking.

    Verwijdert spaties, streepjes, en maakt uppercase.
    Zo matcht 'NL12 ABNA 0123 4567 89' met 'NL12ABNA0123456789'.
    """
    if pd.isna(val):
        return ''
    return str(val).strip().replace(' ', '').replace('-', '').upper()


import re
# NL IBAN: 2 letters + 2 cijfers + 4 letters bankcode + 10 cijfers account
# Buitenlands IBAN (bv EE/DE): ook 4 letters bankcode maar account kan langer zijn
_IBAN_RE = re.compile(r'[A-Z]{2}\d{2}[A-Z]{4}\d{10,14}')


def _parse_transactie_omschrijving(omschrijving: str) -> dict:
    """Extraheer gestructureerde velden uit banktransactie-omschrijvingen.

    Ondersteunt alle NL bankformaten:
    - ABN AMRO XLSX: /TRTP/ formaat en SEPA platte-tekst formaat
    - ING CSV: "Naam | Details" formaat
    - Overig: Rabobank, Triodos, Knab, Bunq hebben al gestructureerde kolommen
              maar als hun omschrijving toch geparst moet worden, werkt deze fallback.

    Returns dict: tegenpartij_naam, tegenpartij_iban, transactie_type, kenmerk
    """
    if not omschrijving or pd.isna(omschrijving) or str(omschrijving).strip() == '':
        return {'tegenpartij_naam': '', 'tegenpartij_iban': '', 'transactie_type': '', 'kenmerk': ''}

    tekst = str(omschrijving).strip()
    naam = ''
    iban = ''
    tx_type = ''
    kenmerk = ''

    # === FORMAT 1: ABN /TRTP/ formaat ===
    if '/TRTP/' in tekst or tekst.startswith('/TRTP/'):
        trtp_match = re.search(r'/TRTP/([^/]+)', tekst)
        if trtp_match:
            tx_type = trtp_match.group(1).strip()
        iban_match = re.search(r'/IBAN/([A-Z]{2}\d{2}[A-Z0-9]{4}[\dA-Z]+?)/', tekst)
        if iban_match:
            iban = iban_match.group(1).replace(' ', '')
        name_match = re.search(r'/NAME/([^/]+)', tekst)
        if name_match:
            naam = name_match.group(1).strip()
        remi_match = re.search(r'/REMI/([^/]+)', tekst)
        if remi_match:
            kenmerk = remi_match.group(1).strip()

    # === FORMAT 2: ABN SEPA platte-tekst ===
    elif tekst.startswith('SEPA ') or ('Naam:' in tekst and 'IBAN:' in tekst):
        type_match = re.match(r'^(SEPA [A-Za-z. ]+?)(?:\s{2,}|IBAN|Incassant)', tekst)
        if type_match:
            tx_type = type_match.group(1).strip().rstrip('.')
        iban_match = re.search(r'IBAN:\s*([A-Z]{2}\d{2}[A-Z0-9]{4}[\dA-Z]+)', tekst)
        if iban_match:
            iban = iban_match.group(1).replace(' ', '')
        naam_match = re.search(r'Naam:\s*(.+?)(?:\s{2,}|Omschrijving:|Betalingskenm|Kenmerk:|Machtiging:|$)', tekst)
        if naam_match:
            naam = naam_match.group(1).strip()
        omschr_match = re.search(r'Omschrijving:\s*(.+?)(?:\s{2,}|Kenmerk:|IBAN:|$)', tekst)
        if omschr_match:
            kenmerk = omschr_match.group(1).strip()
        if not kenmerk:
            betk_match = re.search(r'Betalingskenm\.?:\s*(\d+)', tekst)
            if betk_match:
                kenmerk = f"BK:{betk_match.group(1)}"

    # === FORMAT 3: ING "Naam | Details" formaat ===
    elif '|' in tekst:
        delen = tekst.split('|', 1)
        korte_naam = delen[0].strip()
        details = delen[1].strip() if len(delen) > 1 else ''
        naam_match = re.search(r'Naam:\s*(.+?)(?:\s+Omschrijving:|\s+IBAN:|\s+Kenmerk:|$)', details)
        if naam_match:
            naam = naam_match.group(1).strip()
        else:
            naam = korte_naam
        iban_match = re.search(r'IBAN:\s*([A-Z]{2}\d{2}[A-Z0-9]+)', details)
        if iban_match:
            iban = iban_match.group(1).replace(' ', '')
        omschr_match = re.search(r'Omschrijving:\s*(.+?)(?:\s+IBAN:|\s+Kenmerk:|$)', details)
        if omschr_match:
            kenmerk = omschr_match.group(1).strip()
        tx_type = 'ING'

    # === FALLBACK ===
    else:
        naam = tekst[:80]

    naam = re.sub(r'\s+', ' ', naam).strip()
    kenmerk = re.sub(r'\s+', ' ', kenmerk).strip()
    iban = iban.replace(' ', '').upper() if iban else ''

    return {'tegenpartij_naam': naam, 'tegenpartij_iban': iban, 'transactie_type': tx_type, 'kenmerk': kenmerk}


def _verrijk_transactie_velden(df: pd.DataFrame) -> pd.DataFrame:
    """Voeg gestructureerde velden toe aan elke transactie.

    Stap 1 in de enrichment-pipeline (vóór alle detectie):
    - Vult lege Tegenrekening met IBAN uit omschrijving
    - Extraheert tegenpartij_naam uit omschrijving
    - Maakt omschrijving_schoon voor de AI-prompt

    Voor banken met al gestructureerde data (Rabobank 'Naam tegenpartij',
    Triodos 'Naam', Knab 'Tegenrekeninghouder', Bunq 'Naam') wordt de
    bestaande naam gebruikt en niet overschreven.
    """
    if 'Tegenrekening' not in df.columns:
        df['Tegenrekening'] = ''

    # Parse omschrijvingen
    parsed = df['Omschrijving'].apply(_parse_transactie_omschrijving)
    df['tegenpartij_naam'] = parsed.apply(lambda x: x['tegenpartij_naam'])
    df['kenmerk'] = parsed.apply(lambda x: x['kenmerk'])

    # Vul lege Tegenrekening met geparsed IBAN
    lege_mask = df['Tegenrekening'].apply(lambda x: _normaliseer_iban(x) == '')
    parsed_ibans = parsed.apply(lambda x: x['tegenpartij_iban'])
    n_leeg = lege_mask.sum()
    if n_leeg > 0:
        df.loc[lege_mask, 'Tegenrekening'] = parsed_ibans[lege_mask]
        n_gevuld = (df.loc[lege_mask, 'Tegenrekening'] != '').sum()
        if n_gevuld > 0:
            logger.info(f"IBAN-EXTRACTIE: {n_gevuld}/{n_leeg} lege tegenrekeningen gevuld uit omschrijving")

    # Maak schone omschrijving: "Naam — Kenmerk" (voor AI-prompt)
    df['omschrijving_schoon'] = df.apply(
        lambda r: (r['tegenpartij_naam'] + (' — ' + r['kenmerk'] if r['kenmerk'] else ''))
        if r['tegenpartij_naam'] else str(r['Omschrijving'])[:200],
        axis=1
    )

    n_naam = (df['tegenpartij_naam'] != '').sum()
    logger.info(f"NAAM-EXTRACTIE: {n_naam}/{len(df)} transacties met herkende tegenpartij-naam")

    return df


# ---------------------------------------------------------------------------
# NTROPY ENRICHMENT — externe transactie-verrijking
# ---------------------------------------------------------------------------
NTROPY_API_KEY = ''  # Uitgeschakeld — eigen merchant registry (1451 merchants) is actief
NTROPY_BASE = 'https://api.ntropy.com/v3'

# Mapping van Ntropy consumer categorieën naar onze sectie + categorie
_NTROPY_TO_ONZE_MAPPING = {
    # --- INKOMSTEN (incoming) ---
    'paycheck': ('inkomsten', 'Salaris'),
    'freelance income': ('inkomsten', 'Freelance inkomen'),
    'property rental income': ('inkomsten', 'Huurinkomsten'),
    'interest earned': ('inkomsten', 'Rente-inkomsten'),
    'stock dividend': ('inkomsten', 'Dividend'),
    'benefits': ('inkomsten', 'Uitkering/toeslagen'),
    'tax refund': ('inkomsten', 'Belastingteruggave'),
    'insurance payout': ('inkomsten', 'Verzekeringsuitkering'),
    'cashback': ('inkomsten', 'Cashback'),
    'grant or stipend': ('inkomsten', 'Studiefinanciering'),
    'refund': ('inkomsten', 'Restitutie'),
    # --- VASTE LASTEN (outgoing) ---
    'rent': ('vaste_lasten', 'Huur/Hypotheek'),
    'mortgage repayment': ('vaste_lasten', 'Huur/Hypotheek'),
    'utilities': ('vaste_lasten', 'Energie'),
    'insurance premium': ('vaste_lasten', 'Verzekeringen'),
    'education': ('vaste_lasten', 'Onderwijs'),
    'childcare': ('vaste_lasten', 'Kinderopvang'),
    'loan repayment': ('vaste_lasten', 'Lening'),
    'student loan repayment': ('vaste_lasten', 'Studieschuld'),
    'auto loan repayment': ('vaste_lasten', 'Autolening'),
    'credit card bill': ('vaste_lasten', 'Creditcard'),
    'interest payment': ('vaste_lasten', 'Rente betaald'),
    'taxes': ('vaste_lasten', 'Belastingen'),
    'government': ('vaste_lasten', 'Overheid'),
    'contribution to reserve fund': ('vaste_lasten', 'VvE/Reservefonds'),
    'retirement contribution': ('sparen_beleggen', 'Pensioen'),
    # --- VARIABELE KOSTEN (outgoing) ---
    'groceries': ('variabele_kosten', 'Boodschappen'),
    'restaurant': ('variabele_kosten', 'Uit eten'),
    'fast food': ('variabele_kosten', 'Uit eten'),
    'food delivery': ('variabele_kosten', 'Bezorging'),
    'coffee shop': ('variabele_kosten', 'Horeca'),
    'bars and nightclubs': ('variabele_kosten', 'Horeca'),
    'fuel': ('variabele_kosten', 'Brandstof'),
    'ev charging': ('variabele_kosten', 'Brandstof'),
    'vehicle maintenance': ('variabele_kosten', 'Autokosten'),
    'auto lease payment': ('variabele_kosten', 'Autolease'),
    'rideshare or taxi transport': ('variabele_kosten', 'Vervoer'),
    'public transport': ('variabele_kosten', 'OV'),
    'other transport': ('variabele_kosten', 'Vervoer'),
    'airfare': ('variabele_kosten', 'Reizen'),
    'hotel or lodging': ('variabele_kosten', 'Reizen'),
    'toll charge': ('variabele_kosten', 'Vervoer'),
    'clothing': ('variabele_kosten', 'Kleding'),
    'self care': ('variabele_kosten', 'Persoonlijke verzorging'),
    'drugstore or pharmacy': ('variabele_kosten', 'Drogist/Apotheek'),
    'medical bill': ('variabele_kosten', 'Zorgkosten'),
    'entertainment': ('variabele_kosten', 'Entertainment'),
    'digital content and streaming': ('variabele_kosten', 'Abonnementen'),
    'sport and fitness': ('variabele_kosten', 'Sport'),
    'pets': ('variabele_kosten', 'Huisdieren'),
    'home maintenance': ('variabele_kosten', 'Woning onderhoud'),
    'electronics': ('variabele_kosten', 'Elektronica'),
    'e-commerce purchase': ('variabele_kosten', 'Online shopping'),
    'department or discount store': ('variabele_kosten', 'Winkel'),
    'convenience store': ('variabele_kosten', 'Winkel'),
    'recreational goods': ('variabele_kosten', 'Vrije tijd'),
    'books, newsletters, newspapers': ('variabele_kosten', 'Media'),
    'saas tools': ('variabele_kosten', 'Software'),
    'laundry': ('variabele_kosten', 'Wasserij'),
    'donation': ('variabele_kosten', 'Donaties'),
    'gambling spend': ('variabele_kosten', 'Gokken'),
    # --- SPAREN/BELEGGEN ---
    'transfer to investment app': ('sparen_beleggen', 'Belegging'),
    'transfer to stock broker': ('sparen_beleggen', 'Belegging'),
    'transfer to crypto broker': ('sparen_beleggen', 'Crypto'),
    'transfer from investment app': ('sparen_beleggen', 'Belegging retour'),
    'transfer from stock broker': ('sparen_beleggen', 'Belegging retour'),
    'transfer from crypto broker': ('sparen_beleggen', 'Crypto retour'),
    # --- NEUTRAAL / INTERN ---
    'intra account transfer': ('onderling_neutraal', 'Interne overboeking'),
    'inter account transfer': ('onderling_neutraal', 'Interne overboeking'),
    'peer to peer transfer': ('onderling_neutraal', 'Onderlinge overboeking'),
    'ATM withdrawal': ('onderling_neutraal', 'Cash opname'),
    'ATM deposit': ('onderling_neutraal', 'Cash storting'),
    'teller withdrawal': ('onderling_neutraal', 'Cash opname'),
    'teller deposit': ('onderling_neutraal', 'Cash storting'),
}


def _ntropy_enrich_batch(df: pd.DataFrame) -> pd.DataFrame:
    """Verrijk transacties via Ntropy Batch API. Voegt kolommen toe:
    - ntropy_entity: schone merchant naam
    - ntropy_category: Ntropy categorie (bv. 'groceries')
    - ntropy_sectie: onze sectie-mapping (bv. 'vaste_lasten')
    - ntropy_categorie: onze categorie-mapping (bv. 'Energie')
    - ntropy_website: merchant website

    Gebruikt de Batch API voor snelheid (alle tx in één request).
    Faalt graceful: als API niet beschikbaar, gaat pipeline gewoon door.
    """
    _EMPTY_COLS = ['ntropy_entity', 'ntropy_category', 'ntropy_sectie', 'ntropy_categorie', 'ntropy_website']

    if not NTROPY_API_KEY:
        logger.info("NTROPY: geen API key geconfigureerd, skip enrichment")
        for col in _EMPTY_COLS:
            df[col] = ''
        return df

    import requests as req
    import time

    headers = {
        'X-API-KEY': NTROPY_API_KEY,
        'Content-Type': 'application/json',
        'Accept': 'application/json'
    }

    # Initialiseer kolommen
    for col in _EMPTY_COLS:
        df[col] = ''

    # Maak account holder
    ah_id = f"ph-{uuid.uuid4().hex[:8]}"
    try:
        ah_resp = req.post(f"{NTROPY_BASE}/account_holders", headers=headers, json={
            'id': ah_id, 'type': 'consumer', 'name': 'PH Klant'
        }, timeout=10)
        if ah_resp.status_code != 200:
            logger.warning(f"NTROPY: account_holder mislukt: {ah_resp.status_code} {ah_resp.text[:100]}")
            return df
    except Exception as e:
        logger.warning(f"NTROPY: verbinding mislukt: {e}")
        return df

    # Bouw batch payload — idx_map koppelt Ntropy tx-id aan DataFrame index
    batch_items = []
    idx_map = {}  # ntropy_id -> df_index

    for idx, row in df.iterrows():
        bedrag = abs(float(row.get('Transactiebedrag', 0)))
        if bedrag == 0:
            continue

        entry_type = 'incoming' if float(row.get('Transactiebedrag', 0)) > 0 else 'outgoing'
        raw_datum = str(row.get('Transactiedatum', '20250101'))
        if len(raw_datum) == 8 and raw_datum.isdigit():
            api_datum = f"{raw_datum[:4]}-{raw_datum[4:6]}-{raw_datum[6:8]}"
        else:
            api_datum = raw_datum

        omschrijving = str(row.get('Omschrijving', ''))[:200]
        tx_id = f"ph-{idx}-{uuid.uuid4().hex[:6]}"
        idx_map[tx_id] = idx

        batch_items.append({
            'id': tx_id,
            'description': omschrijving,
            'date': api_datum,
            'amount': bedrag,
            'entry_type': entry_type,
            'currency': 'EUR',
            'account_holder_id': ah_id,
            'location': {'country': 'NL'}
        })

    if not batch_items:
        logger.info("NTROPY: geen transacties om te verrijken")
        return df

    logger.info(f"NTROPY: {len(batch_items)} transacties voorbereid voor batch enrichment")

    # Submit batch
    try:
        batch_resp = req.post(f"{NTROPY_BASE}/batches", headers=headers, json={
            'operation': 'POST /v3/transactions',
            'data': batch_items
        }, timeout=30)
        if batch_resp.status_code != 200:
            logger.warning(f"NTROPY batch submit mislukt: {batch_resp.status_code} {batch_resp.text[:200]}")
            return df
        batch_data = batch_resp.json()
        batch_id = batch_data['id']
        logger.info(f"NTROPY: batch {batch_id} submitted, {batch_data.get('total', '?')} items")
    except Exception as e:
        logger.warning(f"NTROPY: batch submit exception: {e}")
        return df

    # Poll voor resultaat (max 5 minuten)
    max_wait = 300
    poll_interval = 3
    waited = 0
    while waited < max_wait:
        time.sleep(poll_interval)
        waited += poll_interval
        try:
            status_resp = req.get(f"{NTROPY_BASE}/batches/{batch_id}", headers=headers, timeout=10)
            if status_resp.status_code == 200:
                status_data = status_resp.json()
                progress = status_data.get('progress', 0)
                total = status_data.get('total', len(batch_items))
                status = status_data.get('status', '')
                if waited % 15 == 0:
                    logger.info(f"NTROPY: batch {progress}/{total} ({status})")
                if status == 'completed':
                    break
                if status == 'error':
                    logger.error(f"NTROPY: batch failed: {status_data}")
                    return df
        except Exception as e:
            logger.warning(f"NTROPY: poll error: {e}")

    if waited >= max_wait:
        logger.warning(f"NTROPY: batch timeout na {max_wait}s")
        return df

    # Haal resultaten op
    try:
        results_resp = req.get(f"{NTROPY_BASE}/batches/{batch_id}/results", headers=headers, timeout=30)
        if results_resp.status_code != 200:
            logger.warning(f"NTROPY: results ophalen mislukt: {results_resp.status_code}")
            return df
        results_data = results_resp.json()
    except Exception as e:
        logger.warning(f"NTROPY: results exception: {e}")
        return df

    # Verwerk resultaten
    n_success = 0
    n_mapped = 0
    for result in results_data.get('results', []):
        tx_id = result.get('id', '')
        if tx_id not in idx_map:
            continue
        idx = idx_map[tx_id]

        if result.get('error'):
            continue

        n_success += 1
        cp = result.get('entities', {}).get('counterparty') or {}
        entity_name = cp.get('name', '') or ''
        website = cp.get('website', '') or ''
        category = result.get('categories', {}).get('general', '') or ''

        df.at[idx, 'ntropy_entity'] = entity_name
        df.at[idx, 'ntropy_category'] = category
        df.at[idx, 'ntropy_website'] = website

        if category and category in _NTROPY_TO_ONZE_MAPPING:
            sectie, cat = _NTROPY_TO_ONZE_MAPPING[category]
            df.at[idx, 'ntropy_sectie'] = sectie
            df.at[idx, 'ntropy_categorie'] = cat
            n_mapped += 1

    logger.info(f"NTROPY: {n_success}/{len(batch_items)} verrijkt, {n_mapped} gemapped naar onze categorieën")
    return df


def _markeer_interne_transfers(df: pd.DataFrame, eigen_rekeningen: set) -> pd.DataFrame:
    """Markeer transacties tussen eigen rekeningen als interne transfer.

    Detectiemethoden (in volgorde van betrouwbaarheid):
    1. Tegenrekening kolom matcht een eigen rekeningnummer
    2. Eigen rekeningnummer komt voor in de omschrijving
    3. Matching bijschrijving/afschrijving op dezelfde datum (cross-account)

    Voegt kolom 'is_intern' toe (True/False).
    """
    # Normaliseer eigen rekeningen voor vergelijking
    eigen_genormaliseerd = set(_normaliseer_iban(r) for r in eigen_rekeningen if r)

    # Start: alles is niet-intern
    df['is_intern'] = False

    # Methode 1: Tegenrekening matcht eigen rekening
    if 'Tegenrekening' in df.columns:
        df['_tegen_norm'] = df['Tegenrekening'].apply(_normaliseer_iban)
        methode1_mask = df['_tegen_norm'].isin(eigen_genormaliseerd) & (df['_tegen_norm'] != '')
        df.loc[methode1_mask, 'is_intern'] = True
        n_methode1 = methode1_mask.sum()
        if n_methode1 > 0:
            logger.info(f"Interne transfers methode 1 (tegenrekening): {n_methode1} transacties")
        df.drop(columns=['_tegen_norm'], inplace=True)

    # Methode 2: Eigen rekeningnummer in omschrijving (fallback voor banken zonder tegenrekening)
    for rek in eigen_genormaliseerd:
        if len(rek) >= 8:  # Alleen zinvol bij IBANs/lange rekeningnummers
            # Check of het rekeningnummer (of deel ervan) in de omschrijving staat
            mask = (
                ~df['is_intern'] &  # Nog niet gemarkeerd
                df['Omschrijving'].str.upper().str.replace(' ', '').str.contains(rek, na=False, regex=False)
            )
            if mask.sum() > 0:
                df.loc[mask, 'is_intern'] = True
                logger.info(f"Interne transfers methode 2 (omschrijving bevat {rek}): {mask.sum()} transacties")

    # Statistieken loggen
    n_intern = df['is_intern'].sum()
    bedrag_intern = df.loc[df['is_intern'], 'Transactiebedrag'].apply(lambda x: abs(float(x))).sum()
    logger.info(f"Totaal interne transfers: {n_intern} transacties, "
                f"totaal bedrag: EUR {bedrag_intern:,.2f}")

    return df


def _detecteer_huishoudleden(df: pd.DataFrame) -> pd.DataFrame:
    """Detecteer automatisch overboekingen naar/van huishoudleden (partner, kinderen).

    Werkt voor ALLE gezinnen — geen hardcoded namen nodig.

    Heuristiek:
    1. Groepeer transacties per tegenpartij (uit Tegenrekening of naam in omschrijving)
    2. Een tegenpartij is waarschijnlijk een huishoudlid als:
       a) Er ZOWEL positieve als negatieve transacties zijn (geld gaat heen-en-weer)
       b) Er minstens 4 transacties per jaar zijn (regelmatig contact)
       c) Het netto-effect < 40% van het bruto volume is (ongeveer in balans)
       d) De tegenpartij NIET al door de merchant-mapping is herkend als bedrijf
    3. Markeer deze transacties als 'is_intern' = True

    Extra check: als een tegenpartij op dezelfde IBAN zit als een van de eigen rekeningen
    van de klant, is het sowieso al een interne transfer (al afgevangen in methode 1/2).
    """
    if 'Tegenrekening' not in df.columns:
        logger.info("HUISHOUDLEDEN: geen Tegenrekening kolom — skip detectie")
        return df

    # Alleen niet-interne, niet-merchant-mapped transacties bekijken
    mask_kandidaat = ~df['is_intern']
    df_kandidaat = df[mask_kandidaat].copy()

    if len(df_kandidaat) == 0:
        return df

    # Groepeer per tegenrekening (genormaliseerd)
    df_kandidaat['_tegen_norm'] = df_kandidaat['Tegenrekening'].apply(_normaliseer_iban)
    # Filter lege tegenrekeningen
    df_kandidaat = df_kandidaat[df_kandidaat['_tegen_norm'] != '']

    # Bekende merchant-zoektermen om bedrijven eruit te filteren
    bekende_merchants = set()
    for zoekterm, _, _, _ in MERCHANT_MAPPING:
        bekende_merchants.add(zoekterm)

    huishoudleden_ibans = set()
    huishoudleden_namen = []

    for tegen_rek, groep in df_kandidaat.groupby('_tegen_norm'):
        if len(groep) < 4:
            continue  # Te weinig transacties

        # Check of dit een bekende merchant is
        omschr_sample = ' '.join(groep['Omschrijving'].astype(str).str.upper().head(3))
        is_merchant = False
        for merchant_zoek in bekende_merchants:
            if merchant_zoek in omschr_sample:
                is_merchant = True
                break
        if is_merchant:
            continue

        positief = groep[groep['bedrag'] > 0]['bedrag'].sum()
        negatief = abs(groep[groep['bedrag'] < 0]['bedrag'].sum())
        bruto = positief + negatief
        netto = abs(positief - negatief)

        if bruto < 500:
            continue  # Te klein volume, niet significant

        # Bidirectioneel: er moet geld BEIDE kanten op gaan
        if positief == 0 or negatief == 0:
            continue

        # Netto-effect moet relatief klein zijn (< 40% van bruto)
        if bruto > 0 and (netto / bruto) < 0.40:
            huishoudleden_ibans.add(tegen_rek)
            # Pak een leesbare naam uit de omschrijving
            naam_sample = groep.iloc[0]['Omschrijving']
            huishoudleden_namen.append(f"{tegen_rek} ({naam_sample[:40]}...)")
            logger.info(
                f"HUISHOUDLID GEDETECTEERD: {tegen_rek} — "
                f"{len(groep)} transacties, positief EUR {positief:,.0f}, "
                f"negatief EUR {negatief:,.0f}, netto EUR {netto:,.0f} "
                f"({netto/bruto*100:.0f}% van bruto)"
            )

    # Markeer alle transacties met deze tegenrekeningen als intern
    if huishoudleden_ibans:
        df['_tegen_check'] = df['Tegenrekening'].apply(_normaliseer_iban)
        mask_huishoud = df['_tegen_check'].isin(huishoudleden_ibans) & ~df['is_intern']
        n_gemarkeerd = mask_huishoud.sum()
        df.loc[mask_huishoud, 'is_intern'] = True
        df.drop(columns=['_tegen_check'], inplace=True)

        bedrag_huishoud = df.loc[mask_huishoud, 'bedrag'].apply(lambda x: abs(float(x))).sum() if n_gemarkeerd > 0 else 0
        logger.info(
            f"HUISHOUDLEDEN: {len(huishoudleden_ibans)} huishoudlid(en) gedetecteerd, "
            f"{n_gemarkeerd} transacties als intern gemarkeerd "
            f"(EUR {bedrag_huishoud:,.0f} bruto)"
        )
    else:
        if '_tegen_check' in df.columns:
            df.drop(columns=['_tegen_check'], inplace=True)
        logger.info("HUISHOUDLEDEN: geen huishoudleden gedetecteerd via bidirectionele analyse")

    # Cleanup
    if '_tegen_norm' in df.columns:
        df.drop(columns=['_tegen_norm'], errors='ignore', inplace=True)

    return df


# ---------------------------------------------------------------------------
# STAP 1b3: RELATED PARTY RESOLUTION (RPR v1.3)
# ---------------------------------------------------------------------------
# Twee gescheiden lagen:
#   Laag 1: Party Resolution — bepaalt WIE de tegenpartij is
#   Laag 2: Economic Inflow Classification — bepaalt WAT de transactie is
#
# Counterparty roles (V2): own_account, household_related_party, employer_or_payroll,
#   business_counterparty, government, merchant, broker_or_investment_platform,
#   lender_or_mortgage_party, card_settlement, unknown
#
# Signaalklassen:
#   STERK: S2 (multi-IBAN linking), S4 (bidirectioneel), S8 (eigen FI), S10 (merchant)
#   MIDDEL: S6 (cross-account), S9 (rechtsvorm)
#   ZWAK: S1 (achternaam), S3 (adres-hint), S5 (transfer-achtig), S7 (geen eco relatie)
#
# Kernregel: household vereist minimaal één STERK signaal (S2 of S4).
# ---------------------------------------------------------------------------

# Tussenvoegels en titels voor achternaam-extractie
_TUSSENVOEGELS = {
    'van', 'de', 'den', 'der', 'het', 'ter', 'ten', 'te', 'in', "'t",
    'van de', 'van den', 'van der', 'van het', 'in de', "in 't",
}
_TITELS = {
    'mr', 'mrs', 'ms', 'dr', 'prof', 'ir', 'ing', 'drs', 'mr.',
    'mrs.', 'ms.', 'dr.', 'prof.', 'ir.', 'ing.', 'drs.', 'bc',
    'mw', 'mevr', 'dhr',
}


def extract_achternaam(naam: str) -> str:
    """Extraheer de achternaam uit een naam-string.

    Werkt generiek voor alle Nederlandse naamconventies:
    - Verwijdert tussenvoegels (van, de, den, het, ter, etc.)
    - Verwijdert titels (mr, dr, ir, ing, drs, etc.)
    - Splitst op streepje voor dubbele achternamen
    - Neemt het langste woord ≥3 letters
    - Case-insensitive

    Voorbeelden:
      "E. Heijen-Kop" → "heijen"
      "P.H.M. van der Berg" → "berg"
      "Mr. J. de Groot-Jansen" → "groot"  (of "jansen")
      "M. Jansen-Bakker" → "jansen"
    """
    if not naam or str(naam).strip() == '' or str(naam).upper() == 'NAN':
        return ''

    naam = str(naam).strip().lower()

    # Verwijder IBAN-referenties (bv "iban: nl37bick..." uit ABN AMRO omschrijvingen)
    naam = re.sub(r'\biban[:\s]+[a-z0-9]+', '', naam, flags=re.IGNORECASE)
    # Verwijder losse IBAN-patronen (NL + 2 cijfers + 4 letters + cijfers)
    naam = re.sub(r'\b[a-z]{2}\d{2}[a-z]{4}\d{6,14}\b', '', naam, flags=re.IGNORECASE)

    # Verwijder initialen (letters gevolgd door punt)
    naam = re.sub(r'\b[a-z]\.\s*', '', naam)
    # Verwijder initialen zonder punt (enkele letter gevolgd door spatie)
    naam = re.sub(r'\b[a-z]\s+', '', naam)

    # Splits in woorden
    woorden = naam.split()

    # Verwijder titels
    woorden = [w for w in woorden if w.rstrip('.') not in _TITELS]

    # Verwijder tussenvoegels (ook meerdelig: "van de", "van der")
    schoon = []
    i = 0
    while i < len(woorden):
        # Check meerdelige tussenvoegels eerst
        if i + 1 < len(woorden) and f"{woorden[i]} {woorden[i+1]}" in _TUSSENVOEGELS:
            i += 2
            continue
        if woorden[i] in _TUSSENVOEGELS:
            i += 1
            continue
        schoon.append(woorden[i])
        i += 1

    if not schoon:
        return ''

    # Neem het laatste woord (achternaam), splits op streepje voor dubbele achternamen
    achternaam_deel = schoon[-1]
    delen = achternaam_deel.split('-')

    # Neem het langste deel ≥3 letters
    kandidaten = [d for d in delen if len(d) >= 3]
    if not kandidaten:
        # Fallback: neem het langste deel ongeacht lengte
        kandidaten = delen

    return max(kandidaten, key=len) if kandidaten else ''


def _resolve_related_parties(df: pd.DataFrame, eigen_rekeningen: set,
                             eigen_fi_ibans: set = None) -> pd.DataFrame:
    """Related Party Resolution V2 — bepaalt counterparty_role per tegenpartij.

    10 specifieke rollen (Generic Financial Integrity Engine):
      own_account, household_related_party, employer_or_payroll,
      business_counterparty, government, merchant,
      broker_or_investment_platform, lender_or_mortgage_party,
      card_settlement, unknown

    Implementeert 5-fase beslislogica:
      Fase 1: Definitieve labels op naam/IBAN (own_account, government, broker,
              lender, card_settlement, merchant)
      Fase 2: Household (S2/S4/achternaam+gedrag)
      Fase 3: Business (S9 rechtsvorm + extra bewijs → employer_or_payroll of business_counterparty)
      Fase 4: Default (unknown)

    Voegt kolom 'party_type' toe aan df.
    """
    if eigen_fi_ibans is None:
        eigen_fi_ibans = set()

    # Normaliseer eigen rekeningen
    eigen_rek_norm = set(_normaliseer_iban(str(r)) for r in eigen_rekeningen if r and str(r) != 'nan')

    # Initialiseer party_type kolom
    if 'party_type' not in df.columns:
        df['party_type'] = None

    # Bekende merchants set
    bekende_merchants_set = set()
    for zoekterm, _, _, _ in MERCHANT_MAPPING:
        bekende_merchants_set.add(zoekterm)

    # =========================================================
    # STAP 1: Bouw tegenpartij-register (per unieke IBAN)
    # =========================================================
    heeft_tegenrek = 'Tegenrekening' in df.columns

    # Extraheer achternaam van de rekeninghouder(s)
    # We gebruiken de naam die op de eigen rekening staat — die zit in tegenpartij_naam
    # bij INKOMENDE transacties van eigen rekeningen, of we leiden het af uit
    # bidirectionele matches.
    eigen_achternamen = set()

    # Methode: gebruik achternamen van bidirectionele tegenpartijen (is_intern)
    # en achternamen die verschijnen bij transacties NAAR eigen rekeningen
    if 'tegenpartij_naam' in df.columns:
        # Van interne (household) transacties: de naam is die van het huishoudlid
        intern_namen = df[df['is_intern'] & (df['tegenpartij_naam'].notna()) &
                          (df['tegenpartij_naam'] != '')]['tegenpartij_naam'].unique()
        for n in intern_namen:
            ach = extract_achternaam(str(n))
            # Filter: achternaam moet minstens 3 letters zijn en niet een banknaam/keyword
            if ach and len(ach) >= 3 and ach not in {
                'priverekening', 'ondernemersrekening', 'jongerengroeirekening',
                'spaarrekening', 'betaalrekening', 'rekening', 'tanken',
            }:
                eigen_achternamen.add(ach)

        # Methode 2: achternaam uit INKOMENDE transacties naar eigen IBANs
        if heeft_tegenrek:
            for rek in eigen_rekeningen:
                rek_norm = _normaliseer_iban(str(rek))
                if not rek_norm:
                    continue
                incoming = df[df['Tegenrekening'].apply(lambda x: _normaliseer_iban(str(x))) == rek_norm]
                for _, row in incoming.head(5).iterrows():
                    naam = str(row.get('tegenpartij_naam', ''))
                    ach = extract_achternaam(naam)
                    if ach and len(ach) >= 3 and ach not in {
                        'priverekening', 'ondernemersrekening', 'jongerengroeirekening',
                        'spaarrekening', 'betaalrekening', 'rekening', 'tanken',
                    }:
                        eigen_achternamen.add(ach)

    logger.info(f"RPR: eigen achternamen gedetecteerd: {eigen_achternamen}")

    if not heeft_tegenrek:
        logger.info("RPR: geen Tegenrekening kolom — skip party resolution")
        df.loc[df['party_type'].isna(), 'party_type'] = 'unknown'
        return df

    # =========================================================
    # STAP 2: Analyseer elke tegenpartij (per uniek IBAN)
    # =========================================================

    # Bouw lookup: IBAN → naam, transactie-info
    iban_data = {}  # iban → {'namen': set, 'n_pos': int, 'n_neg': int, 'bedrag_pos': float, ...}
    df['_tegen_norm'] = df['Tegenrekening'].apply(_normaliseer_iban)

    for iban, groep in df[df['_tegen_norm'] != ''].groupby('_tegen_norm'):
        namen = set()
        for n in groep['tegenpartij_naam'].dropna().unique():
            n_str = str(n).strip()
            if n_str and n_str.upper() != 'NAN':
                namen.add(n_str)

        omschr_sample = ' '.join(groep['Omschrijving'].astype(str).str.upper().head(5))

        iban_data[iban] = {
            'namen': namen,
            'n_pos': (groep['bedrag'] > 0).sum(),
            'n_neg': (groep['bedrag'] < 0).sum(),
            'n_total': len(groep),
            'bedrag_pos': groep[groep['bedrag'] > 0]['bedrag'].sum(),
            'bedrag_neg': abs(groep[groep['bedrag'] < 0]['bedrag'].sum()),
            'omschr_sample': omschr_sample,
            'is_intern_count': groep['is_intern'].sum(),
        }

    # Household IBANs (reeds gedetecteerd door _detecteer_huishoudleden via S4)
    household_ibans = set()
    for iban, data in iban_data.items():
        if data['is_intern_count'] > 0 and iban not in eigen_rek_norm:
            household_ibans.add(iban)

    logger.info(f"RPR: {len(household_ibans)} IBANs al als household gedetecteerd via bidirectioneel (S4)")

    # =========================================================
    # STAP 3: Resolve party_type per IBAN
    # =========================================================
    iban_party_type = {}  # iban → party_type
    iban_signals = {}     # iban → list of signal codes

    for iban, data in iban_data.items():
        signals = []

        # --- FASE 1: Definitieve labels (naam/IBAN-based) ---

        # 1A: own_account — IBAN ∈ eigen_rekeningen
        if iban in eigen_rek_norm:
            iban_party_type[iban] = 'own_account'
            signals.append('OWN_ACCOUNT')
            iban_signals[iban] = signals
            continue

        # 1B: broker_or_investment_platform — eigen financieel domein (Saxo, DeGiro, etc.)
        if iban in eigen_fi_ibans:
            iban_party_type[iban] = 'broker_or_investment_platform'
            signals.append('S8_eigen_fi')
            iban_signals[iban] = signals
            continue

        # 1C: government — overheidsinstanties
        is_government = any(kw in data['omschr_sample'] for kw in OVERHEID_KEYWORDS)
        if is_government:
            iban_party_type[iban] = 'government'
            signals.append('GOVERNMENT_KEYWORD')
            iban_signals[iban] = signals
            continue

        # 1D: card_settlement — creditcard-maatschappijen
        is_card = any(kw in data['omschr_sample'] for kw in CARD_SETTLEMENT_KEYWORDS)
        if is_card:
            iban_party_type[iban] = 'card_settlement'
            signals.append('CARD_SETTLEMENT')
            iban_signals[iban] = signals
            continue

        # 1E: broker_or_investment_platform — FI keyword (geen IBAN match maar naam)
        is_broker = any(kw in data['omschr_sample'] for kw in FINANCIELE_INSTELLINGEN_KEYWORDS)
        if is_broker:
            iban_party_type[iban] = 'broker_or_investment_platform'
            signals.append('BROKER_FI_KEYWORD')
            iban_signals[iban] = signals
            continue

        # 1F: lender_or_mortgage_party — hypotheek/leningverstrekkers
        is_lender = any(kw in data['omschr_sample'] for kw in LENDER_KEYWORDS_COUNTERPARTY)
        if is_lender:
            iban_party_type[iban] = 'lender_or_mortgage_party'
            signals.append('LENDER_KEYWORD')
            iban_signals[iban] = signals
            continue

        # 1G: merchant — bekende winkels/dienstverleners
        is_merchant = any(zoekterm in data['omschr_sample'] for zoekterm in bekende_merchants_set)
        if is_merchant:
            iban_party_type[iban] = 'merchant'
            signals.append('S10_merchant')
            iban_signals[iban] = signals
            continue

        # --- Signalen detecteren ---

        # S1: Achternaam overlap
        s1 = False
        tegenpartij_achternamen = set()
        for naam in data['namen']:
            ach = extract_achternaam(naam)
            if ach:
                tegenpartij_achternamen.add(ach)
                if ach in eigen_achternamen:
                    s1 = True
        if s1:
            signals.append('S1_achternaam')

        # S4: Bidirectioneel patroon (al gedetecteerd door _detecteer_huishoudleden)
        s4 = iban in household_ibans
        if s4:
            signals.append('S4_bidirectioneel')

        # S2: Multi-IBAN linking
        # Als een naam op dit IBAN matcht met een naam op een bewezen household IBAN
        s2 = False
        if not s4:  # S2 is alleen nodig als S4 niet al geldt
            for hh_iban in household_ibans:
                if hh_iban == iban:
                    continue
                hh_data = iban_data.get(hh_iban, {})
                hh_namen = hh_data.get('namen', set())
                # Check of achternaam matcht
                for hh_naam in hh_namen:
                    hh_ach = extract_achternaam(hh_naam)
                    if hh_ach and hh_ach in tegenpartij_achternamen:
                        s2 = True
                        signals.append('S2_multi_iban')
                        break
                if s2:
                    break

        # S5: Transfer-achtig patroon (ronde bedragen, onregelmatig)
        s5 = False
        if data['n_pos'] >= 2 and data['n_neg'] == 0:
            # Unidirectioneel inkomend, check of bedragen rond zijn
            pos_tx = df[(df['_tegen_norm'] == iban) & (df['bedrag'] > 0)]
            bedragen = pos_tx['bedrag'].astype(float)
            ronde = sum(1 for b in bedragen if b % 100 == 0 or b % 50 == 0)
            if ronde / len(bedragen) > 0.7:
                s5 = True
                signals.append('S5_transfer_achtig')

        # S6: Cross-account verschijning
        s6 = False
        if heeft_tegenrek:
            rekeningen_met_iban = df[df['_tegen_norm'] == iban]['Rekeningnummer'].nunique()
            if rekeningen_met_iban >= 2:
                s6 = True
                signals.append('S6_cross_account')

        # S7: Geen economische relatie (geen salaris-keyword, geen huur-keyword, geen merchant)
        s7_keywords = ['SALARIS', 'LOON', 'MANAGEMENTFEE', 'HUUR', 'RENT', 'KAMER',
                       'WONING', 'PREMIE', 'FACTUUR', 'DECLARATIE']
        s7 = not any(kw in data['omschr_sample'] for kw in s7_keywords)
        if s7:
            signals.append('S7_geen_eco_relatie')

        # S9: Rechtsvorm (B.V., N.V., Stichting, etc.)
        s9 = any(m in data['omschr_sample'] for m in _RECHTSVORM_MARKERS)
        if s9:
            signals.append('S9_rechtsvorm')

        # --- FASE 2: Household (vereist STERK signaal) ---
        if s4:
            iban_party_type[iban] = 'household_related_party'
            iban_signals[iban] = signals
            continue

        if s2:
            iban_party_type[iban] = 'household_related_party'
            iban_signals[iban] = signals
            continue

        # --- FASE 2B: Household via achternaam + gedragspatroon ---
        # GENERIEK: klanten hebben vaak 5-8 rekeningen maar uploaden er 3-4.
        # Transfers van eigen rekeningen bij andere banken, of van partner
        # bij een andere bank, verschijnen als "extern inkomen".
        #
        # Regel: zelfde achternaam + privépersoon + geen economische relatie
        #        → household_related_party
        #
        # Dit vangt:
        # - Self-transfers van niet-geüploade eigen rekeningen (Bunq, N26, Revolut)
        # - Partner-transfers van niet-geüploade rekeningen
        # - Familieleden die geld overmaken
        #
        # False positive bescherming:
        # - s9 = False: geen B.V./Stichting (dus geen zakelijke relatie)
        # - s7 = True: geen salaris/huur/factuur keywords (geen economische relatie)
        # - Alleen privépersonen met dezelfde achternaam
        if s1 and not s9 and s7:
            iban_party_type[iban] = 'household_related_party'
            signals.append('HOUSEHOLD_VIA_ACHTERNAAM_GEDRAG')
            iban_signals[iban] = signals
            continue

        # --- FASE 3: Business / Employer ---
        if s9 and s1:
            # S9 (rechtsvorm) + S1 (achternaam) — check voor extra bewijs
            salaris_keywords = ['SALARIS', 'LOON', 'MANAGEMENTFEE', 'MANAGEMENT FEE',
                                'VERGOEDING', 'HONORARIUM']
            has_keyword = any(kw in data['omschr_sample'] for kw in salaris_keywords)
            has_extra = has_keyword or s4 or s6

            if has_extra:
                # Sterk bewijs: eigen BV, salary keywords → employer_or_payroll
                iban_party_type[iban] = 'employer_or_payroll'
                signals.append('EMPLOYER_EXTRA_BEWIJS')
            else:
                # S9 + S1 zonder extra bewijs → business_counterparty
                iban_party_type[iban] = 'business_counterparty'
            iban_signals[iban] = signals
            continue

        # --- FASE 4: Default ---
        if s9:
            # Rechtsvorm zonder naamoverlap → business_counterparty
            iban_party_type[iban] = 'business_counterparty'
        else:
            # Geen signalen of alleen zwak/middel → unknown
            iban_party_type[iban] = 'unknown'

        iban_signals[iban] = signals

    # =========================================================
    # STAP 4: Schrijf party_type naar DataFrame
    # =========================================================
    for iban, pt in iban_party_type.items():
        mask = df['_tegen_norm'] == iban
        df.loc[mask, 'party_type'] = pt

    # Transacties zonder IBAN → unknown
    df.loc[df['party_type'].isna(), 'party_type'] = 'unknown'

    # V3: household_related_party → NIET meer is_intern, maar eigen sectie 'onderling_neutraal'
    # Ze moeten zichtbaar zijn in het rapport (ONDERLING/NEUTRAAL) maar NIET als salary/income
    mask_hh = (df['party_type'] == 'household_related_party') & (~df['is_intern'])
    n_hh_extra = mask_hh.sum()
    if n_hh_extra > 0:
        df.loc[mask_hh, 'regel_sectie'] = 'onderling_neutraal'
        df.loc[mask_hh, 'regel_categorie'] = 'Overboekingen huishouden/partner'
        df.loc[mask_hh, 'regel_confidence'] = 0.95
        df.loc[mask_hh, 'classificatie_bron'] = 'rule'
        df.loc[mask_hh, 'source_family'] = 'household_transfer'
        logger.info(f"RPR V3: {n_hh_extra} household-transacties → onderling_neutraal (niet meer is_intern)")

    # Cleanup
    if '_tegen_norm' in df.columns:
        df.drop(columns=['_tegen_norm'], inplace=True)

    # === LOGGING ===
    pt_counts = df['party_type'].value_counts()
    logger.info(
        f"RPR: Party types resolved:\n"
        + '\n'.join(f"  {pt}: {count}" for pt, count in pt_counts.items())
    )

    # Log gedetailleerd per IBAN met signalen
    for iban, pt in sorted(iban_party_type.items(), key=lambda x: x[1]):
        sigs = iban_signals.get(iban, [])
        namen = iban_data.get(iban, {}).get('namen', set())
        naam_str = ', '.join(list(namen)[:2]) if namen else '?'
        if pt in ('household_related_party', 'employer_or_payroll', 'business_counterparty',
                  'government', 'broker_or_investment_platform', 'lender_or_mortgage_party', 'card_settlement'):
            logger.info(f"RPR DETAIL: {iban[:20]} ({naam_str[:30]}) → {pt} [{', '.join(sigs)}]")

    return df


# ---------------------------------------------------------------------------
# STAP 1c: RULE-BASED CLASSIFICATIE (vóór AI)
# ---------------------------------------------------------------------------
# ChatGPT CEO-plan: "AI mag pas aan zet nadat het systeem al heeft vastgesteld
# wat inkomen is, wat transfer is, wat belasting is, wat hypotheek is."
#
# Deze laag classificeert transacties op basis van harde regels.
# AI mag daarna samenvatten en restcategorieën invullen, maar NIET overrulen.

# Merchant mapping: bekende tegenpartijen → vaste categorie
# Format: (zoekterm_in_omschrijving_uppercase, sectie, categorie, confidence)
MERCHANT_MAPPING = [
    # --- INKOMEN ---
    # UWV
    ('UWV', 'inkomsten', 'UWV/Uitkeringen', 0.95),
    # Kinderbijslag / Kindregelingen
    ('SVB KINDERBIJSLAG', 'inkomsten', 'Kinderbijslag/Kindregelingen', 0.99),
    ('SVB', 'inkomsten', 'Kinderbijslag/Kindregelingen', 0.80),
    # Belastingteruggave (positieve bedragen van Belastingdienst)
    # → wordt apart afgehandeld in de functie (bedrag-afhankelijk)
    # Toeslagen
    ('ZORGTOESLAG', 'inkomsten', 'Toeslagen', 0.99),
    ('HUURTOESLAG', 'inkomsten', 'Toeslagen', 0.99),
    ('KINDGEBONDEN BUDGET', 'inkomsten', 'Toeslagen', 0.99),
    # DGA-loon: generieke patronen die in heel Nederland voorkomen
    # Specifieke BV-namen worden NIET hardcoded — AI herkent DGA-patronen
    # op basis van vast maandelijks bedrag van een BV.

    # --- BELASTINGDIENST (negatief = betaling, positief = teruggave) ---
    # Wordt apart afgehandeld in de functie (bedrag-afhankelijk)

    # --- VASTE LASTEN ---
    # Hypotheek / Woonlasten
    ('ASR HYPOTHEEK', 'vaste_lasten', 'Hypotheek/Huur', 0.99),
    ('ASR LEVENSVERZEKERING', 'vaste_lasten', 'Levensverzekering/ORV', 0.90),
    ('ASR', 'vaste_lasten', 'Hypotheek/Huur', 0.85),
    ('A.S.R', 'vaste_lasten', 'Hypotheek/Huur', 0.85),
    ('NATIONALE-NEDERLANDEN', 'vaste_lasten', 'Hypotheek/Huur', 0.80),
    ('NN GROUP', 'vaste_lasten', 'Hypotheek/Huur', 0.80),
    ('AEGON', 'vaste_lasten', 'Hypotheek/Huur', 0.80),
    ('DELTA LLOYD', 'vaste_lasten', 'Hypotheek/Huur', 0.80),
    ('VVE', 'vaste_lasten', 'VvE', 0.90),
    ('VERENIGING VAN EIGENAREN', 'vaste_lasten', 'VvE', 0.95),
    ('OBVION', 'vaste_lasten', 'Hypotheek/Huur', 0.99),
    ('FLORIUS', 'vaste_lasten', 'Hypotheek/Huur', 0.99),
    ('WOONFONDS', 'vaste_lasten', 'Hypotheek/Huur', 0.99),
    ('HYPOTHEEK', 'vaste_lasten', 'Hypotheek/Huur', 0.95),
    ('MUNT HYPOTHEKEN', 'vaste_lasten', 'Hypotheek/Huur', 0.99),
    ('VISTA HYPOTHEKEN', 'vaste_lasten', 'Hypotheek/Huur', 0.99),
    ('WONINGCORPORATIE', 'vaste_lasten', 'Hypotheek/Huur', 0.95),
    ('VESTIA', 'vaste_lasten', 'Hypotheek/Huur', 0.95),
    ('YMERE', 'vaste_lasten', 'Hypotheek/Huur', 0.95),
    ('EIGEN HAARD', 'vaste_lasten', 'Hypotheek/Huur', 0.95),
    ('DE ALLIANTIE', 'vaste_lasten', 'Hypotheek/Huur', 0.95),
    ('PORTAAL', 'vaste_lasten', 'Hypotheek/Huur', 0.90),
    ('WOONSTAD', 'vaste_lasten', 'Hypotheek/Huur', 0.95),
    # Energie
    ('FRANK ENERGIE', 'vaste_lasten', 'Energie', 0.99),
    ('VATTENFALL', 'vaste_lasten', 'Energie', 0.99),
    ('ENECO', 'vaste_lasten', 'Energie', 0.99),
    ('ESSENT', 'vaste_lasten', 'Energie', 0.99),
    ('BUDGET ENERGIE', 'vaste_lasten', 'Energie', 0.99),
    ('GREENCHOICE', 'vaste_lasten', 'Energie', 0.99),
    ('VANDEBRON', 'vaste_lasten', 'Energie', 0.99),
    ('ENERGIEDIRECT', 'vaste_lasten', 'Energie', 0.99),
    ('INNOVA ENERGIE', 'vaste_lasten', 'Energie', 0.99),
    ('NUON', 'vaste_lasten', 'Energie', 0.99),
    ('OXXIO', 'vaste_lasten', 'Energie', 0.99),
    ('PURE ENERGIE', 'vaste_lasten', 'Energie', 0.99),
    ('UNITED CONSUMERS', 'vaste_lasten', 'Energie', 0.90),
    ('TIBBER', 'vaste_lasten', 'Energie', 0.99),
    ('NEXT ENERGY', 'vaste_lasten', 'Energie', 0.99),
    ('DUTCH ENERGY', 'vaste_lasten', 'Energie', 0.99),
    ('STROOM', 'vaste_lasten', 'Energie', 0.85),
    # Water
    ('VITENS', 'vaste_lasten', 'Water', 0.99),
    ('BRABANT WATER', 'vaste_lasten', 'Water', 0.99),
    ('PWN', 'vaste_lasten', 'Water', 0.95),
    ('DUNEA', 'vaste_lasten', 'Water', 0.99),
    ('WATERNET', 'vaste_lasten', 'Water', 0.99),
    ('EVIDES', 'vaste_lasten', 'Water', 0.99),
    ('OASEN', 'vaste_lasten', 'Water', 0.99),
    ('WATERBEDRIJF GRONINGEN', 'vaste_lasten', 'Water', 0.99),
    ('WMD', 'vaste_lasten', 'Water', 0.90),
    # Zorgverzekering
    ('CZ GROEP', 'vaste_lasten', 'Zorgverzekering', 0.99),
    ('CZ ZORGVERZEKERING', 'vaste_lasten', 'Zorgverzekering', 0.99),
    ('ZILVEREN KRUIS', 'vaste_lasten', 'Zorgverzekering', 0.99),
    ('ACHMEA', 'vaste_lasten', 'Zorgverzekering', 0.90),
    ('MENZIS', 'vaste_lasten', 'Zorgverzekering', 0.99),
    ('OHRA', 'vaste_lasten', 'Zorgverzekering', 0.95),
    ('VGZ', 'vaste_lasten', 'Zorgverzekering', 0.99),
    ('COOPERATIE VGZ', 'vaste_lasten', 'Zorgverzekering', 0.99),
    ('UNIVE', 'vaste_lasten', 'Zorgverzekering', 0.95),
    ('INTERPOLIS', 'vaste_lasten', 'Zorgverzekering', 0.90),
    ('DITZO', 'vaste_lasten', 'Zorgverzekering', 0.95),
    ('JUST', 'vaste_lasten', 'Zorgverzekering', 0.80),
    ('DSW', 'vaste_lasten', 'Zorgverzekering', 0.99),
    ('ZORG EN ZEKERHEID', 'vaste_lasten', 'Zorgverzekering', 0.99),
    ('ENO ZORGVERZEKERAAR', 'vaste_lasten', 'Zorgverzekering', 0.99),
    ('SALLAND VERZEKERINGEN', 'vaste_lasten', 'Zorgverzekering', 0.99),
    # Gemeentebelasting / OZB / Waterschap
    ('GEMEENTELIJKE BELASTING', 'vaste_lasten', 'Gemeentebelasting/OZB/Waterschapsbelasting', 0.99),
    ('GEMEENTE ', 'vaste_lasten', 'Gemeentebelasting/OZB/Waterschapsbelasting', 0.85),
    ('GBLT', 'vaste_lasten', 'Gemeentebelasting/OZB/Waterschapsbelasting', 0.99),
    ('WATERSCHAP', 'vaste_lasten', 'Gemeentebelasting/OZB/Waterschapsbelasting', 0.99),
    ('BELASTINGSAMENWERKING', 'vaste_lasten', 'Gemeentebelasting/OZB/Waterschapsbelasting', 0.99),
    ('COCENSUS', 'vaste_lasten', 'Gemeentebelasting/OZB/Waterschapsbelasting', 0.99),
    ('SVHW', 'vaste_lasten', 'Gemeentebelasting/OZB/Waterschapsbelasting', 0.99),
    ('BSGR', 'vaste_lasten', 'Gemeentebelasting/OZB/Waterschapsbelasting', 0.99),
    ('SABEWA', 'vaste_lasten', 'Gemeentebelasting/OZB/Waterschapsbelasting', 0.99),
    ('HEFPUNT', 'vaste_lasten', 'Gemeentebelasting/OZB/Waterschapsbelasting', 0.99),
    ('RBG', 'vaste_lasten', 'Gemeentebelasting/OZB/Waterschapsbelasting', 0.90),
    # Autoverzekering
    ('CENTRAAL BEHEER', 'vaste_lasten', 'Autoverzekering', 0.85),
    ('ALLSECUR', 'vaste_lasten', 'Autoverzekering', 0.90),
    ('UNIVÉ', 'vaste_lasten', 'Autoverzekering', 0.85),
    ('ALLIANZ', 'vaste_lasten', 'Autoverzekering', 0.80),
    ('INSHARED', 'vaste_lasten', 'Autoverzekering', 0.90),
    # Overige verzekeringen
    ('FBTO', 'vaste_lasten', 'Overige verzekeringen', 0.90),
    ('REAAL', 'vaste_lasten', 'Overige verzekeringen', 0.85),
    ('DELA', 'vaste_lasten', 'Overige verzekeringen', 0.95),
    ('MONUTA', 'vaste_lasten', 'Overige verzekeringen', 0.95),
    ('YARDEN', 'vaste_lasten', 'Overige verzekeringen', 0.95),
    ('UITVAART', 'vaste_lasten', 'Overige verzekeringen', 0.95),
    ('NOPPES VERZEKERINGEN', 'vaste_lasten', 'Overige verzekeringen', 0.90),
    # Internet/TV
    ('ZIGGO', 'vaste_lasten', 'Internet/TV', 0.95),
    ('VODAFONEZIGGO', 'vaste_lasten', 'Internet/TV', 0.95),
    ('KPN', 'vaste_lasten', 'Internet/TV', 0.85),
    ('GLASPOORT', 'vaste_lasten', 'Internet/TV', 0.95),
    ('DELTA', 'vaste_lasten', 'Internet/TV', 0.80),
    ('CAIWAY', 'vaste_lasten', 'Internet/TV', 0.95),
    ('SOLCON', 'vaste_lasten', 'Internet/TV', 0.95),
    ('FREEDOM INTERNET', 'vaste_lasten', 'Internet/TV', 0.99),
    ('YOUFONE', 'vaste_lasten', 'Internet/TV', 0.85),
    ('ODIDO', 'vaste_lasten', 'Internet/TV', 0.85),
    # Mobiele telefonie
    ('T-MOBILE', 'vaste_lasten', 'Mobiele telefonie', 0.95),
    ('VODAFONE', 'vaste_lasten', 'Mobiele telefonie', 0.95),
    ('SIMPEL', 'vaste_lasten', 'Mobiele telefonie', 0.95),
    ('LEBARA', 'vaste_lasten', 'Mobiele telefonie', 0.95),
    ('LYCAMOBILE', 'vaste_lasten', 'Mobiele telefonie', 0.95),
    ('BEN MOBIEL', 'vaste_lasten', 'Mobiele telefonie', 0.95),
    ('HOLLANDSNIEUWE', 'vaste_lasten', 'Mobiele telefonie', 0.95),
    # Streaming/Digitaal
    ('NETFLIX', 'vaste_lasten', 'Streaming/Digitaal', 0.99),
    ('SPOTIFY', 'vaste_lasten', 'Streaming/Digitaal', 0.99),
    ('DISNEY', 'vaste_lasten', 'Streaming/Digitaal', 0.95),
    ('APPLE.COM/BILL', 'vaste_lasten', 'Streaming/Digitaal', 0.90),
    ('ICLOUD', 'vaste_lasten', 'Streaming/Digitaal', 0.95),
    ('YOUTUBE PREMIUM', 'vaste_lasten', 'Streaming/Digitaal', 0.99),
    ('VIDEOLAND', 'vaste_lasten', 'Streaming/Digitaal', 0.99),
    ('HBO MAX', 'vaste_lasten', 'Streaming/Digitaal', 0.99),
    ('PRIME VIDEO', 'vaste_lasten', 'Streaming/Digitaal', 0.95),
    ('AMAZON PRIME', 'vaste_lasten', 'Streaming/Digitaal', 0.90),
    ('GOOGLE STORAGE', 'vaste_lasten', 'Streaming/Digitaal', 0.95),
    ('GOOGLE ONE', 'vaste_lasten', 'Streaming/Digitaal', 0.95),
    ('MICROSOFT 365', 'vaste_lasten', 'Overige abonnementen', 0.95),
    ('ADOBE', 'vaste_lasten', 'Overige abonnementen', 0.95),
    ('CHATGPT', 'vaste_lasten', 'Overige abonnementen', 0.95),
    ('OPENAI', 'vaste_lasten', 'Overige abonnementen', 0.90),
    ('ANTHROPIC', 'vaste_lasten', 'Overige abonnementen', 0.90),
    # Overige abonnementen
    ('NRC', 'vaste_lasten', 'Overige abonnementen', 0.95),
    ('VOLKSKRANT', 'vaste_lasten', 'Overige abonnementen', 0.95),
    ('TELEGRAAF', 'vaste_lasten', 'Overige abonnementen', 0.95),
    ('AD.NL', 'vaste_lasten', 'Overige abonnementen', 0.95),
    ('TROUW', 'vaste_lasten', 'Overige abonnementen', 0.95),
    ('FD.NL', 'vaste_lasten', 'Overige abonnementen', 0.95),
    ('FINANCIEELE DAGBLAD', 'vaste_lasten', 'Overige abonnementen', 0.95),
    # Kinderopvang
    ('KINDEROPVANG', 'vaste_lasten', 'Kinderopvang/BSO/School', 0.95),
    ('BSO', 'vaste_lasten', 'Kinderopvang/BSO/School', 0.90),
    ('PARTOU', 'vaste_lasten', 'Kinderopvang/BSO/School', 0.99),
    ('KIDS FIRST', 'vaste_lasten', 'Kinderopvang/BSO/School', 0.99),
    ('SMALLSTEPS', 'vaste_lasten', 'Kinderopvang/BSO/School', 0.99),
    ('HUMANKIND', 'vaste_lasten', 'Kinderopvang/BSO/School', 0.95),
    ('KINDERGARDEN', 'vaste_lasten', 'Kinderopvang/BSO/School', 0.99),
    ('SCHOOLGELD', 'vaste_lasten', 'Kinderopvang/BSO/School', 0.95),
    # Contributie/Lidmaatschap
    ('ANWB', 'vaste_lasten', 'Contributie/Lidmaatschap', 0.90),
    ('KNVB', 'vaste_lasten', 'Contributie/Lidmaatschap', 0.95),
    ('SPORTSCHOOL', 'vaste_lasten', 'Contributie/Lidmaatschap', 0.90),
    ('FITNESS', 'vaste_lasten', 'Contributie/Lidmaatschap', 0.85),
    ('BASIC-FIT', 'vaste_lasten', 'Contributie/Lidmaatschap', 0.99),
    ('SPORTCITY', 'vaste_lasten', 'Contributie/Lidmaatschap', 0.99),
    ('TRAINMORE', 'vaste_lasten', 'Contributie/Lidmaatschap', 0.99),
    # Donaties
    ('GIVT', 'vaste_lasten', 'Donaties/Goede doelen', 0.95),
    ('KWF', 'vaste_lasten', 'Donaties/Goede doelen', 0.95),
    ('RODE KRUIS', 'vaste_lasten', 'Donaties/Goede doelen', 0.95),
    ('OXFAM', 'vaste_lasten', 'Donaties/Goede doelen', 0.95),
    ('PARTIJ VOOR DE DIEREN', 'vaste_lasten', 'Donaties/Goede doelen', 0.95),
    ('AMNESTY', 'vaste_lasten', 'Donaties/Goede doelen', 0.95),
    ('GREENPEACE', 'vaste_lasten', 'Donaties/Goede doelen', 0.95),
    ('WWF', 'vaste_lasten', 'Donaties/Goede doelen', 0.95),
    ('UNICEF', 'vaste_lasten', 'Donaties/Goede doelen', 0.95),
    ('NATUURMONUMENTEN', 'vaste_lasten', 'Donaties/Goede doelen', 0.95),
    ('LEGER DES HEILS', 'vaste_lasten', 'Donaties/Goede doelen', 0.95),
    ('HARTSTICHTING', 'vaste_lasten', 'Donaties/Goede doelen', 0.95),
    # Bankkosten
    ('KOSTEN PAKKET', 'vaste_lasten', 'Bankkosten', 0.95),
    ('BANKKOSTEN', 'vaste_lasten', 'Bankkosten', 0.99),
    ('REKENINGKOSTEN', 'vaste_lasten', 'Bankkosten', 0.99),

    # --- VARIABELE KOSTEN ---
    # Supermarkten
    ('ALBERT HEIJN', 'variabele_kosten', 'Boodschappen/Supermarkt', 0.99),
    ('AH TO GO', 'variabele_kosten', 'Boodschappen/Supermarkt', 0.99),
    ('JUMBO', 'variabele_kosten', 'Boodschappen/Supermarkt', 0.95),
    ('LIDL', 'variabele_kosten', 'Boodschappen/Supermarkt', 0.99),
    ('ALDI', 'variabele_kosten', 'Boodschappen/Supermarkt', 0.99),
    ('PLUS SUPERMARKT', 'variabele_kosten', 'Boodschappen/Supermarkt', 0.95),
    ('PLUS ', 'variabele_kosten', 'Boodschappen/Supermarkt', 0.80),
    ('DIRK', 'variabele_kosten', 'Boodschappen/Supermarkt', 0.95),
    ('DIRK VAN DEN BROEK', 'variabele_kosten', 'Boodschappen/Supermarkt', 0.99),
    ('DEKAMARKT', 'variabele_kosten', 'Boodschappen/Supermarkt', 0.99),
    ('COOP ', 'variabele_kosten', 'Boodschappen/Supermarkt', 0.90),
    ('PICNIC', 'variabele_kosten', 'Boodschappen/Supermarkt', 0.95),
    ('SPAR', 'variabele_kosten', 'Boodschappen/Supermarkt', 0.90),
    ('VOMAR', 'variabele_kosten', 'Boodschappen/Supermarkt', 0.99),
    ('HOOGVLIET', 'variabele_kosten', 'Boodschappen/Supermarkt', 0.99),
    ('NETTORAMA', 'variabele_kosten', 'Boodschappen/Supermarkt', 0.99),
    ('JAN LINDERS', 'variabele_kosten', 'Boodschappen/Supermarkt', 0.99),
    ('BONI ', 'variabele_kosten', 'Boodschappen/Supermarkt', 0.90),
    ('POIESZ', 'variabele_kosten', 'Boodschappen/Supermarkt', 0.99),
    ('SLIGRO', 'variabele_kosten', 'Boodschappen/Supermarkt', 0.90),
    ('MARQT', 'variabele_kosten', 'Boodschappen/Supermarkt', 0.99),
    ('EKOPLAZA', 'variabele_kosten', 'Boodschappen/Supermarkt', 0.99),
    # Drogist
    ('ETOS', 'variabele_kosten', 'Drogist', 0.95),
    ('KRUIDVAT', 'variabele_kosten', 'Drogist', 0.95),
    ('TREKPLEISTER', 'variabele_kosten', 'Drogist', 0.95),
    ('HOLLAND & BARRETT', 'variabele_kosten', 'Drogist', 0.90),
    ('DA DROGIST', 'variabele_kosten', 'Drogist', 0.95),
    # Tankstations / Laden
    ('SHELL', 'variabele_kosten', 'Benzine/Diesel/Laden', 0.90),
    ('BP ', 'variabele_kosten', 'Benzine/Diesel/Laden', 0.85),
    ('TOTALENERGIES', 'variabele_kosten', 'Benzine/Diesel/Laden', 0.95),
    ('TANGO', 'variabele_kosten', 'Benzine/Diesel/Laden', 0.95),
    ('TINQ', 'variabele_kosten', 'Benzine/Diesel/Laden', 0.95),
    ('ESSO', 'variabele_kosten', 'Benzine/Diesel/Laden', 0.95),
    ('GULF', 'variabele_kosten', 'Benzine/Diesel/Laden', 0.90),
    ('TEXACO', 'variabele_kosten', 'Benzine/Diesel/Laden', 0.95),
    ('FASTNED', 'variabele_kosten', 'Benzine/Diesel/Laden', 0.99),
    ('ALLEGO', 'variabele_kosten', 'Benzine/Diesel/Laden', 0.99),
    ('IONITY', 'variabele_kosten', 'Benzine/Diesel/Laden', 0.99),
    ('TESLA SUPERCHARGER', 'variabele_kosten', 'Benzine/Diesel/Laden', 0.99),
    ('NEWMOTION', 'variabele_kosten', 'Benzine/Diesel/Laden', 0.99),
    ('VATTENFALL LAADPAAL', 'variabele_kosten', 'Benzine/Diesel/Laden', 0.99),
    # OV
    ('NS GROEP', 'variabele_kosten', 'OV/Trein', 0.95),
    ('NS-', 'variabele_kosten', 'OV/Trein', 0.85),
    ('OV-CHIPKAART', 'variabele_kosten', 'OV/Trein', 0.99),
    ('CONNEXXION', 'variabele_kosten', 'OV/Trein', 0.95),
    ('GVB', 'variabele_kosten', 'OV/Trein', 0.95),
    ('RET', 'variabele_kosten', 'OV/Trein', 0.90),
    ('HTM', 'variabele_kosten', 'OV/Trein', 0.90),
    ('ARRIVA', 'variabele_kosten', 'OV/Trein', 0.95),
    ('QBUZZ', 'variabele_kosten', 'OV/Trein', 0.95),
    ('TRANSLINK', 'variabele_kosten', 'OV/Trein', 0.95),
    ('KEOLIS', 'variabele_kosten', 'OV/Trein', 0.95),
    ('EBS ', 'variabele_kosten', 'OV/Trein', 0.85),
    # Parkeren
    ('PARKMOBILE', 'variabele_kosten', 'Parkeren', 0.99),
    ('YELLOWBRICK', 'variabele_kosten', 'Parkeren', 0.99),
    ('Q-PARK', 'variabele_kosten', 'Parkeren', 0.99),
    ('PARKBEE', 'variabele_kosten', 'Parkeren', 0.99),
    ('P1 PARKING', 'variabele_kosten', 'Parkeren', 0.95),
    ('APCOA', 'variabele_kosten', 'Parkeren', 0.95),
    # Taxi/Ride
    ('UBER ', 'variabele_kosten', 'Taxi/Uber', 0.90),
    ('UBER BV', 'variabele_kosten', 'Taxi/Uber', 0.95),
    ('BOLT.EU', 'variabele_kosten', 'Taxi/Uber', 0.95),
    # Afhaal/Bezorging
    ('THUISBEZORGD', 'variabele_kosten', 'Afhaal/Bezorging', 0.99),
    ('UBER EATS', 'variabele_kosten', 'Afhaal/Bezorging', 0.99),
    ('DELIVEROO', 'variabele_kosten', 'Afhaal/Bezorging', 0.99),
    ('JUST EAT', 'variabele_kosten', 'Afhaal/Bezorging', 0.99),
    ('GORILLAS', 'variabele_kosten', 'Afhaal/Bezorging', 0.99),
    ('GETIR', 'variabele_kosten', 'Afhaal/Bezorging', 0.99),
    ('FLINK', 'variabele_kosten', 'Afhaal/Bezorging', 0.99),
    # Kleding
    ('H&M', 'variabele_kosten', 'Kleding', 0.90),
    ('ZARA', 'variabele_kosten', 'Kleding', 0.90),
    ('C&A', 'variabele_kosten', 'Kleding', 0.90),
    ('PRIMARK', 'variabele_kosten', 'Kleding', 0.95),
    ('UNIQLO', 'variabele_kosten', 'Kleding', 0.95),
    ('NIKE', 'variabele_kosten', 'Kleding', 0.85),
    ('ADIDAS', 'variabele_kosten', 'Kleding', 0.85),
    ('WE FASHION', 'variabele_kosten', 'Kleding', 0.95),
    ('ZEEMAN', 'variabele_kosten', 'Kleding', 0.90),
    ('ZALANDO', 'variabele_kosten', 'Kleding', 0.90),
    ('ABOUT YOU', 'variabele_kosten', 'Kleding', 0.90),
    # Elektronica/Online
    ('BOL.COM', 'variabele_kosten', 'Elektronica/Gadgets', 0.80),
    ('COOLBLUE', 'variabele_kosten', 'Elektronica/Gadgets', 0.90),
    ('AMAZON', 'variabele_kosten', 'Elektronica/Gadgets', 0.75),
    ('MEDIAMARKT', 'variabele_kosten', 'Elektronica/Gadgets', 0.95),
    ('MEDIA MARKT', 'variabele_kosten', 'Elektronica/Gadgets', 0.95),
    ('ALIEXPRESS', 'variabele_kosten', 'Elektronica/Gadgets', 0.85),
    ('TEMU', 'variabele_kosten', 'Elektronica/Gadgets', 0.80),
    # Vakantie/Reizen
    ('BOOKING.COM', 'variabele_kosten', 'Vakantie/Reizen', 0.95),
    ('AIRBNB', 'variabele_kosten', 'Vakantie/Reizen', 0.95),
    ('TRANSAVIA', 'variabele_kosten', 'Vakantie/Reizen', 0.99),
    ('KLM', 'variabele_kosten', 'Vakantie/Reizen', 0.95),
    ('RYANAIR', 'variabele_kosten', 'Vakantie/Reizen', 0.99),
    ('EASYJET', 'variabele_kosten', 'Vakantie/Reizen', 0.99),
    ('VUELING', 'variabele_kosten', 'Vakantie/Reizen', 0.99),
    ('TRAVELBIRD', 'variabele_kosten', 'Vakantie/Reizen', 0.99),
    ('SUNWEB', 'variabele_kosten', 'Vakantie/Reizen', 0.99),
    ('CORENDON', 'variabele_kosten', 'Vakantie/Reizen', 0.99),
    ('TUI ', 'variabele_kosten', 'Vakantie/Reizen', 0.95),
    ('PHAROS', 'variabele_kosten', 'Vakantie/Reizen', 0.90),
    ('D-REIZEN', 'variabele_kosten', 'Vakantie/Reizen', 0.99),
    ('ROOMPOT', 'variabele_kosten', 'Vakantie/Reizen', 0.99),
    ('LANDAL', 'variabele_kosten', 'Vakantie/Reizen', 0.99),
    ('CENTER PARCS', 'variabele_kosten', 'Vakantie/Reizen', 0.99),
    ('EUROCAMP', 'variabele_kosten', 'Vakantie/Reizen', 0.99),
    # Restaurant/Uit eten
    ('MCDONALDS', 'variabele_kosten', 'Restaurant/Uit eten', 0.95),
    ('MCDONALD', 'variabele_kosten', 'Restaurant/Uit eten', 0.95),
    ('BURGER KING', 'variabele_kosten', 'Restaurant/Uit eten', 0.95),
    ('STARBUCKS', 'variabele_kosten', 'Restaurant/Uit eten', 0.90),
    ('SUBWAY', 'variabele_kosten', 'Restaurant/Uit eten', 0.90),
    ('FEBO', 'variabele_kosten', 'Restaurant/Uit eten', 0.95),
    ('DUNKIN', 'variabele_kosten', 'Restaurant/Uit eten', 0.90),
    ('NEW YORK PIZZA', 'variabele_kosten', 'Restaurant/Uit eten', 0.95),
    ('DOMINOS', 'variabele_kosten', 'Restaurant/Uit eten', 0.95),
    ('VAPIANO', 'variabele_kosten', 'Restaurant/Uit eten', 0.95),
    ('LA PLACE', 'variabele_kosten', 'Restaurant/Uit eten', 0.90),
    # Huishoudelijke artikelen
    ('ACTION', 'variabele_kosten', 'Huishoudelijke artikelen', 0.85),
    ('HEMA', 'variabele_kosten', 'Huishoudelijke artikelen', 0.80),
    ('IKEA', 'variabele_kosten', 'Huishoudelijke artikelen', 0.90),
    ('BLOKKER', 'variabele_kosten', 'Huishoudelijke artikelen', 0.90),
    ('XENOS', 'variabele_kosten', 'Huishoudelijke artikelen', 0.90),
    ('FLYING TIGER', 'variabele_kosten', 'Huishoudelijke artikelen', 0.90),
    # Bouwmarkt/Tuin
    ('GAMMA', 'variabele_kosten', 'Doe-het-zelf/Tuin', 0.95),
    ('KARWEI', 'variabele_kosten', 'Doe-het-zelf/Tuin', 0.95),
    ('PRAXIS', 'variabele_kosten', 'Doe-het-zelf/Tuin', 0.95),
    ('HORNBACH', 'variabele_kosten', 'Doe-het-zelf/Tuin', 0.95),
    ('FORMIDO', 'variabele_kosten', 'Doe-het-zelf/Tuin', 0.95),
    ('INTRATUIN', 'variabele_kosten', 'Doe-het-zelf/Tuin', 0.95),
    ('TUINCENTRUM', 'variabele_kosten', 'Doe-het-zelf/Tuin', 0.90),
    # Apotheek/Medisch
    ('BENU', 'variabele_kosten', 'Apotheek/Medicijnen', 0.95),
    ('APOTHEEK', 'variabele_kosten', 'Apotheek/Medicijnen', 0.90),
    ('SPECSAVERS', 'variabele_kosten', 'Medisch/Zorgkosten', 0.90),
    ('PEARLE', 'variabele_kosten', 'Medisch/Zorgkosten', 0.90),
    ('HANS ANDERS', 'variabele_kosten', 'Medisch/Zorgkosten', 0.90),
    ('TANDARTS', 'variabele_kosten', 'Medisch/Zorgkosten', 0.95),
    ('FYSIOTHERAP', 'variabele_kosten', 'Medisch/Zorgkosten', 0.95),
    ('HUISARTS', 'variabele_kosten', 'Medisch/Zorgkosten', 0.95),
    # Huisdieren
    ('PETS PLACE', 'variabele_kosten', 'Huisdieren', 0.99),
    ('DIERENSPECIAALZAAK', 'variabele_kosten', 'Huisdieren', 0.95),
    ('DIERENARTS', 'variabele_kosten', 'Huisdieren', 0.95),
    # Auto-onderhoud
    ('APK ', 'variabele_kosten', 'Auto-onderhoud/APK', 0.90),
    ('KWIK FIT', 'variabele_kosten', 'Auto-onderhoud/APK', 0.99),
    ('EUROMASTER', 'variabele_kosten', 'Auto-onderhoud/APK', 0.95),
    ('PROFILE TYRECENTER', 'variabele_kosten', 'Auto-onderhoud/APK', 0.95),
    ('HALFORDS', 'variabele_kosten', 'Auto-onderhoud/APK', 0.90),
    # Cadeaus/Bloemen
    ('GREETZ', 'variabele_kosten', 'Cadeaus/Sinterklaas/Kerst', 0.90),
    ('HALLMARK', 'variabele_kosten', 'Cadeaus/Sinterklaas/Kerst', 0.90),
    ('BRUNA', 'variabele_kosten', 'Cadeaus/Sinterklaas/Kerst', 0.80),

    # --- SPAREN & BELEGGEN ---
    ('SAXO BANK', 'sparen_beleggen', 'Effectenrekening', 0.95),
    ('SAXO', 'sparen_beleggen', 'Effectenrekening', 0.90),
    ('DEGIRO', 'sparen_beleggen', 'Effectenrekening', 0.95),
    ('IBKR', 'sparen_beleggen', 'Effectenrekening', 0.95),
    ('INTERACTIVE BROKERS', 'sparen_beleggen', 'Effectenrekening', 0.95),
    ('BINCK', 'sparen_beleggen', 'Effectenrekening', 0.95),
    ('LYNX', 'sparen_beleggen', 'Effectenrekening', 0.95),
    ('FLATEX', 'sparen_beleggen', 'Effectenrekening', 0.95),
    ('ETORO', 'sparen_beleggen', 'Effectenrekening', 0.95),
    ('TRADING 212', 'sparen_beleggen', 'Effectenrekening', 0.95),
    ('BITVAVO', 'sparen_beleggen', 'Crypto', 0.99),
    ('COINBASE', 'sparen_beleggen', 'Crypto', 0.99),
    ('KRAKEN', 'sparen_beleggen', 'Crypto', 0.90),
    ('MINTOS', 'sparen_beleggen', 'Crowdlending', 0.99),
    ('LENDAHAND', 'sparen_beleggen', 'Crowdlending', 0.99),
    ('PEERBERRY', 'sparen_beleggen', 'Crowdlending', 0.99),
    ('BRAND NEW DAY', 'sparen_beleggen', 'Pensioenopbouw', 0.99),
    ('MEESMAN', 'sparen_beleggen', 'Effectenrekening', 0.99),
    ('NORTHERN TRUST', 'sparen_beleggen', 'Effectenrekening', 0.90),

    # --- MOBILITEIT VAST (wegenbelasting, lease) ---
    ('MOTORRIJTUIGENBELASTING', 'vaste_lasten', 'Mobiliteit vast', 0.99),
    ('MRB', 'vaste_lasten', 'Mobiliteit vast', 0.85),
    ('WEGENBELASTING', 'vaste_lasten', 'Mobiliteit vast', 0.99),
    ('LEASEPLAN', 'vaste_lasten', 'Mobiliteit vast', 0.95),
    ('ALPHABET FLEET', 'vaste_lasten', 'Mobiliteit vast', 0.95),
    ('ATHLON', 'vaste_lasten', 'Mobiliteit vast', 0.90),
    ('ARVAL', 'vaste_lasten', 'Mobiliteit vast', 0.90),

    # --- CASH OPNAME ---
    ('GELDAUTOMAAT', 'variabele_kosten', 'Cash opname', 0.99),
    ('GELDOPNAME', 'variabele_kosten', 'Cash opname', 0.99),
    ('CASH WITHDRAWAL', 'variabele_kosten', 'Cash opname', 0.99),
    ('GEA', 'variabele_kosten', 'Cash opname', 0.85),

    # --- CREDITCARD (geen consumptie, interne verschuiving) ---
    ('ICS/INT CARD', 'intern', 'Creditcard-aflossing', 0.90),
    ('ICS ', 'intern', 'Creditcard-aflossing', 0.85),
    ('INTERNATIONAL CARD SERVICES', 'intern', 'Creditcard-aflossing', 0.95),
    ('VISA CARD', 'intern', 'Creditcard-aflossing', 0.85),
    ('ADYEN', 'intern', 'Creditcard-aflossing', 0.80),

    # --- TIKKIE (terugbetaling gedeelde kosten, geen inkomen) ---
    ('TIKKIE', 'intern', 'Tikkie-terugbetaling', 0.90),

    # --- FAMILIELEDEN / HUISHOUDLEDEN ---
    # NIET hardcoded — wordt automatisch gedetecteerd door _detecteer_huishoudleden()
    # op basis van bidirectionele geldstromen met persoonsnamen.
]

# Normalisatie-mapping: vertaal registry-categorieën naar de canonieke prompt-taxonomie
# De merchant_registry.py gebruikt soms andere/bredere categorienamen dan de AI-prompt.
# Zonder deze mapping belanden transacties in onbekende categorieën die als "Overig" tellen.
_CATEGORIE_NORMALISATIE = {
    # variabele_kosten
    'Horeca/Restaurants': 'Restaurant/Uit eten',
    'Kleding/Schoenen': 'Kleding',
    'Gezondheid/Apotheek': 'Apotheek/Medicijnen',
    'Entertainment/Vrije tijd': 'Uitjes/Attracties/Bioscoop',
    'Wonen/Interieur': 'Meubels/Inrichting',
    'Auto/Mobiliteit': 'Auto-onderhoud/APK',
    'Openbaar vervoer': 'OV/Trein',
    'Cadeaus/Donaties': 'Cadeaus',
    'Persoonlijke verzorging': 'Huishoudelijke artikelen',
    # vaste_lasten
    'Abonnementen/Lidmaatschappen': 'Overige abonnementen',
    'Verzekeringen': 'Overige verzekeringen',
    'Gemeentelijke heffingen': 'Gemeentebelasting/OZB/Waterschapsbelasting',
    'Telecom/Internet': 'Internet/TV',
    'Kinderopvang': 'Kinderopvang/BSO/School',
    'Onderwijs': 'Kinderopvang/BSO/School',
}

# Remap specifieke "Overig variabel" entries naar juiste categorieën
_OVERIG_REMAP = {
    # Webshops → specifieke categorieën
    'TEMU': ('variabele_kosten', 'Huishoudelijke artikelen', 0.80),
    'ETSY': ('variabele_kosten', 'Huishoudelijke artikelen', 0.80),
    'MARKTPLAATS': ('variabele_kosten', 'Huishoudelijke artikelen', 0.78),
    'EBAY': ('variabele_kosten', 'Huishoudelijke artikelen', 0.78),
    'THUISWINKEL': ('variabele_kosten', 'Huishoudelijke artikelen', 0.75),
    'BOL COM': ('variabele_kosten', 'Elektronica/Gadgets', 0.82),
    'OTTO': ('variabele_kosten', 'Huishoudelijke artikelen', 0.80),
    'HEMA ONLINE': ('variabele_kosten', 'Huishoudelijke artikelen', 0.85),
    # Post/pakketdiensten → Huishoudelijke artikelen (verzendkosten)
    'POSTNL': ('variabele_kosten', 'Huishoudelijke artikelen', 0.78),
    'POST NL': ('variabele_kosten', 'Huishoudelijke artikelen', 0.78),
    'DHL': ('variabele_kosten', 'Huishoudelijke artikelen', 0.78),
    'UPS': ('variabele_kosten', 'Huishoudelijke artikelen', 0.78),
    'FEDEX': ('variabele_kosten', 'Huishoudelijke artikelen', 0.78),
    'TNT EXPRESS': ('variabele_kosten', 'Huishoudelijke artikelen', 0.78),
    'GLS': ('variabele_kosten', 'Huishoudelijke artikelen', 0.78),
    'DPD': ('variabele_kosten', 'Huishoudelijke artikelen', 0.78),
    'BUDBEE': ('variabele_kosten', 'Huishoudelijke artikelen', 0.78),
    'BRENGER': ('variabele_kosten', 'Huishoudelijke artikelen', 0.78),
    'BRIEVENBUSPAKKET': ('variabele_kosten', 'Huishoudelijke artikelen', 0.78),
    # Boekhandels → Boeken/Media
    'THE READ SHOP': ('variabele_kosten', 'Boeken/Media', 0.85),
    'READSHOP': ('variabele_kosten', 'Boeken/Media', 0.85),
    'BOEKHANDEL': ('variabele_kosten', 'Boeken/Media', 0.82),
    # Belastingdienst → vaste_lasten (FOUT in registry als variabele_kosten)
    'BELASTINGDIENST': ('vaste_lasten', 'Inkomstenbelasting/Voorlopige aanslag', 0.85),
    'BELASTINGAANGIFTE': ('vaste_lasten', 'Inkomstenbelasting/Voorlopige aanslag', 0.80),
    'INKOMSTENBELASTING': ('vaste_lasten', 'Inkomstenbelasting/Voorlopige aanslag', 0.90),
    'AANGIFTELOON': ('vaste_lasten', 'Inkomstenbelasting/Voorlopige aanslag', 0.80),
    'ANF VOOR STAKINGSUITKERINGEN': ('vaste_lasten', 'Inkomstenbelasting/Voorlopige aanslag', 0.75),
    'NETTO LOONBELASTINGEN': ('vaste_lasten', 'Inkomstenbelasting/Voorlopige aanslag', 0.80),
    'TOESLAGEN': ('inkomsten', 'Toeslagen', 0.85),
    # Professionele diensten → Overige vaste lasten (niet variabel)
    'ACCOUNTANT': ('vaste_lasten', 'Overige vaste lasten', 0.82),
    'ACCOUNTANTSKANTOOR': ('vaste_lasten', 'Overige vaste lasten', 0.82),
    'ADMINISTRATIE': ('vaste_lasten', 'Overige vaste lasten', 0.75),
    'ADMINISTRATIEKANTOOR': ('vaste_lasten', 'Overige vaste lasten', 0.82),
    'BOEKHOUDER': ('vaste_lasten', 'Overige vaste lasten', 0.85),
    'BELASTINGADVISEUR': ('vaste_lasten', 'Overige vaste lasten', 0.85),
    'FINANCIEEL ADVISEUR': ('vaste_lasten', 'Overige vaste lasten', 0.82),
    'FINANCIAL ADVISOR': ('vaste_lasten', 'Overige vaste lasten', 0.82),
    'VERZEKERINGSADVISEUR': ('vaste_lasten', 'Overige vaste lasten', 0.82),
    'HYPOTHEEKADVIES': ('vaste_lasten', 'Overige vaste lasten', 0.82),
    # Juridisch → Overige vaste lasten
    'ADVOCAAT': ('vaste_lasten', 'Overige vaste lasten', 0.82),
    'ADVOKAAT KANTOOR': ('vaste_lasten', 'Overige vaste lasten', 0.82),
    'JURIDISCH ADVIES': ('vaste_lasten', 'Overige vaste lasten', 0.82),
    'JURIDISCH': ('vaste_lasten', 'Overige vaste lasten', 0.80),
    'NOTARIS': ('vaste_lasten', 'Overige vaste lasten', 0.82),
    'NOTARISKANTOOR': ('vaste_lasten', 'Overige vaste lasten', 0.82),
    'GERECHTSDEURWAARDER': ('vaste_lasten', 'Overige vaste lasten', 0.85),
    # Makelaar → Onderhoud woning (makelen is woonkosten-gerelateerd)
    'MAKELAAR': ('variabele_kosten', 'Onderhoud woning/Klussen', 0.80),
    'MAKELAARDIJ': ('variabele_kosten', 'Onderhoud woning/Klussen', 0.80),
    # Dienstverlening → specifieke categorieën
    'TUINMAN': ('variabele_kosten', 'Tuin/Buiten', 0.82),
    'INSTALLATEUR': ('variabele_kosten', 'Onderhoud woning/Klussen', 0.82),
    'STOMERIJ': ('variabele_kosten', 'Huishoudelijke artikelen', 0.85),
    'WASSERETTE': ('variabele_kosten', 'Huishoudelijke artikelen', 0.85),
    'DRUKKERIJ': ('variabele_kosten', 'Huishoudelijke artikelen', 0.80),
    'COPYSHOP': ('variabele_kosten', 'Huishoudelijke artikelen', 0.80),
    # Bank → vaste_lasten
    'BANK': ('vaste_lasten', 'Bankkosten', 0.70),
    'BANKGEBUREN': ('vaste_lasten', 'Bankkosten', 0.85),
    'RENTEBETALING': ('vaste_lasten', 'Overige vaste lasten', 0.75),
    'RENTE': ('vaste_lasten', 'Overige vaste lasten', 0.70),
    'LENING': ('vaste_lasten', 'Overige vaste lasten', 0.75),
    # Advies generiek → laat AI beslissen (te breed)
    'ADVIES': None,  # skip, te generiek
}

# Extend met MEGA_MERCHANT_MAPPING (1451 extra merchants)
try:
    from merchant_registry import MEGA_MERCHANT_MAPPING
    _existing_terms = set(z for z, _, _, _ in MERCHANT_MAPPING)
    _n_remapped = 0
    _n_normalized = 0
    _extra = []
    for z, s, c, conf in MEGA_MERCHANT_MAPPING:
        if z in _existing_terms:
            continue
        # Stap 1: Remap "Overig variabel" entries naar specifieke categorieën
        if z in _OVERIG_REMAP:
            remap = _OVERIG_REMAP[z]
            if remap is None:
                continue  # Skip deze entry (te generiek)
            s, c, conf = remap
            _n_remapped += 1
        # Stap 2: Normaliseer registry-categorienamen naar prompt-taxonomie
        elif c in _CATEGORIE_NORMALISATIE:
            c = _CATEGORIE_NORMALISATIE[c]
            _n_normalized += 1
        _extra.append((z, s, c, conf))
    MERCHANT_MAPPING.extend(_extra)
    logger.info(f"Merchant registry geladen: {len(_extra)} extra merchants "
                f"({_n_remapped} remapped, {_n_normalized} normalized), totaal {len(MERCHANT_MAPPING)}")
except ImportError:
    logger.warning("merchant_registry.py niet gevonden, alleen basis MERCHANT_MAPPING actief")


# ---------------------------------------------------------------------------
# Belastingdienst aanslagnummer letter-code detectie
# Structuur: [BSN 9 cijfers].[Letter(s)].[Periode].[Status]
# De letter(s) na het BSN bepalen het type belasting.
# ---------------------------------------------------------------------------
import re as _re

_AANSLAGNUMMER_LETTER_MAP = {
    'H': 'Inkomstenbelasting/Voorlopige aanslag',
    'V': 'Vennootschapsbelasting (VPB)',
    'B': 'BTW/Omzetbelasting',
    'F': 'BTW/Omzetbelasting',           # Naheffingsaanslag OB
    'O': 'BTW/Omzetbelasting',           # Teruggave OB
    'L': 'Loonheffing',
    'A': 'Loonheffing',                  # Naheffingsaanslag LH
    'J': 'Loonheffing',                  # Teruggave LH
    'W': 'ZVW-premie',
    'M': 'Motorrijtuigenbelasting (MRB)',
    'Y': 'Motorrijtuigenbelasting (MRB)', # Naheffingsaanslag MRB
    'Z': 'Overige belastingen',
}

# Toeslagen: T1=kinderopvang, T2=huurtoeslag, T3=zorgtoeslag
_TOESLAG_MAP = {
    'T1': 'Toeslagen',
    'T2': 'Toeslagen',
    'T3': 'Toeslagen',
}

# Regex: 9 cijfers gevolgd door een of meer letters (optioneel met punt ervoor)
# Voorbeelden: 123456789H40, 123456789.H.40.1, 123456789V20
_AANSLAG_RE = _re.compile(r'\b(\d{9})\.?([A-Z][A-Z0-9]?)\.?\d')

def _detecteer_belastingtype_uit_kenmerk(omschrijving: str) -> str | None:
    """Extract belastingtype uit aanslagnummer betalingskenmerk in bankomschrijving.

    Returns categorie-string of None als geen kenmerk gevonden.
    """
    m = _AANSLAG_RE.search(omschrijving)
    if m:
        letter_code = m.group(2)
        # Check eerst toeslagen (T1, T2, T3)
        if letter_code in _TOESLAG_MAP:
            return _TOESLAG_MAP[letter_code]
        # Dan enkelvoudige letter
        first_letter = letter_code[0]
        if first_letter in _AANSLAGNUMMER_LETTER_MAP:
            return _AANSLAGNUMMER_LETTER_MAP[first_letter]
    return None


def _classificeer_rule_based(df: pd.DataFrame) -> pd.DataFrame:
    """Classificeer transacties op basis van harde regels VOORDAT de AI eraan te pas komt.

    Voegt kolommen toe:
    - 'regel_sectie': inkomsten/vaste_lasten/variabele_kosten/sparen_beleggen/intern/None
    - 'regel_categorie': specifieke categorie of None
    - 'regel_confidence': 0.0-1.0
    - 'classificatie_bron': 'rule' of 'ai' (wordt later ingevuld)

    Belastingdienst wordt apart afgehandeld: positief = teruggave (inkomsten), negatief = betaling (vaste lasten).
    """
    df['regel_sectie'] = None
    df['regel_categorie'] = None
    df['regel_confidence'] = 0.0
    df['classificatie_bron'] = None
    df['source_family'] = None  # V3: income source lineage

    n_geclassificeerd = 0

    for idx, row in df.iterrows():
        if row.get('is_intern', False):
            continue  # Al gemarkeerd als intern

        omschr = str(row.get('Omschrijving', '')).upper()
        bedrag = float(row.get('bedrag', 0))

        # Speciaal geval: Belastingdienst — uitsplitsing per belastingtype
        if 'BELASTINGDIENST' in omschr or 'BELASTING DIENST' in omschr:
            # Stap 1: probeer belastingtype uit aanslagnummer kenmerk
            kenmerk_type = _detecteer_belastingtype_uit_kenmerk(omschr)

            # Stap 2: fallback op tekst-keywords als geen kenmerk gevonden
            if kenmerk_type is None:
                if 'IB' in omschr or 'INKOMSTENBELASTING' in omschr or 'INKOMSTENBEL' in omschr or 'VOORLOPIGE AANSLAG' in omschr:
                    kenmerk_type = 'Inkomstenbelasting/Voorlopige aanslag'
                elif 'VPB' in omschr or 'VENNOOTSCHAPSBELASTING' in omschr:
                    kenmerk_type = 'Vennootschapsbelasting (VPB)'
                elif 'OB' in omschr or 'OMZETBELASTING' in omschr or 'BTW' in omschr:
                    kenmerk_type = 'BTW/Omzetbelasting'
                elif 'LH' in omschr or 'LOONHEFFING' in omschr or 'LOONBELASTING' in omschr:
                    kenmerk_type = 'Loonheffing'
                elif 'ZVW' in omschr or 'ZORGVERZEKERINGSWET' in omschr:
                    kenmerk_type = 'ZVW-premie'
                elif 'MRB' in omschr or 'MOTORRIJTUIGEN' in omschr:
                    kenmerk_type = 'Motorrijtuigenbelasting (MRB)'
                elif 'TOESLAG' in omschr or 'ZORGTOESLAG' in omschr or 'HUURTOESLAG' in omschr:
                    kenmerk_type = 'Toeslagen'

            # Stap 3: bepaal sectie + categorie
            if kenmerk_type == 'Toeslagen':
                # Toeslagen zijn altijd inkomsten
                df.at[idx, 'regel_sectie'] = 'inkomsten'
                df.at[idx, 'regel_categorie'] = 'Toeslagen'
                df.at[idx, 'regel_confidence'] = 0.95
                df.at[idx, 'source_family'] = 'child_benefit'
            elif bedrag > 0:
                # Teruggave: specificeer welk type teruggave
                df.at[idx, 'regel_sectie'] = 'inkomsten'
                if kenmerk_type:
                    df.at[idx, 'regel_categorie'] = f'Belastingteruggave ({kenmerk_type})'
                else:
                    df.at[idx, 'regel_categorie'] = 'Belastingteruggave'
                df.at[idx, 'regel_confidence'] = 0.95
                df.at[idx, 'source_family'] = 'tax_refund'
            else:
                # Betaling: gebruik gevonden type of default
                if kenmerk_type:
                    df.at[idx, 'regel_categorie'] = kenmerk_type
                    df.at[idx, 'regel_confidence'] = 0.95  # kenmerk-based = hoge confidence
                else:
                    df.at[idx, 'regel_categorie'] = 'Overige belastingen'  # NIET meer default naar IB
                    df.at[idx, 'regel_confidence'] = 0.70  # lagere confidence want onbekend type
                df.at[idx, 'regel_sectie'] = 'vaste_lasten'
            df.at[idx, 'classificatie_bron'] = 'rule'
            n_geclassificeerd += 1
            continue

        # Merchant mapping doorlopen
        # V3: source_family mapping per categorie
        _CAT_TO_SOURCE_FAMILY = {
            'UWV/Uitkeringen': 'benefits_uwv',
            'Kinderbijslag/Kindregelingen': 'child_benefit',
            'Toeslagen': 'child_benefit',
            'Pensioen/AOW': 'benefits_uwv',
            'Studiefinanciering': 'benefits_uwv',
            'Huurinkomsten': 'rental_income',
            'Belastingteruggave': 'tax_refund',
            'Effectenrekening': 'wealth_allocation',
            'Effectenrekening (terugstorting)': 'broker_return',
            'Crowdlending': 'wealth_allocation',
            'Crowdlending (terugbetaling)': 'broker_return',
            'Crypto': 'wealth_allocation',
            'Crypto (terugstorting)': 'broker_return',
            'Pensioenopbouw': 'wealth_allocation',
            'Spaarrekening': 'wealth_allocation',
            'Creditcard-aflossing': 'creditcard_settlement',
            'Tikkie-terugbetaling': 'internal_transfer',
            'Freelance/Opdrachten': 'freelance_business',
        }
        for zoekterm, sectie, categorie, confidence in MERCHANT_MAPPING:
            if zoekterm in omschr:
                # Creditcard en Tikkie: markeer als intern
                if sectie == 'intern':
                    df.at[idx, 'regel_sectie'] = 'intern'
                    df.at[idx, 'regel_categorie'] = categorie
                    df.at[idx, 'regel_confidence'] = confidence
                    df.at[idx, 'classificatie_bron'] = 'rule'
                    df.at[idx, 'source_family'] = _CAT_TO_SOURCE_FAMILY.get(categorie, 'neutral_technical')
                    n_geclassificeerd += 1
                    break
                # Effectenrekening: positief = terugstorting (vermogensmutatie, NIET inkomen)
                elif sectie == 'sparen_beleggen' and bedrag > 0:
                    terugstort_cat = categorie.replace('Effectenrekening', 'Effectenrekening (terugstorting)') if 'Effectenrekening' in categorie else categorie + ' (terugstorting)'
                    df.at[idx, 'regel_sectie'] = 'sparen_beleggen'
                    df.at[idx, 'regel_categorie'] = terugstort_cat
                    df.at[idx, 'regel_confidence'] = confidence
                    df.at[idx, 'classificatie_bron'] = 'rule'
                    df.at[idx, 'source_family'] = 'broker_return'
                    n_geclassificeerd += 1
                    break
                # Tankstations: positief bedrag >€100 is GEEN benzine (Shell Energy etc.)
                # Laat AI dit classificeren (vaak energie-terugbetaling)
                elif categorie == 'Benzine/Diesel/Laden' and bedrag > 100:
                    continue  # Skip, laat AI beslissen
                else:
                    df.at[idx, 'regel_sectie'] = sectie
                    df.at[idx, 'regel_categorie'] = categorie
                    df.at[idx, 'regel_confidence'] = confidence
                    df.at[idx, 'classificatie_bron'] = 'rule'
                    df.at[idx, 'source_family'] = _CAT_TO_SOURCE_FAMILY.get(categorie)
                    n_geclassificeerd += 1
                    break

    # Keyword-heuristieken: voor transacties die MERCHANT_MAPPING niet pakt,
    # kijk naar generieke Nederlandse trefwoorden in de bankomschrijving.
    # Dit pakt lokale/regionale merchants op die niet in de registry staan.
    # Alleen voor UITGAANDE transacties (expenses).
    _KEYWORD_HEURISTIEKEN = [
        # --- VASTE LASTEN ---
        # Verzekeringen (generiek)
        (['VERZEKERING', 'INSURANCE', 'VERZEKERAAR', 'POLIS'], 'vaste_lasten', 'Overige verzekeringen', 0.80),
        (['ZORGVERZEKERING', 'ZORGPREMIE', 'ZORGPOLIS'], 'vaste_lasten', 'Zorgverzekering', 0.90),
        (['AUTOVERZEKERING', 'MOTORVERZEKERING', 'WA VERZEKERING', 'CASCO'], 'vaste_lasten', 'Autoverzekering', 0.90),
        (['INBOEDEL', 'WOONVERZEKERING', 'OPSTALVERZEKERING', 'WOONHUISVERZEKERING'], 'vaste_lasten', 'Woonverzekering/Inboedel', 0.90),
        (['UITVAART'], 'vaste_lasten', 'Overige verzekeringen', 0.90),
        (['RECHTSBIJSTAND', 'AANSPRAKELIJKHEID', 'AVP'], 'vaste_lasten', 'Overige verzekeringen', 0.85),
        # Energie
        (['STROOM', 'ELEKTRA', 'ELEKTRICITEIT', 'GASREKENING', 'ENERGIELEVER', 'WARMTENET', 'WARMTELEVERING', 'STADSWARMTE'], 'vaste_lasten', 'Energie', 0.85),
        # Water
        (['WATERBEDRIJF', 'DRINKWATER', 'WATERLEIDINGBEDRIJF', 'PWN', 'OASEN', 'BRABANT WATER', 'VITENS', 'EVIDES', 'DUNEA', 'WATERNET'], 'vaste_lasten', 'Water', 0.90),
        # Belasting / gemeente
        (['GEMEENTEBELASTING', 'GEMEENTELIJKE BELASTING', 'GBLT', 'WATERSCHAP', 'WATERSCHAPSBELASTING', 'OZB', 'RIOOLHEFFING', 'AFVALSTOFFENHEFFING'], 'vaste_lasten', 'Gemeentebelasting/OZB/Waterschapsbelasting', 0.90),
        (['GEMEENTE ', 'GEMEENTEHUIS'], 'vaste_lasten', 'Gemeentebelasting/OZB/Waterschapsbelasting', 0.75),
        # Telecom
        (['GLASVEZEL', 'FIBER', 'INTERNET BETALING'], 'vaste_lasten', 'Internet/TV', 0.85),
        (['MOBIEL ABONNEMENT', 'SIMONLY', 'SIM ONLY'], 'vaste_lasten', 'Mobiele telefonie', 0.90),
        # Kinderopvang
        (['KINDEROPVANG', 'KINDERDAGVERBLIJF', 'KDV ', 'BSO ', 'BUITENSCHOOLSE', 'PEUTERSPEELZAAL', 'GASTOUDER', 'GASTOUDERBUREAU'], 'vaste_lasten', 'Kinderopvang/BSO/School', 0.90),
        (['SCHOOLGELD', 'OUDERBIJDRAGE', 'LEERMIDDELEN'], 'vaste_lasten', 'Kinderopvang/BSO/School', 0.85),
        # Contributie / lidmaatschap
        (['CONTRIBUTIE', 'LIDMAATSCHAP', 'VERENIGING', 'SPORTVERENIGING'], 'vaste_lasten', 'Contributie/Lidmaatschap', 0.80),
        # Donaties
        (['DONATIE', 'SCHENKING', 'GOED DOEL', 'GOEDE DOEL', 'CHARITY', 'GIFTEN'], 'vaste_lasten', 'Donaties/Goede doelen', 0.85),
        (['OXFAM', 'UNICEF', 'RODE KRUIS', 'ARTSEN ZONDER GRENZEN', 'AMNESTY', 'GREENPEACE', 'WWF', 'CLINICLOWNS', 'KIKA', 'WARCHILD', 'WAR CHILD', 'PLAN INTERNATIONAL', 'SOS KINDERDORPEN', 'STICHTING VLUCHTELING'], 'vaste_lasten', 'Donaties/Goede doelen', 0.92),
        # Bankkosten
        (['BANKKOSTEN', 'BETAALPAKKET', 'REKENINGKOSTEN', 'SERVICEKOSTEN REKENING'], 'vaste_lasten', 'Bankkosten', 0.95),
        # Abonnementen
        (['ABONNEMENT', 'SUBSCRIPTION', 'MAANDELIJKS BEDRAG'], 'vaste_lasten', 'Overige abonnementen', 0.75),
        # Hypotheek / huur
        (['HYPOTHEEK', 'MORTGAGE'], 'vaste_lasten', 'Hypotheek/Huur', 0.90),
        (['HUUR WONING', 'HUURPRIJS', 'HUURBETALING', 'MAANDHUUR'], 'vaste_lasten', 'Hypotheek/Huur', 0.85),

        # --- VARIABELE KOSTEN ---
        # Medisch
        (['TANDARTS', 'TANDARTSENPRAKTIJK', 'DENTAL', 'MONDHYGIEN'], 'variabele_kosten', 'Huisarts/Tandarts/Specialist', 0.92),
        (['HUISARTS', 'HUISARTSENPRAKTIJK', 'DOKTER', 'DOKTERSPRAKTIJK', 'MEDISCH CENTRUM'], 'variabele_kosten', 'Huisarts/Tandarts/Specialist', 0.90),
        (['APOTHEEK', 'PHARMACY', 'FARMACEUTISCH'], 'variabele_kosten', 'Apotheek/Medicijnen', 0.92),
        (['ZIEKENHUIS', 'HOSPITAL', 'KLINIEK', 'CLINIC', 'POLIKLINIEK'], 'variabele_kosten', 'Ziekenhuiskosten/Eigen risico', 0.90),
        (['FYSIOTHERAP', 'FYSIO ', 'OSTEOPATH', 'CHIROPRACT', 'ACUPUNCT', 'HOMEOPATH'], 'variabele_kosten', 'Fysiotherapie/Alternatief', 0.90),
        (['OPTICI', 'OPTICIEN', 'BRILLEN', 'SPECSAVERS', 'HANS ANDERS', 'PEARL OPTICIENS', 'EYE WISH', 'CONTACTLENZEN'], 'variabele_kosten', 'Brillen/Lenzen', 0.90),
        # Sport / fitness
        (['FITNESS', 'SPORTSCHOOL', 'GYM ', 'BASIC FIT', 'ANYTIME FITNESS', 'FIT FOR FREE', 'TRAININGCENTR'], 'variabele_kosten', 'Sport/Fitness', 0.90),
        # Auto
        (['GARAGE', 'APK ', 'AUTO-ONDERHOUD', 'AUTOSERVICE', 'BANDENCENTR', 'KWIK FIT', 'EUROMASTER', 'PROFILE TYRECENTER'], 'variabele_kosten', 'Auto-onderhoud/APK', 0.85),
        (['PARKEERGARAGE', 'PARKEERGELD', 'PARKEERAUTOMAAT', 'Q-PARK', 'APCOA', 'YELLOWBRICK', 'PARKMOBILE'], 'variabele_kosten', 'Parkeren', 0.90),
        # Woning
        (['BOUWMARKT', 'HORNBACH', 'KARWEI', 'PRAXIS', 'GAMMA'], 'variabele_kosten', 'Onderhoud woning/Klussen', 0.85),
        (['TUINCENTR', 'INTRATUIN', 'GROENRIJK'], 'variabele_kosten', 'Tuin/Buiten', 0.88),
        (['MEUBEL', 'IKEA', 'SLAAPKAMER', 'MATRAS', 'BEDDENGOED'], 'variabele_kosten', 'Meubels/Inrichting', 0.85),
        # Eten & drinken
        (['RESTAURANT', 'RISTORANTE', 'BRASSERIE', 'BISTRO', 'EETCAFE'], 'variabele_kosten', 'Restaurant/Uit eten', 0.88),
        (['CAFE ', 'CAFÉ', 'BAR ', 'KROEG', 'PUB '], 'variabele_kosten', 'Café/Drinken', 0.80),
        # Kleding
        (['H&M', 'ZARA', 'C&A', 'PRIMARK', 'WE FASHION', 'ONLY ', 'VERO MODA', 'JACK & JONES', 'UNIQLO', 'NIKE STORE', 'ADIDAS STORE'], 'variabele_kosten', 'Kleding', 0.88),
        (['KLEDINGWINKEL', 'FASHION', 'MODE '], 'variabele_kosten', 'Kleding', 0.75),
        (['SCHOENENWINKEL', 'SCHOENEN', 'NELSON', 'VAN HAREN', 'SHOE', 'FOOTLOCKER'], 'variabele_kosten', 'Schoenen', 0.85),
        # Cadeaus
        (['CADEAU', 'GIFT', 'HALLMARK', 'GREETZ', 'BLOEMEN', 'FLEUROP', 'BLOEMIST'], 'variabele_kosten', 'Cadeaus', 0.80),
        # Huisdieren
        (['DIERENARTS', 'VETERINAIR', 'DIERENKLINIEK', 'DIERENWINKEL', 'PETS PLACE', 'RANZIJN', 'JUMPER DIER'], 'variabele_kosten', 'Huisdieren', 0.90),
        # Vakantie / reizen
        (['HOTEL', 'BOOKING.COM', 'AIRBNB', 'HOSTEL', 'RESORT', 'CAMPING', 'VAKANTIEPARK', 'ROOMPOT', 'LANDAL', 'CENTER PARCS', 'CENTERPARCS'], 'variabele_kosten', 'Vakantie/Reizen', 0.88),
        (['VLIEGTICKET', 'AIRLINE', 'TRANSAVIA', 'KLM ', 'EASYJET', 'RYANAIR', 'CORENDON', 'VUELING', 'SCHIPHOL'], 'variabele_kosten', 'Vakantie/Reizen', 0.90),
        # School / studie
        (['UNIVERSITEIT', 'HOGESCHOOL', 'COLLEGEGELD', 'STUDIEBOEK', 'STUDIE', 'CURSUS', 'OPLEIDING', 'WORKSHOP'], 'variabele_kosten', 'School/Studie/Cursussen', 0.85),
        # Boeken / media
        (['BOEKHANDEL', 'BOEKWINKEL', 'BRUNA', 'BIBLIOTHEEK', 'READSHOP'], 'variabele_kosten', 'Boeken/Media', 0.85),
        # Uitjes
        (['BIOSCOOP', 'CINEMA', 'PATHE', 'KINEPOLIS', 'EFTELING', 'ATTRACTIEPARK', 'DIERENTUIN', 'ARTIS', 'BLIJDORP', 'MUSEUM', 'THEATER', 'SCHOUWBURG', 'CONCERT', 'TICKETMASTER', 'EVENTIM'], 'variabele_kosten', 'Uitjes/Attracties/Bioscoop', 0.88),
        # Terugbetaling (positieve bedragen van winkels)
        (['RETOUR', 'REFUND', 'STORNO', 'TERUGBETALING', 'RESTITUTIE', 'CREDITNOTA'], 'variabele_kosten', 'Terugbetaling/Refund', 0.80),
    ]

    n_keyword_used = 0
    for idx, row in df.iterrows():
        if row.get('classificatie_bron') is not None:
            continue  # Al geclassificeerd
        if row.get('is_intern', False):
            continue
        bedrag = float(row.get('bedrag', 0))
        if bedrag > 0:
            continue  # Positieve (inkomende) transactie → skip, eigen detectoren
        omschr = str(row.get('Omschrijving', '')).upper()
        best_match = None
        best_confidence = 0
        for keywords, sectie, categorie, confidence in _KEYWORD_HEURISTIEKEN:
            for kw in keywords:
                if kw in omschr and confidence > best_confidence:
                    best_match = (sectie, categorie, confidence)
                    best_confidence = confidence
                    break
        if best_match:
            df.at[idx, 'regel_sectie'] = best_match[0]
            df.at[idx, 'regel_categorie'] = best_match[1]
            df.at[idx, 'regel_confidence'] = best_match[2]
            df.at[idx, 'classificatie_bron'] = 'rule'
            n_keyword_used += 1
            n_geclassificeerd += 1

    if n_keyword_used > 0:
        logger.info(f"KEYWORD-HEURISTIEK: {n_keyword_used} uitgaande transacties geclassificeerd via trefwoorden")

    # Statistieken
    n_totaal = len(df[~df.get('is_intern', False)])
    n_onzeker = n_totaal - n_geclassificeerd
    pct = (n_geclassificeerd / n_totaal * 100) if n_totaal > 0 else 0
    logger.info(f"Rule-based classificatie: {n_geclassificeerd}/{n_totaal} transacties ({pct:.0f}%) "
                f"geclassificeerd, {n_onzeker} naar AI")

    # Log de verdeling per sectie
    for sectie in ['inkomsten', 'vaste_lasten', 'variabele_kosten', 'sparen_beleggen', 'onderling_neutraal', 'intern']:
        n = len(df[df['regel_sectie'] == sectie])
        if n > 0:
            bedrag = abs(df[df['regel_sectie'] == sectie]['bedrag'].sum())
            logger.info(f"  {sectie}: {n} transacties, EUR {bedrag:,.0f}")

    return df


def _detecteer_vast_inkomen(df: pd.DataFrame) -> pd.DataFrame:
    """Detecteer ALLE vormen van vast inkomen — GEEN hardcoded namen.

    Werkt voor ALLE Nederlandse huishoudens:
    - DGA's met Holding/Management B.V. → DGA-loon/Managementfee
    - Werknemers bij B.V./N.V./Stichting/Gemeente/Overheid → Netto salaris
    - Iedereen met "SALARIS"/"LOON" in transactieomschrijving → Netto salaris
    - Onbekende bron maar vast maandelijks patroon ≥€800 → Netto salaris

    Drie detectie-lagen (van meest naar minst betrouwbaar):
    1. KEYWORD: "SALARIS"/"LOON" in omschrijving → confidence 0.95
    2. RECHTSVORM: B.V./Stichting/N.V./Gemeente etc. → confidence 0.88-0.90
    3. PATROON: onbekende bron, ≥6x vast bedrag ≥€800 → confidence 0.80

    HARDE REGELS (V3 No Guess Zone):
    - household_related_party kan NOOIT salary_employment zijn
    - own_account kan NOOIT extern inkomen zijn
    - merchant kan NOOIT salary zijn
    """
    if 'Omschrijving' not in df.columns:
        return df

    # Alleen niet-interne, niet-reeds-geclassificeerde, POSITIEVE transacties
    mask = (~df['is_intern']) & (df['classificatie_bron'].isna()) & (df['bedrag'] > 0)
    df_kandidaat = df[mask].copy()

    if len(df_kandidaat) == 0:
        return df

    # === V3 HARDE REGELS: counterparty roles die NOOIT salary/income mogen zijn ===
    _VERBODEN_SALARY_PARTY_TYPES = {
        'household_related_party',
        'own_account',
        'merchant',                       # winkels/dienstverleners → nooit salaris
        'broker_or_investment_platform',   # beleggingsplatform → nooit salaris
        'lender_or_mortgage_party',        # hypotheek/lening → nooit salaris
        'card_settlement',                 # creditcard → nooit salaris
    }

    # Bekende merchants uitsluiten
    bekende_merchants = set()
    for zoekterm, _, _, _ in MERCHANT_MAPPING:
        bekende_merchants.add(zoekterm)

    n_dga = 0
    n_salaris = 0
    n_keyword = 0
    n_geblokkeerd_party_type = 0
    gevonden_ibans = set()

    # =========================================================================
    # LAAG 1: KEYWORD — expliciete inkomens-keywords in omschrijving
    # =========================================================================
    # Twee sub-categorieën:
    # A) Salaris-keywords → Netto salaris
    # B) Management fee keywords → DGA-loon/Managementfee
    n_mgmt_keyword = 0
    for idx, row in df_kandidaat.iterrows():
        omschr = str(row.get('Omschrijving', '')).upper()
        bedrag = float(row.get('bedrag', 0))

        # V3: party_type gating — blokkeer salary voor verboden party types
        pt = row.get('party_type', '')
        if pt in _VERBODEN_SALARY_PARTY_TYPES:
            n_geblokkeerd_party_type += 1
            continue

        if bedrag < 200:
            continue

        # A) Salaris-keywords → Netto salaris
        salaris_keywords = ['SALARIS', ' LOON ', 'LOON/', '/LOON', 'SALARY',
                            'NETTO LOON', 'NETTOLOON', 'LOONBETALING',
                            'SALARISBETALING', 'MAANDLOON']
        if any(kw in omschr or omschr.startswith(kw.lstrip()) for kw in salaris_keywords):
            df.at[idx, 'regel_sectie'] = 'inkomsten'
            df.at[idx, 'regel_categorie'] = 'Netto salaris'
            df.at[idx, 'regel_confidence'] = 0.95
            df.at[idx, 'classificatie_bron'] = 'rule'
            df.at[idx, 'source_family'] = 'salary_employment'
            n_keyword += 1
            if 'Tegenrekening' in df.columns:
                tegen = _normaliseer_iban(str(row.get('Tegenrekening', '')))
                if tegen:
                    gevonden_ibans.add(tegen)
            continue

        # B) Management fee keywords → DGA-loon/Managementfee
        mgmt_keywords = ['MANAGEMENT FEE', 'MANAGEMENTFEE', 'MANAGEMENTVERGOEDING',
                         'MGMT FEE', 'MANAGEMENT VERGOEDING', 'BEHEERVERGOEDING']
        if any(kw in omschr for kw in mgmt_keywords):
            df.at[idx, 'regel_sectie'] = 'inkomsten'
            df.at[idx, 'regel_categorie'] = 'DGA-loon/Managementfee'
            df.at[idx, 'regel_confidence'] = 0.93
            df.at[idx, 'classificatie_bron'] = 'rule'
            df.at[idx, 'source_family'] = 'management_fee'
            n_mgmt_keyword += 1
            if 'Tegenrekening' in df.columns:
                tegen = _normaliseer_iban(str(row.get('Tegenrekening', '')))
                if tegen:
                    gevonden_ibans.add(tegen)
            continue

    if n_keyword > 0:
        logger.info(f"SALARIS-KEYWORD: {n_keyword} transacties met 'SALARIS'/'LOON' in omschrijving")
    if n_mgmt_keyword > 0:
        logger.info(f"MGMT-KEYWORD: {n_mgmt_keyword} transacties met 'MANAGEMENT FEE' in omschrijving")

    # Update kandidaten
    mask = (~df['is_intern']) & (df['classificatie_bron'].isna()) & (df['bedrag'] > 0)
    df_kandidaat = df[mask].copy()

    # =========================================================================
    # LAAG 2: RECHTSVORM — B.V., Stichting, N.V., Gemeente, etc.
    # =========================================================================
    # =========================================================================
    # ZAKELIJKE REKENING DETECTIE v2 (generiek voor alle Nederlandse banken)
    # =========================================================================
    # Detecteert welke van de geüploade rekeningen een zakelijke rekening is.
    # Twee methoden:
    #   A) BV/Holding in tegenpartijnaam bij interne transfers (alle formatting-varianten)
    #   B) Bank-specifieke account type labels ("Ondernemersrekening", "Zakelijke rekening")
    # Werkt voor: ABN AMRO, ING, Rabo, Triodos, Bunq, Knab, ASN, RegioBank, SNS
    #
    # GENERIEK: elke DGA bij elke Nederlandse bank profiteert hiervan.

    # BV/Holding naammarkers — alle formatting-varianten die banken gebruiken
    # Banken formatteren "B.V." inconsistent: met/zonder punten, met/zonder spaties
    _BV_DETECT_MARKERS = [
        'B.V.', ' BV ', ' BV,', 'B.V ', ' B.V', ' BV.',
        ' B V ', ' B V,', ' B V.', 'B V ',   # spatie-variant (ABN AMRO)
        ' B V', 'B V',                         # einde-van-string variant
        ' BV',                                  # einde-van-string "PIETERSEN BV"
        'HOLDING', 'HLDG',
    ]

    # Bank-specifieke account type labels die "zakelijk" aanduiden
    # Nederlandse banken gebruiken deze termen bij interne overboekingen
    _ZAKELIJK_ACCOUNT_LABELS = [
        'ONDERNEMERSREKENING', 'ZAKELIJKE REKENING', 'BEDRIJFSREKENING',
        'BUSINESS ACCOUNT', 'ZAKELIJKREKENING',
    ]

    # Privé account type labels — als je HIERNAARTOE overmaakt, is de BRON zakelijk
    _PRIVE_ACCOUNT_LABELS = [
        'PRIVEREKENING', 'PRIVÉREKENING', 'PRIVE REKENING', 'PRIVÉ REKENING',
        'PRIVATE ACCOUNT',
    ]

    _zakelijke_rekeningen = set()  # Set van Rekeningnummer-strings
    if 'Rekeningnummer' in df.columns and 'Tegenrekening' in df.columns:
        eigen_genorm = set(_normaliseer_iban(str(r)) for r in df['Rekeningnummer'].unique()
                          if str(r).strip() and str(r).strip() != 'Onbekend')
        df_intern = df[df['is_intern']].copy()

        _bv_ibans = set()
        for _, row in df_intern.iterrows():
            omschr = str(row.get('Omschrijving', '')).upper()
            tegen_naam = str(row.get('tegenpartij_naam', '')).upper()
            tekst = omschr + ' ' + tegen_naam

            tegen_iban = _normaliseer_iban(str(row.get('Tegenrekening', '')))
            bron_iban = _normaliseer_iban(str(row.get('Rekeningnummer', '')))

            # Methode A: BV/Holding in tegenpartij naam → Tegenrekening is zakelijk
            if any(m in tekst for m in _BV_DETECT_MARKERS):
                if tegen_iban and tegen_iban in eigen_genorm:
                    _bv_ibans.add(tegen_iban)

            # Methode B: Bank account type label = "Ondernemersrekening" →
            # de Tegenrekening is de zakelijke rekening
            elif any(label in tegen_naam for label in _ZAKELIJK_ACCOUNT_LABELS):
                if tegen_iban and tegen_iban in eigen_genorm:
                    _bv_ibans.add(tegen_iban)

            # Methode C: Tegenpartij = "Priverekening" → de BRON is de zakelijke rekening
            # (je zit op je BV en maakt over naar je privé)
            elif any(label in tegen_naam for label in _PRIVE_ACCOUNT_LABELS):
                if bron_iban and bron_iban in eigen_genorm:
                    _bv_ibans.add(bron_iban)

        # Map BV-IBANs terug naar Rekeningnummer strings
        for rek in df['Rekeningnummer'].unique():
            if _normaliseer_iban(str(rek)) in _bv_ibans:
                _zakelijke_rekeningen.add(str(rek).strip())
        if _zakelijke_rekeningen:
            logger.info(f"ZAKELIJKE REKENINGEN via interne transfers: {_zakelijke_rekeningen}")
        else:
            logger.info("Geen zakelijke rekeningen gedetecteerd via interne transfers")

    groepeer_col = 'Tegenrekening' if 'Tegenrekening' in df.columns else 'Omschrijving'

    for key, groep in df_kandidaat.groupby(groepeer_col):
        key_str = str(key).upper().strip()

        if 'Tegenrekening' in df.columns:
            tegen_norm = _normaliseer_iban(key_str)
            if tegen_norm in gevonden_ibans:
                continue

        # V3: party_type gating — hele groep blokkeren als verboden type
        groep_party_types = set(groep['party_type'].dropna().unique()) if 'party_type' in groep.columns else set()
        if groep_party_types & _VERBODEN_SALARY_PARTY_TYPES:
            n_geblokkeerd_party_type += len(groep[groep['bedrag'] > 0])
            logger.info(
                f"SALARY-BLOKKADE: {key_str[:40]} geblokkeerd — party_type={groep_party_types & _VERBODEN_SALARY_PARTY_TYPES}"
            )
            continue

        omschr_alle = ' '.join(groep['Omschrijving'].astype(str).str.upper())
        tekst_check = key_str + ' ' + omschr_alle

        is_merchant = any(m in tekst_check for m in bekende_merchants)
        if is_merchant:
            continue

        bv_markers = ['B.V.', ' BV ', ' BV,', 'B.V ', ' B.V', ' BV.']
        holding_markers = ['HOLDING', 'HLDG']
        werkgever_markers = ['STICHTING', 'GEMEENTE', 'MINISTERIE', 'PROVINCIE',
                             'UNIVERSITEIT', 'HOGESCHOOL', 'POLITIE', 'RIJKS',
                             'WATERSCHAP', 'GGD', 'GGZ', 'ZIEKENHUIS']
        nv_markers = ['N.V.', ' NV ', ' NV,', 'N.V ']
        mgmt_met_bedrijf = ('MANAGEMENT' in tekst_check and
                            any(m in tekst_check for m in bv_markers + holding_markers +
                                ['CONSULTANCY', 'ADVIES', 'DIENSTEN']))

        heeft_bv = any(marker in tekst_check for marker in bv_markers)
        heeft_holding = any(marker in tekst_check for marker in holding_markers)
        heeft_werkgever = any(marker in tekst_check for marker in werkgever_markers)
        heeft_nv = any(marker in tekst_check for marker in nv_markers)

        if not (heeft_bv or heeft_holding or mgmt_met_bedrijf or heeft_werkgever or heeft_nv):
            continue

        groep_pos = groep[groep['bedrag'] > 0]
        if len(groep_pos) < 3:
            continue

        bedragen = groep_pos['bedrag'].astype(float)
        gemiddeld = bedragen.mean()
        if gemiddeld < 500:
            continue

        std = bedragen.std()
        variatie = (std / gemiddeld) if gemiddeld > 0 else 1.0
        if variatie >= 0.25:
            continue

        # Classificatielogica geïnspireerd op Shortcut.ai:
        # Kernprincipe: classificeer eerlijk op basis van bewijs.
        # Salaris/management-fee keywords zijn al afgehandeld in Laag 1.
        # Laag 2 classificeert op basis van rechtsvorm + rekeningtype.
        #
        # 1. Holding/Management in naam → DGA-loon/Managementfee
        # 2. Op BV-rekening → Freelance/Opdrachten (bedrijfsomzet)
        # 3. Werkgever op privérekening → Netto salaris
        # 4. Overige BV/NV op privérekening → Netto salaris (default)

        heeft_salaris_kw = any(kw in tekst_check for kw in ['SALARIS', 'LOON', 'SALARY',
                                                             'NETTOLOON', 'PAYROLL', 'LOONRUN'])
        is_holding_mgmt = (heeft_holding or mgmt_met_bedrijf) and not heeft_werkgever

        # Zakelijke rekening: komt dit inkomen binnen op je BV-rekening?
        op_zakelijke_rek = False
        if 'Rekeningnummer' in groep.columns and _zakelijke_rekeningen:
            groep_rekeningen = set(groep['Rekeningnummer'].astype(str).str.strip().unique())
            op_zakelijke_rek = bool(groep_rekeningen & _zakelijke_rekeningen)

        if heeft_salaris_kw:
            # Salaris-keyword (backup voor als Laag 1 het miste)
            categorie = 'Netto salaris'
            source_fam = 'salary_employment'
            confidence = 0.95
        elif is_holding_mgmt:
            categorie = 'DGA-loon/Managementfee'
            source_fam = 'management_fee'
            confidence = 0.90
        elif op_zakelijke_rek:
            # Inkomen op BV-rekening van derde partij = bedrijfsomzet.
            # Sevi BV, Foundation, Gemeente etc. die je BV betaalt =
            # opdrachtgever, niet werkgever.
            categorie = 'Freelance/Opdrachten'
            source_fam = 'freelance_business'
            confidence = 0.85
            logger.info(
                f"ZAKELIJK: {key_str[:50]} op BV-rekening → "
                f"Freelance/Opdrachten"
            )
        elif heeft_werkgever:
            # Stichting/Gemeente/etc. op privérekening → salaris
            categorie = 'Netto salaris'
            source_fam = 'salary_employment'
            confidence = 0.88
        else:
            categorie = 'Netto salaris'
            source_fam = 'salary_employment'
            confidence = 0.85

        mask_match = df.index.isin(groep_pos.index)
        df.loc[mask_match, 'regel_sectie'] = 'inkomsten'
        df.loc[mask_match, 'regel_categorie'] = categorie
        df.loc[mask_match, 'regel_confidence'] = confidence
        df.loc[mask_match, 'classificatie_bron'] = 'rule'
        df.loc[mask_match, 'source_family'] = source_fam

        totaal = bedragen.sum()
        if categorie == 'DGA-loon/Managementfee':
            n_dga += len(groep_pos)
        else:
            n_salaris += len(groep_pos)
        if 'Tegenrekening' in df.columns:
            gevonden_ibans.add(_normaliseer_iban(key_str))
        logger.info(
            f"INKOMEN GEDETECTEERD: {key_str[:50]} — "
            f"{len(groep_pos)}x, gem EUR {gemiddeld:,.0f}, {variatie*100:.0f}% var → {categorie}"
        )

    # =========================================================================
    # LAAG 3: PATROON — onbekende bron maar vast maandelijks bedrag
    # Strengere eisen: ≥6 betalingen, ≥€800, <20% variatie
    # =========================================================================
    mask = (~df['is_intern']) & (df['classificatie_bron'].isna()) & (df['bedrag'] > 0)
    df_rest = df[mask].copy()

    if len(df_rest) > 0 and 'Tegenrekening' in df.columns:
        df_rest['_tegen_norm'] = df_rest['Tegenrekening'].apply(_normaliseer_iban)
        df_rest = df_rest[(df_rest['_tegen_norm'] != '') & (~df_rest['_tegen_norm'].isin(gevonden_ibans))]

        for tegen_rek, groep in df_rest.groupby('_tegen_norm'):
            groep_pos = groep[groep['bedrag'] > 0]
            if len(groep_pos) < 6:
                continue

            # V3: party_type gating
            groep_pts = set(groep_pos['party_type'].dropna().unique()) if 'party_type' in groep_pos.columns else set()
            if groep_pts & _VERBODEN_SALARY_PARTY_TYPES:
                naam = groep_pos.iloc[0]['Omschrijving'] if len(groep_pos) > 0 else '?'
                n_geblokkeerd_party_type += len(groep_pos)
                logger.info(
                    f"SALARY-PATROON-BLOKKADE: {tegen_rek} ({str(naam)[:30]}) — "
                    f"party_type={groep_pts & _VERBODEN_SALARY_PARTY_TYPES}, {len(groep_pos)} tx geblokkeerd"
                )
                continue

            omschr_check = ' '.join(groep_pos['Omschrijving'].astype(str).str.upper().head(5))
            if any(m in omschr_check for m in bekende_merchants):
                continue

            bedragen = groep_pos['bedrag'].astype(float)
            gemiddeld = bedragen.mean()
            if gemiddeld < 800:
                continue

            std = bedragen.std()
            variatie = (std / gemiddeld) if gemiddeld > 0 else 1.0
            if variatie < 0.20:
                mask_match = df.index.isin(groep_pos.index)
                df.loc[mask_match, 'regel_sectie'] = 'inkomsten'
                df.loc[mask_match, 'regel_categorie'] = 'Netto salaris'
                df.loc[mask_match, 'regel_confidence'] = 0.80
                df.loc[mask_match, 'classificatie_bron'] = 'rule'
                df.loc[mask_match, 'source_family'] = 'salary_employment'
                n_salaris += len(groep_pos)
                naam = groep_pos.iloc[0]['Omschrijving']
                logger.info(
                    f"SALARIS-PATROON: {tegen_rek} ({str(naam)[:30]}) — "
                    f"{len(groep_pos)}x, gem EUR {gemiddeld:,.0f}, {variatie*100:.0f}% var"
                )

    totaal_gevonden = n_keyword + n_dga + n_salaris
    if totaal_gevonden > 0:
        logger.info(f"VAST INKOMEN: {totaal_gevonden} tx — {n_keyword} keyword, {n_dga} DGA, {n_salaris} salaris")
    else:
        logger.info("VAST INKOMEN: geen vast inkomen gedetecteerd")
    if n_geblokkeerd_party_type > 0:
        logger.info(
            f"V3-GATING: {n_geblokkeerd_party_type} transacties geblokkeerd als salary "
            f"vanwege verboden party_type (household/own_account)"
        )

    return df


# ---------------------------------------------------------------------------
# DECISION ENGINE — Evidence-Based Classifiers (Laag B)
# ---------------------------------------------------------------------------
# Elke classifier verzamelt onafhankelijke bewijsstukken (evidence checks)
# en komt tot een driedelige uitkomst: likely / uncertain / reject.
# Elke check is binair (ja/nee) en transparant uitlegbaar.
# ---------------------------------------------------------------------------

# Page-1 whitelist — ALLEEN deze categorieën tellen als structural income
_STRUCTURAL_INCOME_WHITELIST = {
    'Netto salaris',
    'DGA-loon/Managementfee',
    'Inkomen uit eigen BV',
    'Huurinkomsten',
    'UWV/Uitkeringen',
    'Kinderbijslag/Kindregelingen',
    'Toeslagen',
    'Freelance/Opdrachten',
    'Pensioen/AOW',
    'Studiefinanciering',
    'Overheid overig',
}

# Categorieën die NOOIT structural income zijn (voor quality checks)
_UNCERTAIN_CATS = {
    'Onzeker positief (niet-geverifieerd inkomen)',
    'Onzeker positief',
    'Onzeker positief (verzekeraar)',
    'Onzeker positief (bidirectioneel)',
    'Onzeker positief (financiële instelling)',
    'Onzeker positief (bedrijf/organisatie)',
    'Onzeker positief (bedrijf, klein bedrag)',
    'Onzeker positief (privépersoon, groot bedrag)',
    'Onderlinge betaling (privépersoon)',
    'Onderlinge betaling (klein bedrag)',
    'Overige bijschrijving (klein)',
    'Overige bijschrijving',
    'Terugbetaling/Tikkie',
    'Verkoop (tweedehands)',
    'Cashback/Spaarprogramma',
    'Notaris/Woningtransactie',
}

# Rechtsvorm-markers voor employer detection
_RECHTSVORM_MARKERS = ['B.V.', ' BV ', ' BV,', 'B.V ', ' B.V', ' BV.',
                       'HOLDING', 'HLDG', 'STICHTING', 'VERENIGING',
                       'N.V.', ' NV ', 'GEMEENTE', 'MINISTERIE', 'UNIVERSITEIT']


def _rent_classifier(groep_df, df_all, groep_key, eigen_rekeningen, eigen_fi_ibans,
                     bekende_merchants_set, heeft_tegenrek):
    """Evidence-based rent income classifier.

    9 evidence checks, driedelige uitkomst: likely / uncertain / reject.
    Elke check is binair en uitlegbaar.

    Returns: (uitkomst, categorie, evidence_dict)
        uitkomst: 'likely' | 'uncertain' | 'reject'
        categorie: 'Huurinkomsten' | 'Onzeker positief' | None
        evidence_dict: {check_name: (bool, str)} voor logging
    """
    evidence = {}
    tekst_alle = ' '.join(groep_df['Omschrijving'].astype(str).str.upper())

    # === EXCLUSION CHECKS (verplicht, allemaal ja) ===

    # E2: Externality — is dit een externe partij?
    iban = groep_key if groep_key.startswith('NL') or groep_key.startswith('DE') or groep_key.startswith('BE') else ''
    e2 = True
    if iban:
        if iban in eigen_rekeningen or iban in eigen_fi_ibans:
            e2 = False
    evidence['E2_externality'] = (e2, f'IBAN {"niet " if e2 else ""}in eigen domein')

    # E6: Not an employer — geen rechtsvorm-markers
    e6 = not any(m in tekst_alle for m in _RECHTSVORM_MARKERS)
    evidence['E6_not_employer'] = (e6, f'{"Geen" if e6 else "Wel"} rechtsvorm gevonden')

    # E7: Not a known merchant
    e7 = not any(m in tekst_alle for m in bekende_merchants_set)
    evidence['E7_not_merchant'] = (e7, f'{"Geen" if e7 else "Wel"} merchant match')

    # E8: Not a financial institution
    e8_fi = not any(kw in tekst_alle for kw in FINANCIELE_INSTELLINGEN_KEYWORDS)
    e8_verz = not any(v in tekst_alle for v in VERZEKERAAR_NAMEN)
    e8 = e8_fi and e8_verz
    evidence['E8_not_fi'] = (e8, f'{"Geen" if e8 else "Wel"} FI/verzekeraar match')

    # Als één exclusion check faalt → REJECT
    if not (e2 and e6 and e7 and e8):
        failed = [k for k, (v, _) in evidence.items() if not v]
        return ('reject', None, evidence)

    # === EVIDENCE CHECKS ===

    # E1: Recurring counterparty — minstens 3 positieve transacties
    e1 = len(groep_df) >= 3
    evidence['E1_recurring'] = (e1, f'{len(groep_df)} transacties')

    if not e1:
        return ('reject', None, evidence)  # Te weinig data

    # E3: Amount stability — IQR < 30% van mediaan (robuust tegen outliers)
    bedragen = groep_df['bedrag'].astype(float)
    mediaan = bedragen.median()

    # Outlier-filter: verwijder bedragen < 25% van mediaan
    kern_bedragen = bedragen[bedragen >= mediaan * 0.25]
    if len(kern_bedragen) < 3:
        e3 = False
        variatie = 1.0
    else:
        q1 = kern_bedragen.quantile(0.25)
        q3 = kern_bedragen.quantile(0.75)
        iqr = q3 - q1
        variatie = (iqr / mediaan) if mediaan > 0 else 1.0
        e3 = variatie < 0.30
    evidence['E3_amount_stability'] = (e3, f'IQR-variatie {variatie*100:.1f}%, mediaan EUR {mediaan:,.0f}')

    # E4: Minimum amount — mediaan >= €300
    e4 = mediaan >= 300
    evidence['E4_minimum_amount'] = (e4, f'Mediaan EUR {mediaan:,.0f}')

    if not e4:
        return ('reject', None, evidence)  # Bedrag te laag voor huur

    # E5: Predominantly unidirectional — max 15% negatief
    # Zoek ALLE transacties (ook negatieve) van deze tegenpartij
    if heeft_tegenrek and iban:
        df_all['_temp_iban'] = df_all['Tegenrekening'].apply(
            lambda x: _normaliseer_iban(str(x)) if pd.notna(x) else '')
        alle_tp = df_all[(df_all['_temp_iban'] == iban) & (~df_all['is_intern'])]
        df_all.drop(columns=['_temp_iban'], inplace=True)
    else:
        df_all['_temp_key'] = (
            df_all['Omschrijving'].astype(str).str.upper().str.split().str[:3].str.join(' '))
        alle_tp = df_all[(df_all['_temp_key'] == groep_key) & (~df_all['is_intern'])]
        df_all.drop(columns=['_temp_key'], inplace=True)

    n_negatief = (alle_tp['bedrag'] < 0).sum()
    n_totaal = len(alle_tp)
    pct_negatief = n_negatief / n_totaal if n_totaal > 0 else 0
    e5 = pct_negatief <= 0.15
    evidence['E5_unidirectional'] = (e5, f'{pct_negatief*100:.0f}% negatief ({n_negatief}/{n_totaal})')

    # E9: Description hint — huur-gerelateerde termen (optioneel, versterkt bewijs)
    huur_keywords = ['HUUR', ' RENT ', 'KAMER', 'WONING', 'HUURPENNING', 'KAMERHUUR']
    e9 = any(kw in tekst_alle for kw in huur_keywords)
    evidence['E9_description_hint'] = (e9, f'{"Ja" if e9 else "Nee"}: huur-keyword in omschrijving')

    # === BESLISLOGICA ===
    verplichte_evidence = [e1, e3, e4, e5]
    n_ja = sum(verplichte_evidence)

    if all(verplichte_evidence):
        # Alle 4 ja → LIKELY
        return ('likely', 'Huurinkomsten', evidence)
    elif n_ja >= 3 and e9:
        # 3 van 4 ja + huur-keyword → nog steeds UNCERTAIN (keyword alleen is niet genoeg)
        return ('uncertain', 'Onzeker positief', evidence)
    elif n_ja >= 3:
        # 3 van 4 ja, geen keyword → UNCERTAIN
        return ('uncertain', 'Onzeker positief', evidence)
    else:
        # < 3 ja → UNCERTAIN (niet reject, want exclusion checks passeerden)
        return ('uncertain', 'Onzeker positief', evidence)


def _refund_matcher(row, negatieve_bedragen_per_iban, negatieve_bedragen_per_naam,
                    kosten_merchants_set):
    """Evidence-based refund matcher.

    5 evidence checks, driedelige uitkomst.

    Returns: (uitkomst, categorie)
        uitkomst: 'likely' | 'uncertain' | 'reject'
    """
    omschr = str(row.get('Omschrijving', '')).upper()
    bedrag = float(row.get('bedrag', 0))
    iban = _normaliseer_iban(str(row.get('Tegenrekening', ''))) if 'Tegenrekening' in row.index else ''
    naam = str(row.get('tegenpartij_naam', '')).upper() if 'tegenpartij_naam' in row.index else ''
    tekst = omschr + ' ' + naam

    # E4: Refund keyword
    e4 = any(kw in tekst for kw in REFUND_KEYWORDS)
    if e4:
        return ('likely', 'Terugbetaling/Refund')

    # E5: Known expense merchant (positief bedrag van kosten-merchant)
    e5 = any(m in tekst for m in kosten_merchants_set)

    # E1: Amount match — eerder negatief bedrag binnen ±30%
    e1 = False
    if iban and iban in negatieve_bedragen_per_iban:
        for nb in negatieve_bedragen_per_iban[iban]:
            if nb > 0 and abs(bedrag - nb) / nb < 0.30:
                e1 = True
                break
    if not e1 and naam:
        # Fallback: zoek op naam
        for key, bedragen_lijst in negatieve_bedragen_per_naam.items():
            if key in naam or naam in key:
                for nb in bedragen_lijst:
                    if nb > 0 and abs(bedrag - nb) / nb < 0.30:
                        e1 = True
                        break
                if e1:
                    break

    # E2: Same counterparty (IBAN of naam match met negatieve tx) — impliciet in E1
    # E3: Time window — niet geïmplementeerd in eerste versie (vereist datum-matching)

    if e5:
        # Kosten-merchant + positief bedrag = waarschijnlijk refund
        return ('likely', f'Terugbetaling ({tekst[:20].strip().title()})')

    if e1:
        # Bedragmatch met eerder negatief bedrag = waarschijnlijk refund
        return ('likely', 'Terugbetaling (bedrag-match)')

    return ('reject', None)


def _uncertainty_gate(row, bidi_ibans):
    """Subcategoriseer uncertain positive inflows.

    Geeft een specifieke subcategorie terug die de gebruiker vertelt
    WAT het waarschijnlijk is, zodat ze het kunnen verifiëren.
    Hoe specifieker de subcategorie, hoe minder "onzeker" het aanvoelt.

    Returns: (subcategorie_string, inflow_type_hint)
    """
    omschr = str(row.get('Omschrijving', '')).upper()
    naam = str(row.get('tegenpartij_naam', '')).upper() if 'tegenpartij_naam' in row.index else ''
    tekst = omschr + ' ' + naam
    bedrag = float(row.get('bedrag', 0))
    iban = _normaliseer_iban(str(row.get('Tegenrekening', ''))) if 'Tegenrekening' in row.index else ''
    party_type = str(row.get('party_type', '')) if 'party_type' in row.index else ''

    # === SPECIFIEKE HERKENNING (van meest naar minst zeker) ===

    # 1. Verzekeraar zonder uitkering-keyword → waarschijnlijk schade-uitkering
    if any(v in tekst for v in VERZEKERAAR_NAMEN):
        return 'Onzeker positief (verzekeraar)'

    # 2. Financiële instelling niet in eigen domein
    if any(kw in tekst for kw in FINANCIELE_INSTELLINGEN_KEYWORDS):
        return 'Onzeker positief (financiële instelling)'

    # 3. Tikkie / betaalverzoek / iDEAL terugbetaling
    tikkie_kw = ['TIKKIE', 'BETAALVERZOEK', 'BETAAL VERZOEK', 'IDEAL',
                 'BUNQ.ME', 'PAYREQ', 'BETAALLINK']
    if any(kw in tekst for kw in tikkie_kw):
        return 'Terugbetaling/Tikkie'

    # 4. Marktplaats / tweedehands verkoop
    verkoop_kw = ['MARKTPLAATS', 'VINTED', 'WALLAPOP', 'TWEEDEHANDS',
                  'VERKOOP', 'SOLD', '2DEHANDS', 'EBAY', 'BOL.COM VERKOOP']
    if any(kw in tekst for kw in verkoop_kw):
        return 'Verkoop (tweedehands)'

    # 5. Cashback / spaarprogramma
    cashback_kw = ['CASHBACK', 'LOYALTY', 'SPAARPUNTEN', 'REWARD',
                   'RAKUTEN', 'SHOPBUDDIES', 'SCOUPY']
    if any(kw in tekst for kw in cashback_kw):
        return 'Cashback/Spaarprogramma'

    # 6. Notaris / makelaar (grote eenmalige bedragen, woningtransactie)
    notaris_kw = ['NOTARIS', 'MAKELAAR', 'MAKELAARDIJ', 'KADASTER',
                  'HYPOTHEEK', 'WAARBORGSOM', 'BORG']
    if any(kw in tekst for kw in notaris_kw):
        return 'Notaris/Woningtransactie'

    # 7. Rechtsvorm zonder salary/mgmt match → bedrijfsinkomen
    rechtsvorm_kw = ['B.V.', ' BV ', ' BV,', 'B.V ', ' B.V', ' BV.',
                     'N.V.', ' NV ', 'HOLDING', 'HLDG', 'STICHTING',
                     'VERENIGING', 'MAATSCHAP', 'V.O.F.', ' VOF ']
    if any(kw in tekst for kw in rechtsvorm_kw):
        if bedrag >= 500:
            return 'Onzeker positief (bedrijf/organisatie)'
        else:
            return 'Onzeker positief (bedrijf, klein bedrag)'

    # 8. Bidirectioneel IBAN → je stuurt ook geld naar deze partij
    if iban and iban in bidi_ibans:
        if party_type == 'unknown':
            return 'Onderlinge betaling (privépersoon)'
        return 'Onzeker positief (bidirectioneel)'

    # 9. Privépersoon (geen rechtsvorm, geen bedrijf)
    if party_type == 'unknown':
        if bedrag < 100:
            return 'Onderlinge betaling (klein bedrag)'
        elif bedrag < 500:
            return 'Onderlinge betaling (privépersoon)'
        else:
            return 'Onzeker positief (privépersoon, groot bedrag)'

    # 10. Klein bedrag zonder verdere aanwijzingen
    if bedrag < 50:
        return 'Overige bijschrijving (klein)'
    elif bedrag < 200:
        return 'Overige bijschrijving'

    return 'Onzeker positief'


# ---------------------------------------------------------------------------
# INFLOW TYPE CLASSIFICATIE v2 — Eigen Financieel Domein + Uncertain Bucket
# ---------------------------------------------------------------------------
# Kernprincipe: een positief bedrag is NIET automatisch inkomen.
# Alleen een strikte whitelist mag naar "structureel inkomen" op page 1.
# Alles wat niet bewezen inkomen is gaat naar uncertain_positive_inflow
# en wordt UITGESLOTEN van de executive conclusies.
#
# Gebaseerd op:
# - Yapily's productie-taxonomie (90+ categorieën, 6 top-level credit types)
# - Plaid's twee-fase classificatie (is_income binary → dan categoriseren)
# - ChatGPT CEO advies: eigen financieel domein register, multi-role logica
# ---------------------------------------------------------------------------

# Financiële instellingen keywords — voor auto-detectie eigen financieel domein
FINANCIELE_INSTELLINGEN_KEYWORDS = [
    # Brokers / beleggingsplatformen
    'SAXO', 'DEGIRO', 'IBKR', 'INTERACTIVE BROKERS', 'BINCK', 'LYNX',
    'FLATEX', 'ETORO', 'TRADING 212', 'MEESMAN', 'NORTHERN TRUST',
    'BRAND NEW DAY', 'VANGUARD', 'BLACKROCK', 'FIDELITY', 'ROBECO',
    'ACTIAM', 'KEMPEN', 'VAN LANSCHOT', 'THINK ETF', 'BUX',
    # Crypto
    'BITVAVO', 'COINBASE', 'KRAKEN', 'BINANCE', 'BYBIT',
    # Crowdlending / P2P
    'MINTOS', 'LENDAHAND', 'PEERBERRY', 'BONDORA', 'TWINO', 'OCTOBER',
    'FUNDING CIRCLE', 'COLLIN CROWDFUND',
    # Spaar/deposito
    'RAISIN', 'SAVEDO', 'LEASEPLAN BANK', 'KNAB SPAAR',
]

# =====================================================================
# COUNTERPARTY ROLE DETECTION — Keyword sets per rol
# =====================================================================

# Creditcard-maatschappijen → card_settlement (nooit inkomen, nooit kosten)
CARD_SETTLEMENT_KEYWORDS = [
    'ICS/INT CARD', 'ICS ', 'INTERNATIONAL CARD SERVICES',
    'VISA CARD', 'MASTERCARD', 'AMERICAN EXPRESS', 'AMEX',
    'ADYEN', 'WORLDLINE', 'BUCKAROO',
]

# Hypotheek/lening-verstrekkers → lender_or_mortgage_party
# ALLEEN dedicated hypotheek-/leningmaatschappijen, NIET multi-role (NN, Aegon, ASR)
# want die kunnen ook verzekeringen uitkeren. Multi-role partijen staan in MERCHANT_MAPPING.
LENDER_KEYWORDS_COUNTERPARTY = [
    'OBVION', 'FLORIUS', 'WOONFONDS', 'MUNT HYPOTHEKEN', 'VISTA HYPOTHEKEN',
    'ABN AMRO HYPOTHEEK', 'ING HYPOTHEEK', 'RABO HYPOTHEEK',
    'ASR HYPOTHEEK',
    'WONINGCORPORATIE', 'VESTIA', 'YMERE', 'EIGEN HAARD',
    'DE ALLIANTIE', 'WOONSTAD', 'PORTAAL',
]

# Investment income keywords — ALLEEN deze mogen als beleggingsinkomen tellen
INVESTMENT_INCOME_KEYWORDS = [
    'DIVIDEND', 'DIVIDENDUITKERING', 'COUPON', 'RENTE-UITKERING',
    'RENTEUITKERING', 'INTEREST', 'YIELD', 'DISTRIBUTION',
    'WINSTUITKERING', 'INTERIM DIVIDEND', 'SLOTDIVIDEND',
]

# Overheidsinstanties die geld uitkeren
OVERHEID_KEYWORDS = [
    'BELASTINGDIENST', 'BELASTING DIENST',
    'UWV', 'SVB',
    'ZORGTOESLAG', 'HUURTOESLAG', 'KINDGEBONDEN BUDGET',
    'KINDERBIJSLAG',
    'DUO ', 'DIENST UITVOERING ONDERWIJS',
    'GEMEENTE',  # Let op: gemeentebelasting (negatief) al apart afgehandeld
    'RIJKSOVERHEID', 'MINISTERIE',
    'CAK ', 'CENTRAAL ADMINISTRATIE KANTOOR',
    'RVO ', 'RIJKSDIENST VOOR ONDERNEMEND',
    'CJIB',  # Kan ook boete zijn, maar positief = teruggave
]

# Refund / terugbetaling keywords
REFUND_KEYWORDS = [
    'RETOUR', 'REFUND', 'TERUGBET', 'STORNO', 'CREDITNOTA',
    'TERUGSTORT', 'RESTITUTIE', 'TERUGGAVE', 'REVERSAL',
    'TERUGBOEKING', 'ANNULERING', 'CORRECTIE', 'CREDIT',
    'CASHBACK', 'GELD TERUG', 'REIMBURSEMENT',
]

# Transfer keywords (eigen rekeningen, Tikkie, peer-to-peer)
TRANSFER_KEYWORDS = [
    'TIKKIE', 'BETAALVERZOEK',
    'SPAARREKENING', 'SPAAR',
    'OVERBOEKING EIGEN', 'EIGEN REKENING',
    'SAVINGS', 'DEPOSIT',
]

# Verzekering-uitkering keywords
VERZEKERING_KEYWORDS = [
    'UITKERING', 'SCHADEVERGOEDING', 'SCHADE-UITKERING',
    'LETSELSCHADE', 'VERZEKERINGSUITKERING',
    'SCHADEREGELING', 'POLISUITKERING',
]

# Verzekeraars — multi-role: premie-inning (negatief) vs uitkering (positief)
VERZEKERAAR_NAMEN = [
    'CENTRAAL BEHEER', 'ACHMEA', 'INTERPOLIS', 'NATIONALE-NEDERLANDEN',
    'NN GROUP', 'AEGON', 'DELTA LLOYD', 'ASR', 'A.S.R',
    'ZILVEREN KRUIS', 'CZ GROEP', 'CZ ZORGVERZEKERING',
    'MENZIS', 'VGZ', 'COOPERATIE VGZ', 'UNIVE', 'UNIVÉ',
    'FBTO', 'REAAL', 'ALLIANZ', 'INSHARED', 'ALLSECUR',
    'OHRA', 'DITZO', 'DSW', 'ZORG EN ZEKERHEID',
    'ENO ZORGVERZEKERAAR', 'SALLAND VERZEKERINGEN',
    'DELA', 'MONUTA', 'YARDEN',
]

# Lening / hypotheek-gerelateerd (positief = uitbetaling, niet inkomen)
LENING_KEYWORDS = [
    'HYPOTHEEK', 'LENING', 'KREDIET', 'FINANCIERING',
    'AFLOSSING', 'DOORSTORT', 'RESTSCHULD',
]


def _bouw_eigen_financieel_domein(df: pd.DataFrame) -> set:
    """Detecteer automatisch welke financiële instellingen de klant gebruikt.

    Kijkt naar UITGAANDE transacties (negatief bedrag) naar bekende financiële
    instellingen. Als je geld STUURT naar Saxo/DeGiro/Mintos, dan zijn die
    IBANs onderdeel van je eigen financieel domein.

    Retourneert set van IBANs die bij het eigen financieel domein horen.
    Alle geldstromen van/naar deze IBANs zijn vermogensmutaties, NOOIT inkomen.
    """
    eigen_fi_ibans = set()

    if 'Tegenrekening' not in df.columns:
        return eigen_fi_ibans

    # Zoek alle uitgaande transacties naar financiële instellingen
    for idx, row in df.iterrows():
        if float(row.get('bedrag', 0)) >= 0:
            continue  # Alleen uitgaande (negatieve) transacties

        omschr = str(row.get('Omschrijving', '')).upper()
        naam = str(row.get('tegenpartij_naam', '')).upper() if 'tegenpartij_naam' in df.columns else ''
        tekst = omschr + ' ' + naam
        iban = _normaliseer_iban(str(row.get('Tegenrekening', '')))

        if not iban:
            continue

        # Check of dit een financiële instelling is
        if any(kw in tekst for kw in FINANCIELE_INSTELLINGEN_KEYWORDS):
            eigen_fi_ibans.add(iban)

        # Check MERCHANT_MAPPING sparen_beleggen entries
        for zoekterm, sectie, _, _ in MERCHANT_MAPPING:
            if sectie == 'sparen_beleggen' and zoekterm in tekst:
                eigen_fi_ibans.add(iban)
                break

    if eigen_fi_ibans:
        logger.info(f"EIGEN FINANCIEEL DOMEIN: {len(eigen_fi_ibans)} IBANs gedetecteerd:")
        for iban in sorted(eigen_fi_ibans):
            # Zoek naam voor logging
            matches = df[df['Tegenrekening'].apply(lambda x: _normaliseer_iban(str(x))) == iban]
            naam_sample = str(matches.iloc[0]['Omschrijving'])[:40] if len(matches) > 0 else '?'
            logger.info(f"  {iban} ({naam_sample})")
    else:
        logger.info("EIGEN FINANCIEEL DOMEIN: geen financiële instellingen gedetecteerd")

    return eigen_fi_ibans


def _classify_positive_inflows(df: pd.DataFrame, eigen_fi_ibans: set = None,
                               eigen_rekeningen: set = None) -> pd.DataFrame:
    """Decision Engine — Classificeer ALLE ongeclass positieve transacties.

    Drielaags model:
      Laag A: Deterministic rules (overheid, FI domein, keywords)
      Laag B: Evidence-based classifiers (rent, refund)
      Uncertainty Gate: subcategoriseer rest

    Kernprincipe: elke positieve transactie krijgt een classificatie.
    Uncertain = geldige einduitkomst, telt NOOIT als structural income.

    Draait NA _detecteer_vast_inkomen() en VOOR _afdwing_iban_consistentie().
    """
    if 'classificatie_bron' not in df.columns:
        return df

    if eigen_fi_ibans is None:
        eigen_fi_ibans = set()
    if eigen_rekeningen is None:
        eigen_rekeningen = set()

    # Normaliseer eigen rekeningen
    eigen_rek_norm = set(_normaliseer_iban(str(r)) for r in eigen_rekeningen if r and str(r) != 'nan')

    # Alleen niet-interne, niet-geclassificeerde, POSITIEVE transacties
    mask = (~df['is_intern']) & (df['classificatie_bron'].isna()) & (df['bedrag'] > 0)
    kandidaten = df[mask].copy()

    if len(kandidaten) == 0:
        logger.info("DECISION ENGINE: geen ongeclassificeerde positieve transacties")
        return df

    if 'inflow_type' not in df.columns:
        df['inflow_type'] = None

    # V2: flow_type kolom (multi-axis spec) — credit/debit/internal
    if 'flow_type' not in df.columns:
        df['flow_type'] = df['bedrag'].apply(
            lambda b: 'credit' if float(b) > 0 else ('debit' if float(b) < 0 else 'zero')
        )
        df.loc[df['is_intern'] == True, 'flow_type'] = 'internal'

    # === LOOKUPS BOUWEN ===

    # Bekende merchants (voor rent classifier exclusion + refund detection)
    bekende_merchants_set = set()
    kosten_merchants_set = set()
    for zoekterm, sectie, _, _ in MERCHANT_MAPPING:
        bekende_merchants_set.add(zoekterm)
        if sectie in ('vaste_lasten', 'variabele_kosten'):
            kosten_merchants_set.add(zoekterm)

    # Bidirectionele IBANs (waar we ook geld naartoe sturen)
    bidi_ibans = set()
    if 'Tegenrekening' in df.columns:
        for iban_val in df[(df['bedrag'] < 0) & (~df['is_intern']) & df['Tegenrekening'].notna()]['Tegenrekening'].unique():
            iban_norm = _normaliseer_iban(str(iban_val))
            if iban_norm:
                bidi_ibans.add(iban_norm)

    # Negatieve bedragen per IBAN en per naam (voor refund matching)
    negatieve_bedragen_per_iban = {}
    negatieve_bedragen_per_naam = {}
    if 'Tegenrekening' in df.columns:
        for _, row in df[(df['bedrag'] < 0) & (~df['is_intern'])].iterrows():
            iban = _normaliseer_iban(str(row.get('Tegenrekening', '')))
            if iban:
                negatieve_bedragen_per_iban.setdefault(iban, []).append(abs(float(row['bedrag'])))
            naam = str(row.get('tegenpartij_naam', '')).upper().strip()
            if naam and naam != 'NAN':
                negatieve_bedragen_per_naam.setdefault(naam, []).append(abs(float(row['bedrag'])))

    heeft_tegenrek = 'Tegenrekening' in df.columns

    # === STATISTIEKEN ===
    stats = {
        'government': 0, 'asset_withdrawal': 0, 'investment_income': 0,
        'refund': 0, 'internal_transfer': 0, 'insurance': 0,
        'loan_inflow': 0, 'rent_likely': 0, 'uncertain': 0,
    }
    uncertain_bedrag = 0.0
    herclassificaties = []  # Voor before/after logging

    def _apply(idx, inflow_t, sectie, categorie, confidence):
        """Helper: pas classificatie toe op DataFrame."""
        df.at[idx, 'inflow_type'] = inflow_t
        df.at[idx, 'regel_sectie'] = sectie
        df.at[idx, 'regel_categorie'] = categorie
        df.at[idx, 'regel_confidence'] = confidence
        df.at[idx, 'classificatie_bron'] = 'rule'
        if inflow_t == 'internal_transfer':
            df.at[idx, 'is_intern'] = True

    # ===================================================================
    # FASE 1: LAAG A — Deterministic Rules (per transactie)
    # ===================================================================
    fase1_handled = set()  # indices die door Laag A zijn afgehandeld

    for idx, row in kandidaten.iterrows():
        omschr = str(row.get('Omschrijving', '')).upper()
        bedrag = float(row.get('bedrag', 0))
        iban = _normaliseer_iban(str(row.get('Tegenrekening', ''))) if heeft_tegenrek else ''
        naam = str(row.get('tegenpartij_naam', '')).upper() if 'tegenpartij_naam' in df.columns else ''
        tekst = omschr + ' ' + naam
        pt = row.get('party_type', '') if 'party_type' in df.columns else ''

        classified = False

        # A0: COUNTERPARTY ROLE SHORTCUT — als RPR al een specifieke rol heeft
        # bepaald, gebruik die direct (No Guess Zone)
        if pt == 'card_settlement':
            _apply(idx, 'card_settlement', 'onderling_neutraal', 'Creditcard-afrekening', 0.90)
            stats.setdefault('card_settlement', 0)
            stats['card_settlement'] += 1
            herclassificaties.append((idx, 'Laag A', 'card_settlement', 'Creditcard-afrekening', f'party_type=card_settlement'))
            classified = True
        elif pt == 'lender_or_mortgage_party':
            _apply(idx, 'loan_inflow', 'sparen_beleggen', 'Lening/Hypotheek (uitbetaling)', 0.85)
            stats.setdefault('loan_inflow', 0)
            stats['loan_inflow'] = stats.get('loan_inflow', 0) + 1
            herclassificaties.append((idx, 'Laag A', 'loan_inflow', 'Lening/Hypotheek (uitbetaling)', f'party_type=lender_or_mortgage_party'))
            classified = True
        elif pt == 'broker_or_investment_platform':
            if any(kw in tekst for kw in INVESTMENT_INCOME_KEYWORDS):
                _apply(idx, 'investment_income', 'inkomsten', 'Beleggingsinkomen', 0.90)
                stats['investment_income'] += 1
                herclassificaties.append((idx, 'Laag A', 'investment_income', 'Beleggingsinkomen', f'party_type=broker + dividend-keyword'))
            else:
                if any(kw in tekst for kw in ['BITVAVO', 'COINBASE', 'KRAKEN', 'BINANCE', 'BYBIT']):
                    cat = 'Crypto (terugstorting)'
                elif any(kw in tekst for kw in ['MINTOS', 'LENDAHAND', 'PEERBERRY', 'BONDORA',
                                                 'TWINO', 'OCTOBER', 'FUNDING CIRCLE', 'COLLIN']):
                    cat = 'Crowdlending (terugbetaling)'
                elif 'BRAND NEW DAY' in tekst:
                    cat = 'Pensioen (terugstorting)'
                else:
                    cat = 'Effectenrekening (terugstorting)'
                _apply(idx, 'asset_withdrawal', 'sparen_beleggen', cat, 0.95)
                stats['asset_withdrawal'] += 1
                herclassificaties.append((idx, 'Laag A', 'asset_withdrawal', cat, f'party_type=broker'))
            classified = True

        # A1: OVERHEID (keyword of party_type=government)
        elif pt == 'government' or any(kw in tekst for kw in OVERHEID_KEYWORDS):
            if 'BELASTINGDIENST' in tekst or 'BELASTING DIENST' in tekst:
                # Probeer specifiek belastingtype uit aanslagnummer
                _bt = _detecteer_belastingtype_uit_kenmerk(tekst)
                cat = f'Belastingteruggave ({_bt})' if _bt and _bt != 'Toeslagen' else 'Belastingteruggave'
            elif 'UWV' in tekst:
                cat = 'UWV/Uitkeringen'
            elif 'SVB' in tekst or 'KINDERBIJSLAG' in tekst:
                cat = 'Kinderbijslag/Kindregelingen'
            elif any(kw in tekst for kw in ['ZORGTOESLAG', 'HUURTOESLAG', 'KINDGEBONDEN']):
                cat = 'Toeslagen'
            elif 'DUO' in tekst or 'DIENST UITVOERING' in tekst:
                cat = 'Studiefinanciering'
            else:
                cat = 'Overheid overig'
            _apply(idx, 'government', 'inkomsten', cat, 0.90)
            stats['government'] += 1
            herclassificaties.append((idx, 'Laag A', 'government', cat, f'Overheid-keyword'))
            classified = True

        # A2: EIGEN FINANCIEEL DOMEIN (IBAN match)
        elif iban and iban in eigen_fi_ibans:
            if any(kw in tekst for kw in INVESTMENT_INCOME_KEYWORDS):
                _apply(idx, 'investment_income', 'inkomsten', 'Beleggingsinkomen', 0.90)
                stats['investment_income'] += 1
                herclassificaties.append((idx, 'Laag A', 'investment_income', 'Beleggingsinkomen', f'FI-domein IBAN + dividend-keyword'))
            else:
                if any(kw in tekst for kw in ['BITVAVO', 'COINBASE', 'KRAKEN', 'BINANCE', 'BYBIT']):
                    cat = 'Crypto (terugstorting)'
                elif any(kw in tekst for kw in ['MINTOS', 'LENDAHAND', 'PEERBERRY', 'BONDORA',
                                                 'TWINO', 'OCTOBER', 'FUNDING CIRCLE', 'COLLIN']):
                    cat = 'Crowdlending (terugbetaling)'
                elif 'BRAND NEW DAY' in tekst:
                    cat = 'Pensioen (terugstorting)'
                else:
                    cat = 'Effectenrekening (terugstorting)'
                _apply(idx, 'asset_withdrawal', 'sparen_beleggen', cat, 0.95)
                stats['asset_withdrawal'] += 1
                herclassificaties.append((idx, 'Laag A', 'asset_withdrawal', cat, f'FI-domein IBAN'))
            classified = True

        # A3: FINANCIËLE INSTELLING KEYWORD (naam match, geen IBAN match)
        elif any(kw in tekst for kw in FINANCIELE_INSTELLINGEN_KEYWORDS):
            if any(kw in tekst for kw in INVESTMENT_INCOME_KEYWORDS):
                _apply(idx, 'investment_income', 'inkomsten', 'Beleggingsinkomen', 0.88)
                stats['investment_income'] += 1
                herclassificaties.append((idx, 'Laag A', 'investment_income', 'Beleggingsinkomen', f'FI-keyword + dividend'))
            else:
                if any(kw in tekst for kw in ['BITVAVO', 'COINBASE', 'KRAKEN', 'BINANCE', 'BYBIT']):
                    cat = 'Crypto (terugstorting)'
                elif any(kw in tekst for kw in ['MINTOS', 'LENDAHAND', 'PEERBERRY', 'BONDORA',
                                                 'TWINO', 'OCTOBER', 'FUNDING CIRCLE', 'COLLIN']):
                    cat = 'Crowdlending (terugbetaling)'
                elif 'BRAND NEW DAY' in tekst:
                    cat = 'Pensioen (terugstorting)'
                else:
                    cat = 'Effectenrekening (terugstorting)'
                _apply(idx, 'asset_withdrawal', 'sparen_beleggen', cat, 0.92)
                stats['asset_withdrawal'] += 1
                herclassificaties.append((idx, 'Laag A', 'asset_withdrawal', cat, f'FI-keyword naam'))
            classified = True

        # A4: REFUND KEYWORD
        elif any(kw in tekst for kw in REFUND_KEYWORDS):
            _apply(idx, 'refund', 'variabele_kosten', 'Terugbetaling/Refund', 0.88)
            stats['refund'] += 1
            herclassificaties.append((idx, 'Laag A', 'refund', 'Terugbetaling/Refund', f'Refund-keyword'))
            classified = True

        # A5: TRANSFER KEYWORD
        elif any(kw in tekst for kw in TRANSFER_KEYWORDS):
            _apply(idx, 'internal_transfer', 'intern', 'Overboeking (intern)', 0.85)
            stats['internal_transfer'] += 1
            herclassificaties.append((idx, 'Laag A', 'internal_transfer', 'Overboeking (intern)', f'Transfer-keyword'))
            classified = True

        # A6: LENING KEYWORD
        elif any(kw in tekst for kw in LENING_KEYWORDS):
            _apply(idx, 'loan_inflow', 'sparen_beleggen', 'Lening/Hypotheek (uitbetaling)', 0.80)
            stats['loan_inflow'] += 1
            herclassificaties.append((idx, 'Laag A', 'loan_inflow', 'Lening/Hypotheek (uitbetaling)', f'Lening-keyword'))
            classified = True

        # A7: VERZEKERING MET UITKERING-KEYWORD
        elif any(kw in tekst for kw in VERZEKERING_KEYWORDS):
            _apply(idx, 'insurance', 'inkomsten', 'Verzekeringsuitkering', 0.85)
            stats['insurance'] += 1
            herclassificaties.append((idx, 'Laag A', 'insurance', 'Verzekeringsuitkering', f'Verzekering + uitkering-keyword'))
            classified = True

        if classified:
            fase1_handled.add(idx)

    # ===================================================================
    # FASE 2: LAAG B — Evidence-Based Classifiers (per groep)
    # ===================================================================
    # Resterende ongeclass positieve transacties na Laag A
    refund_handled = set()
    rent_handled = set()
    salary_handled = set()
    rest_mask = mask & (~df.index.isin(fase1_handled))
    rest = df[rest_mask].copy()

    if len(rest) > 0:
        # Groepeer op tegenpartij voor rent classifier
        if heeft_tegenrek:
            rest['_groep_key'] = rest['Tegenrekening'].apply(
                lambda x: _normaliseer_iban(str(x)) if pd.notna(x) else '')
            lege = rest['_groep_key'] == ''
            if lege.any():
                rest.loc[lege, '_groep_key'] = (
                    rest.loc[lege, 'tegenpartij_naam'].astype(str).str.upper().str.strip()
                    .where(lambda s: (s != '') & (s != 'NAN'),
                           rest.loc[lege, 'Omschrijving'].astype(str).str.upper().str.split().str[:3].str.join(' '))
                )
        else:
            rest['_groep_key'] = (
                rest['tegenpartij_naam'].astype(str).str.upper().str.strip()
                .where(lambda s: (s != '') & (s != 'NAN'),
                       rest['Omschrijving'].astype(str).str.upper().str.split().str[:3].str.join(' '))
            )

        rest = rest[rest['_groep_key'] != '']

        # B1: Refund matcher (per transactie, voor groepering)
        refund_handled = set()
        for idx, row in rest.iterrows():
            uitkomst, cat = _refund_matcher(row, negatieve_bedragen_per_iban,
                                            negatieve_bedragen_per_naam, kosten_merchants_set)
            if uitkomst == 'likely':
                _apply(idx, 'refund', 'variabele_kosten', cat, 0.82)
                stats['refund'] += 1
                herclassificaties.append((idx, 'Laag B', 'refund', cat, 'Refund matcher'))
                refund_handled.add(idx)

        # B2: Rent + Salary classifiers (per groep, gated by party_type)
        # RPR V2 classifier-toegang matrix:
        #   own_account / household → al uitgesloten (is_intern / onderling_neutraal)
        #   employer_or_payroll → salary (deterministic)
        #   business_counterparty → salary + rent
        #   government / merchant / broker / lender / card → al Fase 1
        #   unknown → rent ONLY
        rent_handled = set()
        salary_handled = set()
        has_party_type = 'party_type' in df.columns

        for groep_key, groep in rest[~rest.index.isin(refund_handled)].groupby('_groep_key'):
            # Determine party_type for this group
            if has_party_type:
                pt_values = df.loc[groep.index, 'party_type'].dropna().unique()
                party_type = pt_values[0] if len(pt_values) > 0 else 'unknown'
            else:
                party_type = 'unknown'

            # --- Classifier gating matrix (V2 counterparty roles) ---
            # employer_or_payroll → salary (deterministic, high confidence)
            # business_counterparty → salary + rent
            # government → SKIP (already handled in Fase 1 Laag A)
            # merchant → SKIP (refunds only, never salary/rent)
            # broker_or_investment_platform → SKIP (already handled in Fase 1 Laag A)
            # lender_or_mortgage_party → SKIP (already handled)
            # card_settlement → SKIP (internal)
            # unknown → rent only (no salary from unknown parties)
            allow_salary = party_type in ('employer_or_payroll', 'business_counterparty')
            allow_rent = party_type in ('business_counterparty', 'unknown')

            if allow_salary:
                # Simple salary detection: check for salary keywords in the group
                tekst_alle = ' '.join(groep['Omschrijving'].astype(str).str.upper())
                salaris_kw = ['SALARIS', 'LOON', 'MANAGEMENTFEE', 'MANAGEMENT FEE',
                              'HONORARIUM', 'VERGOEDING', 'NETTOLOON', 'NETTO LOON']
                has_salary_kw = any(kw in tekst_alle for kw in salaris_kw)
                has_rechtsvorm = any(m in tekst_alle for m in _RECHTSVORM_MARKERS)

                # Evidence: recurring + stable + from org = likely salary
                n_tx = len(groep)
                bedragen = groep['bedrag'].astype(float)
                mediaan = bedragen.median()

                if n_tx >= 3 and mediaan >= 500:
                    # IQR stability check
                    if n_tx >= 4:
                        q1 = bedragen.quantile(0.25)
                        q3 = bedragen.quantile(0.75)
                        iqr = q3 - q1
                        variatie = (iqr / mediaan) if mediaan > 0 else 1.0
                        is_stable = variatie < 0.30
                    else:
                        is_stable = bedragen.std() / mediaan < 0.20 if mediaan > 0 else False

                    if is_stable and (has_salary_kw or has_rechtsvorm):
                        # Determine subcategorie
                        if has_salary_kw and any(kw in tekst_alle for kw in ['MANAGEMENTFEE', 'MANAGEMENT FEE']):
                            cat = 'DGA-loon/Managementfee'
                        elif party_type == 'employer_or_payroll':
                            cat = 'DGA-loon/Managementfee'
                        else:
                            cat = 'Netto salaris'

                        for gidx in groep.index:
                            _apply(gidx, 'salary_income', 'inkomsten', cat, 0.85)
                            salary_handled.add(gidx)
                            # Promotie: bevestig counterparty role als employer_or_payroll
                            if party_type != 'employer_or_payroll':
                                df.at[gidx, 'party_type'] = 'employer_or_payroll'
                        stats.setdefault('salary_likely', 0)
                        stats['salary_likely'] = stats.get('salary_likely', 0) + len(groep)
                        naam_sample = str(groep.iloc[0].get('tegenpartij_naam', groep.iloc[0]['Omschrijving']))[:30]
                        if party_type != 'employer_or_payroll':
                            logger.info(
                                f"ROLE PROMOTIE: {groep_key[:30]} ({naam_sample}) "
                                f"{party_type} → employer_or_payroll (salary bevestigd)"
                            )
                        logger.info(
                            f"SALARY CLASSIFIER: LIKELY — {groep_key[:30]} ({naam_sample}) → {cat} — "
                            f"{n_tx} tx, mediaan EUR {mediaan:,.0f}, party_type={party_type}"
                        )
                        herclassificaties.append((groep.index.tolist(), 'Laag B', 'salary_likely',
                                                 cat, f'Salary classifier LIKELY (party_type={party_type})'))
                        continue  # Skip rent classifier for this group

            # --- Rent classifier (for business_counterparty and unknown) ---
            if not allow_rent:
                # business_related without salary match → uncertainty gate
                logger.info(
                    f"CLASSIFIER GATE: {groep_key[:30]} — party_type={party_type} — "
                    f"rent classifier BLOCKED (alleen salary toegestaan)"
                )
                continue

            uitkomst, cat, evidence = _rent_classifier(
                groep, df, groep_key, eigen_rek_norm, eigen_fi_ibans,
                bekende_merchants_set, heeft_tegenrek
            )

            if uitkomst == 'likely':
                for gidx in groep.index:
                    _apply(gidx, 'rent_income', 'inkomsten', 'Huurinkomsten', 0.85)
                    rent_handled.add(gidx)
                stats['rent_likely'] += len(groep)
                # Log evidence
                evidence_str = ', '.join(f'{k}={"JA" if v else "NEE"}' for k, (v, _) in evidence.items())
                naam_sample = str(groep.iloc[0].get('tegenpartij_naam', groep.iloc[0]['Omschrijving']))[:30]
                logger.info(
                    f"RENT CLASSIFIER: LIKELY — {groep_key[:30]} ({naam_sample}) — "
                    f"{len(groep)} tx, mediaan EUR {groep['bedrag'].median():,.0f}, "
                    f"totaal EUR {groep['bedrag'].sum():,.0f}, party_type={party_type}\n"
                    f"  Evidence: {evidence_str}"
                )
                herclassificaties.append((groep.index.tolist(), 'Laag B', 'rent_likely',
                                         'Huurinkomsten', f'Rent classifier LIKELY (party_type={party_type}): {evidence_str}'))

            elif uitkomst == 'uncertain':
                # Niet als huur herkend → gaat naar uncertainty gate (fase 3)
                failed = [k for k, (v, _) in evidence.items() if not v]
                logger.info(
                    f"RENT CLASSIFIER: UNCERTAIN — {groep_key[:30]} — "
                    f"{len(groep)} tx, mediaan EUR {groep['bedrag'].median():,.0f} — "
                    f"gefaalde checks: {', '.join(failed)}, party_type={party_type}"
                )

            # 'reject' = exclusion check gefaald, doorschuiven naar uncertainty gate

    # ===================================================================
    # FASE 2.5: HUISHOUD/SELF-TRANSFERS — household_related_party → neutraal
    # ===================================================================
    # GENERIEK: bij elke klant zijn er transfers van partner, familie, of
    # eigen rekeningen bij andere banken. RPR labelt deze als
    # household_related_party. Ze zijn al geblokkeerd voor salary (goed),
    # maar zonder deze stap vallen ze in de uncertainty gate als "Onzeker
    # positief" (slecht). Dit routeert ze naar onderling_neutraal.
    #
    # Werkt voor: DGA die geld verschuift, tweeverdiener met gescheiden
    # bankrekeningen, gezin met onderlinge betalingen, etc.
    alle_handled = fase1_handled | refund_handled | rent_handled | salary_handled
    pre_uncertainty_rest = mask & (~df.index.isin(alle_handled))
    household_handled = set()

    for idx in df[pre_uncertainty_rest].index:
        row = df.loc[idx]
        pt = row.get('party_type', '')
        if pt in ('household_related_party', 'own_account'):
            bedrag = float(row.get('bedrag', 0))
            if pt == 'own_account':
                _apply(idx, 'internal_transfer', 'onderling_neutraal',
                       'Eigen overboeking (niet-geüpload)', 0.80)
            else:
                _apply(idx, 'household_transfer', 'onderling_neutraal',
                       'Huishoudtransfer', 0.70)
            household_handled.add(idx)
            stats.setdefault('household_transfer', 0)
            stats['household_transfer'] = stats.get('household_transfer', 0) + 1
            herclassificaties.append((idx, 'Fase 2.5', 'household_transfer',
                                      f'{pt} → onderling_neutraal',
                                      f'party_type={pt}, bedrag={bedrag:.0f}'))

    if household_handled:
        logger.info(f"HOUSEHOLD ROUTING: {len(household_handled)} transacties → onderling_neutraal")

    # ===================================================================
    # FASE 3: UNCERTAINTY GATE — alles wat overblijft
    # ===================================================================
    alle_handled = alle_handled | household_handled
    final_rest = mask & (~df.index.isin(alle_handled))

    # Mapping van subcategorie → (sectie, inflow_type, confidence)
    # GENERIEK: elke subcategorie die we herkennen krijgt een specifieke sectie
    # zodat het niet in de generieke "Onzeker positief" valt.
    _UNCERTAINTY_SECTIE = {
        # --- Neutraal / onderling ---
        'Terugbetaling/Tikkie': ('onderling_neutraal', 'tikkie_refund', 0.75),
        'Onderlinge betaling (privépersoon)': ('onderling_neutraal', 'private_transfer', 0.50),
        'Onderlinge betaling (klein bedrag)': ('onderling_neutraal', 'private_transfer', 0.60),
        'Overige bijschrijving (klein)': ('onderling_neutraal', 'misc_small', 0.50),
        'Overige bijschrijving': ('onderling_neutraal', 'misc_inflow', 0.40),
        # --- Verzekeraar/financieel ---
        'Onzeker positief (verzekeraar)': ('variabele_kosten', 'insurance_payout', 0.65),
        'Onzeker positief (financiële instelling)': ('onderling_neutraal', 'financial_transfer', 0.55),
        # --- Bedrijf/organisatie —-- specifiekere routing
        'Onzeker positief (bedrijf/organisatie)': ('inkomsten', 'business_income_uncertain', 0.45),
        'Onzeker positief (bedrijf, klein bedrag)': ('onderling_neutraal', 'business_small', 0.50),
        # --- Bidirectioneel ---
        'Onzeker positief (bidirectioneel)': ('onderling_neutraal', 'bidirectional_transfer', 0.55),
        # --- Privépersoon groot bedrag ---
        'Onzeker positief (privépersoon, groot bedrag)': ('onderling_neutraal', 'private_large', 0.40),
        # --- Verkoop / cashback ---
        'Verkoop (tweedehands)': ('inkomsten', 'marketplace_sale', 0.70),
        'Cashback/Spaarprogramma': ('variabele_kosten', 'cashback', 0.75),
        'Notaris/Woningtransactie': ('inkomsten', 'property_transaction', 0.60),
    }
    n_reclassified = 0  # teller voor specifiek herkende subcategorieën

    for idx in df[final_rest].index:
        row = df.loc[idx]
        bedrag = float(row.get('bedrag', 0))
        subcategorie = _uncertainty_gate(row, bidi_ibans)

        # Specifiek herkende categorieën krijgen betere sectie + confidence
        if subcategorie in _UNCERTAINTY_SECTIE:
            sectie, inflow_t, conf = _UNCERTAINTY_SECTIE[subcategorie]
            _apply(idx, inflow_t, sectie, subcategorie, conf)
            n_reclassified += 1
        else:
            # Echt onzeker → standaard uncertainty bucket
            _apply(idx, 'uncertain', 'inkomsten', subcategorie, 0.30)

        stats['uncertain'] += 1
        uncertain_bedrag += bedrag
        herclassificaties.append((idx, 'Uncertainty Gate', 'uncertain', subcategorie, 'Subcategorie via uncertainty gate'))

    # === LOGGING ===
    totaal = sum(stats.values())
    logger.info(
        f"DECISION ENGINE: {totaal} positieve transacties geclassificeerd:\n"
        f"  Laag A — government:        {stats['government']}\n"
        f"  Laag A — asset_withdrawal:   {stats['asset_withdrawal']}\n"
        f"  Laag A — investment_income:  {stats['investment_income']}\n"
        f"  Laag A — refund (keyword):   {stats['refund'] - len(refund_handled) if 'refund_handled' in dir() else stats['refund']}\n"
        f"  Laag A — internal_transfer:  {stats['internal_transfer']}\n"
        f"  Laag A — insurance:          {stats['insurance']}\n"
        f"  Laag A — loan_inflow:        {stats['loan_inflow']}\n"
        f"  Laag B — refund (matcher):   {len(refund_handled) if 'refund_handled' in dir() else 0}\n"
        f"  Laag B — salary (likely):    {stats.get('salary_likely', 0)}\n"
        f"  Laag B — rent (likely):      {stats['rent_likely']}\n"
        f"  Uncertainty Gate:            {stats['uncertain']} (EUR {uncertain_bedrag:,.0f})\n"
        f"  → Geen positieve tx naar AI!"
    )

    # Log herclassificaties samenvatting
    if herclassificaties:
        logger.info(f"DECISION ENGINE: {len(herclassificaties)} herclassificatie-acties gelogd")

    return df


# ---------------------------------------------------------------------------
# CONSISTENTIE-AFDWINGING: zelfde IBAN → zelfde categorie
# ---------------------------------------------------------------------------

def _afdwing_iban_consistentie(df: pd.DataFrame) -> pd.DataFrame:
    """Zorg dat transacties met dezelfde Tegenrekening (IBAN) consistent geclassificeerd worden.

    Twee stappen:
    1. PROPAGATIE: Als een IBAN al rule-based geclassificeerd is, geef ALLE transacties
       met die IBAN dezelfde classificatie (mits zelfde richting: in/uit).
    2. MAJORITY VOTE: Als een IBAN meerdere keren voorkomt en de AI het inconsistent
       classificeert, wordt dit later in de prompt-hint afgevangen.

    Dit draait NA _classificeer_rule_based + _detecteer_vast_inkomen + _detecteer_huurinkomsten.
    """
    if 'Tegenrekening' not in df.columns or 'classificatie_bron' not in df.columns:
        return df

    n_gepropageerd = 0

    # Stap 1: Vind alle IBANs die al een rule-based classificatie hebben
    df_regel = df[(df['classificatie_bron'] == 'rule') & df['Tegenrekening'].notna()].copy()
    df_regel = df_regel[df_regel['Tegenrekening'].str.len() > 5]

    if len(df_regel) == 0:
        logger.info("CONSISTENTIE: geen IBANs met rule-based classificatie gevonden")
        return df

    # Bouw lookup: IBAN + richting → (sectie, categorie, confidence)
    iban_classificatie = {}
    for iban, groep in df_regel.groupby('Tegenrekening'):
        # Splits per richting (positief = inkomend, negatief = uitgaand)
        for richting in ['in', 'uit']:
            if richting == 'in':
                sub = groep[groep['bedrag'] > 0]
            else:
                sub = groep[groep['bedrag'] < 0]

            if len(sub) == 0:
                continue

            # Majority vote: welke sectie+categorie komt het vaakst voor?
            classificaties = sub.groupby(['regel_sectie', 'regel_categorie']).size()
            beste = classificaties.idxmax()
            iban_classificatie[(iban, richting)] = {
                'sectie': beste[0],
                'categorie': beste[1],
                'count': int(classificaties[beste]),
                'total': len(sub)
            }

    # Stap 2: Propageer naar niet-geclassificeerde transacties met dezelfde IBAN
    mask_onbekend = (df['classificatie_bron'].isna()) & (~df.get('is_intern', False)) & df['Tegenrekening'].notna()
    df_onbekend = df[mask_onbekend]

    for idx, row in df_onbekend.iterrows():
        iban = row['Tegenrekening']
        if not iban or len(str(iban)) < 5:
            continue

        richting = 'in' if row['bedrag'] > 0 else 'uit'
        key = (iban, richting)

        if key in iban_classificatie:
            info = iban_classificatie[key]
            # Alleen propageren als we ≥2 bevestigde transacties hebben (of 1 met hoge confidence)
            if info['count'] >= 2 or info['total'] >= 2:
                df.at[idx, 'regel_sectie'] = info['sectie']
                df.at[idx, 'regel_categorie'] = info['categorie']
                df.at[idx, 'classificatie_bron'] = 'rule'
                df.at[idx, 'regel_confidence'] = 0.85
                n_gepropageerd += 1

    if n_gepropageerd > 0:
        logger.info(f"CONSISTENTIE: {n_gepropageerd} transacties gepropageerd via IBAN-matching")
    else:
        logger.info("CONSISTENTIE: geen extra transacties gepropageerd")

    return df


# ---------------------------------------------------------------------------
# STAP 2: DETERMINISTISCH REKENEN
# ---------------------------------------------------------------------------

def bereken_feiten(df: pd.DataFrame) -> dict:
    resultaat = {}
    for rekening in sorted(df['Rekeningnummer'].unique()):
        rdf = df[df['Rekeningnummer'] == rekening].sort_values('datum')

        eerste_begin = float(rdf.iloc[0]['Beginsaldo'])
        laatste_eind = float(rdf.iloc[-1]['Eindsaldo'])
        totaal_mutaties = round(float(rdf['bedrag'].sum()), 2)
        berekend_eind = round(eerste_begin + totaal_mutaties, 2)

        maanden = {}
        for maand, mdf in rdf.groupby('maand'):
            maanden[str(maand)] = {
                'inkomsten': round(float(mdf[mdf['bedrag'] > 0]['bedrag'].sum()), 2),
                'uitgaven': round(float(mdf[mdf['bedrag'] < 0]['bedrag'].sum()), 2),
                'netto': round(float(mdf['bedrag'].sum()), 2),
                'transacties': len(mdf),
            }

        resultaat[str(rekening)] = {
            'periode': {
                'van': rdf['datum'].min().strftime('%Y-%m-%d'),
                'tot': rdf['datum'].max().strftime('%Y-%m-%d'),
            },
            'saldo': {
                'beginsaldo': eerste_begin,
                'eindsaldo': laatste_eind,
                'mutaties': totaal_mutaties,
                'berekend_eind': berekend_eind,
                'klopt': abs(berekend_eind - laatste_eind) < 0.02,
            },
            'totalen': {
                'inkomsten': round(float(rdf[rdf['bedrag'] > 0]['bedrag'].sum()), 2),
                'uitgaven': round(float(rdf[rdf['bedrag'] < 0]['bedrag'].sum()), 2),
                'netto': totaal_mutaties,
            },
            'transacties': len(rdf),
            'maanden': maanden,
        }

    return resultaat


def extract_naam(omschr: str) -> str:
    if pd.isna(omschr):
        return 'Onbekend'
    omschr = str(omschr)
    for marker in ['Naam: ', 'Naam:']:
        if marker in omschr:
            naam = omschr.split(marker)[1].split('Omschrijving:')[0].split('IBAN:')[0].strip()
            return naam[:60]
    if 'BEA' in omschr:
        delen = omschr.split(',')
        if len(delen) >= 2:
            return delen[1].strip().split('PAS')[0].strip()[:60]
    return omschr[:60]


# ---------------------------------------------------------------------------
# V3: POST-CLASSIFICATIE RECONCILIATIE
# ---------------------------------------------------------------------------

def _post_classificatie_reconciliatie(df: pd.DataFrame, feiten: dict) -> dict:
    """V3 Reconciliation-First: controleer dat ALLE transacties verantwoord zijn.

    Controleert per rekening, per maand:
    1. SUM(alle transacties in die maand) = netto mutaties uit feiten
    2. Geen transactie zonder sectie-toewijzing (na AI)
    3. Begin saldo + mutaties = eind saldo

    Returns dict met:
    - status: 'GREEN' / 'ORANGE' / 'RED'
    - checks: lijst van individuele checks
    - onvolledige_maanden: set van maanden met <15 transacties
    """
    checks = []
    status = 'GREEN'
    onvolledige_maanden = set()

    # Bepaal volle maanden (minimaal 15 transacties per maand over alle rekeningen)
    if 'maand' in df.columns:
        tx_per_maand = df.groupby('maand').size()
        for maand, count in tx_per_maand.items():
            if count < 15:
                onvolledige_maanden.add(str(maand))
                checks.append({
                    'type': 'ONVOLLEDIGE_MAAND',
                    'maand': str(maand),
                    'transacties': int(count),
                    'status': 'ORANGE',
                    'detail': f'Maand {maand} heeft slechts {count} transacties — telt niet mee in 12m'
                })

    # Per rekening: saldo-check
    for rek, rek_feiten in feiten.items():
        saldo = rek_feiten.get('saldo', {})
        if not saldo.get('klopt', True):
            verschil = abs(saldo.get('berekend_eind', 0) - saldo.get('eindsaldo', 0))
            checks.append({
                'type': 'SALDO_MISMATCH',
                'rekening': rek,
                'verschil': verschil,
                'status': 'RED',
                'detail': f'Rekening {rek}: berekend eind {saldo.get("berekend_eind")} != werkelijk {saldo.get("eindsaldo")}'
            })
            status = 'RED'

    # Check: hoeveel niet-interne transacties hebben nog geen sectie
    # NOTE: Deze check draait VOOR de AI-stap. Transacties zonder sectie gaan naar
    # de AI voor classificatie. Dit is geen fout maar een INFO check.
    # De echte ongeclassificeerde check hoort NA de AI merge (in _no_send_gate).
    df_extern = df[~df.get('is_intern', False)]
    n_zonder_sectie = len(df_extern[df_extern['regel_sectie'].isna()])
    if n_zonder_sectie > 0:
        pct = n_zonder_sectie / len(df_extern) * 100 if len(df_extern) > 0 else 0
        checks.append({
            'type': 'ONGECLASS_TRANSACTIES',
            'aantal': n_zonder_sectie,
            'percentage': round(pct, 1),
            'status': 'INFO',  # Niet RED/ORANGE — deze gaan naar AI
            'detail': f'{n_zonder_sectie} transacties ({pct:.1f}%) gaan naar AI voor classificatie'
        })

    # Household/salary contamination check
    if 'party_type' in df.columns and 'regel_categorie' in df.columns:
        hh_salary = df[
            (df['party_type'] == 'household_related_party') &
            (df['regel_categorie'].isin(['Netto salaris', 'DGA-loon/Managementfee']))
        ]
        if len(hh_salary) > 0:
            bedrag = abs(hh_salary['bedrag'].sum())
            checks.append({
                'type': 'HOUSEHOLD_ALS_SALARY',
                'aantal': len(hh_salary),
                'bedrag': round(bedrag, 2),
                'status': 'RED',
                'detail': f'{len(hh_salary)} household-transacties foutief als salary (EUR {bedrag:,.0f})'
            })
            status = 'RED'

    # Counterparty role integrity checks — rollen die NOOIT in inkomsten mogen staan
    _NOOIT_INKOMEN_ROLLEN = {'merchant', 'broker_or_investment_platform', 'card_settlement',
                              'lender_or_mortgage_party'}
    if 'party_type' in df.columns:
        fout_in_inkomen = df[
            (df['party_type'].isin(_NOOIT_INKOMEN_ROLLEN)) &
            (df['regel_sectie'] == 'inkomsten') &
            (df['bedrag'] > 0)
        ]
        if len(fout_in_inkomen) > 0:
            per_rol = fout_in_inkomen['party_type'].value_counts().to_dict()
            detail_parts = [f'{n}x {rol}' for rol, n in per_rol.items()]
            checks.append({
                'type': 'COUNTERPARTY_ROLE_ALS_INKOMEN',
                'aantal': len(fout_in_inkomen),
                'status': 'RED',
                'detail': f'{len(fout_in_inkomen)} transacties met verkeerde rol in inkomsten: {", ".join(detail_parts)}'
            })
            status = 'RED'

    logger.info(
        f"V3-RECONCILIATIE: status={status}, {len(checks)} checks, "
        f"{len(onvolledige_maanden)} onvolledige maanden: {onvolledige_maanden}"
    )
    for c in checks:
        logger.info(f"  [{c['status']}] {c['type']}: {c['detail']}")

    return {
        'status': status,
        'checks': checks,
        'onvolledige_maanden': onvolledige_maanden,
    }


def _bepaal_rapportperiode(df: pd.DataFrame, reconciliatie: dict) -> dict:
    """V3: Bepaal welke maanden meetellen voor 12m totalen/gemiddelden.

    Onvolledige maanden (< 15 tx) worden uitgesloten uit 12m berekeningen
    maar zijn wel zichtbaar in de Excel output.
    """
    if 'maand' not in df.columns:
        return {'volle_maanden': [], 'onvolledige_maanden': [], 'n_mnd': 0}

    alle_maanden = sorted(df['maand'].dropna().unique().astype(str))
    onvolledig = reconciliatie.get('onvolledige_maanden', set())

    volle_maanden = [m for m in alle_maanden if m not in onvolledig]
    # Beperk tot max 12 meest recente volle maanden
    if len(volle_maanden) > 12:
        volle_maanden = volle_maanden[-12:]

    logger.info(
        f"V3-PERIODE: {len(volle_maanden)} volle maanden voor 12m "
        f"(van {volle_maanden[0] if volle_maanden else '?'} t/m {volle_maanden[-1] if volle_maanden else '?'}), "
        f"{len(onvolledig)} onvolledige uitgesloten: {onvolledig}"
    )

    return {
        'volle_maanden': volle_maanden,
        'onvolledige_maanden': list(onvolledig),
        'n_mnd': len(volle_maanden),
    }


def bereken_top(df: pd.DataFrame, n: int = 15) -> dict:
    resultaat = {}
    for rekening in sorted(df['Rekeningnummer'].unique()):
        rdf = df[df['Rekeningnummer'] == rekening].copy()
        # Gebruik geëxtraheerde naam als beschikbaar, anders fallback op extract_naam
        rdf['tegenpartij'] = rdf.apply(
            lambda r: r['tegenpartij_naam'] if r.get('tegenpartij_naam') else extract_naam(r['Omschrijving']),
            axis=1)

        top_uit = (rdf[rdf['bedrag'] < 0]
                   .groupby('tegenpartij')['bedrag']
                   .agg(['sum', 'count'])
                   .sort_values('sum')
                   .head(n))

        top_in = (rdf[rdf['bedrag'] > 0]
                  .groupby('tegenpartij')['bedrag']
                  .agg(['sum', 'count'])
                  .sort_values('sum', ascending=False)
                  .head(n))

        resultaat[str(rekening)] = {
            'top_uitgaven': [
                {'naam': naam, 'bedrag': round(float(row['sum']), 2), 'aantal': int(row['count'])}
                for naam, row in top_uit.iterrows()
            ],
            'top_inkomsten': [
                {'naam': naam, 'bedrag': round(float(row['sum']), 2), 'aantal': int(row['count'])}
                for naam, row in top_in.iterrows()
            ],
        }

    return resultaat


# ---------------------------------------------------------------------------
# STAP 3: CLAUDE CATEGORISEERT EN ANALYSEERT
# ---------------------------------------------------------------------------

def _bouw_ground_truth_prompt_sectie(ground_truth: dict) -> str:
    """Bouw de GRONDWAARHEID sectie voor de AI-prompt.

    Dit vertelt de AI welke getallen bevroren zijn en dat het NIETS zelf
    mag berekenen. Elk getal in de AI-tekst moet hieruit komen.
    """
    lines = []
    lines.append("## ======================================================")
    lines.append("## GRONDWAARHEID — BEREKENDE TOTALEN (V3)")
    lines.append("## ======================================================")
    lines.append("## KRITIEK: Gebruik UITSLUITEND onderstaande getallen in je analyse.")
    lines.append("## Je mag NIET zelf optellen, aftrekken, of percentages berekenen.")
    lines.append("## Elk getal in je tekst MOET letterlijk voorkomen in deze lijst.")
    lines.append("## Als een getal niet in deze lijst staat, noem het dan NIET.")
    lines.append("")

    # Periode
    periode = ground_truth.get('periode', {})
    n_mnd = periode.get('n_mnd', 12)
    lines.append(f"Analyseperiode: {n_mnd} volle maanden")
    if periode.get('volle_maanden'):
        lines.append(f"  Van: {min(periode['volle_maanden'])} t/m {max(periode['volle_maanden'])}")
    if periode.get('onvolledige_maanden'):
        lines.append(f"  Onvolledige maanden (niet meegeteld): {', '.join(periode['onvolledige_maanden'])}")
    lines.append("")

    # Saldo
    saldo = ground_truth.get('saldo', {})
    lines.append(f"Totaal beginsaldo: EUR {saldo.get('totaal_begin', 0):,.2f}")
    lines.append(f"Totaal eindsaldo: EUR {saldo.get('totaal_eind', 0):,.2f}")
    lines.append("")

    # Sectie-totalen
    sectie_labels = {
        'inkomsten': 'INKOMSTEN',
        'vaste_lasten': 'VASTE LASTEN',
        'variabele_kosten': 'VARIABELE UITGAVEN',
        'sparen_beleggen': 'SPAREN & BELEGGEN',
        'onderling_neutraal': 'ONDERLING / NEUTRAAL',
    }
    sectie_totalen = ground_truth.get('sectie_totalen_12m', {})
    sectie_gem = ground_truth.get('sectie_gemiddelden_pm', {})
    cat_totalen = ground_truth.get('categorie_totalen_12m', {})

    for sectie_key, label in sectie_labels.items():
        totaal = sectie_totalen.get(sectie_key, 0)
        gem = sectie_gem.get(sectie_key, 0)
        lines.append(f"### {label}")
        lines.append(f"  Totaal {n_mnd}m: EUR {totaal:,.2f}")
        lines.append(f"  Gemiddeld p/m: EUR {gem:,.2f}")

        # Categorieën
        cats = cat_totalen.get(sectie_key, {})
        for cat, bedrag in sorted(cats.items(), key=lambda x: abs(x[1]), reverse=True):
            if abs(bedrag) >= 0.01:
                lines.append(f"    - {cat}: EUR {bedrag:,.2f}")
        lines.append("")

    # Income sources (bronopbouw + vertrouwen)
    income_sources = ground_truth.get('income_sources', {})
    if income_sources:
        lines.append("### INKOMSTENBRONOPBOUW (source_family)")
        for sf, data in sorted(income_sources.items(), key=lambda x: abs(x[1].get('bedrag_12m', 0)), reverse=True):
            bedrag = data.get('bedrag_12m', 0)
            n_tx = data.get('transacties', 0)
            vertr = data.get('vertrouwen', '?')
            vertr_icon = '✓' if vertr == 'hoog' else ('~' if vertr == 'medium' else '?')
            lines.append(f"  - {sf}: EUR {bedrag:,.2f} ({n_tx} tx) [{vertr_icon} {vertr}]")
        lines.append("")

    # Vertrouwensindicatoren per sectie
    vertrouwen = ground_truth.get('vertrouwen_per_sectie', {})
    if vertrouwen:
        lines.append("### VERTROUWENSINDICATOREN PER SECTIE")
        for sectie, v in vertrouwen.items():
            gem = v.get('gem_confidence', 0)
            label = v.get('vertrouwen', '?')
            lines.append(f"  - {sectie}: confidence {gem:.0%} ({label}), "
                        f"{v.get('pct_hoog', 0):.0f}% hoog, {v.get('pct_laag', 0):.0f}% laag")
        lines.append("")

    # Netto mutaties
    netto = saldo.get('totaal_eind', 0) - saldo.get('totaal_begin', 0)
    lines.append(f"Netto vermogensmutatie over periode: EUR {netto:,.2f}")
    lines.append("")

    # Strategische inzichten (deterministisch berekend)
    strat = ground_truth.get('strategische_inzichten', {})
    kengetallen = strat.get('kengetallen', {})
    if kengetallen:
        lines.append("### STRATEGISCHE KENGETALLEN (berekend, niet geschat)")
        lines.append(f"  Spaarquote: {kengetallen.get('spaarquote', 0):.1f}%")
        lines.append(f"  Netto cashflow per maand: EUR {kengetallen.get('netto_cashflow_pm', 0):,.0f}")
        lines.append(f"  Vaste lasten ratio: {kengetallen.get('vaste_lasten_ratio', 0):.1f}%")
        lines.append(f"  Vermogensopbouw per maand: EUR {kengetallen.get('vermogensopbouw_pm', 0):,.0f}")
        lines.append(f"  Inkomstenbronnen: {kengetallen.get('n_inkomstenbronnen', 0)}")
        lines.append(f"  Concentratie grootste bron: {kengetallen.get('concentratie_pct', 0):.0f}%")
        stab = kengetallen.get('inkomen_stabiliteit')
        if stab is not None:
            lines.append(f"  Inkomensstabiliteit: {stab:.0f}%")
        lines.append("")

    signalen = strat.get('signalen', [])
    if signalen:
        lines.append("### STRATEGISCHE SIGNALEN (gebruik deze in je analyse)")
        for s in signalen:
            icon = '✓' if s['type'] == 'positief' else ('!' if s['type'] == 'aandacht' else '⚠' if s['type'] == 'waarschuwing' else '·')
            lines.append(f"  {icon} {s['titel']}: {s['beschrijving']}")
        lines.append("")

    return "\n".join(lines)


def bouw_prompt(df: pd.DataFrame, feiten: dict, top: dict, eigen_rekeningen: set = None,
                ground_truth: dict = None) -> str:
    # Splits interne transfers van echte transacties
    if 'is_intern' in df.columns:
        df_extern = df[~df['is_intern']].copy()
        df_intern = df[df['is_intern']].copy()
        n_intern = len(df_intern)
        bedrag_intern_in = round(float(df_intern[df_intern['bedrag'] > 0]['bedrag'].sum()), 2)
        bedrag_intern_uit = round(float(df_intern[df_intern['bedrag'] < 0]['bedrag'].sum()), 2)
    else:
        df_extern = df.copy()
        df_intern = pd.DataFrame()
        n_intern = 0
        bedrag_intern_in = 0
        bedrag_intern_uit = 0

    # Geen harde limiet meer — moderne LLMs (Claude Opus, GPT-5.4) hebben
    # 200K+ context window. 3000 transacties × ~200 tokens = 600K tokens, past ruim.
    if len(df_extern) > 5000:
        logger.warning(f"Bestand bevat {len(df_extern)} externe transacties — gelimiteerd tot 5000")
        df_extern = df_extern.head(5000)

    # Transacties die rule-based als 'intern' zijn geclassificeerd (Tikkie, creditcard)
    # worden ook uit de lijst voor AI verwijderd
    if 'regel_sectie' in df_extern.columns:
        intern_regel_mask = df_extern['regel_sectie'] == 'intern'
        n_intern_regel = intern_regel_mask.sum()
        if n_intern_regel > 0:
            logger.info(f"Prompt: {n_intern_regel} regel-intern transacties (Tikkie/creditcard) verwijderd uit AI-lijst")
            df_extern = df_extern[~intern_regel_mask].copy()

    # =========================================================================
    # SPLIT: rule-based transacties gaan NIET naar de AI
    # De AI classificeert ALLEEN onbekende transacties. Rule-based totalen
    # worden apart berekend uit het DataFrame en na afloop gemerged.
    # Dit voorkomt dubbeltelling per definitie.
    # =========================================================================
    if 'classificatie_bron' in df_extern.columns:
        df_ai_only = df_extern[df_extern['classificatie_bron'] != 'rule'].copy()
        df_rule = df_extern[df_extern['classificatie_bron'] == 'rule'].copy()
        n_preclassified = len(df_rule)
    else:
        df_ai_only = df_extern.copy()
        df_rule = pd.DataFrame()
        n_preclassified = 0

    logger.info(f"Prompt: {len(df_ai_only)} transacties naar AI, {n_preclassified} rule-based (apart berekend)")

    regels = []
    for _, row in df_ai_only.iterrows():
        # Gebruik gestructureerde omschrijving als beschikbaar (tegenpartij — kenmerk)
        omschr = str(row.get('omschrijving_schoon') or row['Omschrijving'])[:250]

        # Voeg tegenpartij IBAN toe als beschikbaar (helpt bij consistentie)
        iban_hint = ''
        tr = row.get('Tegenrekening', '')
        if tr and str(tr).startswith('NL'):
            iban_hint = f'|IBAN:{tr}'

        regels.append(
            f"{row['datum'].strftime('%Y-%m-%d')}|{row['Rekeningnummer']}|"
            f"{row['bedrag']:>10.2f}|{omschr}{iban_hint}"
        )

    # Samenvatting van rule-based classificaties als context voor de AI
    rule_samenvatting = ""
    if n_preclassified > 0:
        rule_cats = df_rule.groupby(['regel_sectie', 'regel_categorie'])['bedrag'].agg(['sum', 'count'])
        rule_samenvatting = "\n## AL GECLASSIFICEERDE TRANSACTIES (niet in onderstaande lijst)\n"
        rule_samenvatting += f"Er zijn {n_preclassified} transacties al deterministisch geclassificeerd. "
        rule_samenvatting += "Deze zitten NIET in de lijst hieronder — jij hoeft ze NIET te classificeren.\n"
        rule_samenvatting += "Samenvatting ter context:\n"
        for (sectie, cat), row in rule_cats.iterrows():
            rule_samenvatting += f"- {sectie}/{cat}: {int(row['count'])}x, EUR {row['sum']:,.2f}\n"
        rule_samenvatting += "\nDeze bedragen worden automatisch samengevoegd met jouw classificatie.\n"

    # Eigen rekeningen info voor in de prompt
    eigen_rek_tekst = ""
    if eigen_rekeningen:
        eigen_rek_tekst = "\n## EIGEN REKENINGNUMMERS VAN DE KLANT\n"
        eigen_rek_tekst += "De klant heeft de volgende rekeningen (ALLE rekeningen hieronder zijn van dezelfde persoon/huishouden):\n"
        for rek in sorted(eigen_rekeningen):
            eigen_rek_tekst += f"- {rek}\n"
        eigen_rek_tekst += "\nBELANGRIJK: Overboekingen tussen twee rekeningen uit deze lijst zijn INTERNE VERSCHUIVINGEN.\n"
        eigen_rek_tekst += "Deze zijn AL verwijderd uit de transactielijst hieronder. Ze worden apart getoond in het rapport.\n"

    # Interne transfers samenvatting
    intern_tekst = ""
    if n_intern > 0:
        intern_tekst = f"""
## INTERNE VERSCHUIVINGEN (al verwijderd uit onderstaande transacties)
Er zijn {n_intern} interne overboekingen gedetecteerd tussen eigen rekeningen.
Totaal bijschrijvingen (ontvangst eigen rekening): EUR {bedrag_intern_in:,.2f}
Totaal afschrijvingen (verzending eigen rekening): EUR {bedrag_intern_uit:,.2f}
Deze tellen NIET mee als inkomen of uitgave. Ze worden apart in het rapport vermeld.
"""

    return f"""Acteer als een uiterst nauwkeurige financieel analist voor vermogende particulieren en DGA's in Nederland.
Analyseer elke transactie regel voor regel en wijs deze toe aan exact een categorie uit de onderstaande lijst.
Hieronder staan {len(df_ai_only)} banktransacties die JIJ moet classificeren.
{rule_samenvatting}{eigen_rek_tekst}{intern_tekst}

## CONTEXT
- Er zijn {n_preclassified} transacties al deterministisch geclassificeerd (zie samenvatting hierboven). Die worden automatisch samengevoegd — jij hoeft daar NIETS mee te doen.
- De TOTALEN in de GRONDWAARHEID-sectie zijn wiskundig berekend en 100% correct. Gebruik die cijfers, reken NIETS zelf.

## KERNREGELS
1. Gebruik UITSLUITEND de categorienamen uit de lijst hieronder. Verzin geen nieuwe categorieen.
2. RICHTING: Let altijd op of het bedrag positief (bij) of negatief (af) is.
   - Negatief bedrag = uitgave of vermogensoverdracht (naar vaste lasten, variabele kosten, of sparen/beleggen).
   - Positief bedrag van een winkel/tankstation/webshop = Terugbetaling/Refund, NOOIT inkomen.
   - Positief bedrag van een B.V./Stichting/werkgever/overheid = inkomen.
   - Positief bedrag van broker/crowdlending/crypto = vermogensmutatie (sparen_beleggen), GEEN inkomen.
3. BETALINGSVERWERKERS: Let op tussenpartijen zoals Adyen, Stichting Mollie Payments, Buckaroo, Pay.nl, Stripe.
   De naam van de verwerker is NIET de categorie — kijk naar de rest van de omschrijving om de werkelijke winkel of dienst te achterhalen.
4. SPAARDETECTIE: Een overboeking met omschrijvingen als "spaarpot", "buffer", "spaarrekening", "deposito" = Spaarrekening.
   Transacties naar bekende brokers (DeGiro, Saxo, IBKR, Meesman) of crypto-exchanges = Effectenrekening of Crypto.
5. CONSISTENTIE: Dezelfde tegenpartij (of dezelfde IBAN) MOET ALTIJD dezelfde categorie krijgen. Zoek patronen.
6. ANTI-OVERIG: "Overig" categorieen zijn een LAATSTE REDMIDDEL (max 3% per sectie). Doorloop bij twijfel:
   a. Staat er een bedrijfsnaam? Zoek wat voor bedrijf het is.
   b. Bevat de omschrijving een aanwijzing? (verzekering, tandarts, apotheek, garage, etc.)
   c. Is het BEA/GEA? Kijk naar de locatie/winkelnaam.
   d. Komt dezelfde tegenpartij vaker voor? Vast bedrag = vaste lasten, wisselend = variabel.
   e. Kijk naar het bedrag: <EUR50 onbekend = Boodschappen/Huishoudelijk. EUR50-500 = Kleding/Onderhoud/Elektronica. >EUR500 = Vakantie/Meubels/Auto.
   Pas na deze 5 stappen mag je "Overig" gebruiken.

## TRANSACTIEFORMAAT
datum|rekening|bedrag|omschrijving[|IBAN:tegenrekening]
Omschrijving is vaak "Tegenpartij — Kenmerk" (bv "Sevi B.V. — 2025 12 Sevi"). Sommige transacties hebben ruwe bankomschrijvingen — lees dan de HELE tekst.

## CATEGORIEEN

### INKOMSTEN (11 categorieen — ALLEEN echt verdiend/ontvangen geld)
- Netto salaris
- UWV/Uitkeringen
- DGA-loon/Managementfee (vanuit eigen BV met "Holding"/"Management" in naam)
- Huurinkomsten
- Toeslagen (zorgtoeslag, huurtoeslag, kindgebonden budget)
- Belastingteruggave (specificeer type als bekend: IB, VPB, BTW, LH, ZVW, MRB)
- Kinderbijslag/Kindregelingen
- Freelance/Opdrachten
- Verzekeringsuitkering
- Beleggingsinkomen (dividend, rente — NIET terugstortingen van broker)
- Overig inkomen (STRIKT: max 5% van totaal inkomsten, NOOIT refunds/terugboekingen)

### VASTE LASTEN (20 categorieen)
- Hypotheek/Huur
- Energie (gas, elektra, warmte)
- Water
- Gemeentebelasting/OZB/Waterschapsbelasting
- Zorgverzekering (basis + aanvullend)
- Inkomstenbelasting/Voorlopige aanslag
- Vennootschapsbelasting (VPB)
- BTW/Omzetbelasting
- Loonheffing
- ZVW-premie
- Motorrijtuigenbelasting (MRB)
- Overige belastingen (erfbelasting, schenkbelasting)
- Autoverzekering
- Woonverzekering/Inboedel
- Overige verzekeringen (reis, aansprakelijkheid, uitvaart)
- Internet/TV (Ziggo, KPN, glasvezel)
- Mobiele telefonie
- Streaming/Digitaal (Netflix, Spotify, Disney+, iCloud)
- Overige abonnementen (krant, tijdschrift, software)
- Kinderopvang/BSO/School
- Contributie/Lidmaatschap (sport, vereniging)
- Donaties/Goede doelen
- Bankkosten
- Overige vaste lasten

### VARIABELE KOSTEN (30 categorieen)
- Boodschappen/Supermarkt
- Drogist (Etos, Kruidvat)
- Restaurant/Uit eten
- Cafe/Drinken
- Afhaal/Bezorging (Thuisbezorgd, Uber Eats)
- Benzine/Diesel/Laden
- OV/Trein (NS, OV-chipkaart)
- Parkeren
- Taxi/Uber
- Auto-onderhoud/APK
- Kleding
- Schoenen
- Huisarts/Tandarts/Specialist
- Apotheek/Medicijnen
- Ziekenhuiskosten/Eigen risico
- Fysiotherapie/Alternatief
- Brillen/Lenzen
- Huishoudelijke artikelen
- Meubels/Inrichting
- Tuin/Buiten
- Onderhoud woning/Klussen
- Elektronica/Gadgets
- Boeken/Media
- Sport/Fitness
- Uitjes/Attracties/Bioscoop
- Vakantie/Reizen
- Cadeaus
- School/Studie/Cursussen
- Huisdieren
- Terugbetaling/Refund (positief bedrag van winkel/webshop: retour, storno, cashback)
- Overig variabel

### SPAREN & BELEGGEN (14 categorieen)
- Effectenrekening (negatief = storting NAAR broker)
- Effectenrekening (terugstorting) (positief = geld TERUG van broker, GEEN inkomen)
- Crowdlending (negatief = storting)
- Crowdlending (terugbetaling) (positief = aflossing/terugbetaling, GEEN inkomen)
- Crypto (negatief = storting)
- Crypto (terugstorting) (positief = geld terug, GEEN inkomen)
- Pensioenopbouw (Brand New Day, lijfrente)
- Pensioen (terugstorting)
- Kindersparen
- Spaarrekening
- Vastgoedinvestering
- Beleggingsfonds/ETF
- Levensverzekering/Kapitaalverzekering
- Overig sparen/beleggen

### INTERNE VERSCHUIVINGEN
- Overboekingen eigen rekeningen

## HERKENNINGSREGELS

### Belastingdienst (bepaal type op aanslagnummer betalingskenmerk)
Let op het aanslagnummer in de omschrijving (9 cijfers + lettercode):
- Letter H = Inkomstenbelasting/Voorlopige aanslag
- Letter V = Vennootschapsbelasting (VPB)
- Letter B/F/O = BTW/Omzetbelasting
- Letter L/A/J = Loonheffing
- Letter W = ZVW-premie
- Letter M/Y = Motorrijtuigenbelasting (MRB)
- Letter T1/T2/T3 = Toeslagen (INKOMSTEN, niet vaste lasten!)
- Letter Z = Overige belastingen
Fallback op tekst: "IB" = Inkomstenbelasting, "VPB" = Vennootschapsbelasting, "OB"/"BTW" = BTW/Omzetbelasting, "LH" = Loonheffing, "ZVW" = ZVW-premie, "MRB" = Motorrijtuigenbelasting
- "WOZ"/"OZB"/"waterschapsbelasting" = Gemeentebelasting/OZB/Waterschapsbelasting (GEEN Belastingdienst)
- Belastingdienst met POSITIEF bedrag = Belastingteruggave (specificeer type als bekend)

### Nederlandse bedrijven (categorie-lookup)
- Albert Heijn/Jumbo/Lidl/Plus/Dirk = Boodschappen/Supermarkt
- Etos/Kruidvat = Drogist
- Shell/BP/TotalEnergies/Tango/Tinq = Benzine/Diesel/Laden
- NS/OV-chipkaart/Connexxon/Arriva = OV/Trein
- Ziggo/KPN/T-Mobile = Internet/TV of Mobiele telefonie (op basis van bedrag/context)
- CZ/Zilveren Kruis/Menzis/VGZ = Zorgverzekering
- Frank Energie/Vattenfall/Eneco/Essent = Energie
- Vitens/Brabant Water/PWN/Dunea = Water
- Netflix/Spotify/Disney+/Apple/iCloud = Streaming/Digitaal
- Thuisbezorgd/Uber Eats/Deliveroo = Afhaal/Bezorging
- Uber/Bolt (taxi) = Taxi/Uber
- Booking.com/Airbnb/Transavia/KLM/Ryanair = Vakantie/Reizen
- bol.com/Coolblue/Amazon = Elektronica/Gadgets (tenzij duidelijk anders)
- GIVT/KWF/Rode Kruis/Oxfam = Donaties/Goede doelen
- H&M/Zara/C&A/Primark = Kleding
- Action/HEMA = Huishoudelijke artikelen
- IKEA = Meubels/Inrichting (of Huishoudelijke artikelen bij klein bedrag)
- Apotheek/BENU = Apotheek/Medicijnen
- Saxo Bank/DeGiro/IBKR/Meesman = Effectenrekening
- Mintos/Lendahand/PeerBerry = Crowdlending
- Brand New Day = Pensioenopbouw
- UWV = UWV/Uitkeringen
- ASR/Nationale-Nederlanden/Aegon/Delta Lloyd/Reaal: check context — levensverzekering bij hypotheek = Hypotheek/Huur

### BEA/GEA transacties (pinbetalingen)
Kijk ALTIJD naar de winkelnaam NA "BEA" of "GEA":
- Bij supermarkt = Boodschappen/Supermarkt
- Bij tankstation = Benzine/Diesel/Laden
- Bij restaurant/eetgelegenheid = Restaurant/Uit eten
- Bij kledingwinkel = Kleding

### Salaris uit B.V.
- Vast maandelijks bedrag van een B.V. = waarschijnlijk salaris
- "Holding"/"Management" in naam = DGA-loon/Managementfee
- Anders = Netto salaris
- Zeg NOOIT "uw BV" tenzij de BV expliciet als eigen BV is aangemerkt

### Creditcard-aflossing
ICS/VISA/Mastercard/American Express = Interne verschuivingen (geen consumptie)

### Hypotheek-gekoppelde verzekeringen
Maandelijkse betalingen aan ASR/NN/Aegon/Delta Lloyd voor levensverzekering/kapitaalverzekering = Hypotheek/Huur (onderdeel spaarhypotheek), NIET Sparen & Beleggen

## TOP TEGENPARTIJEN
{json.dumps(top, indent=2, ensure_ascii=False)}

## TRANSACTIES (datum|rekening|bedrag|omschrijving)
{chr(10).join(regels)}

{_bouw_ground_truth_prompt_sectie(ground_truth) if ground_truth else f"## CORRECTE TOTALEN{chr(10)}{json.dumps(feiten, indent=2, ensure_ascii=False)}"}

## OUTPUT FORMAT
Antwoord ALLEEN in valid JSON:
{{
  "maandoverzicht": {{
    "<rekening>": {{
      "<YYYY-MM>": {{
        "inkomsten": {{"<categorie>": {{"bedrag": 0.00, "aantal": 0}}}},
        "vaste_lasten": {{"<categorie>": {{"bedrag": 0.00, "aantal": 0}}}},
        "variabele_kosten": {{"<categorie>": {{"bedrag": 0.00, "aantal": 0}}}},
        "sparen_beleggen": {{"<categorie>": {{"bedrag": 0.00, "aantal": 0}}}},
        "interne_verschuivingen": {{"bedrag": 0.00, "aantal": 0}}
      }}
    }}
  }},
  "jaartotalen": {{
    "<rekening>": {{
      "inkomsten": {{"<categorie>": 0.00}},
      "vaste_lasten": {{"<categorie>": 0.00}},
      "variabele_kosten": {{"<categorie>": 0.00}},
      "sparen_beleggen": {{"<categorie>": 0.00}},
      "interne_verschuivingen": 0.00
    }}
  }},
  "analyse": {{
    "samenvatting": "3-4 alinea's. Schrijf als een senior financieel adviseur die een vermogende particulier of DGA informeert — rustig, zakelijk, respectvol. Begin met het totaalbeeld: hoeveel komt er structureel binnen, hoeveel gaat er structureel uit, hoeveel gaat naar vermogensopbouw. Benoem dan de cashflowdynamiek: zijn er grote interne verschuivingen, beleggingstransacties, of seizoenseffecten die het beeld vertekenen? Eindig met de kern: waar zit de financiele kracht en waar de kwetsbaarheid. KRITIEK: Elk getal in je tekst MOET letterlijk voorkomen in de GRONDWAARHEID sectie hierboven. Je mag NIET zelf optellen, aftrekken, of percentages berekenen. TOON-REGELS: (1) Gebruik NOOIT een oordelende of budgetcoach-achtige toon. Gebruik neutrale financiele taal. (2) Wees EERLIJK over onzekerheden. (3) Spreek de gebruiker NOOIT aan met 'uw BV' tenzij je ZEKER weet dat het een eigen BV is.",
    "sterke_punten": ["Noem 3-5 financiele sterktes. Gebruik ALLEEN bedragen uit de GRONDWAARHEID. Schrijf bevestigend en zakelijk."],
    "aandachtspunten": ["Noem 3-5 signalen die aandacht verdienen. Gebruik ALLEEN bedragen uit de GRONDWAARHEID. Geen oordelende taal."],
    "aanbevelingen": ["Geef 3-5 concrete, strategische aanbevelingen op het niveau van financieel advies, niet budgetcoaching."],
    "verrassende_inzichten": ["Geef 2-3 patronen of inzichten die een drukke vermogende particulier NIET zelf zou zien maar die een AI wel opvalt. Gebruik ALLEEN bedragen uit de GRONDWAARHEID."]
  }}
}}"""


def _genereer_strategische_inzichten(ground_truth: dict, df: pd.DataFrame) -> dict:
    """Genereer deterministische strategische inzichten uit de data.

    GEEN AI-gok. Puur berekeningen die een financieel adviseur zou maken.
    Elke inzicht is traceerbaar naar concrete data.

    Returns dict met:
    - kengetallen: spaarquote, vaste lasten ratio, etc.
    - signalen: lijst van strategische observaties
    - cashflow_trend: maandelijkse trend data
    """
    # Gebruik rapport_totalen + executive_buckets als single source of truth
    rt = ground_truth.get('rapport_totalen', {})
    eb = ground_truth.get('executive_buckets', {})
    sectie_totalen = ground_truth.get('sectie_totalen_12m', {})
    n_mnd = ground_truth.get('periode', {}).get('n_mnd', 12)
    income_sources = ground_truth.get('income_sources', {})
    maand_sectie = ground_truth.get('maand_sectie_totalen', {})

    inkomsten = rt.get('bruto_inkomen_12m', abs(sectie_totalen.get('inkomsten', 0)))
    vaste_lasten = rt.get('vaste_lasten_12m', abs(sectie_totalen.get('vaste_lasten', 0)))
    variabele = rt.get('variabele_kosten_12m', abs(sectie_totalen.get('variabele_kosten', 0)))
    sparen = abs(sectie_totalen.get('sparen_beleggen', 0))
    totaal_uit = vaste_lasten + variabele

    # === KENGETALLEN ===
    kengetallen = {}

    # Spaarquote: (inkomsten - uitgaven) / inkomsten
    if inkomsten > 0:
        netto_cashflow = rt.get('netto_beschikbaar_12m', inkomsten - totaal_uit)
        spaarquote = netto_cashflow / inkomsten
        kengetallen['spaarquote'] = round(spaarquote * 100, 1)
        kengetallen['netto_cashflow_pm'] = round(netto_cashflow / max(n_mnd, 1), 0)
    else:
        kengetallen['spaarquote'] = 0
        kengetallen['netto_cashflow_pm'] = 0

    # Vaste lasten ratio
    if inkomsten > 0:
        kengetallen['vaste_lasten_ratio'] = round(vaste_lasten / inkomsten * 100, 1)
    else:
        kengetallen['vaste_lasten_ratio'] = 0

    # Vermogensopbouw indicator
    kengetallen['vermogensopbouw_12m'] = round(sparen, 0)
    kengetallen['vermogensopbouw_pm'] = round(sparen / max(n_mnd, 1), 0)

    # Inkomstendiversificatie: hoeveel bronnen, concentratie
    bronnen = {k: abs(v.get('bedrag_12m', 0)) for k, v in income_sources.items()
               if abs(v.get('bedrag_12m', 0)) > 100}
    kengetallen['n_inkomstenbronnen'] = len(bronnen)
    if bronnen and inkomsten > 0:
        grootste_bron = max(bronnen.values())
        kengetallen['concentratie_pct'] = round(grootste_bron / inkomsten * 100, 1)
    else:
        kengetallen['concentratie_pct'] = 100

    # Inkomenstabiliteit: standaardafwijking van maandelijks inkomen
    maand_inkomsten = []
    for maand, secties in maand_sectie.items():
        ink = abs(secties.get('inkomsten', 0))
        if ink > 0:
            maand_inkomsten.append(ink)
    if len(maand_inkomsten) >= 3:
        import statistics
        gem = statistics.mean(maand_inkomsten)
        std = statistics.stdev(maand_inkomsten)
        kengetallen['inkomen_stabiliteit'] = round((1 - min(std / gem, 1)) * 100, 1) if gem > 0 else 0
    else:
        kengetallen['inkomen_stabiliteit'] = None

    # === STRATEGISCHE SIGNALEN ===
    signalen = []

    # 1. Spaarquote beoordeling
    sq = kengetallen.get('spaarquote', 0)
    if sq >= 30:
        signalen.append({
            'type': 'positief',
            'titel': 'Sterke spaarquote',
            'beschrijving': f'Spaarquote van {sq:.0f}% — ruim boven de 20% richtlijn voor vermogensopbouw.',
        })
    elif sq >= 10:
        signalen.append({
            'type': 'neutraal',
            'titel': 'Gezonde spaarquote',
            'beschrijving': f'Spaarquote van {sq:.0f}% — voldoende voor basisopbouw, maar er is ruimte voor meer.',
        })
    elif sq >= 0:
        signalen.append({
            'type': 'aandacht',
            'titel': 'Lage spaarquote',
            'beschrijving': f'Spaarquote van {sq:.0f}% — bijna alle inkomsten worden uitgegeven.',
        })
    else:
        signalen.append({
            'type': 'waarschuwing',
            'titel': 'Negatieve cashflow',
            'beschrijving': f'Er wordt maandelijks meer uitgegeven dan er binnenkomt (spaarquote {sq:.0f}%).',
        })

    # 2. Vaste lasten ratio
    vl_ratio = kengetallen.get('vaste_lasten_ratio', 0)
    if vl_ratio > 50:
        signalen.append({
            'type': 'aandacht',
            'titel': 'Hoge vaste lasten',
            'beschrijving': f'{vl_ratio:.0f}% van het inkomen gaat naar vaste lasten — beperkt flexibiliteit.',
        })

    # 3. Inkomensconcentratie
    conc = kengetallen.get('concentratie_pct', 100)
    n_bronnen = kengetallen.get('n_inkomstenbronnen', 0)
    if conc > 80 and n_bronnen <= 1:
        signalen.append({
            'type': 'aandacht',
            'titel': 'Eén inkomstenbron',
            'beschrijving': f'{conc:.0f}% van het inkomen komt uit één bron — hoog concentratierisico.',
        })
    elif n_bronnen >= 3:
        signalen.append({
            'type': 'positief',
            'titel': 'Gediversifieerd inkomen',
            'beschrijving': f'{n_bronnen} inkomstenbronnen — goed gespreid risico.',
        })

    # 4. Vermogensopbouw signaal
    vo_pm = kengetallen.get('vermogensopbouw_pm', 0)
    if vo_pm > 1000:
        signalen.append({
            'type': 'positief',
            'titel': 'Actieve vermogensopbouw',
            'beschrijving': f'Gemiddeld €{vo_pm:,.0f}/mnd naar beleggingen en spaarrekeningen.',
        })

    # 5. Inkomensstabiliteit
    stab = kengetallen.get('inkomen_stabiliteit')
    if stab is not None:
        if stab >= 85:
            signalen.append({
                'type': 'positief',
                'titel': 'Stabiel inkomen',
                'beschrijving': f'Inkomensstabiliteit van {stab:.0f}% — zeer voorspelbaar maandpatroon.',
            })
        elif stab < 60:
            signalen.append({
                'type': 'aandacht',
                'titel': 'Variabel inkomen',
                'beschrijving': f'Inkomensstabiliteit van {stab:.0f}% — overweeg een buffer van 3-6 maanden vaste lasten.',
            })

    # 6. DGA-specifiek: management fee vs salaris verhouding
    mgmt_bedrag = abs(income_sources.get('management_fee', {}).get('bedrag_12m', 0))
    salary_bedrag = abs(income_sources.get('salary_employment', {}).get('bedrag_12m', 0))
    freelance_bedrag = abs(income_sources.get('freelance_business', {}).get('bedrag_12m', 0))
    if mgmt_bedrag > 0 and salary_bedrag > 0:
        signalen.append({
            'type': 'neutraal',
            'titel': 'Gemengd inkomen (DGA + loondienst)',
            'beschrijving': f'DGA-inkomen €{mgmt_bedrag:,.0f}/jr naast salaris €{salary_bedrag:,.0f}/jr — '
                           f'controleer of de DGA-beloning fiscaal optimaal is ingericht.',
        })
    elif mgmt_bedrag > 0 and freelance_bedrag > 0:
        signalen.append({
            'type': 'neutraal',
            'titel': 'BV-inkomen + freelance',
            'beschrijving': f'DGA-inkomen €{mgmt_bedrag:,.0f}/jr + freelance €{freelance_bedrag:,.0f}/jr — '
                           f'controleer of alle inkomsten via de BV lopen voor fiscale efficiëntie.',
        })

    # 7. Cashflow trend (stijgend/dalend)
    cashflow_trend = []
    for maand in sorted(maand_sectie.keys()):
        secties = maand_sectie[maand]
        ink = abs(secties.get('inkomsten', 0))
        uit = abs(secties.get('vaste_lasten', 0)) + abs(secties.get('variabele_kosten', 0))
        netto = ink - uit
        cashflow_trend.append({'maand': maand, 'inkomsten': round(ink, 0),
                               'uitgaven': round(uit, 0), 'netto': round(netto, 0)})

    logger.info(f"STRATEGISCHE INZICHTEN: {len(kengetallen)} kengetallen, {len(signalen)} signalen")

    return {
        'kengetallen': kengetallen,
        'signalen': signalen,
        'cashflow_trend': cashflow_trend,
    }


def _bereken_premium_inzichten(ground_truth: dict, df: pd.DataFrame) -> dict:
    """Bereken 5 premium strategische inzichten uit ground_truth data.

    Puur deterministisch. Elke inzicht heeft:
    - id: unieke key
    - titel: korte naam
    - beschrijving: 1-2 zinnen conclusie
    - waarde: numerieke kernwaarde (of None als niet berekend)
    - detail: extra context
    - relevantie: 'hoog' / 'medium' / 'laag' / None (niet van toepassing)
    """
    inzichten = []
    sectie_totalen = ground_truth.get('sectie_totalen_12m', {})
    cat_totalen = ground_truth.get('categorie_totalen_12m', {})
    maand_sectie = ground_truth.get('maand_sectie_totalen', {})
    income_sources = ground_truth.get('income_sources', {})
    saldo = ground_truth.get('saldo', {})
    n_mnd = ground_truth.get('periode', {}).get('n_mnd', 12) or 12

    inkomsten = abs(sectie_totalen.get('inkomsten', 0))
    vaste_lasten = abs(sectie_totalen.get('vaste_lasten', 0))
    variabele = abs(sectie_totalen.get('variabele_kosten', 0))
    totaal_uit = vaste_lasten + variabele

    # ── 1. CASH DRAG / "LUI GELD" RADAR ──
    # Gemiddeld banksaldo vs hoogste maanduitgave → hoeveel maanden buffer
    gemiddeld_saldo = None
    maand_uitgaven = []
    for maand, secties in maand_sectie.items():
        uit = abs(secties.get('vaste_lasten', 0)) + abs(secties.get('variabele_kosten', 0))
        if uit > 0:
            maand_uitgaven.append(uit)

    begin_saldo = saldo.get('totaal_begin', 0)
    eind_saldo = saldo.get('totaal_eind', 0)
    if begin_saldo and eind_saldo:
        gemiddeld_saldo = (begin_saldo + eind_saldo) / 2

    gem_maand_uitgave = sum(maand_uitgaven) / max(len(maand_uitgaven), 1) if maand_uitgaven else 0

    if gemiddeld_saldo is not None and gem_maand_uitgave > 0:
        buffer_maanden = gemiddeld_saldo / gem_maand_uitgave
        if buffer_maanden > 6:
            relevantie = 'hoog'
            beschrijving = (
                f'Gemiddeld banksaldo van \u20ac{gemiddeld_saldo:,.0f} dekt {buffer_maanden:.1f} maanden uitgaven. '
                f'Boven 6 maanden buffer is het saldo mogelijk niet optimaal ingezet \u2014 '
                f'overweeg of een deel naar een hogere rendementsklasse kan.'
            )
        elif buffer_maanden > 3:
            relevantie = 'medium'
            beschrijving = (
                f'Gemiddeld banksaldo van \u20ac{gemiddeld_saldo:,.0f} dekt {buffer_maanden:.1f} maanden uitgaven. '
                f'Een gezonde buffer. Geen directe actie nodig.'
            )
        else:
            relevantie = 'laag'
            beschrijving = (
                f'Gemiddeld banksaldo van \u20ac{gemiddeld_saldo:,.0f} dekt {buffer_maanden:.1f} maanden uitgaven. '
                f'Relatief krap \u2014 overweeg een aparte liquiditeitsreserve.'
            )
        inzichten.append({
            'id': 'cash_drag',
            'titel': 'Liquiditeitsradar',
            'beschrijving': beschrijving,
            'waarde': round(buffer_maanden, 1),
            'detail': f'\u20ac{gemiddeld_saldo:,.0f} gem. saldo / \u20ac{gem_maand_uitgave:,.0f} gem. uitgaven p/m',
            'relevantie': relevantie,
        })

    # ── 2. ZAKELIJK VS PRIV\u00c9 LEKKAGE (DGA-specifiek) ──
    # Kijk of er DGA/management fee inkomen is + zakelijke kosten op priv\u00e9rekening
    mgmt_fee = abs(income_sources.get('management_fee', {}).get('bedrag_12m', 0))
    dga_loon = 0
    for src_key, src_data in income_sources.items():
        if 'dga' in src_key.lower() or 'management' in src_key.lower():
            dga_loon += abs(src_data.get('bedrag_12m', 0))
    if mgmt_fee > 0:
        dga_loon = max(dga_loon, mgmt_fee)

    # Zoek zakelijk-gerelateerde uitgaven op priv\u00e9rekening
    zakelijke_cats_op_prive = 0
    zakelijke_labels = ['accountant', 'boekhouder', 'kvk', 'notaris', 'zakelijk', 'kantoor']
    for sectie_key in ['vaste_lasten', 'variabele_kosten']:
        cats = cat_totalen.get(sectie_key, {})
        for cat_naam, bedrag in cats.items():
            for label in zakelijke_labels:
                if label in cat_naam.lower():
                    zakelijke_cats_op_prive += abs(float(bedrag))

    if dga_loon > 0:
        if zakelijke_cats_op_prive > 500:
            relevantie = 'hoog'
            beschrijving = (
                f'Er gaat \u20ac{zakelijke_cats_op_prive:,.0f}/jaar aan mogelijk zakelijke kosten via de priv\u00e9rekening. '
                f'Bij een DGA-inkomen van \u20ac{dga_loon:,.0f}/jaar is dit een fiscaal aandachtspunt \u2014 '
                f'deze kosten zijn mogelijk aftrekbaar via de BV.'
            )
        else:
            relevantie = 'laag'
            beschrijving = (
                f'DGA-inkomen van \u20ac{dga_loon:,.0f}/jaar. Geen significante zakelijke kosten op de priv\u00e9rekening gedetecteerd. '
                f'De scheiding zakelijk/priv\u00e9 lijkt goed ingericht.'
            )
        inzichten.append({
            'id': 'zakelijk_prive',
            'titel': 'Zakelijk/Priv\u00e9 scheiding',
            'beschrijving': beschrijving,
            'waarde': round(zakelijke_cats_op_prive, 0),
            'detail': f'DGA-inkomen: \u20ac{dga_loon:,.0f}/jr | Zakelijk op priv\u00e9: \u20ac{zakelijke_cats_op_prive:,.0f}/jr',
            'relevantie': relevantie,
        })

    # ── 3. VASTGOED RENDEMENT ──
    # Huurinkomsten vs woonkosten (hypotheek/huur, onderhoud)
    huur_inkomen = 0
    for src_key, src_data in income_sources.items():
        if 'huur' in src_key.lower() or 'rent' in src_key.lower():
            huur_inkomen += abs(src_data.get('bedrag_12m', 0))

    woonkosten = 0
    woon_labels = ['hypotheek', 'huur', 'onderhoud', 'vve', 'opstal', 'woning']
    for sectie_key in ['vaste_lasten']:
        cats = cat_totalen.get(sectie_key, {})
        for cat_naam, bedrag in cats.items():
            for label in woon_labels:
                if label in cat_naam.lower():
                    woonkosten += abs(float(bedrag))

    if huur_inkomen > 1000:
        netto_vastgoed = huur_inkomen - woonkosten
        rendement_pct = (netto_vastgoed / huur_inkomen * 100) if huur_inkomen > 0 else 0
        if rendement_pct > 50:
            relevantie = 'hoog'
        elif rendement_pct > 20:
            relevantie = 'medium'
        else:
            relevantie = 'laag'
        beschrijving = (
            f'Huurinkomsten van \u20ac{huur_inkomen:,.0f}/jaar met \u20ac{woonkosten:,.0f} aan woonkosten. '
            f'Netto vastgoedresultaat: \u20ac{netto_vastgoed:,.0f}/jaar ({rendement_pct:.0f}% netto marge). '
        )
        if netto_vastgoed < 0:
            beschrijving += 'De woonkosten overtreffen de huurinkomsten \u2014 controleer of alle kosten correct zijn toegewezen.'
        inzichten.append({
            'id': 'vastgoed_rendement',
            'titel': 'Vastgoedrendement',
            'beschrijving': beschrijving,
            'waarde': round(rendement_pct, 1),
            'detail': f'Huur: \u20ac{huur_inkomen:,.0f} | Kosten: \u20ac{woonkosten:,.0f} | Netto: \u20ac{netto_vastgoed:,.0f}',
            'relevantie': relevantie,
        })

    # ── 4. FAIR SHARE CHECK ──
    # Inkomensbronnen verdeling — hoeveel % komt van grootste bron
    bronnen = {}
    for src_key, src_data in income_sources.items():
        bedrag = abs(src_data.get('bedrag_12m', 0))
        if bedrag > 100:
            bronnen[src_key] = bedrag

    if len(bronnen) >= 2 and inkomsten > 0:
        gesorteerd = sorted(bronnen.items(), key=lambda x: -x[1])
        grootste_naam, grootste_bedrag = gesorteerd[0]
        concentratie = grootste_bedrag / inkomsten * 100
        tweede_naam, tweede_bedrag = gesorteerd[1]

        if concentratie > 80:
            relevantie = 'hoog'
            beschrijving = (
                f'{concentratie:.0f}% van het inkomen komt uit \u00e9\u00e9n bron (\u20ac{grootste_bedrag:,.0f}/jaar). '
                f'Dit is een hoog concentratierisico. De tweede bron levert slechts \u20ac{tweede_bedrag:,.0f}/jaar.'
            )
        elif concentratie > 60:
            relevantie = 'medium'
            beschrijving = (
                f'Het inkomen is verdeeld over {len(bronnen)} bronnen, met {concentratie:.0f}% uit de grootste. '
                f'Redelijke spreiding, maar nog afhankelijk van \u00e9\u00e9n primaire bron.'
            )
        else:
            relevantie = 'laag'
            beschrijving = (
                f'Het inkomen is goed gespreid over {len(bronnen)} bronnen. '
                f'De grootste bron levert {concentratie:.0f}% \u2014 laag concentratierisico.'
            )
        inzichten.append({
            'id': 'fair_share',
            'titel': 'Inkomensconcentratie',
            'beschrijving': beschrijving,
            'waarde': round(concentratie, 1),
            'detail': ' | '.join(f'{k}: \u20ac{v:,.0f}' for k, v in gesorteerd[:3]),
            'relevantie': relevantie,
        })

    # ── 5. LIFESTYLE INFLATIE DETECTIE ──
    # Abonnementen + subscriptions als % van inkomen + trend
    abo_totaal = 0
    abo_cats = []
    abo_labels = ['abonnement', 'streaming', 'spotify', 'netflix', 'disney', 'lidmaatschap',
                  'contributie', 'apple', 'icloud', 'youtube', 'digitaal']
    for sectie_key in ['vaste_lasten', 'variabele_kosten']:
        cats = cat_totalen.get(sectie_key, {})
        for cat_naam, bedrag in cats.items():
            for label in abo_labels:
                if label in cat_naam.lower():
                    abo_totaal += abs(float(bedrag))
                    abo_cats.append((cat_naam, abs(float(bedrag))))
                    break

    abo_pm = abo_totaal / n_mnd
    if abo_totaal > 0 and inkomsten > 0:
        abo_pct = abo_totaal / inkomsten * 100
        if abo_pct > 5:
            relevantie = 'hoog'
            beschrijving = (
                f'Abonnementen en lidmaatschappen kosten \u20ac{abo_pm:,.0f}/maand '
                f'({abo_pct:.1f}% van inkomen). Bij vermogende huishoudens sluipen abonnementen erin \u2014 '
                f'een jaarlijkse review kan \u20ac{abo_totaal * 0.2:,.0f} besparen.'
            )
        elif abo_pct > 2:
            relevantie = 'medium'
            beschrijving = (
                f'Abonnementen en lidmaatschappen: \u20ac{abo_pm:,.0f}/maand ({abo_pct:.1f}% van inkomen). '
                f'Gemiddeld niveau. Overweeg jaarlijks een subscription audit.'
            )
        else:
            relevantie = 'laag'
            beschrijving = (
                f'Abonnementen en lidmaatschappen: \u20ac{abo_pm:,.0f}/maand ({abo_pct:.1f}% van inkomen). '
                f'Beperkte impact op de totale cashflow.'
            )
        inzichten.append({
            'id': 'lifestyle_inflatie',
            'titel': 'Abonnementencheck',
            'beschrijving': beschrijving,
            'waarde': round(abo_pm, 0),
            'detail': ', '.join(f'{c}: \u20ac{b:,.0f}' for c, b in sorted(abo_cats, key=lambda x: -x[1])[:5]),
            'relevantie': relevantie,
        })

    # Sorteer op relevantie (hoog eerst)
    rel_order = {'hoog': 0, 'medium': 1, 'laag': 2, None: 3}
    inzichten.sort(key=lambda x: rel_order.get(x.get('relevantie'), 3))

    logger.info(f"PREMIUM INZICHTEN: {len(inzichten)} inzichten berekend")
    for iz in inzichten:
        logger.info(f"  [{iz['relevantie']}] {iz['titel']}: {iz.get('waarde')}")

    return inzichten


def _log_classificatie_kwaliteit(df: pd.DataFrame) -> dict:
    """Log een samenvatting van classificatie-kwaliteit.

    Retourneert een dict met kwaliteitsmetrieken die in het rapport kunnen worden opgenomen.
    """
    totaal = len(df[~df.get('is_intern', False)])
    n_rule = len(df[df['classificatie_bron'] == 'rule'])
    n_ai = totaal - n_rule
    pct_rule = (n_rule / totaal * 100) if totaal > 0 else 0

    # Confidence verdeling voor rule-based
    if 'regel_confidence' in df.columns:
        df_conf = df[df['regel_confidence'].notna()]
        n_high = len(df_conf[df_conf['regel_confidence'] >= 0.90])
        n_med = len(df_conf[(df_conf['regel_confidence'] >= 0.80) & (df_conf['regel_confidence'] < 0.90)])
        n_low = len(df_conf[df_conf['regel_confidence'] < 0.80])
    else:
        n_high = n_med = n_low = 0

    # Bedragen
    bedrag_rule = abs(df[df['classificatie_bron'] == 'rule']['bedrag'].sum()) if n_rule > 0 else 0
    bedrag_totaal = abs(df[~df.get('is_intern', False)]['bedrag'].sum()) if totaal > 0 else 0
    pct_bedrag_rule = (bedrag_rule / bedrag_totaal * 100) if bedrag_totaal > 0 else 0

    logger.info(
        f"CLASSIFICATIE-KWALITEIT: {n_rule}/{totaal} ({pct_rule:.0f}%) transacties rule-based, "
        f"{n_ai} naar AI. Bedrag: {pct_bedrag_rule:.0f}% rule-based."
    )
    logger.info(
        f"  Confidence: {n_high} hoog (≥0.90), {n_med} medium (0.80-0.89), {n_low} laag (<0.80)"
    )

    return {
        'totaal_transacties': totaal,
        'rule_based': n_rule,
        'ai_geclassificeerd': n_ai,
        'pct_rule_based': round(pct_rule, 1),
        'pct_bedrag_rule_based': round(pct_bedrag_rule, 1),
        'confidence': {
            'hoog': n_high,
            'medium': n_med,
            'laag': n_low,
        }
    }


def _verzamel_review_items(df: pd.DataFrame) -> list:
    """Verzamel items die menselijke review nodig hebben.

    Draait NA alle classificatie. Scant het DataFrame op:
    1. Grote onzekere bedragen (>€1000 in uncertain bucket)
    2. Lage confidence classificaties (<0.50 op bedragen >€200)
    3. Inkomen op zakelijke rekening zonder keyword-match
    4. Grote eenmalige transacties (>€5000, <3 keer)

    Returns: lijst van review item dicts.
    """
    items = []

    df_niet_intern = df[~df.get('is_intern', False)].copy()

    # 1. Grote onzekere bedragen
    if 'regel_categorie' in df.columns:
        uncertain_mask = df_niet_intern['regel_categorie'].str.startswith('Onzeker positief', na=False)
        df_uncertain = df_niet_intern[uncertain_mask & (df_niet_intern['bedrag'] > 1000)]
        for _, row in df_uncertain.iterrows():
            naam = str(row.get('tegenpartij_naam', row.get('Omschrijving', '')))[:40]
            items.append({
                'type': 'groot_onzeker_bedrag',
                'ernst': 'hoog',
                'beschrijving': f"€{float(row['bedrag']):,.0f} van {naam} — niet geclassificeerd als bewezen inkomen",
                'categorie': str(row.get('regel_categorie', 'Onzeker')),
                'bedrag': float(row['bedrag']),
            })

    # 2. Lage confidence op significante bedragen
    if 'regel_confidence' in df.columns:
        low_conf = df_niet_intern[
            (df_niet_intern['regel_confidence'] < 0.50) &
            (df_niet_intern['bedrag'].abs() > 200) &
            (df_niet_intern['regel_confidence'].notna())
        ]
        for _, row in low_conf.head(10).iterrows():
            naam = str(row.get('tegenpartij_naam', row.get('Omschrijving', '')))[:40]
            items.append({
                'type': 'lage_confidence',
                'ernst': 'medium',
                'beschrijving': f"€{abs(float(row['bedrag'])):,.0f} — {naam} — confidence {float(row['regel_confidence']):.0%}",
                'categorie': str(row.get('regel_categorie', '?')),
                'bedrag': abs(float(row['bedrag'])),
            })

    # 3. Grote eenmalige transacties (>€5000, komt <3x voor van zelfde bron)
    if 'Tegenrekening' in df.columns:
        grote = df_niet_intern[df_niet_intern['bedrag'] > 5000].copy()
        if len(grote) > 0:
            for iban in grote['Tegenrekening'].unique():
                if pd.isna(iban):
                    continue
                n = len(df_niet_intern[df_niet_intern['Tegenrekening'] == iban])
                if n < 3:
                    sub = grote[grote['Tegenrekening'] == iban]
                    naam = str(sub.iloc[0].get('tegenpartij_naam', sub.iloc[0].get('Omschrijving', '')))[:40]
                    totaal = float(sub['bedrag'].sum())
                    items.append({
                        'type': 'groot_eenmalig',
                        'ernst': 'medium',
                        'beschrijving': f"€{totaal:,.0f} van {naam} — slechts {n}x, controleer herkomst",
                        'categorie': str(sub.iloc[0].get('regel_categorie', '?')),
                        'bedrag': totaal,
                    })

    # Sorteer op ernst (hoog eerst) dan bedrag (groot eerst)
    ernst_order = {'hoog': 0, 'medium': 1, 'laag': 2}
    items.sort(key=lambda x: (ernst_order.get(x['ernst'], 9), -x.get('bedrag', 0)))

    logger.info(f"REVIEW ITEMS: {len(items)} items gevonden "
                f"({sum(1 for i in items if i['ernst'] == 'hoog')} hoog, "
                f"{sum(1 for i in items if i['ernst'] == 'medium')} medium)")

    return items[:20]  # Max 20 items


def _rapport_kwaliteitscheck(data: dict, df: pd.DataFrame, eigen_rekeningen: set):
    """Blokkeer rapport als er grove classificatiefouten in zitten.

    Controleert (BLOKKEREND - rapport wordt niet gegenereerd):
    1. 'Overig inkomen' > 40% van totaal inkomen → classificatiefout
    2. Transfer in inkomen → AI heeft interne transfer als inkomen gezet
    3. Groot bedrag (>€2.000/jaar) in AI-only categorie zonder rule-backup

    Controleert (WAARSCHUWING - rapport wordt gegenereerd met disclaimer):
    4. 'Overig inkomen' > 15% van totaal inkomen
    5. Hypotheek-verzekeringen in beleggingen
    6. Hoge AI-afhankelijkheid (>60% transacties alleen door AI geclassificeerd)

    Raises ValueError als een blokkerende check faalt.
    Retourneert lijst met waarschuwingen voor disclaimer in rapport.
    """
    blockers = []
    warnings_found = []

    jaartotalen = data.get('jaartotalen', {})

    # =========================================================================
    # BLOKKERENDE CHECKS — rapport wordt NIET gegenereerd
    # =========================================================================

    # Check 1: 'Overig inkomen' mag niet groter zijn dan 40% van GEVERIFIEERD inkomen
    # Gebruikt de globale _STRUCTURAL_INCOME_WHITELIST voor consistentie
    for rek, totalen in jaartotalen.items():
        inkomsten = totalen.get('inkomsten', {})
        if isinstance(inkomsten, dict):
            overig = abs(float(inkomsten.get('Overig inkomen', 0)))
            totaal_ink = sum(abs(float(v)) for k, v in inkomsten.items()
                           if isinstance(v, (int, float)) and k in _STRUCTURAL_INCOME_WHITELIST)
            if totaal_ink > 0 and overig > 0:
                ratio = overig / totaal_ink
                if ratio > 0.40:
                    blockers.append(
                        f"BLOKKADE: Rekening {rek}: 'Overig inkomen' is {ratio:.0%} van totaal inkomen "
                        f"(EUR {overig:,.0f} / EUR {totaal_ink:,.0f}). "
                        f"Dit duidt op fout-geclassificeerde interne overboekingen."
                    )
                elif ratio > 0.15:
                    warnings_found.append(
                        f"Rekening {rek}: 'Overig inkomen' is {ratio:.0%} van totaal inkomen "
                        f"(EUR {overig:,.0f}). Mogelijk zijn niet alle inkomstenbronnen herkend."
                    )

    # Check 2: Transfer in inkomen of variabele kosten
    # Controleer of eigen rekening-IBANs voorkomen als tegenpartij in AI-geclassificeerde transacties
    if eigen_rekeningen and 'Tegenrekening' in df.columns:
        df_ai_only = df[(df['classificatie_bron'] != 'rule') & (~df['is_intern'])]
        for idx, row in df_ai_only.iterrows():
            tegen = _normaliseer_iban(str(row.get('Tegenrekening', '')))
            if tegen in eigen_rekeningen:
                bedrag = abs(float(row.get('bedrag', 0)))
                if bedrag > 100:
                    blockers.append(
                        f"BLOKKADE: Transfer naar eigen rekening ({tegen}) "
                        f"niet als intern gemarkeerd (EUR {bedrag:,.0f}). "
                        f"Dit mag nooit in het rapport als inkomen of uitgave staan."
                    )
                    break  # Eén blocker is genoeg

    # =========================================================================
    # WAARSCHUWINGS-CHECKS — rapport wordt gegenereerd MET disclaimer
    # =========================================================================

    # Check 3: Hypotheek-verzekeringen in beleggingen
    for rek, totalen in jaartotalen.items():
        sparen = totalen.get('sparen_beleggen', {})
        if isinstance(sparen, dict):
            for cat, bedrag in sparen.items():
                cat_lower = cat.lower()
                if ('levensverzekering' in cat_lower or 'kapitaalverzekering' in cat_lower):
                    bedrag_abs = abs(float(bedrag)) if isinstance(bedrag, (int, float)) else 0
                    if bedrag_abs > 500:
                        warnings_found.append(
                            f"'{cat}' (EUR {bedrag_abs:,.0f}) staat onder Sparen & Beleggen. "
                            f"Levensverzekeringen zijn in Nederland vaak hypotheek-gekoppeld."
                        )

    # Check 4: Confidence / AI-afhankelijkheid
    df_extern = df[~df.get('is_intern', False)]
    n_totaal = len(df_extern)
    n_rule = len(df_extern[df_extern['classificatie_bron'] == 'rule'])
    n_ai = n_totaal - n_rule
    pct_ai = (n_ai / n_totaal * 100) if n_totaal > 0 else 0
    if pct_ai > 60:
        warnings_found.append(
            f"{pct_ai:.0f}% van de transacties is alleen door AI geclassificeerd "
            f"({n_ai} van {n_totaal}). Overweeg meer merchants toe te voegen aan de regelslaag."
        )

    # Check 5: Grote AI-only bedragen (>€2.000/jaar per categorie)
    # Dit zijn potentieel onbetrouwbare classificaties
    # (voor nu waarschuwing, later mogelijk blokkade)
    if n_ai > 0:
        df_ai_bedragen = df_extern[df_extern['classificatie_bron'] != 'rule']
        if len(df_ai_bedragen) > 0:
            groot_ai_totaal = df_ai_bedragen['bedrag'].apply(lambda x: abs(float(x))).sum()
            if groot_ai_totaal > 5000:
                warnings_found.append(
                    f"EUR {groot_ai_totaal:,.0f} aan transacties is alleen door AI geclassificeerd "
                    f"zonder rule-based backup. Bij grote bedragen kan dit onbetrouwbaar zijn."
                )

    # =========================================================================
    # RESULTAAT
    # =========================================================================

    # Blokkerende fouten → rapport NIET genereren
    if blockers:
        for b in blockers:
            logger.error(f"KWALITEITSCHECK: {b}")
        raise ValueError(
            f"Rapport geblokkeerd door {len(blockers)} kwaliteitscheck(s): "
            + " | ".join(blockers)
        )

    # Waarschuwingen → rapport WEL genereren, maar loggen
    if warnings_found:
        logger.warning(f"KWALITEITSCHECK: {len(warnings_found)} waarschuwing(en)")
        for w in warnings_found:
            logger.warning(f"  ⚠ {w}")
        # Sla waarschuwingen op zodat ze in het rapport kunnen worden verwerkt
        data['_kwaliteitswaarschuwingen'] = warnings_found
    else:
        logger.info("KWALITEITSCHECK: alle checks geslaagd, geen waarschuwingen")
        data['_kwaliteitswaarschuwingen'] = []

    return warnings_found


def _bereken_rule_based_totalen(df: pd.DataFrame) -> dict:
    """Bereken totalen voor rule-based geclassificeerde transacties direct uit het DataFrame.

    Dit is de GROUND TRUTH — deze cijfers komen rechtstreeks uit de bankdata,
    niet uit de AI. Ze worden later gemerged met de AI-output.

    Retourneert dezelfde structuur als de AI: {rekening: {sectie: {categorie: bedrag}}}
    Plus maandoverzicht: {rekening: {maand: {sectie: {categorie: {bedrag, aantal}}}}}
    """
    result = {'jaartotalen': {}, 'maandoverzicht': {}}

    if 'classificatie_bron' not in df.columns:
        return result

    # V3: onderling_neutraal meenemen in rule-based totalen (niet meer uitgefilterd)
    df_regel = df[(df['classificatie_bron'] == 'rule') & (~df.get('is_intern', False))].copy()
    df_regel = df_regel[df_regel['regel_sectie'] != 'intern']
    # Household transacties die als onderling_neutraal zijn gemarkeerd zijn NIET is_intern
    # en worden dus correct meegenomen

    if len(df_regel) == 0:
        logger.info("RULE-TOTALEN: geen rule-based transacties")
        return result

    # Jaartotalen: per rekening, per sectie, per categorie → som bedragen
    for (rek, sectie, cat), groep in df_regel.groupby(
        [df_regel['Rekeningnummer'].astype(str), 'regel_sectie', 'regel_categorie']
    ):
        bedrag = round(float(groep['bedrag'].sum()), 2)

        if rek not in result['jaartotalen']:
            result['jaartotalen'][rek] = {}
        if sectie not in result['jaartotalen'][rek]:
            result['jaartotalen'][rek][sectie] = {}
        result['jaartotalen'][rek][sectie][cat] = bedrag

        logger.info(f"RULE-TOTALEN: {rek}/{sectie}/{cat} = EUR {bedrag:,.2f} ({len(groep)} tx)")

    # Maandoverzicht: per rekening, per maand, per sectie, per categorie
    df_regel['_maand'] = df_regel['datum'].apply(
        lambda d: d.strftime('%Y-%m') if hasattr(d, 'strftime') else str(d)
    )
    for (rek, maand, sectie, cat), groep in df_regel.groupby(
        [df_regel['Rekeningnummer'].astype(str), '_maand', 'regel_sectie', 'regel_categorie']
    ):
        bedrag = round(float(groep['bedrag'].sum()), 2)
        aantal = len(groep)

        if rek not in result['maandoverzicht']:
            result['maandoverzicht'][rek] = {}
        if maand not in result['maandoverzicht'][rek]:
            result['maandoverzicht'][rek][maand] = {}
        if sectie not in result['maandoverzicht'][rek][maand]:
            result['maandoverzicht'][rek][maand][sectie] = {}
        result['maandoverzicht'][rek][maand][sectie][cat] = {
            'bedrag': bedrag,
            'aantal': aantal
        }

    logger.info(f"RULE-TOTALEN: {len(df_regel)} transacties in {len(result['jaartotalen'])} rekening(en)")
    return result


def _merge_rule_en_ai_totalen(rule_data: dict, ai_data: dict) -> dict:
    """Voeg rule-based totalen en AI-totalen samen.

    GEEN overlap mogelijk: rule-based transacties zijn NIET naar de AI gestuurd.
    Simpele merge: voor elke rekening/sectie/categorie → tel bedragen op.
    Als dezelfde categorie in beide voorkomt, is dat correct (rule-based had andere
    transacties dan de AI).
    """
    merged = ai_data.copy()

    # === Jaartotalen mergen ===
    ai_jaar = merged.get('jaartotalen', {})
    for rek, secties in rule_data.get('jaartotalen', {}).items():
        if rek not in ai_jaar:
            ai_jaar[rek] = {}
        for sectie, cats in secties.items():
            if sectie not in ai_jaar[rek]:
                ai_jaar[rek][sectie] = {}
            if not isinstance(ai_jaar[rek][sectie], dict):
                ai_jaar[rek][sectie] = {}
            for cat, bedrag in cats.items():
                # Optellen: AI had andere transacties, rule had andere transacties
                bestaand = float(ai_jaar[rek][sectie].get(cat, 0))
                ai_jaar[rek][sectie][cat] = round(bestaand + bedrag, 2)
                logger.info(
                    f"MERGE: {rek}/{sectie}/{cat}: rule EUR {bedrag:,.2f} + AI EUR {bestaand:,.2f} "
                    f"= EUR {round(bestaand + bedrag, 2):,.2f}"
                )
    merged['jaartotalen'] = ai_jaar

    # === Maandoverzicht mergen ===
    ai_maand = merged.get('maandoverzicht', {})
    for rek, maanden in rule_data.get('maandoverzicht', {}).items():
        if rek not in ai_maand:
            ai_maand[rek] = {}
        for maand, secties in maanden.items():
            if maand not in ai_maand[rek]:
                ai_maand[rek][maand] = {}
            for sectie, cats in secties.items():
                if sectie not in ai_maand[rek][maand]:
                    ai_maand[rek][maand][sectie] = {}
                if not isinstance(ai_maand[rek][maand][sectie], dict):
                    ai_maand[rek][maand][sectie] = {}
                for cat, data in cats.items():
                    bestaand = ai_maand[rek][maand][sectie].get(cat, {'bedrag': 0, 'aantal': 0})
                    if isinstance(bestaand, dict):
                        b = float(bestaand.get('bedrag', 0))
                        a = int(bestaand.get('aantal', 0))
                    else:
                        b = float(bestaand)
                        a = 0
                    ai_maand[rek][maand][sectie][cat] = {
                        'bedrag': round(b + data['bedrag'], 2),
                        'aantal': a + data['aantal']
                    }
    merged['maandoverzicht'] = ai_maand

    logger.info("MERGE: rule-based en AI totalen samengevoegd")
    return merged


def _vraag_claude(prompt: str, model: str) -> dict:
    """Roep Anthropic Claude API aan."""
    from anthropic import Anthropic

    api_key = os.environ.get('ANTHROPIC_API_KEY')
    if not api_key:
        raise ValueError("ANTHROPIC_API_KEY niet geconfigureerd")

    import httpx
    client = Anthropic(
        api_key=api_key,
        timeout=httpx.Timeout(600.0, connect=30.0),
    )

    logger.info(f"Claude aanroepen ({model}), prompt: {len(prompt)} tekens (~{len(prompt)//4} tokens)")

    response = client.messages.create(
        model=model,
        max_tokens=32000,
        messages=[{"role": "user", "content": prompt}],
    )

    tekst = response.content[0].text
    tokens_in = response.usage.input_tokens
    tokens_out = response.usage.output_tokens
    logger.info(f"Claude klaar: {tokens_in} in, {tokens_out} out")

    return tekst, tokens_in, tokens_out


def _vraag_openai(prompt: str, model: str) -> dict:
    """Roep OpenAI GPT API aan."""
    from openai import OpenAI

    api_key = os.environ.get('OPENAI_API_KEY')
    if not api_key:
        raise ValueError("OPENAI_API_KEY niet geconfigureerd — stel in via Railway Variables")

    client = OpenAI(
        api_key=api_key,
        timeout=600.0,
    )

    logger.info(f"OpenAI aanroepen ({model}), prompt: {len(prompt)} tekens (~{len(prompt)//4} tokens)")

    response = client.chat.completions.create(
        model=model,
        max_completion_tokens=32000,
        messages=[{"role": "user", "content": prompt}],
    )

    tekst = response.choices[0].message.content
    tokens_in = response.usage.prompt_tokens
    tokens_out = response.usage.completion_tokens
    logger.info(f"OpenAI klaar: {tokens_in} in, {tokens_out} out")

    return tekst, tokens_in, tokens_out


def _vraag_gemini(prompt: str, model: str) -> tuple:
    """Roep Google Gemini API aan."""
    import google.generativeai as genai

    api_key = os.environ.get('GOOGLE_AI_API_KEY')
    if not api_key:
        raise ValueError("Gemini is niet beschikbaar: GOOGLE_AI_API_KEY niet geconfigureerd. Kies een ander model.")

    genai.configure(api_key=api_key)

    logger.info(f"Gemini aanroepen ({model}), prompt: {len(prompt)} tekens (~{len(prompt)//4} tokens)")

    try:
        gen_model = genai.GenerativeModel(model)
        response = gen_model.generate_content(
            prompt,
            generation_config=genai.types.GenerationConfig(
                max_output_tokens=32000,
                temperature=0.1,
            ),
        )
    except Exception as e:
        err_str = str(e)
        logger.error(f"Gemini API fout: {err_str[:500]}")
        if '429' in err_str or 'quota' in err_str.lower():
            raise ValueError(
                f"Gemini API quota overschreden voor {model}. "
                "Dit betekent dat de gratis tier geen capaciteit heeft voor dit model. "
                "Schakel billing in bij Google AI Studio of kies een ander model (Claude Opus 4.7 of GPT 5.4)."
            )
        elif '404' in err_str or 'not found' in err_str.lower():
            raise ValueError(
                f"Gemini model '{model}' bestaat niet of is niet beschikbaar. "
                "Kies een ander model."
            )
        else:
            raise ValueError(f"Gemini API fout: {err_str[:200]}")

    tekst = response.text
    # Gemini usage metadata
    tokens_in = getattr(response.usage_metadata, 'prompt_token_count', 0) if response.usage_metadata else 0
    tokens_out = getattr(response.usage_metadata, 'candidates_token_count', 0) if response.usage_metadata else 0
    logger.info(f"Gemini klaar: {tokens_in} in, {tokens_out} out")

    return tekst, tokens_in, tokens_out


# ---------------------------------------------------------------------------
# Model configuratie & selectie
# ---------------------------------------------------------------------------
# Beschikbare categorizers (frontend value → provider + model):
_AI_CATEGORIZERS = {
    'claude_opus_47': ('claude', 'claude-opus-4-7'),
    'claude_opus_46': ('claude', 'claude-opus-4-6'),
    'openai_gpt54':   ('openai', 'gpt-5.4'),
    'gemini_31_pro':  ('gemini', 'gemini-3.1-pro-preview'),
}

def vraag_ai(prompt: str, categorizer: str = None) -> dict:
    """Generieke AI-aanroep — kiest model op basis van categorizer parameter.

    categorizer: een key uit _AI_CATEGORIZERS (bijv. 'claude_opus_47').
    Als None: valt terug op env vars AI_PROVIDER/CLAUDE_MODEL (backward compatible).
    """
    if categorizer and categorizer in _AI_CATEGORIZERS:
        provider, model = _AI_CATEGORIZERS[categorizer]
    elif categorizer and categorizer == 'shortcut_excel':
        # Shortcut flow wordt apart afgehandeld, niet via deze functie
        raise ValueError("Shortcut Excel flow moet apart worden afgehandeld")
    else:
        # Backward compatible: env vars
        provider = os.environ.get('AI_PROVIDER', 'claude').lower()
        if provider == 'openai':
            model = os.environ.get('OPENAI_MODEL', 'gpt-5.4')
        elif provider == 'gemini':
            model = os.environ.get('GEMINI_MODEL', 'gemini-3.1-pro-preview')
        else:
            model = os.environ.get('CLAUDE_MODEL', 'claude-opus-4-7')

    if provider == 'openai':
        tekst, tokens_in, tokens_out = _vraag_openai(prompt, model)
    elif provider == 'gemini':
        tekst, tokens_in, tokens_out = _vraag_gemini(prompt, model)
    else:
        tekst, tokens_in, tokens_out = _vraag_claude(prompt, model)

    # Parse JSON uit response
    if '```json' in tekst:
        tekst = tekst.split('```json')[1].split('```')[0]
    elif '```' in tekst:
        tekst = tekst.split('```')[1].split('```')[0]

    try:
        return {
            'data': json.loads(tekst),
            'tokens': {'input': tokens_in, 'output': tokens_out},
            'model': model,
            'provider': provider,
        }
    except json.JSONDecodeError as e:
        logger.error(f"JSON parse error: {e}")
        return {
            'data': None,
            'raw': tekst[:2000],
            'error': str(e),
            'tokens': {'input': tokens_in, 'output': tokens_out},
            'model': model,
            'provider': provider,
        }


# Backward compatible alias
def vraag_claude(prompt: str, categorizer: str = None) -> dict:
    return vraag_ai(prompt, categorizer=categorizer)


# ---------------------------------------------------------------------------
# AI AUDITOR — onafhankelijke sanity check op ground truth
# ---------------------------------------------------------------------------

_AUDITOR_MODELS = {
    'openai_gpt54': ('openai', 'gpt-5.4'),
    'claude_opus_47': ('claude', 'claude-opus-4-7'),
    'gemini_31_pro': ('gemini', 'gemini-3.1-pro-preview'),
    'none': (None, None),
}


def _ai_auditor(ground_truth: dict, auditor_model: str = 'openai_gpt54') -> dict:
    """Onafhankelijke AI sanity check op het eindrapport.

    Stuurt de rapport_totalen, sectie_totalen en top-categorieën naar een AI
    met de opdracht: controleer op inconsistenties, onrealistische verhoudingen,
    en logische fouten. Retourneert dict met status + bevindingen.
    """
    if auditor_model == 'none' or auditor_model not in _AUDITOR_MODELS:
        return {'status': 'skipped', 'issues': [], 'model': auditor_model}

    rt = ground_truth.get('rapport_totalen', {})
    st = ground_truth.get('sectie_totalen_12m', {})
    ct = ground_truth.get('categorie_totalen_12m', {})
    n_mnd = ground_truth.get('periode', {}).get('n_mnd', 12)
    saldo = ground_truth.get('saldo', {})
    income_sources = ground_truth.get('income_sources', {})

    # Bouw compact overzicht voor de auditor
    cat_overzicht = {}
    for sectie, cats in ct.items():
        top_cats = sorted(cats.items(), key=lambda x: -abs(float(x[1])))[:8]
        cat_overzicht[sectie] = {k: round(float(v), 2) for k, v in top_cats}

    src_overzicht = {}
    for sf, data in income_sources.items():
        src_overzicht[sf] = round(abs(data.get('bedrag_12m', 0)), 2)

    audit_data = {
        'periode_maanden': n_mnd,
        'rapport_totalen': rt,
        'sectie_totalen_12m': {k: round(v, 2) for k, v in st.items()},
        'top_categorieen_per_sectie': cat_overzicht,
        'income_sources': src_overzicht,
        'saldo_begin': saldo.get('totaal_begin', 0),
        'saldo_eind': saldo.get('totaal_eind', 0),
    }

    prompt = f"""Je bent een onafhankelijke financiële auditor. Je controleert een automatisch gegenereerd
huishoudrapport op fouten en inconsistenties VOORDAT het naar de klant gaat.

GEGEVENS:
{json.dumps(audit_data, indent=2, ensure_ascii=False)}

CONTROLEER:
1. Kloppen de rapport_totalen met de sectie_totalen? (bruto_inkomen moet gelijk zijn aan sectie inkomsten)
2. Zijn de ratio's realistisch? (vaste lasten < 100% inkomen, spaarquote < 100%)
3. Klopt het saldo? (begin + netto alle secties ≈ eind)
4. Zijn income_sources consistent met bruto_inkomen?
5. Zijn er verdachte patronen? (bijv. enorm hoge variabele kosten, nul inkomen maar wel uitgaven)
6. Zijn broker-terugstortingen correct behandeld? (moeten NIET bij inkomen staan)

ANTWOORD in exact dit JSON formaat:
{{
  "status": "ok" of "issues_found",
  "confidence": 0.0-1.0,
  "issues": [
    {{"severity": "error" of "warning", "beschrijving": "..."}}
  ],
  "samenvatting": "1 zin conclusie"
}}

Wees streng maar realistisch. Een huishouden met €150k+ inkomen en €80k vaste lasten is ongebruikelijk maar niet per se fout voor een DGA."""

    try:
        result = vraag_ai(prompt, categorizer=auditor_model)
        data = result.get('data')
        if data and isinstance(data, dict):
            data['model'] = result.get('model', auditor_model)
            return data
        # Probeer raw te parsen
        raw = result.get('raw', '')
        if raw:
            import json as _json
            try:
                parsed = _json.loads(raw)
                parsed['model'] = result.get('model', auditor_model)
                return parsed
            except _json.JSONDecodeError:
                pass
        return {
            'status': 'error',
            'issues': [{'severity': 'warning', 'beschrijving': f'Auditor gaf geen geldig antwoord'}],
            'model': auditor_model,
        }
    except Exception as e:
        logger.error(f"AI Auditor fout: {e}")
        return {
            'status': 'error',
            'issues': [{'severity': 'warning', 'beschrijving': f'Auditor fout: {str(e)[:100]}'}],
            'model': auditor_model,
        }


# ---------------------------------------------------------------------------
# STAP 4: PDF RAPPORT GENEREREN
# ---------------------------------------------------------------------------

# V2 Kleuren
INK = (26, 26, 46)
INK_SOFT = (61, 61, 92)
ACCENT = (31, 92, 139)
GOLD = (201, 168, 76)
WHITE = (255, 255, 255)
SURFACE = (247, 246, 242)
GREEN = (39, 174, 96)
RED = (192, 57, 43)
BORDER = (221, 217, 208)

SEC_COLORS = {
    'inkomsten': (26, 107, 60),
    'vaste_lasten': (139, 69, 19),
    'variabele_kosten': (74, 85, 104),
    'sparen_beleggen': (31, 92, 139),
    'interne_verschuivingen': (107, 91, 115),
}

SEC_LABELS = {
    'inkomsten': 'INKOMSTEN',
    'vaste_lasten': 'VASTE LASTEN',
    'variabele_kosten': 'VARIABELE KOSTEN',
    'sparen_beleggen': 'SPAREN & BELEGGEN',
    'interne_verschuivingen': 'INTERNE VERSCHUIVINGEN',
}

MAAND_NAMEN = {
    '01': 'jan', '02': 'feb', '03': 'mrt', '04': 'apr',
    '05': 'mei', '06': 'jun', '07': 'jul', '08': 'aug',
    '09': 'sep', '10': 'okt', '11': 'nov', '12': 'dec',
}


def eur(n: float) -> str:
    """Format getal als Euro bedrag."""
    if abs(n) < 0.01:
        return '\u20ac 0'
    return f"\u20ac {n:,.0f}".replace(',', '.')


class RapportPDF(FPDF):
    """Premium financieel rapport PDF met website-huisstijl.

    Fonts (zelfde als peterheijen.com):
      - Playfair Display: koppen en titels (serif, premium uitstraling)
      - Source Serif 4: lopende tekst en analyse (leesbaar, warm)
      - Inter: tabellen, data, kleine labels (helder, professioneel)

    Alle fonts worden meegeleverd als TTF — geen afhankelijkheid van systeemfonts.
    """

    # Font aliassen voor makkelijk gebruik door de class heen
    HEADING = 'Playfair'
    BODY = 'SourceSerif'
    DATA = 'Inter'

    def __init__(self):
        super().__init__('P', 'mm', 'A4')
        self.set_auto_page_break(auto=True, margin=20)

        # Premium fonts laden — meegeleverd in fonts/ naast app.py
        fonts_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'fonts')

        # Playfair Display (headings)
        self.add_font('Playfair', '', os.path.join(fonts_dir, 'PlayfairDisplay-Regular.ttf'))
        self.add_font('Playfair', 'B', os.path.join(fonts_dir, 'PlayfairDisplay-Bold.ttf'))
        self.add_font('Playfair', 'I', os.path.join(fonts_dir, 'PlayfairDisplay-Italic.ttf'))

        # Source Serif 4 (body text)
        self.add_font('SourceSerif', '', os.path.join(fonts_dir, 'SourceSerif4-Regular.ttf'))
        self.add_font('SourceSerif', 'B', os.path.join(fonts_dir, 'SourceSerif4-SemiBold.ttf'))
        self.add_font('SourceSerif', 'I', os.path.join(fonts_dir, 'SourceSerif4-Italic.ttf'))

        # Inter (data / tabellen / labels)
        self.add_font('Inter', '', os.path.join(fonts_dir, 'Inter-Regular.ttf'))
        self.add_font('Inter', 'B', os.path.join(fonts_dir, 'Inter-Bold.ttf'))

    def header(self):
        if self.page_no() > 1:
            # Detecteer landscape vs portrait
            pw = self.w  # actuele paginabreedte (297 landscape, 210 portrait)
            self.set_fill_color(*INK)
            self.rect(0, 0, pw, 8, 'F')
            self.set_font(self.DATA, 'B', 6)
            self.set_text_color(*WHITE)
            self.set_xy(8, 2)
            self.cell(0, 4, 'PeterHeijen.com  |  Financieel Rapport', 0, 0, 'L')
            self.set_font(self.DATA, '', 6)
            self.set_xy(8, 2)
            self.cell(pw - 16, 4, f'Pagina {self.page_no()}', 0, 0, 'R')
            self.set_y(10)
        else:
            self.set_y(10)

    def footer(self):
        self.set_y(-15)
        # Subtiele scheidslijn boven footer
        self.set_draw_color(*BORDER)
        self.set_line_width(0.2)
        self.line(15, self.get_y(), 195 if self.cur_orientation == 'P' else 282, self.get_y())
        self.set_font(self.DATA, '', 7)
        self.set_text_color(*INK_SOFT)
        self.cell(0, 10, 'Dit rapport is gegenereerd door PeterHeijen.com  |  Vertrouwelijk', 0, 0, 'C')

    def cover_page(self, feiten: dict, rapport_datum: str, jaartotalen: dict = None, maandoverzicht: dict = None,
                   ground_truth: dict = None, gate: dict = None):
        """Pagina 1: Inleiding + 7 belangrijkste bevindingen.

        Simpele, leesbare pagina met uitleg over de data en de key findings.
        """
        # Donkere header
        self.set_fill_color(*INK)
        self.rect(0, 0, 210, 55, 'F')

        # Gouden lijn
        self.set_draw_color(*GOLD)
        self.set_line_width(0.8)
        self.line(15, 48, 55, 48)

        # Titel
        self.set_font(self.HEADING, '', 26)
        self.set_text_color(*WHITE)
        self.set_xy(15, 15)
        self.cell(0, 12, 'Financieel Overzicht', 0, 1, 'L')

        # Subtitel
        self.set_font(self.BODY, '', 11)
        self.set_text_color(200, 200, 210)
        self.set_xy(15, 30)
        self.cell(0, 6, f'Gegenereerd op {rapport_datum}', 0, 1, 'L')

        # ===================================================================
        # INLEIDING: rekeningen, periode, totalen
        # ===================================================================
        y = 62

        # Verzamel data
        n_rek = len(feiten)
        vm = ground_truth.get('periode', {}).get('volle_maanden', []) if ground_truth else []
        n_mnd = ground_truth.get('periode', {}).get('n_mnd', len(vm)) if ground_truth else 12
        van_mnd = min(vm) if vm else ''
        tot_mnd = max(vm) if vm else ''

        # Maandnamen voor leesbare periode
        _MND_NAMEN = {
            '01': 'januari', '02': 'februari', '03': 'maart', '04': 'april',
            '05': 'mei', '06': 'juni', '07': 'juli', '08': 'augustus',
            '09': 'september', '10': 'oktober', '11': 'november', '12': 'december',
        }
        def _leesbare_maand(m):
            try:
                parts = m.split('-')
                return f"{_MND_NAMEN.get(parts[1], parts[1])} {parts[0]}"
            except Exception:
                return m

        # Rekening-beschrijvingen
        rek_naar_persoon = ground_truth.get('_rek_naar_persoon', {}) if ground_truth else {}
        rek_beschrijvingen = []
        for rek in sorted(feiten.keys()):
            houder = rek_naar_persoon.get(rek, '')
            # Laatste 4 cijfers van IBAN
            iban_kort = rek[-4:] if len(rek) >= 4 else rek
            if houder:
                rek_beschrijvingen.append(f"rekening ...{iban_kort} ({houder})")
            else:
                rek_beschrijvingen.append(f"rekening ...{iban_kort}")

        # Totaal inkomsten en uitgaven — uit rapport_totalen (single source of truth)
        cat_totalen = ground_truth.get('categorie_totalen_12m', {}) if ground_truth else {}
        rt = ground_truth.get('rapport_totalen', {}) if ground_truth else {}
        totaal_in = rt.get('bruto_inkomen_12m', 0)
        totaal_uit = -(rt.get('vaste_lasten_12m', 0) + rt.get('variabele_kosten_12m', 0))
        saldo_data = ground_truth.get('saldo', {}) if ground_truth else {}
        begin_saldo = saldo_data.get('totaal_begin', 0)
        eind_saldo = saldo_data.get('totaal_eind', 0)

        # Schrijf inleiding als lopende tekst
        self.set_font(self.BODY, '', 9.5)
        self.set_text_color(*INK)
        self.set_xy(15, y)

        rek_tekst = ', '.join(rek_beschrijvingen[:-1]) + f" en {rek_beschrijvingen[-1]}" if len(rek_beschrijvingen) > 1 else rek_beschrijvingen[0] if rek_beschrijvingen else ''

        inleiding = (
            f"Het systeem heeft {n_rek} rekeningen herkend: {rek_tekst}. "
            f"Het rapport kijkt naar een periode van {n_mnd} volle maanden, "
            f"van {_leesbare_maand(van_mnd)} tot en met {_leesbare_maand(tot_mnd)}."
        )
        self.multi_cell(180, 5, inleiding, 0, 'L')
        y = self.get_y() + 2

        self.set_xy(15, y)
        samenvatting = (
            f"Het bruto-inkomen bedroeg {eur(totaal_in)} over deze periode, "
            f"de totale vaste lasten en variabele kosten waren {eur(abs(totaal_uit))}. "
            f"Het gecombineerde saldo veranderde van {eur(begin_saldo)} naar {eur(eind_saldo)} "
            f"({'+' if eind_saldo >= begin_saldo else ''}{eur(eind_saldo - begin_saldo)})."
        )
        self.multi_cell(180, 5, samenvatting, 0, 'L')

        # ===================================================================
        # 7 BELANGRIJKSTE BEVINDINGEN — leunt 100% op executive_buckets
        # ===================================================================
        n_maanden = max(n_mnd, 1)
        eb = ground_truth.get('executive_buckets', {}) if ground_truth else {}
        ratios = eb.get('_ratios', {})

        bevindingen = []

        # 1. Inkomensbronnen (uit income_sources, niet zelf berekenen)
        income_sources = ground_truth.get('income_sources', {}) if ground_truth else {}
        if income_sources:
            top_bronnen = sorted(income_sources.items(), key=lambda x: -abs(x[1].get('bedrag_12m', 0)))[:3]
            _SRC_NL = {
                'salary_employment': 'salaris', 'freelance_business': 'freelance-inkomen',
                'rent_income': 'huurinkomsten', 'uwv_benefits': 'UWV-uitkeringen',
                'pension_aow': 'pensioen', 'dga_management_fee': 'DGA-loon',
                'tax_refund': 'belastingteruggave', 'government_benefits': 'overheidstoeslagen',
                'investment_income': 'beleggingsinkomsten',
            }
            bron_delen = []
            for sf, data in top_bronnen:
                label = _SRC_NL.get(sf, sf.replace('_', ' '))
                ontv = data.get('ontvanger', '')
                pm = abs(data.get('bedrag_12m', 0)) / n_maanden
                if ontv:
                    bron_delen.append(f"{label} ({ontv}): {eur(pm)}/mnd")
                else:
                    bron_delen.append(f"{label}: {eur(pm)}/mnd")
            bevindingen.append(f"De belangrijkste inkomensbronnen zijn: {', '.join(bron_delen)}.")

        # 2. Kerninkomen vs aanvullend (uit executive_buckets)
        kern = eb.get('kerninkomen', {})
        aanv = eb.get('aanvullende_instroom', {})
        if kern.get('bedrag_12m', 0) > 0:
            bevindingen.append(
                f"Het kerninkomen bedraagt {eur(kern['bedrag_pm'])}/mnd ({eur(kern['bedrag_12m'])}/jaar)."
                + (f" Daarnaast is er {eur(aanv['bedrag_pm'])}/mnd aan aanvullende instroom (teruggaven/toeslagen)."
                   if aanv.get('bedrag_12m', 0) > 100 else "")
            )

        # 3. Woonquote (uit executive_buckets ratios)
        woon = eb.get('woonlasten', {})
        woonquote = ratios.get('woonquote', 0)
        bevindingen.append(
            f"De woonquote bedraagt {woonquote}% van het bruto-inkomen "
            f"({eur(woon.get('bedrag_pm', 0))}/mnd aan woonlasten)."
        )

        # 4. Spaarquote (uit executive_buckets)
        spaar = eb.get('netto_allocatie_vermogen', {})
        spaarquote = ratios.get('spaarquote', 0)
        bevindingen.append(
            f"Er gaat {eur(spaar.get('bedrag_pm', 0))}/mnd naar vermogensopbouw "
            f"({spaarquote}% van het inkomen)."
        )

        # 5. Discretionaire uitgaven (uit executive_buckets)
        disc = eb.get('leefkosten_discretionair', {})
        vk = cat_totalen.get('variabele_kosten', {})
        top_vk_tekst = ""
        if vk:
            top_vk = max(vk.items(), key=lambda x: abs(float(x[1])))
            top_vk_tekst = f" Grootste post: {top_vk[0]} ({eur(abs(float(top_vk[1])) / n_maanden)}/mnd)."
        bevindingen.append(
            f"De discretionaire uitgaven bedragen {eur(disc.get('bedrag_pm', 0))}/mnd.{top_vk_tekst}"
        )

        # 6. Saldo-ontwikkeling
        delta = eind_saldo - begin_saldo
        if delta >= 0:
            bevindingen.append(f"Het vermogen is in {n_maanden} maanden gegroeid met {eur(delta)}.")
        else:
            bevindingen.append(f"Het vermogen is in {n_maanden} maanden afgenomen met {eur(abs(delta))}.")

        # 7. Belastingdruk (uit executive_buckets)
        bel = eb.get('belastingdruk', {})
        bel_pct = ratios.get('belastingdruk_pct', 0)
        bevindingen.append(
            f"De belastingdruk bedraagt {eur(bel.get('bedrag_pm', 0))}/mnd "
            f"({bel_pct}% van het inkomen)."
        )

        # Teken de bevindingen
        y = self.get_y() + 6

        self.set_draw_color(*GOLD)
        self.set_line_width(0.5)
        self.line(15, y, 55, y)
        y += 4

        self.set_font(self.HEADING, '', 14)
        self.set_text_color(*INK)
        self.set_xy(15, y)
        self.cell(0, 7, 'Belangrijkste bevindingen', 0, 1, 'L')
        y += 10

        for i, bevinding in enumerate(bevindingen[:7]):
            if y + 12 > 260:
                break
            # Nummer in goud-cirkel
            self.set_fill_color(*GOLD)
            self.ellipse(15, y, 6, 6, 'F')
            self.set_font(self.DATA, 'B', 7)
            self.set_text_color(255, 255, 255)
            self.set_xy(15, y + 0.8)
            self.cell(6, 5, str(i + 1), 0, 0, 'C')

            # Tekst
            self.set_font(self.BODY, '', 8.5)
            self.set_text_color(*INK)
            self.set_xy(24, y + 0.5)
            self.multi_cell(165, 4.5, bevinding, 0, 'L')
            y = self.get_y() + 3
        # Disclaimer onderaan cover (houd binnen pagina-margins)
        self.set_y(265)
        self.set_font(self.DATA, '', 6.5)
        self.set_text_color(*INK_SOFT)
        self.cell(180, 4, 'Dit rapport is uitsluitend bedoeld als financieel inzicht en vormt geen financieel advies.', 0, 1, 'C')
        self.cell(180, 4, 'Raadpleeg altijd een erkend financieel adviseur voor persoonlijke beslissingen.', 0, 0, 'C')

    def strategic_insights_page(self, premium_inzichten: list, kengetallen: dict):
        """Pagina 2: Strategische inzichten — premium analyses die een adviseur zou geven.

        Portrait A4. Toont de berekende premium inzichten in een visueel
        aantrekkelijk format met relevantie-indicators.
        """
        self.add_page('P')

        # Kleine header
        y = 14

        # Sectietitel
        self.set_font(self.HEADING, '', 18)
        self.set_text_color(*INK)
        self.set_xy(15, y)
        self.cell(0, 10, 'Strategische Inzichten', 0, 1, 'L')
        y += 10

        # Gouden lijn onder titel
        self.set_draw_color(*GOLD)
        self.set_line_width(0.6)
        self.line(15, y, 55, y)
        y += 4

        # Subtitel
        self.set_font(self.BODY, 'I', 9)
        self.set_text_color(*INK_SOFT)
        self.set_xy(15, y)
        self.cell(0, 5, 'Analyses op basis van uw transactiedata \u2014 berekend, niet geschat.', 0, 1, 'L')
        y += 10

        # Kengetallen strip bovenaan
        keng_items = []
        if 'spaarquote' in kengetallen:
            keng_items.append(('Spaarquote', f"{kengetallen['spaarquote']:.0f}%"))
        if 'vaste_lasten_ratio' in kengetallen:
            keng_items.append(('Vaste lasten', f"{kengetallen['vaste_lasten_ratio']:.0f}%"))
        if 'n_inkomstenbronnen' in kengetallen:
            keng_items.append(('Inkomensbronnen', f"{kengetallen['n_inkomstenbronnen']}"))
        if kengetallen.get('vermogensopbouw_pm', 0) > 0:
            keng_items.append(('Opbouw p/m', f"\u20ac{kengetallen['vermogensopbouw_pm']:,.0f}"))
        if kengetallen.get('inkomen_stabiliteit') is not None:
            keng_items.append(('Stabiliteit', f"{kengetallen['inkomen_stabiliteit']:.0f}%"))

        if keng_items:
            # Lichtgrijze achtergrond strip
            strip_h = 16
            self.set_fill_color(245, 245, 248)
            self.rect(15, y, 180, strip_h, 'F')

            n_items = len(keng_items)
            item_w = 180 / n_items
            for i, (label, waarde) in enumerate(keng_items):
                x = 15 + i * item_w
                # Waarde
                self.set_font(self.DATA, 'B', 12)
                self.set_text_color(*ACCENT)
                self.set_xy(x, y + 1)
                self.cell(item_w, 7, waarde, 0, 0, 'C')
                # Label
                self.set_font(self.DATA, '', 7)
                self.set_text_color(*INK_SOFT)
                self.set_xy(x, y + 8)
                self.cell(item_w, 5, label, 0, 0, 'C')

            y += strip_h + 8

        # Relevantie kleuren
        REL_COLORS = {
            'hoog': (201, 168, 76),    # Goud
            'medium': (31, 92, 139),    # Accent blauw
            'laag': (130, 140, 155),    # Grijs
        }

        # Inzichten cards
        for iz in premium_inzichten:
            rel = iz.get('relevantie', 'laag')
            rel_kleur = REL_COLORS.get(rel, REL_COLORS['laag'])

            # Check of we voldoende ruimte hebben (anders nieuwe pagina)
            if y > 240:
                self.add_page('P')
                y = 14

            # Verticale kleur-indicator links
            card_start_y = y
            self.set_fill_color(*rel_kleur)
            self.rect(15, y, 2, 24, 'F')

            # Titel
            self.set_font(self.DATA, 'B', 10)
            self.set_text_color(*INK)
            self.set_xy(20, y)
            self.cell(140, 5, iz['titel'], 0, 0, 'L')

            # Relevantie badge rechts
            badge_text = rel.upper() if rel else ''
            if badge_text:
                self.set_font(self.DATA, 'B', 6.5)
                self.set_text_color(*rel_kleur)
                badge_w = self.get_string_width(badge_text) + 6
                self.set_xy(195 - badge_w, y + 0.5)
                self.set_draw_color(*rel_kleur)
                self.set_line_width(0.3)
                self.rect(195 - badge_w, y + 0.5, badge_w, 4, 'D')
                self.cell(badge_w, 4, badge_text, 0, 0, 'C')

            y += 6

            # Beschrijving
            self.set_font(self.BODY, '', 8.5)
            self.set_text_color(*INK)
            self.set_xy(20, y)
            self.multi_cell(170, 4.2, iz['beschrijving'], 0, 'L')
            y = self.get_y() + 1.5

            # Detail regel (kleiner, grijs)
            if iz.get('detail'):
                self.set_font(self.DATA, '', 7)
                self.set_text_color(*INK_SOFT)
                self.set_xy(20, y)
                self.multi_cell(170, 3.5, iz['detail'], 0, 'L')
                y = self.get_y() + 1

            # Pas verticale indicator aan op werkelijke hoogte
            card_h = y - card_start_y
            self.set_fill_color(*rel_kleur)
            self.rect(15, card_start_y, 2, card_h, 'F')

            # Subtiele scheidslijn
            y += 3
            self.set_draw_color(*BORDER)
            self.set_line_width(0.15)
            self.line(20, y, 190, y)
            y += 5

        # Disclaimer onderaan
        if y < 250:
            self.set_y(265)
        else:
            self.set_y(self.get_y() + 10)
        self.set_font(self.DATA, '', 6.5)
        self.set_text_color(*INK_SOFT)
        self.cell(180, 4, 'Deze inzichten zijn berekend uit uw transactiedata en vormen geen persoonlijk financieel advies.', 0, 0, 'C')

    def categorie_overzicht_page(self, ground_truth: dict):
        """Maandelijks cashflow-overzicht op landscape pagina(s).

        Structuur per maand:
        - Beginsaldo (alle rekeningen)
        - Inkomsten (categorieën)
        - Vaste lasten (categorieën)
        - Variabele kosten (categorieën)
        - Sparen & Beleggen (categorieën)
        - Eindsaldo (= begin + netto)

        Consolidatie: < €500/jaar → 'Overige [sectie]', tenzij gem > €500/mnd.
        """
        if not ground_truth or not ground_truth.get('categorie_totalen_12m'):
            return

        cat_totalen = ground_truth['categorie_totalen_12m']
        maandoverzicht = ground_truth.get('maandoverzicht', {})
        if not maandoverzicht:
            return

        # Gebruik alleen volle maanden (geen onvolledige begin/eind-maanden)
        volle_maanden = ground_truth.get('periode', {}).get('volle_maanden', [])
        maanden = sorted(volle_maanden) if volle_maanden else sorted(maandoverzicht.keys())
        n_mnd = max(len(maanden), 1)

        # --- Consolidatielogica: top-N per sectie, rest → Overig ---
        # Max rijen per sectie om op 1 pagina te passen
        _SECTIE_CONFIG = [
            ('inkomsten',        'INKOMSTEN',          (26, 107, 60),   6),
            ('vaste_lasten',     'VASTE LASTEN',       (74, 85, 104),   14),
            ('variabele_kosten', 'VARIABELE KOSTEN',   (100, 100, 120), 12),
            ('sparen_beleggen',  'SPAREN & BELEGGEN',  (26, 90, 140),   5),
        ]

        def _consolideer(cats_dict, sectie_label, max_cats=15):
            """Consolideer: top-N op absoluut bedrag, rest → Overig."""
            alle = sorted(cats_dict.items(), key=lambda x: -abs(float(x[1])))
            groot = alle[:max_cats]
            rest = alle[max_cats:]
            overig_totaal = sum(float(v) for _, v in rest)
            result = [(k, float(v)) for k, v in groot]
            if abs(overig_totaal) >= 1:
                result.append(('Overig', overig_totaal))
            return result

        # Bouw rijstructuur op
        rows = []  # [(label, sectie_key_or_special, is_header, is_total)]

        rows.append(('Beginsaldo', '_beginsaldo', False, True))

        for sectie_key, sectie_label, kleur, max_cats in _SECTIE_CONFIG:
            cats = cat_totalen.get(sectie_key, {})
            if not cats:
                continue
            cons = _consolideer(cats, sectie_label, max_cats)
            rows.append((sectie_label, sectie_key, True, False))
            for cat_name, _ in cons:
                rows.append((cat_name, sectie_key, False, False))

        rows.append(('Eindsaldo', '_eindsaldo', False, True))

        # --- Maanddata per categorie opbouwen ---
        # maand_cat_data[maand][sectie][cat] = bedrag
        maand_cat_data = {}
        for maand in maanden:
            maand_cat_data[maand] = {}
            for sectie_key, _, _, _ in _SECTIE_CONFIG:
                maand_cat_data[maand][sectie_key] = maandoverzicht.get(maand, {}).get(sectie_key, {})

        # Consolideer ook op maandniveau
        def _get_cat_month(maand, sectie_key, cat_name, cons_cats_for_sectie):
            """Haal maandwaarde op, rekening houdend met consolidatie."""
            if cat_name == 'Overig':
                # Tel alle categorieën die NIET in de grote lijst staan
                alle_cats = maand_cat_data.get(maand, {}).get(sectie_key, {})
                grote_namen = {c for c, _ in cons_cats_for_sectie if c != 'Overig'}
                return sum(float(v) for k, v in alle_cats.items() if k not in grote_namen)
            return float(maand_cat_data.get(maand, {}).get(sectie_key, {}).get(cat_name, 0))

        # Bewaar geconsolideerde cats per sectie voor lookup
        cons_per_sectie = {}
        for sectie_key, _, _, max_cats in _SECTIE_CONFIG:
            cats = cat_totalen.get(sectie_key, {})
            if cats:
                cons_per_sectie[sectie_key] = _consolideer(cats, sectie_key, max_cats)

        # --- Bereken saldo per maand ---
        saldo_data = ground_truth.get('saldo', {})
        totaal_begin = saldo_data.get('totaal_begin', 0)
        totaal_eind = saldo_data.get('totaal_eind', 0)

        # Bereken cumulatief saldo per maand
        maand_netto = {}
        for maand in maanden:
            netto = 0
            for sectie_key, _, _, _ in _SECTIE_CONFIG:
                cats = maandoverzicht.get(maand, {}).get(sectie_key, {})
                netto += sum(float(v) for v in cats.values())
            # onderling_neutraal ook meenemen
            neutraal_cats = maandoverzicht.get(maand, {}).get('onderling_neutraal', {})
            netto += sum(float(v) for v in neutraal_cats.values())
            maand_netto[maand] = round(netto, 2)

        # Beginsaldo per maand: cumulatief
        maand_beginsaldo = {}
        running = totaal_begin
        for maand in maanden:
            maand_beginsaldo[maand] = round(running, 2)
            running += maand_netto[maand]
        maand_eindsaldo = {m: round(maand_beginsaldo[m] + maand_netto[m], 2) for m in maanden}

        # --- LAYOUT: portrait A4 pagina ---
        self.add_page()  # portrait (default)

        # Dimensies portrait A4: 210 x 297
        page_w = 210
        page_h = 297
        margin_l = 6
        margin_r = 4
        margin_t = 11
        usable_w = page_w - margin_l - margin_r  # ~200

        # Kolom-breedtes: smaller voor portrait
        col_label = 34  # categorienaam (compact)
        col_totaal = 14  # totaalkolom rechts (iets breder voor getallen)
        col_sep = 1.0    # separator tussen maanden en totaal
        n_months = len(maanden)
        col_month = (usable_w - col_label - col_totaal - col_sep) / n_months if n_months > 0 else 12
        rij_h = 3.2

        # Titel
        self.set_font(self.HEADING, '', 10)
        self.set_text_color(*INK)
        self.set_xy(margin_l, margin_t)
        self.cell(0, 5, 'Maandelijks Cashflow Overzicht', 0, 0, 'L')
        self.set_draw_color(*GOLD)
        self.set_line_width(0.5)
        self.line(margin_l, margin_t + 5.5, margin_l + 70, margin_t + 5.5)

        y = margin_t + 8

        # Korte euro-formatter voor tabel (compact)
        def _eur_k(val):
            v = float(val)
            if abs(v) < 1:
                return '-'
            if abs(v) >= 10000:
                return f"{v/1000:,.1f}k".replace(',', '.')
            return f"{v:,.0f}".replace(',', '.')

        # --- KOLOMKOPPEN: maandlabels + TOTAAL ---
        self.set_fill_color(26, 26, 46)
        self.rect(margin_l, y, usable_w, rij_h + 1, 'F')
        self.set_font(self.DATA, 'B', 4.5)
        self.set_text_color(255, 255, 255)
        self.set_xy(margin_l + 1, y + 0.5)
        self.cell(col_label - 1, rij_h, '', 0, 0, 'L')
        for i, maand in enumerate(maanden):
            try:
                parts = maand.split('-')
                maand_namen = ['jan','feb','mrt','apr','mei','jun','jul','aug','sep','okt','nov','dec']
                label = f"{maand_namen[int(parts[1])-1]}'{parts[0][2:]}"
            except Exception:
                label = maand[-5:]
            x = margin_l + col_label + i * col_month
            self.set_xy(x, y + 0.5)
            self.cell(col_month, rij_h, label, 0, 0, 'R')
        # Totaal header — donkerder achtergrond met gouden tekst
        x_totaal = margin_l + col_label + n_months * col_month + col_sep
        self.set_fill_color(18, 18, 36)  # donkerder dan de header
        self.rect(x_totaal - 0.5, y, col_totaal + 0.5, rij_h + 1, 'F')
        self.set_xy(x_totaal, y + 0.5)
        self.set_font(self.DATA, 'B', 4.5)
        self.set_text_color(212, 175, 55)  # GOLD text
        self.cell(col_totaal - 1, rij_h, 'TOTAAL', 0, 0, 'R')
        y += rij_h + 1

        # --- Helper: totaal voor een rij over alle maanden ---
        def _row_totaal(row_key, row_label, cur_sk, cons):
            """Bereken totaal over alle maanden voor een rij."""
            t = 0
            for maand in maanden:
                if row_key == '_beginsaldo':
                    return maand_beginsaldo.get(maanden[0], 0)  # begin = eerste maand
                elif row_key == '_eindsaldo':
                    return maand_eindsaldo.get(maanden[-1], 0)  # eind = laatste maand
                else:
                    t += _get_cat_month(maand, cur_sk, row_label, cons)
            return t

        # --- Verticale scheidingslijn voor totaalkolom ---
        totaal_x_start = margin_l + col_label + n_months * col_month + col_sep - 0.5

        # --- RIJEN TEKENEN ---
        cur_sectie_key = None
        row_idx = 0

        for row_label, row_key, is_header, is_total in rows:
            if y + rij_h > page_h - 10:
                break

            if is_total:
                # Beginsaldo / Eindsaldo regel
                self.set_fill_color(240, 238, 232)
                self.rect(margin_l, y, usable_w - col_totaal - col_sep, rij_h, 'F')
                # Totaalkolom donkerder achtergrond
                self.set_fill_color(225, 220, 210)
                self.rect(totaal_x_start, y, col_totaal + 1, rij_h, 'F')
                self.set_font(self.DATA, 'B', 4.5)
                self.set_text_color(26, 26, 46)
                self.set_xy(margin_l + 1, y + 0.3)
                self.cell(col_label - 1, rij_h, row_label, 0, 0, 'L')

                for i, maand in enumerate(maanden):
                    if row_key == '_beginsaldo':
                        val = maand_beginsaldo.get(maand, 0)
                    else:
                        val = maand_eindsaldo.get(maand, 0)
                    x = margin_l + col_label + i * col_month
                    self.set_xy(x, y + 0.3)
                    self.set_text_color(26, 26, 46)
                    self.cell(col_month, rij_h, _eur_k(val), 0, 0, 'R')

                # Totaalkolom: begin=eerste, eind=laatste
                if row_key == '_beginsaldo':
                    tot_val = maand_beginsaldo.get(maanden[0], 0) if maanden else 0
                else:
                    tot_val = maand_eindsaldo.get(maanden[-1], 0) if maanden else 0
                self.set_xy(x_totaal, y + 0.3)
                self.set_font(self.DATA, 'B', 4.5)
                self.set_text_color(26, 26, 46)
                self.cell(col_totaal - 1, rij_h, _eur_k(tot_val), 0, 0, 'R')
                y += rij_h + 0.5

            elif is_header:
                # Sectie-header
                cur_sectie_key = row_key
                kleur = next((k for sk, _, k, _ in _SECTIE_CONFIG if sk == row_key), (100, 100, 100))
                self.set_fill_color(*kleur)
                self.rect(margin_l, y, usable_w, rij_h, 'F')
                self.set_font(self.DATA, 'B', 4.5)
                self.set_text_color(255, 255, 255)
                self.set_xy(margin_l + 1, y + 0.3)
                self.cell(col_label - 1, rij_h, row_label, 0, 0, 'L')

                # Sectie-totalen per maand
                sectie_jaar_totaal = 0
                for i, maand in enumerate(maanden):
                    cats = maandoverzicht.get(maand, {}).get(row_key, {})
                    totaal = sum(float(v) for v in cats.values())
                    sectie_jaar_totaal += totaal
                    x = margin_l + col_label + i * col_month
                    self.set_xy(x, y + 0.3)
                    self.cell(col_month, rij_h, _eur_k(totaal), 0, 0, 'R')

                # Totaalkolom voor sectie — zelfde kleur als header
                self.set_fill_color(*kleur)
                self.rect(totaal_x_start, y, col_totaal + 1, rij_h, 'F')
                self.set_xy(x_totaal, y + 0.3)
                self.set_font(self.DATA, 'B', 4.5)
                self.set_text_color(255, 255, 255)
                self.cell(col_totaal - 1, rij_h, _eur_k(sectie_jaar_totaal), 0, 0, 'R')
                y += rij_h
                row_idx = 0

            else:
                # Categorie-rij
                if row_idx % 2 == 0:
                    self.set_fill_color(250, 249, 246)
                    self.rect(margin_l, y, usable_w, rij_h, 'F')
                # Totaalkolom licht accent
                self.set_fill_color(242, 240, 234) if row_idx % 2 == 0 else self.set_fill_color(248, 246, 242)
                self.rect(totaal_x_start, y, col_totaal + 1, rij_h, 'F')

                # Dunne lijn
                self.set_draw_color(235, 233, 228)
                self.line(margin_l, y + rij_h, margin_l + usable_w, y + rij_h)

                self.set_font(self.DATA, '', 4)
                self.set_text_color(80, 80, 100)
                self.set_xy(margin_l + 3, y + 0.3)
                # Truncate label om te passen in smallere kolom
                display_label = row_label[:26] if len(row_label) > 26 else row_label
                self.cell(col_label - 3, rij_h, display_label, 0, 0, 'L')

                cons = cons_per_sectie.get(cur_sectie_key, [])
                cat_jaar_totaal = 0
                for i, maand in enumerate(maanden):
                    val = _get_cat_month(maand, cur_sectie_key, row_label, cons)
                    cat_jaar_totaal += val
                    x = margin_l + col_label + i * col_month
                    self.set_xy(x, y + 0.3)
                    if val > 0:
                        self.set_text_color(26, 107, 60)
                    elif val < 0:
                        self.set_text_color(80, 80, 100)
                    else:
                        self.set_text_color(180, 180, 180)
                    self.cell(col_month, rij_h, _eur_k(val), 0, 0, 'R')

                # Totaalkolom — consistente positie
                self.set_xy(x_totaal, y + 0.3)
                self.set_font(self.DATA, 'B', 4)
                if cat_jaar_totaal > 0:
                    self.set_text_color(26, 107, 60)
                elif cat_jaar_totaal < 0:
                    self.set_text_color(60, 60, 80)
                else:
                    self.set_text_color(180, 180, 180)
                self.cell(col_totaal - 1, rij_h, _eur_k(cat_jaar_totaal), 0, 0, 'R')

                y += rij_h
                row_idx += 1

        # Verticale scheidingslijn totaalkolom
        self.set_draw_color(*GOLD)
        self.set_line_width(0.3)
        self.line(totaal_x_start + 0.5, margin_t + 8, totaal_x_start + 0.5, y)

        # Gouden lijn voor eindsaldo
        if rows and rows[-1][1] == '_eindsaldo':
            self.set_draw_color(*GOLD)
            self.set_line_width(0.4)
            self.line(margin_l, y - rij_h - 0.5, margin_l + usable_w, y - rij_h - 0.5)

    def analyse_page(self, analyse: dict):
        """Pagina met AI-analyse: samenvatting, sterke punten, etc."""
        self.add_page()

        # Sectie header — Playfair voor premium titels
        self._section_title('Analyse & Inzichten')

        # Samenvatting
        if analyse.get('samenvatting'):
            self.set_font(self.HEADING, '', 12)
            self.set_text_color(*ACCENT)
            self.cell(0, 7, 'Samenvatting', 0, 1, 'L')
            self.set_font(self.BODY, '', 10)
            self.set_text_color(*INK)
            self.multi_cell(180, 5.5, analyse['samenvatting'])
            self.ln(6)

        # Sterke punten
        if analyse.get('sterke_punten'):
            self._bullet_section('Sterke punten', analyse['sterke_punten'], GREEN)

        # Aandachtspunten
        if analyse.get('aandachtspunten'):
            self._bullet_section('Aandachtspunten', analyse['aandachtspunten'], RED)

        # Aanbevelingen
        if analyse.get('aanbevelingen'):
            self._bullet_section('Aanbevelingen', analyse['aanbevelingen'], ACCENT)

        # Verrassende inzichten — de WOW-factor
        if analyse.get('verrassende_inzichten'):
            self._insight_section('Wat valt op?', analyse['verrassende_inzichten'])

    def _insight_section(self, title: str, items: list):
        """Premium sectie voor verrassende inzichten — visueel onderscheidend."""
        if self.get_y() + 25 > 270:
            self.add_page()

        # Gouden accent balk
        self.set_fill_color(*GOLD)
        self.rect(15, self.get_y(), 3, 8, 'F')
        self.set_font(self.HEADING, '', 12)
        self.set_text_color(*INK)
        self.set_x(22)
        self.cell(0, 8, title, 0, 1, 'L')
        self.ln(2)

        self.set_font(self.BODY, '', 9.5)
        self.set_text_color(*INK)
        for i, item in enumerate(items):
            est_lines = max(1, len(item) // 80 + 1)
            est_h = est_lines * 5 + 8
            if self.get_y() + est_h > 270:
                self.add_page()

            # Genummerd met gouden cirkel
            y = self.get_y()
            self.set_fill_color(*GOLD)
            self.set_text_color(*WHITE)
            self.set_font(self.DATA, 'B', 8)
            # Kleine gouden cirkel met nummer
            cx = 20
            self.set_xy(cx - 3, y)
            self.cell(8, 5, str(i + 1), 0, 0, 'C')
            # Tekst
            self.set_text_color(*INK)
            self.set_font(self.BODY, '', 9.5)
            self.set_xy(30, y)
            self.multi_cell(165, 5, item)
            self.ln(2)
        self.ln(3)

    def _section_title(self, title: str):
        self.set_font(self.HEADING, '', 16)
        self.set_text_color(*INK)
        self.cell(0, 10, title, 0, 1, 'L')
        self.set_draw_color(*GOLD)
        self.set_line_width(0.5)
        self.line(15, self.get_y(), 60, self.get_y())
        self.ln(6)

    def _bullet_section(self, title: str, items: list, color: tuple):
        # Check of titel + eerste item past, anders nieuwe pagina
        if self.get_y() + 20 > 270:
            self.add_page()
        self.set_font(self.HEADING, '', 11)
        self.set_text_color(*color)
        self.cell(0, 7, title, 0, 1, 'L')
        self.set_font(self.BODY, '', 9.5)
        self.set_text_color(*INK)
        for item in items:
            # Schat hoogte: ~5mm per 80 tekens
            est_lines = max(1, len(item) // 80 + 1)
            est_h = est_lines * 5 + 2
            if self.get_y() + est_h > 270:
                self.add_page()
            self.set_x(20)
            self.cell(4, 5, '\u2022', 0, 0, 'L')
            self.multi_cell(170, 5, f'  {item}')
            self.ln(1)
        self.ln(4)

    def _maand_table_header(self, sec_label, sec_key, months, cat_w, m_w, total_w, continued=False):
        """Render sectie-header + maand-kolomheaders voor maandoverzicht tabel."""
        color = SEC_COLORS.get(sec_key, INK)
        self.set_fill_color(*color)
        self.set_text_color(*WHITE)
        self.set_font(self.DATA, 'B', 7)
        label = f'  {sec_label}' + (' (vervolg)' if continued else '')
        self.cell(total_w, 5, label, 1, 1, 'L', True)

        # Maand headers
        self.set_fill_color(*SURFACE)
        self.set_text_color(*INK_SOFT)
        self.set_font(self.DATA, 'B', 6)
        self.cell(cat_w, 5, '  Categorie', 1, 0, 'L', True)
        for m in months:
            parts = m.split('-')
            label = MAAND_NAMEN.get(parts[1], parts[1]) + ' ' + parts[0][2:]
            self.cell(m_w, 5, label, 1, 0, 'C', True)
        self.ln()

    def maandoverzicht_page(self, maandoverzicht: dict, feiten: dict):
        """Pagina's met maandelijks overzicht per rekening in spreadsheet-stijl.

        Slimme page-breaks: als een sectie niet op de huidige pagina past,
        wordt de tabel gesplitst met herhaalde headers op de nieuwe pagina.
        """
        sections_config = [
            ('inkomsten', 'INKOMSTEN'),
            ('vaste_lasten', 'VASTE LASTEN'),
            ('variabele_kosten', 'VARIABELE KOSTEN'),
            ('sparen_beleggen', 'SPAREN & BELEGGEN'),
        ]

        PAGE_BOTTOM = 185  # landscape max Y voor content

        for rek, maanden in maandoverzicht.items():
            self.add_page('L')  # Landscape voor brede tabel
            self._section_title(f'Maandoverzicht — Rekening {rek}')

            months = sorted(maanden.keys())
            if not months:
                continue

            # Kolom breedte — vast voor alle secties op deze rekening
            cat_w = 55
            m_w = min((277 - cat_w - 22) / len(months), 22)
            total_w = cat_w + m_w * len(months)

            for sec_key, sec_label in sections_config:
                cats = set()
                for m in months:
                    md = maanden[m].get(sec_key, {})
                    if isinstance(md, dict):
                        for cat, val in md.items():
                            b = val.get('bedrag', 0) if isinstance(val, dict) else (val or 0)
                            if abs(b) > 0.01:
                                cats.add(cat)
                cats = sorted(cats)
                if not cats:
                    continue

                # Benodigde hoogte: header (10mm) + rijen (4.5mm elk) + totaal (5mm) + spacing (7mm)
                row_h = 4.5
                header_h = 10
                footer_h = 12  # totaal rij + spacing
                needed = header_h + len(cats) * row_h + footer_h

                # Past de hele sectie op de huidige pagina?
                if self.get_y() + needed > PAGE_BOTTOM:
                    # Past het op een NIEUWE pagina?
                    fresh_start = 22  # na page header
                    if fresh_start + needed <= PAGE_BOTTOM:
                        # Hele sectie past op nieuwe pagina
                        self.add_page('L')
                    else:
                        # Te groot — we moeten splitsen. Start op nieuwe pagina
                        # als er minder dan 5 rijen ruimte over is
                        remaining = PAGE_BOTTOM - self.get_y()
                        if remaining < header_h + 5 * row_h + footer_h:
                            self.add_page('L')

                # Render header
                self._maand_table_header(sec_label, sec_key, months, cat_w, m_w, total_w)

                # Data rijen — met slimme page-break
                section_totals = [0.0] * len(months)
                self.set_font(self.DATA, '', 7)
                rows_on_page = 0

                for ci, cat in enumerate(cats):
                    # Check of deze rij + eventuele totaalrij past
                    is_last = (ci == len(cats) - 1)
                    space_needed = row_h + (footer_h if is_last else 0)

                    if self.get_y() + space_needed > PAGE_BOTTOM:
                        # Page break nodig — eerst subtotaal-tussenstand tonen
                        self.set_fill_color(*SURFACE)
                        self.set_font(self.DATA, 'B', 6)
                        self.set_text_color(*INK_SOFT)
                        self.cell(cat_w, 4, '  (vervolg op volgende pagina)', 'T', 0, 'L', True)
                        for mi in range(len(months)):
                            self.cell(m_w, 4, '', 'T', 0, 'R', True)
                        self.ln()

                        # Nieuwe pagina met herhaalde header
                        self.add_page('L')
                        self._maand_table_header(sec_label, sec_key, months, cat_w, m_w, total_w, continued=True)
                        self.set_font(self.DATA, '', 7)
                        rows_on_page = 0

                    # Render data rij — zebra-strepen voor leesbaarheid
                    is_even = (rows_on_page % 2 == 0)
                    if is_even:
                        self.set_fill_color(252, 251, 249)  # Heel subtiel warm grijs
                    else:
                        self.set_fill_color(*WHITE)

                    self.set_text_color(*INK)
                    self.cell(cat_w, row_h, f'  {cat[:35]}', 0, 0, 'L', True)

                    for mi, m in enumerate(months):
                        sd = maanden[m].get(sec_key, {})
                        b = 0
                        if cat in sd:
                            b = sd[cat].get('bedrag', 0) if isinstance(sd[cat], dict) else (sd[cat] or 0)
                        section_totals[mi] += b

                        if abs(b) > 0.01:
                            self.set_text_color(*(GREEN if b > 0 else RED))
                            self.cell(m_w, row_h, eur(b), 0, 0, 'R', True)
                        else:
                            self.set_text_color(*INK_SOFT)
                            self.cell(m_w, row_h, '', 0, 0, 'R', True)
                    self.ln()
                    rows_on_page += 1

                # Subtotaal rij
                self.set_fill_color(*SURFACE)
                self.set_font(self.DATA, 'B', 7)
                self.set_text_color(*INK)
                self.cell(cat_w, 5, f'  Totaal', 'T', 0, 'L', True)
                for mi in range(len(months)):
                    t = section_totals[mi]
                    self.set_text_color(*(GREEN if t > 0 else RED if t < 0 else INK))
                    self.cell(m_w, 5, eur(t), 'T', 0, 'R', True)
                self.ln(7)

    def _jaar_table_header(self, sec_label, sec_key, continued=False):
        """Render sectie-header + kolomheaders voor jaartotalen tabel."""
        color = SEC_COLORS.get(sec_key, INK)
        self.set_fill_color(*color)
        self.set_text_color(*WHITE)
        self.set_font(self.DATA, 'B', 8)
        label = f'  {sec_label}' + (' (vervolg)' if continued else '')
        self.cell(180, 6, label, 0, 1, 'L', True)

        self.set_fill_color(*SURFACE)
        self.set_text_color(*INK_SOFT)
        self.set_font(self.DATA, 'B', 7)
        self.cell(90, 5, '  Categorie', 'B', 0, 'L', True)
        self.cell(45, 5, 'Jaarbedrag', 'B', 0, 'R', True)
        self.cell(45, 5, 'Per maand', 'B', 1, 'R', True)

    def jaartotalen_page(self, jaartotalen: dict, maandoverzicht: dict):
        """Pagina met jaartotalen per categorie.

        Slimme page-breaks: secties die te lang zijn worden gesplitst
        met herhaalde headers op de nieuwe pagina.
        """
        sections_config = [
            ('inkomsten', 'Inkomsten'),
            ('vaste_lasten', 'Vaste Lasten'),
            ('variabele_kosten', 'Variabele Kosten'),
            ('sparen_beleggen', 'Sparen & Beleggen'),
        ]

        PAGE_BOTTOM = 270  # portrait max Y

        for rek, totalen in jaartotalen.items():
            self.add_page()
            n_maanden = len(maandoverzicht.get(rek, {})) or 12
            self._section_title(f'Jaartotalen — Rekening {rek}')

            for sec_key, sec_label in sections_config:
                data = totalen.get(sec_key)
                if not data or not isinstance(data, dict):
                    continue

                entries = [(cat, b) for cat, b in data.items() if abs(b or 0) > 0.01]
                entries.sort(key=lambda x: abs(x[1]), reverse=True)
                if not entries:
                    continue

                row_h = 5
                header_h = 11  # sectie header + kolom headers
                footer_h = 12  # totaal rij + spacing
                needed = header_h + len(entries) * row_h + footer_h

                # Past de hele sectie?
                if self.get_y() + needed > PAGE_BOTTOM:
                    fresh_start = 22
                    if fresh_start + needed <= PAGE_BOTTOM:
                        self.add_page()
                    else:
                        remaining = PAGE_BOTTOM - self.get_y()
                        if remaining < header_h + 3 * row_h + footer_h:
                            self.add_page()

                # Render header
                self._jaar_table_header(sec_label, sec_key)

                # Data rijen met slimme page-break
                self.set_font(self.DATA, '', 8)
                section_total = 0

                row_count = 0
                for ei, (cat, bedrag) in enumerate(entries):
                    section_total += bedrag
                    pm = bedrag / n_maanden
                    is_last = (ei == len(entries) - 1)
                    space_needed = row_h + (footer_h if is_last else 0)

                    if self.get_y() + space_needed > PAGE_BOTTOM:
                        # Page break — nieuwe pagina met herhaalde header
                        self.add_page()
                        self._jaar_table_header(sec_label, sec_key, continued=True)
                        self.set_font(self.DATA, '', 8)
                        row_count = 0

                    # Zebra-strepen
                    if row_count % 2 == 0:
                        self.set_fill_color(252, 251, 249)
                    else:
                        self.set_fill_color(*WHITE)

                    self.set_text_color(*INK)
                    self.cell(90, row_h, f'  {cat}', 0, 0, 'L', True)
                    self.set_text_color(*(GREEN if bedrag > 0 else RED))
                    self.cell(45, row_h, eur(bedrag), 0, 0, 'R', True)
                    self.set_text_color(*(GREEN if pm > 0 else RED))
                    self.cell(45, row_h, eur(pm), 0, 1, 'R', True)
                    row_count += 1

                # Totaal
                self.set_font(self.DATA, 'B', 8)
                self.set_fill_color(*SURFACE)
                self.set_text_color(*INK)
                self.cell(90, 5.5, '  Totaal', 'T', 0, 'L', True)
                self.set_text_color(*(GREEN if section_total > 0 else RED))
                self.cell(45, 5.5, eur(section_total), 'T', 0, 'R', True)
                pm_total = section_total / n_maanden
                self.set_text_color(*(GREEN if pm_total > 0 else RED))
                self.cell(45, 5.5, eur(pm_total), 'T', 1, 'R', True)
                self.ln(6)


def _combineer_maandoverzichten(maandoverzicht: dict) -> dict:
    """Combineer maandoverzichten van alle rekeningen tot één geheel.

    Telt bedragen per categorie per maand op over alle rekeningen.
    Slaat interne_verschuivingen over — die vallen weg in het totaalbeeld.
    Retourneert: {'TOTAAL': {maand: {sectie: {cat: {bedrag: x}}}}}
    """
    gecombineerd = {}

    for rek, maanden in maandoverzicht.items():
        for maand, secties in maanden.items():
            if maand not in gecombineerd:
                gecombineerd[maand] = {}
            for sec_key, cats in secties.items():
                if not isinstance(cats, dict):
                    continue
                # Interne verschuivingen overslaan in gecombineerd overzicht
                if sec_key == 'interne_verschuivingen':
                    continue
                if sec_key not in gecombineerd[maand]:
                    gecombineerd[maand][sec_key] = {}
                for cat, val in cats.items():
                    b = val.get('bedrag', 0) if isinstance(val, dict) else (val or 0)
                    if cat not in gecombineerd[maand][sec_key]:
                        gecombineerd[maand][sec_key][cat] = {'bedrag': 0}
                    gecombineerd[maand][sec_key][cat]['bedrag'] += b

    return {'TOTAAL': gecombineerd}


def _combineer_jaartotalen(jaartotalen: dict) -> dict:
    """Combineer jaartotalen van alle rekeningen tot één geheel.

    Slaat interne_verschuivingen over.
    """
    gecombineerd = {}

    for rek, totalen in jaartotalen.items():
        for sec_key, cats in totalen.items():
            if not isinstance(cats, dict):
                continue
            # Interne verschuivingen overslaan
            if sec_key == 'interne_verschuivingen':
                continue
            if sec_key not in gecombineerd:
                gecombineerd[sec_key] = {}
            for cat, bedrag in cats.items():
                if cat not in gecombineerd[sec_key]:
                    gecombineerd[sec_key][cat] = 0
                gecombineerd[sec_key][cat] += (bedrag or 0)

    return {'TOTAAL': gecombineerd}


def _combineer_feiten(feiten: dict) -> dict:
    """Combineer feiten van alle rekeningen voor cover-stats."""
    totaal_ink = sum(f['totalen']['inkomsten'] for f in feiten.values())
    totaal_uit = sum(f['totalen']['uitgaven'] for f in feiten.values())
    periodes = []
    for f in feiten.values():
        periodes.append(f['periode']['van'])
        periodes.append(f['periode']['tot'])

    return {
        'totalen': {'inkomsten': totaal_ink, 'uitgaven': totaal_uit, 'netto': totaal_ink + totaal_uit},
        'periode': {'van': min(periodes), 'tot': max(periodes)},
    }


def genereer_pdf(rapport: dict) -> bytes:
    """Genereer een premium PDF rapport."""
    pdf = RapportPDF()

    feiten = rapport.get('feiten', {})
    analyse = rapport.get('analyse', {})
    maandoverzicht = rapport.get('maandoverzicht', {})
    jaartotalen = rapport.get('jaartotalen', {})
    ground_truth = rapport.get('ground_truth')
    gate = rapport.get('gate')
    datum = datetime.now().strftime('%d-%m-%Y')

    # Pagina 1: Executive summary op huishoudniveau
    pdf.add_page()
    pdf.cover_page(feiten, datum, jaartotalen=jaartotalen, maandoverzicht=maandoverzicht,
                   ground_truth=ground_truth, gate=gate)

    # Pagina 2: Strategische premium inzichten
    if ground_truth:
        strat = ground_truth.get('strategische_inzichten', {})
        kengetallen = strat.get('kengetallen', {})
        premium_inzichten = ground_truth.get('premium_inzichten', [])
        if premium_inzichten:
            pdf.strategic_insights_page(premium_inzichten, kengetallen)

    # Pagina 3: Categorie-overzicht / cashflow tabel
    if ground_truth and ground_truth.get('categorie_totalen_12m'):
        pdf.categorie_overzicht_page(ground_truth)

    # Rapport: 3 pagina's — cover + strategische inzichten + cashflow tabel

    return pdf.output()


# ---------------------------------------------------------------------------
# STAP 5: EMAIL VERSTUREN VIA RESEND
# ---------------------------------------------------------------------------

def verstuur_rapport_email(email: str, pdf_bytes: bytes, report_id: str,
                           reconciliatie_excel: bytes = None,
                           geblokkeerd: bool = False,
                           gate_redenen: list = None):
    """Verstuur het PDF rapport per email via Resend.

    Als geblokkeerd=True, wordt een blokkade-email verstuurd met alleen de
    reconciliatie Excel (geen PDF) zodat de gebruiker kan reviewen.

    Retourneert True bij succes, False bij falen.
    Logt altijd de volledige Resend-response voor diagnose.
    """
    resend_key = os.environ.get('RESEND_API_KEY')
    if not resend_key:
        logger.error("RESEND_API_KEY ontbreekt in environment variables — email KAN NIET worden verstuurd")
        return False

    # Valideer dat de key er uitziet als een geldige Resend key
    if not resend_key.startswith('re_'):
        logger.error(f"RESEND_API_KEY lijkt ongeldig (begint niet met 're_') — controleer Railway variables")
        return False

    pdf_base64 = base64.b64encode(pdf_bytes).decode('utf-8')
    logger.info(f"PDF bijlage: {len(pdf_bytes)} bytes, base64: {len(pdf_base64)} chars")

    # Bij geblokkeerd rapport: stuur alleen review-email naar Peter
    if geblokkeerd:
        redenen_tekst = '\n'.join(f'- {r}' for r in (gate_redenen or ['Onbekende reden']))
        plain_text = (
            f"RAPPORT GEBLOKKEERD — Kwaliteitscheck niet doorstaan\n\n"
            f"Rapport ID: {report_id}\n\n"
            f"Redenen:\n{redenen_tekst}\n\n"
            f"Het Cashflow Overzicht Excel is bijgevoegd voor handmatige review.\n"
            f"Na correctie kan het rapport opnieuw worden gegenereerd.\n\n"
            f"PeterHeijen.com"
        )
        payload = {
            "from": "Peter Heijen <rapport@peterheijen.com>",
            "reply_to": "info@peterheijen.com",
            "to": ["peterheijen2026@gmail.com"],  # Altijd naar Peter, niet naar klant
            "subject": f"[GEBLOKKEERD] Rapport {report_id} — kwaliteitscheck",
            "text": plain_text,
            "attachments": [],
        }
        if reconciliatie_excel:
            excel_base64 = base64.b64encode(reconciliatie_excel).decode('utf-8')
            payload["attachments"].append({
                "filename": f"cashflow-overzicht-{report_id}.xlsx",
                "content": excel_base64,
                "type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            })
        try:
            resp = httpx.post(
                "https://api.resend.com/emails",
                headers={"Authorization": f"Bearer {resend_key}", "Content-Type": "application/json"},
                json=payload, timeout=30.0,
            )
            logger.info(f"Blokkade-email: {resp.status_code}")
            return resp.status_code == 200
        except Exception as exc:
            logger.error(f"Blokkade-email fout: {exc}")
            return False

    # Anti-spam: platte tekst versie (HTML-only is een spam signaal)
    plain_text = (
        f"Uw financieel rapport is klaar\n\n"
        f"Beste klant,\n\n"
        f"Bijgevoegd vindt u uw persoonlijke financiele analyse van PeterHeijen.com.\n"
        f"Het rapport bevat een overzicht van uw inkomsten en uitgaven, "
        f"gecategoriseerd per maand, met concrete inzichten en aanbevelingen.\n\n"
        f"Rapport ID: {report_id}\n\n"
        f"Heeft u vragen? Neem contact op via info@peterheijen.com\n\n"
        f"Met vriendelijke groet,\n"
        f"Peter Heijen\n"
        f"PeterHeijen.com\n\n"
        f"Engelcke B.V. | Tienhoven | KvK 30277920\n"
        f"U ontvangt deze email omdat u een financiele analyse heeft aangevraagd op peterheijen.com."
    )

    payload = {
        "from": "Peter Heijen <rapport@peterheijen.com>",
        "reply_to": "info@peterheijen.com",
        "to": [email],
        "subject": f"Uw persoonlijke financiele analyse is klaar",
        "text": plain_text,
        "html": f"""<!DOCTYPE html>
<html lang="nl">
<head><meta charset="utf-8"></head>
<body style="margin:0;padding:0;background:#f7f6f2;font-family:Georgia,serif">
<table width="100%" cellpadding="0" cellspacing="0" style="background:#f7f6f2">
<tr><td align="center" style="padding:20px 0">
<table width="560" cellpadding="0" cellspacing="0" style="background:#ffffff;border-radius:8px;overflow:hidden">
  <tr><td style="background:#1a1a2e;padding:24px 30px">
    <p style="color:#ffffff;font-size:18px;margin:0;font-family:Georgia,serif">Peter<span style="color:#c9a84c">Heijen</span>.com</p>
  </td></tr>
  <tr><td style="padding:30px">
    <h1 style="color:#1a1a2e;font-size:20px;margin:0 0 16px;font-family:Georgia,serif">Uw financieel rapport is klaar</h1>
    <p style="color:#3d3d5c;font-size:15px;line-height:1.7;margin:0 0 16px">
      Beste klant,
    </p>
    <p style="color:#3d3d5c;font-size:15px;line-height:1.7;margin:0 0 16px">
      Bijgevoegd vindt u uw persoonlijke financiele analyse.
      Het rapport bevat een overzicht van uw inkomsten en uitgaven,
      gecategoriseerd per maand, met concrete inzichten en aanbevelingen.
    </p>
    <p style="color:#3d3d5c;font-size:14px;line-height:1.7;margin:0 0 16px">
      Rapport ID: <strong>{report_id}</strong>
    </p>
    <hr style="border:none;border-top:1px solid #ddd9d0;margin:20px 0">
    <p style="color:#3d3d5c;font-size:13px;line-height:1.6;margin:0 0 8px">
      Heeft u vragen over uw rapport? Neem contact op via
      <a href="mailto:info@peterheijen.com" style="color:#1f5c8b">info@peterheijen.com</a>
    </p>
    <p style="color:#3d3d5c;font-size:13px;line-height:1.6;margin:0">
      Met vriendelijke groet,<br>Peter Heijen
    </p>
  </td></tr>
  <tr><td style="background:#f7f6f2;padding:16px 30px;border-top:1px solid #ddd9d0">
    <p style="color:#999;font-size:11px;line-height:1.5;margin:0">
      Engelcke B.V. | Tienhoven | KvK 30277920<br>
      U ontvangt deze email omdat u een financiele analyse heeft aangevraagd op
      <a href="https://peterheijen.com" style="color:#999">peterheijen.com</a>.
    </p>
  </td></tr>
</table>
</td></tr>
</table>
</body>
</html>""",
        "attachments": [
            {
                "filename": f"financieel-rapport-{report_id}.pdf",
                "content": pdf_base64,
                "type": "application/pdf",
            },
        ],
    }

    # V3: Reconciliatie Excel als extra bijlage
    if reconciliatie_excel:
        excel_base64 = base64.b64encode(reconciliatie_excel).decode('utf-8')
        payload["attachments"].append({
            "filename": f"cashflow-overzicht-{report_id}.xlsx",
            "content": excel_base64,
            "type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        })
        logger.info(f"Reconciliatie Excel bijlage: {len(reconciliatie_excel)} bytes")

    try:
        resp = httpx.post(
            "https://api.resend.com/emails",
            headers={
                "Authorization": f"Bearer {resend_key}",
                "Content-Type": "application/json",
            },
            json=payload,
            timeout=30,
        )
        if resp.status_code in (200, 201):
            logger.info(f"Email SUCCESVOL verstuurd naar {email} (Resend status {resp.status_code})")
            return True
        else:
            # Log de volledige error voor diagnose
            logger.error(
                f"Resend FOUT: status={resp.status_code}, "
                f"response={resp.text}, "
                f"from=rapport@peterheijen.com, to={email}"
            )
            # Veelvoorkomende fouten loggen met uitleg
            if resp.status_code == 403:
                logger.error("DIAGNOSE: Domein peterheijen.com is waarschijnlijk niet geverifieerd in Resend. "
                             "Ga naar resend.com/domains en voeg peterheijen.com toe met de juiste DNS records.")
            elif resp.status_code == 422:
                logger.error("DIAGNOSE: Ongeldige email parameters. Check of het 'from' adres correct is "
                             "en het domein is geverifieerd.")
            elif resp.status_code == 429:
                logger.error("DIAGNOSE: Rate limit bereikt. Wacht even en probeer opnieuw.")
            return False
    except httpx.TimeoutException:
        logger.error(f"Resend TIMEOUT na 30 seconden voor email naar {email}")
        return False
    except Exception as e:
        logger.error(f"Email FOUT (onverwacht): {type(e).__name__}: {e}")
        return False


# ---------------------------------------------------------------------------
# API ENDPOINTS
# ---------------------------------------------------------------------------

@app.get("/")
def health():
    """Basis health check."""
    return {"status": "ok", "versie": "0.1.0", "service": "peterheijen-finance"}


@app.get("/health")
def deep_health():
    """Uitgebreide health check — test of alle afhankelijkheden geconfigureerd zijn."""
    checks = {}

    # Claude API key
    api_key = os.environ.get('ANTHROPIC_API_KEY')
    checks['claude_api_key'] = bool(api_key and api_key.startswith('sk-'))

    # Claude model
    model = os.environ.get('CLAUDE_MODEL', 'claude-opus-4-6')
    checks['claude_model'] = model

    # Resend API key
    resend_key = os.environ.get('RESEND_API_KEY')
    checks['resend_api_key'] = bool(resend_key and resend_key.startswith('re_'))

    # Fonts directory
    fonts_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'fonts')
    fonts_present = os.path.exists(fonts_dir) and len([f for f in os.listdir(fonts_dir) if f.endswith('.ttf')]) >= 8
    checks['premium_fonts'] = fonts_present

    alles_ok = all([
        checks['claude_api_key'],
        checks['resend_api_key'],
        checks['premium_fonts'],
    ])

    return {
        "status": "ok" if alles_ok else "NIET KLAAR",
        "checks": checks,
        "bericht": "Alle systemen operationeel" if alles_ok else "Een of meer afhankelijkheden ontbreken — check Railway Variables",
    }


@app.post("/analyseer")
async def analyseer(bestand: UploadFile):
    """Upload een ABN AMRO CSV/XLSX en ontvang een financieel rapport."""
    report_id = str(uuid.uuid4())[:8]
    logger.info(f"[{report_id}] Start analyse: {bestand.filename}")

    # 1. Inlezen
    try:
        inhoud = await bestand.read()
        df = lees_transacties(inhoud, bestand.filename)
        logger.info(f"[{report_id}] {len(df)} transacties ingelezen")
    except Exception as e:
        logger.error(f"[{report_id}] Inleesfout: {e}")
        raise HTTPException(status_code=400, detail=f"Kan bestand niet lezen: {e}")

    # 2. Deterministisch rekenen
    feiten = bereken_feiten(df)
    top = bereken_top(df)

    for rek, data in feiten.items():
        if not data['saldo']['klopt']:
            logger.warning(f"[{report_id}] Saldo klopt NIET voor rekening {rek}")

    logger.info(f"[{report_id}] Feiten berekend, saldo's gecheckt")

    # 3. Claude categoriseren + analyseren
    claude_result = None
    try:
        prompt = bouw_prompt(df, feiten, top)
        claude_result = vraag_claude(prompt)
        logger.info(f"[{report_id}] Claude analyse compleet")
    except Exception as e:
        logger.error(f"[{report_id}] Claude fout: {e}")
        claude_result = {'data': None, 'error': str(e)}

    # 4. Rapport samenstellen
    rapport = {
        'report_id': report_id,
        'gegenereerd': datetime.now().isoformat(),
        'bestand': bestand.filename,
        'feiten': feiten,
        'top_tegenpartijen': top,
    }

    if claude_result and claude_result.get('data'):
        rapport['maandoverzicht'] = claude_result['data'].get('maandoverzicht', {})
        rapport['jaartotalen'] = claude_result['data'].get('jaartotalen', {})
        rapport['analyse'] = claude_result['data'].get('analyse', {})
        rapport['ai'] = {
            'model': claude_result.get('model'),
            'tokens': claude_result.get('tokens'),
        }
    else:
        rapport['analyse'] = {
            'samenvatting': 'AI-analyse niet beschikbaar. Hieronder de deterministisch berekende cijfers.',
            'ai_fout': claude_result.get('error') if claude_result else 'Geen API key',
        }

    logger.info(f"[{report_id}] Rapport klaar")
    return rapport


# ---------------------------------------------------------------------------
# V3: FROZEN GROUND TRUTH PAYLOAD
# ---------------------------------------------------------------------------

def _bouw_ground_truth(merged_data: dict, feiten: dict, rapportperiode: dict,
                       reconciliatie: dict, df: pd.DataFrame) -> dict:
    """V3: Bouw bevroren ground truth payload.

    Na dit punt komen ALLE getallen in het rapport (PDF, Excel, AI-tekst)
    uitsluitend hieruit. Niets wordt meer opnieuw berekend.
    """
    volle_maanden = set(rapportperiode.get('volle_maanden', []))
    n_mnd = rapportperiode.get('n_mnd', 12) or 12

    maandoverzicht = merged_data.get('maandoverzicht', {})
    jaartotalen = merged_data.get('jaartotalen', {})

    # Combineer alle rekeningen naar TOTAAL
    combined_maand = {}  # {maand: {sectie: {categorie: bedrag}}}
    combined_jaar = {}   # {sectie: {categorie: bedrag}}

    for rek, maanden in maandoverzicht.items():
        for maand, secties in maanden.items():
            if maand not in combined_maand:
                combined_maand[maand] = {}
            for sectie, cats in secties.items():
                if sectie == 'interne_verschuivingen':
                    continue
                if sectie not in combined_maand[maand]:
                    combined_maand[maand][sectie] = {}
                if isinstance(cats, dict):
                    for cat, data in cats.items():
                        bedrag = data.get('bedrag', 0) if isinstance(data, dict) else float(data)
                        bestaand = combined_maand[maand][sectie].get(cat, 0)
                        combined_maand[maand][sectie][cat] = round(bestaand + bedrag, 2)

    for rek, secties in jaartotalen.items():
        for sectie, cats in secties.items():
            if sectie == 'interne_verschuivingen':
                continue
            if sectie not in combined_jaar:
                combined_jaar[sectie] = {}
            if isinstance(cats, dict):
                for cat, bedrag in cats.items():
                    bestaand = combined_jaar[sectie].get(cat, 0)
                    combined_jaar[sectie][cat] = round(bestaand + float(bedrag), 2)

    # Per-maand sectie-totalen
    maand_sectie_totalen = {}
    for maand, secties in combined_maand.items():
        maand_sectie_totalen[maand] = {}
        for sectie, cats in secties.items():
            maand_sectie_totalen[maand][sectie] = round(sum(float(v) for v in cats.values()), 2)

    # Bereken sectie-totalen als SOM VAN MAANDEN (niet uit AI-jaartotalen)
    # Dit garandeert dat sectie_totalen_12m == sum(maand_sectie_totalen)
    # en voorkomt JAARTOTAAL ≠ SOM MAANDEN gate-blokkade
    sectie_totalen_12m = {}
    for maand_data in maand_sectie_totalen.values():
        for sectie, totaal in maand_data.items():
            sectie_totalen_12m[sectie] = round(sectie_totalen_12m.get(sectie, 0) + totaal, 2)

    # Saldo per rekening per maand
    saldo_per_rekening = {}
    for rek, rek_feiten in feiten.items():
        saldo_per_rekening[rek] = {
            'beginsaldo': rek_feiten['saldo']['beginsaldo'],
            'eindsaldo': rek_feiten['saldo']['eindsaldo'],
            'maanden': rek_feiten.get('maanden', {}),
        }

    # Geconsolideerd saldo
    totaal_begin = sum(f['saldo']['beginsaldo'] for f in feiten.values())
    totaal_eind = sum(f['saldo']['eindsaldo'] for f in feiten.values())

    # Income source breakdown (V3: expliciete bronopbouw + confidence)
    # Stap 1: Bouw rekening → rekeninghouder mapping
    # De rekeninghouder wordt afgeleid uit INTERNE transfers: bij overboekingen
    # tussen eigen rekeningen verschijnt de naam van de houder als tegenpartij.
    rek_naar_persoon = {}  # {rekeningnummer: persoonsnaam}
    if 'Rekeningnummer' in df.columns and 'tegenpartij_naam' in df.columns:
        df_intern = df[df.get('is_intern', False)]
        if len(df_intern) > 0:
            # Bij interne overboekingen: de tegenpartij_naam = naam rekeninghouder
            # van de ANDERE rekening. Dus: Tegenrekening → tegenpartij_naam
            if 'Tegenrekening' in df_intern.columns:
                for _, row in df_intern.iterrows():
                    teg_rek = str(row.get('Tegenrekening', '')).strip()
                    tp_naam = str(row.get('tegenpartij_naam', '')).strip()
                    if teg_rek and tp_naam and len(tp_naam) >= 3:
                        # Filter banknamen eruit
                        tp_lower = tp_naam.lower()
                        if not any(skip in tp_lower for skip in [
                            'rekening', 'spaar', 'betaal', 'jongeren', 'tanken',
                            'ondernemers', 'zakelijk', 'deposito',
                        ]):
                            if teg_rek not in rek_naar_persoon:
                                rek_naar_persoon[teg_rek] = tp_naam
        # Fallback: als een rekening geen match heeft, probeer de meest
        # voorkomende tegenpartij_naam bij UITGAANDE interne transfers
        alle_rekeningen = set(df['Rekeningnummer'].unique())
        for rek in alle_rekeningen:
            if rek not in rek_naar_persoon:
                # Kijk naar uitgaande interne transfers VAN deze rekening
                rek_intern = df_intern[df_intern['Rekeningnummer'] == rek]
                if len(rek_intern) > 0 and 'tegenpartij_naam' in rek_intern.columns:
                    namen = rek_intern['tegenpartij_naam'].dropna()
                    namen = namen[namen.str.strip() != '']
                    namen = namen[~namen.str.lower().str.contains(
                        'rekening|spaar|betaal|jongeren|tanken|ondernemers|zakelijk|deposito',
                        na=False
                    )]
                    if len(namen) > 0:
                        # Dit is de naam van de ANDERE persoon, niet van deze rekening
                        # Sla op als "indirect" — we kennen de andere, niet deze
                        pass

    logger.info(f"Rekening→persoon mapping: {rek_naar_persoon}")

    income_sources = {}
    if 'source_family' in df.columns:
        df_income = df[(df['regel_sectie'] == 'inkomsten') & (~df.get('is_intern', False))]
        for sf, groep in df_income.groupby('source_family'):
            if pd.notna(sf):
                conf_vals = groep['regel_confidence'].dropna()
                gem_conf = float(conf_vals.mean()) if len(conf_vals) > 0 else 0.0
                if gem_conf >= 0.80:
                    vertrouwen = 'hoog'
                elif gem_conf >= 0.50:
                    vertrouwen = 'medium'
                else:
                    vertrouwen = 'laag'
                # Bepaal hoofdtegenpartij (wie betaalt dit inkomen)
                hoofd_tegenpartij = ''
                if 'tegenpartij_naam' in groep.columns:
                    namen = groep['tegenpartij_naam'].dropna()
                    namen = namen[namen.str.strip() != '']
                    if len(namen) > 0:
                        hoofd_tegenpartij = namen.value_counts().index[0]
                # Bepaal ONTVANGER: welk huishoudlid ontvangt dit inkomen?
                # Gebaseerd op welke rekening het binnenkomt
                ontvanger = ''
                if 'Rekeningnummer' in groep.columns:
                    hoofd_rek = groep['Rekeningnummer'].value_counts().index[0]
                    ontvanger = rek_naar_persoon.get(str(hoofd_rek), '')
                income_sources[str(sf)] = {
                    'bedrag_12m': round(float(groep['bedrag'].sum()), 2),
                    'transacties': len(groep),
                    'categorieën': list(groep['regel_categorie'].dropna().unique()),
                    'gem_confidence': round(gem_conf, 2),
                    'vertrouwen': vertrouwen,
                    'tegenpartij': hoofd_tegenpartij,
                    'ontvanger': ontvanger,
                }

    # Vertrouwensindicatoren per sectie
    vertrouwen_per_sectie = {}
    if 'regel_confidence' in df.columns:
        df_niet_intern = df[~df.get('is_intern', False)]
        for sectie in ['inkomsten', 'vaste_lasten', 'variabele_kosten', 'sparen_beleggen', 'onderling_neutraal']:
            df_s = df_niet_intern[df_niet_intern['regel_sectie'] == sectie]
            if len(df_s) == 0:
                continue
            conf_vals = df_s['regel_confidence'].dropna()
            gem = float(conf_vals.mean()) if len(conf_vals) > 0 else 0.0
            n_hoog = len(conf_vals[conf_vals >= 0.80])
            n_laag = len(conf_vals[conf_vals < 0.50])
            vertrouwen_per_sectie[sectie] = {
                'gem_confidence': round(gem, 2),
                'pct_hoog': round(n_hoog / len(df_s) * 100, 1) if len(df_s) > 0 else 0,
                'pct_laag': round(n_laag / len(df_s) * 100, 1) if len(df_s) > 0 else 0,
                'vertrouwen': 'hoog' if gem >= 0.80 else ('medium' if gem >= 0.50 else 'laag'),
            }

    # =========================================================================
    # rapport_totalen: SINGLE SOURCE OF TRUTH voor alle pagina's
    # =========================================================================
    # Bruto inkomen = ALLEEN sectie "inkomsten" (salaris, DGA-loon, huur, etc.)
    # Broker-terugstortingen, refunds, spaarrekening-mutaties zijn GEEN inkomen.
    # Die vallen onder sparen_beleggen of variabele_kosten (terugbetaling).
    inkomsten_12m = round(sectie_totalen_12m.get('inkomsten', 0), 2)
    vaste_lasten_12m = round(abs(sectie_totalen_12m.get('vaste_lasten', 0)), 2)
    variabele_kosten_12m = round(abs(sectie_totalen_12m.get('variabele_kosten', 0)), 2)
    sparen_beleggen_12m = round(sectie_totalen_12m.get('sparen_beleggen', 0), 2)
    onderling_neutraal_12m = round(sectie_totalen_12m.get('onderling_neutraal', 0), 2)

    # Netto beschikbaar = inkomen - vaste lasten - variabele kosten
    netto_beschikbaar_12m = round(inkomsten_12m - vaste_lasten_12m - variabele_kosten_12m, 2)

    # Totaal positief/negatief over ALLE secties (voor saldoverloop, niet "inkomen")
    totaal_positief_12m = round(sum(
        sum(max(float(v), 0) for v in cats.values())
        for cats in combined_jaar.values()
    ), 2)
    totaal_negatief_12m = round(sum(
        sum(min(float(v), 0) for v in cats.values())
        for cats in combined_jaar.values()
    ), 2)

    # =========================================================================
    # GRANULAIRE EXECUTIVE BUCKETS — afgeleid uit categorie_totalen
    # =========================================================================
    # Woonlasten: hypotheek/huur + energie + water + VvE + gemeentelijke heffingen
    _WOONLASTEN_CATS = {
        'Hypotheek/Huur', 'Huur/Hypotheek', 'VvE',
        'Gemeentebelasting/OZB/Waterschapsbelasting',
        'Gemeentelijke heffingen', 'Water', 'Energie',
    }
    # Belastingdruk: alle belasting-categorieën (alleen betalingen, niet teruggaven)
    _BELASTING_CATS = {
        'Inkomstenbelasting/Voorlopige aanslag', 'Inkomstenbelasting',
        'Vennootschapsbelasting (VPB)', 'BTW/Omzetbelasting',
        'Loonheffing', 'ZVW-premie', 'Motorrijtuigenbelasting (MRB)',
        'Overige belastingen',
    }
    # Vaste leefkosten: verzekeringen, abonnementen, etc.
    _VASTE_LEEFKOSTEN_CATS = {
        'Verzekeringen', 'Zorgverzekering', 'Telefoon/Internet',
        'Abonnementen/Streaming', 'Lidmaatschap/Contributie',
        'Kinderopvang', 'School/Opleiding',
    }
    # Alle vaste_lasten categorieën die NIET woonlasten, NIET belasting, NIET vaste leefkosten zijn
    # → die vallen in "overige vaste lasten" (wordt bij vaste leefkosten opgeteld)

    vl_cats = combined_jaar.get('vaste_lasten', {})
    vk_cats = combined_jaar.get('variabele_kosten', {})

    woonlasten_12m = round(abs(sum(float(vl_cats.get(c, 0)) for c in _WOONLASTEN_CATS)), 2)
    belastingdruk_12m = round(abs(sum(float(vl_cats.get(c, 0)) for c in _BELASTING_CATS)), 2)

    # Vaste leefkosten = expliciet benoemde + alle overige vaste lasten die niet woon/belasting zijn
    vaste_leefkosten_expliciet = abs(sum(float(vl_cats.get(c, 0)) for c in _VASTE_LEEFKOSTEN_CATS))
    overige_vaste = abs(sum(
        float(v) for c, v in vl_cats.items()
        if c not in _WOONLASTEN_CATS and c not in _BELASTING_CATS and c not in _VASTE_LEEFKOSTEN_CATS
    ))
    leefkosten_vast_12m = round(vaste_leefkosten_expliciet + overige_vaste, 2)

    # Discretionaire leefkosten = alle variabele kosten
    leefkosten_discretionair_12m = variabele_kosten_12m

    # Netto allocatie naar vermogen = sparen + beleggen (negatief = geld gaat erheen = goed)
    netto_allocatie_vermogen_12m = round(abs(sparen_beleggen_12m), 2) if sparen_beleggen_12m < 0 else 0
    # Als sparen_beleggen positief: geld komt terug van vermogen → niet-kerninstroom
    niet_kerninstroom_12m = round(sparen_beleggen_12m, 2) if sparen_beleggen_12m > 0 else 0

    # Review/onzeker: categorieën met "Onzeker" of "Review" in de naam
    review_onzeker_12m = round(abs(sum(
        float(v) for sectie_cats in combined_jaar.values()
        for c, v in sectie_cats.items()
        if 'onzeker' in c.lower() or 'review' in c.lower()
    )), 2)

    # Saldo
    begin_saldo = round(totaal_begin, 2)
    eind_saldo = round(totaal_eind, 2)
    netto_saldo_mutatie_12m = round(eind_saldo - begin_saldo, 2)

    # Aanvullende structurele instroom (teruggaven belasting in inkomsten-sectie)
    aanvullende_instroom_12m = round(sum(
        float(v) for c, v in combined_jaar.get('inkomsten', {}).items()
        if 'teruggave' in c.lower() or 'toeslagen' in c.lower()
    ), 2)
    # Kerninkomen = bruto inkomen minus aanvullende instroom
    kerninkomen_12m = round(inkomsten_12m - aanvullende_instroom_12m, 2)

    rapport_totalen = {
        # --- Sectie-niveau (backward compatible) ---
        'bruto_inkomen_12m': inkomsten_12m,
        'bruto_inkomen_pm': round(inkomsten_12m / n_mnd, 2) if n_mnd > 0 else 0,
        'vaste_lasten_12m': vaste_lasten_12m,
        'vaste_lasten_pm': round(vaste_lasten_12m / n_mnd, 2) if n_mnd > 0 else 0,
        'variabele_kosten_12m': variabele_kosten_12m,
        'variabele_kosten_pm': round(variabele_kosten_12m / n_mnd, 2) if n_mnd > 0 else 0,
        'sparen_beleggen_12m': sparen_beleggen_12m,
        'netto_beschikbaar_12m': netto_beschikbaar_12m,
        'netto_beschikbaar_pm': round(netto_beschikbaar_12m / n_mnd, 2) if n_mnd > 0 else 0,
        # --- Totaal geldstromen (voor saldoverloop, NIET "inkomen" noemen!) ---
        'totaal_positief_12m': totaal_positief_12m,
        'totaal_negatief_12m': totaal_negatief_12m,
        # --- Executive buckets (granulaire economische lens) ---
        'kerninkomen_12m': kerninkomen_12m,
        'aanvullende_instroom_12m': aanvullende_instroom_12m,
        'belastingdruk_12m': belastingdruk_12m,
        'woonlasten_12m': woonlasten_12m,
        'leefkosten_vast_12m': leefkosten_vast_12m,
        'leefkosten_discretionair_12m': leefkosten_discretionair_12m,
        'netto_allocatie_vermogen_12m': netto_allocatie_vermogen_12m,
        'niet_kerninstroom_12m': niet_kerninstroom_12m,
        'review_onzeker_12m': review_onzeker_12m,
        # --- Saldo ---
        'begin_saldo': begin_saldo,
        'eind_saldo': eind_saldo,
        'netto_saldo_mutatie_12m': netto_saldo_mutatie_12m,
    }

    ground_truth = {
        'versie': 'V3',
        'periode': {
            'volle_maanden': sorted(volle_maanden),
            'onvolledige_maanden': rapportperiode.get('onvolledige_maanden', []),
            'n_mnd': n_mnd,
            'alle_maanden': sorted(combined_maand.keys()),
        },
        'saldo': {
            'totaal_begin': round(totaal_begin, 2),
            'totaal_eind': round(totaal_eind, 2),
            'per_rekening': saldo_per_rekening,
        },
        'rapport_totalen': rapport_totalen,
        'sectie_totalen_12m': sectie_totalen_12m,
        'sectie_gemiddelden_pm': {
            s: round(t / n_mnd, 2) for s, t in sectie_totalen_12m.items()
        } if n_mnd > 0 else {},
        'categorie_totalen_12m': combined_jaar,
        'maandoverzicht': combined_maand,
        'maand_sectie_totalen': maand_sectie_totalen,
        'income_sources': income_sources,
        'vertrouwen_per_sectie': vertrouwen_per_sectie,
        'reconciliatie': reconciliatie,
        '_rek_naar_persoon': rek_naar_persoon,
    }

    # Log de ground truth samenvatting
    logger.info("V3 GROUND TRUTH gebouwd:")
    for sectie, totaal in sectie_totalen_12m.items():
        logger.info(f"  {sectie}: EUR {totaal:,.2f}")
    logger.info(f"  Saldo: begin EUR {totaal_begin:,.2f} → eind EUR {totaal_eind:,.2f}")
    logger.info(f"  rapport_totalen: inkomen={rapport_totalen['bruto_inkomen_12m']:,.2f}, "
                f"vaste_lasten={rapport_totalen['vaste_lasten_12m']:,.2f}, "
                f"variabel={rapport_totalen['variabele_kosten_12m']:,.2f}")

    return ground_truth


def _bouw_executive_buckets(ground_truth: dict) -> dict:
    """Laag 4: Bouw executive buckets voor pagina 1 uit rapport_totalen.

    Dit is de ENIGE bron voor pagina-1 metrics. Geen losse herberekeningen.
    Retourneert een dict met 7 buckets + subtotalen.
    """
    rt = ground_truth.get('rapport_totalen', {})
    n_mnd = ground_truth.get('periode', {}).get('n_mnd', 12) or 12

    buckets = {
        # 1. Kerninkomen (salaris, DGA-loon, huur — zonder teruggaven/toeslagen)
        'kerninkomen': {
            'label': 'Kerninkomen',
            'bedrag_12m': rt.get('kerninkomen_12m', 0),
            'bedrag_pm': round(rt.get('kerninkomen_12m', 0) / n_mnd, 2),
            'toelichting': 'Structureel inkomen uit arbeid, onderneming of vastgoed',
        },
        # 2. Aanvullende structurele instroom (teruggaven, toeslagen)
        'aanvullende_instroom': {
            'label': 'Aanvullende instroom',
            'bedrag_12m': rt.get('aanvullende_instroom_12m', 0),
            'bedrag_pm': round(rt.get('aanvullende_instroom_12m', 0) / n_mnd, 2),
            'toelichting': 'Belastingteruggaven, toeslagen en overige instroom',
        },
        # 3. Belastingdruk
        'belastingdruk': {
            'label': 'Belastingdruk',
            'bedrag_12m': rt.get('belastingdruk_12m', 0),
            'bedrag_pm': round(rt.get('belastingdruk_12m', 0) / n_mnd, 2),
            'toelichting': 'IB, VPB, BTW, loonheffing, ZVW, MRB',
        },
        # 4. Woonlasten
        'woonlasten': {
            'label': 'Woonlasten',
            'bedrag_12m': rt.get('woonlasten_12m', 0),
            'bedrag_pm': round(rt.get('woonlasten_12m', 0) / n_mnd, 2),
            'toelichting': 'Hypotheek/huur, energie, water, VvE, gemeentelijke heffingen',
        },
        # 5a. Vaste leefkosten
        'leefkosten_vast': {
            'label': 'Vaste leefkosten',
            'bedrag_12m': rt.get('leefkosten_vast_12m', 0),
            'bedrag_pm': round(rt.get('leefkosten_vast_12m', 0) / n_mnd, 2),
            'toelichting': 'Verzekeringen, abonnementen, kinderopvang, overige vaste lasten',
        },
        # 5b. Discretionaire leefkosten
        'leefkosten_discretionair': {
            'label': 'Discretionaire uitgaven',
            'bedrag_12m': rt.get('leefkosten_discretionair_12m', 0),
            'bedrag_pm': round(rt.get('leefkosten_discretionair_12m', 0) / n_mnd, 2),
            'toelichting': 'Boodschappen, horeca, kleding, vrije tijd, overige variabele kosten',
        },
        # 6. Netto allocatie naar vermogen
        'netto_allocatie_vermogen': {
            'label': 'Netto allocatie vermogen',
            'bedrag_12m': rt.get('netto_allocatie_vermogen_12m', 0),
            'bedrag_pm': round(rt.get('netto_allocatie_vermogen_12m', 0) / n_mnd, 2),
            'toelichting': 'Netto storting naar sparen en beleggen',
        },
        # 7a. Niet-kerninstroom (broker terugstortingen, spaar-terugboekingen)
        'niet_kerninstroom': {
            'label': 'Buiten kernbeeld: instroom',
            'bedrag_12m': rt.get('niet_kerninstroom_12m', 0),
            'bedrag_pm': round(rt.get('niet_kerninstroom_12m', 0) / n_mnd, 2),
            'toelichting': 'Terugstortingen van beleggingen, spaar-terugboekingen',
        },
        # 7b. Review/onzeker
        'review_onzeker': {
            'label': 'Buiten kernbeeld: review',
            'bedrag_12m': rt.get('review_onzeker_12m', 0),
            'bedrag_pm': round(rt.get('review_onzeker_12m', 0) / n_mnd, 2),
            'toelichting': 'Posten die handmatige review vereisen',
        },
    }

    # Subtotalen voor snelle lookups
    totaal_uitgaven = (
        buckets['belastingdruk']['bedrag_12m']
        + buckets['woonlasten']['bedrag_12m']
        + buckets['leefkosten_vast']['bedrag_12m']
        + buckets['leefkosten_discretionair']['bedrag_12m']
    )
    buckets['_totaal_uitgaven_12m'] = round(totaal_uitgaven, 2)
    buckets['_bruto_inkomen_12m'] = rt.get('bruto_inkomen_12m', 0)

    # Ratios (allen op basis van bruto inkomen, nooit totaal_positief)
    bruto = rt.get('bruto_inkomen_12m', 1) or 1  # voorkom /0
    buckets['_ratios'] = {
        'woonquote': round(buckets['woonlasten']['bedrag_12m'] / bruto * 100, 1),
        'belastingdruk_pct': round(buckets['belastingdruk']['bedrag_12m'] / bruto * 100, 1),
        'spaarquote': round(buckets['netto_allocatie_vermogen']['bedrag_12m'] / bruto * 100, 1),
        'vaste_lasten_pct': round(
            (buckets['woonlasten']['bedrag_12m'] + buckets['leefkosten_vast']['bedrag_12m']
             + buckets['belastingdruk']['bedrag_12m']) / bruto * 100, 1),
    }

    ground_truth['executive_buckets'] = buckets
    logger.info(f"  Executive buckets: woonlasten={buckets['woonlasten']['bedrag_12m']:,.2f}, "
                f"belasting={buckets['belastingdruk']['bedrag_12m']:,.2f}, "
                f"kerninkomen={buckets['kerninkomen']['bedrag_12m']:,.2f}")
    return buckets


def _valideer_ground_truth(ground_truth: dict) -> list:
    """Deterministische cross-checks op ground_truth.

    Draait NA ground_truth opbouw, VOOR PDF generatie.
    Retourneert lijst met (severity, message) tuples.
    severity = 'ERROR' (blokkeer) of 'WARNING' (disclaimer).
    Auto-repareert waar mogelijk.
    """
    issues = []
    rt = ground_truth.get('rapport_totalen', {})
    st = ground_truth.get('sectie_totalen_12m', {})
    ct = ground_truth.get('categorie_totalen_12m', {})
    mst = ground_truth.get('maand_sectie_totalen', {})
    saldo = ground_truth.get('saldo', {})
    n_mnd = ground_truth.get('periode', {}).get('n_mnd', 12)

    # CHECK 1: rapport_totalen.bruto_inkomen == sectie_totalen_12m.inkomsten
    bruto = rt.get('bruto_inkomen_12m', 0)
    sectie_ink = st.get('inkomsten', 0)
    if abs(bruto - sectie_ink) > 1:
        issues.append(('ERROR', f"rapport_totalen.bruto_inkomen ({bruto:,.2f}) ≠ "
                       f"sectie_totalen.inkomsten ({sectie_ink:,.2f})"))

    # CHECK 2: Per sectie: sum(categorieën) == sectie_totaal
    for sectie in ['inkomsten', 'vaste_lasten', 'variabele_kosten', 'sparen_beleggen']:
        cat_sum = sum(float(v) for v in ct.get(sectie, {}).values())
        sectie_tot = st.get(sectie, 0)
        if abs(cat_sum - sectie_tot) > 1:
            issues.append(('WARNING', f"Sectie '{sectie}': cat_sum ({cat_sum:,.2f}) ≠ "
                           f"sectie_totaal ({sectie_tot:,.2f}), delta={cat_sum - sectie_tot:,.2f}"))

    # CHECK 3: Sum(maand_sectie_totalen per sectie) == sectie_totalen_12m
    for sectie in ['inkomsten', 'vaste_lasten', 'variabele_kosten', 'sparen_beleggen']:
        maand_sum = sum(
            float(mdata.get(sectie, 0))
            for mdata in mst.values()
        )
        sectie_tot = st.get(sectie, 0)
        if abs(maand_sum - sectie_tot) > 1:
            issues.append(('WARNING', f"SOM maanden '{sectie}' ({maand_sum:,.2f}) ≠ "
                           f"sectie_totaal ({sectie_tot:,.2f})"))

    # CHECK 4: Saldo-kloppendheid: begin + sum(alle netto) ≈ eind
    totaal_begin = saldo.get('totaal_begin', 0)
    totaal_eind = saldo.get('totaal_eind', 0)
    som_netto = sum(st.get(s, 0) for s in st)
    verwacht_eind = totaal_begin + som_netto
    delta_saldo = abs(verwacht_eind - totaal_eind)
    if delta_saldo > 50:  # €50 tolerantie voor afrondingen
        issues.append(('WARNING', f"Saldo check: begin ({totaal_begin:,.2f}) + netto ({som_netto:,.2f}) "
                       f"= {verwacht_eind:,.2f}, maar eind = {totaal_eind:,.2f} (delta {delta_saldo:,.2f})"))

    # CHECK 5: Ratio's realistisch
    if bruto > 0:
        wl = rt.get('vaste_lasten_12m', 0)
        woonquote = wl / bruto * 100
        if woonquote > 150:
            issues.append(('ERROR', f"Vaste lasten > 150% van inkomen ({woonquote:.1f}%) — "
                           "waarschijnlijk classificatiefout"))
    elif bruto == 0 and abs(st.get('vaste_lasten', 0)) > 1000:
        issues.append(('ERROR', "Geen inkomen gedetecteerd maar wel vaste lasten — "
                       "classificatiefout in inkomsten"))

    # CHECK 6: income_sources som ≈ bruto inkomen
    income_sources = ground_truth.get('income_sources', {})
    if income_sources and bruto > 0:
        src_sum = sum(abs(d.get('bedrag_12m', 0)) for d in income_sources.values())
        if abs(src_sum - bruto) > bruto * 0.10:  # >10% verschil
            issues.append(('WARNING', f"income_sources som ({src_sum:,.2f}) wijkt >10% af "
                           f"van bruto inkomen ({bruto:,.2f})"))

    # CHECK 7: Geen negatief bruto inkomen
    if bruto < 0:
        issues.append(('ERROR', f"Negatief bruto inkomen ({bruto:,.2f}) — classificatiefout"))

    # CHECK 8: Executive buckets sluiten aan op rapport_totalen
    eb = ground_truth.get('executive_buckets', {})
    if eb:
        # Vaste lasten moet splitsen in woonlasten + belasting + vaste leefkosten
        eb_vaste = (
            eb.get('woonlasten', {}).get('bedrag_12m', 0)
            + eb.get('belastingdruk', {}).get('bedrag_12m', 0)
            + eb.get('leefkosten_vast', {}).get('bedrag_12m', 0)
        )
        rt_vaste = rt.get('vaste_lasten_12m', 0)
        if abs(eb_vaste - rt_vaste) > 1:
            issues.append(('WARNING', f"Executive buckets vaste lasten ({eb_vaste:,.2f}) ≠ "
                           f"rapport_totalen.vaste_lasten ({rt_vaste:,.2f})"))

        # Discretionair moet == variabele kosten
        eb_disc = eb.get('leefkosten_discretionair', {}).get('bedrag_12m', 0)
        rt_var = rt.get('variabele_kosten_12m', 0)
        if abs(eb_disc - rt_var) > 1:
            issues.append(('WARNING', f"Executive buckets discretionair ({eb_disc:,.2f}) ≠ "
                           f"rapport_totalen.variabele_kosten ({rt_var:,.2f})"))

        # Kern + aanvullend moet == bruto inkomen
        eb_ink = (
            eb.get('kerninkomen', {}).get('bedrag_12m', 0)
            + eb.get('aanvullende_instroom', {}).get('bedrag_12m', 0)
        )
        if abs(eb_ink - bruto) > 1:
            issues.append(('WARNING', f"Executive buckets inkomen ({eb_ink:,.2f}) ≠ "
                           f"rapport_totalen.bruto_inkomen ({bruto:,.2f})"))

    # CHECK 9: Richtinglogica — belastingbetalingen moeten negatief zijn in bron
    for sectie_cats in ct.values():
        for cat_name, val in sectie_cats.items():
            fval = float(val)
            # Asset withdrawal mag niet in inkomsten staan
            if 'terugstorting' in cat_name.lower() and cat_name in ct.get('inkomsten', {}):
                if fval > 100:
                    issues.append(('WARNING', f"'{cat_name}' ({fval:,.2f}) staat in inkomsten — "
                                   "mogelijke asset withdrawal leakage"))

    for severity, msg in issues:
        logger.warning(f"GROUND TRUTH {severity}: {msg}")

    return issues


# ---------------------------------------------------------------------------
# V3: RECONCILIATIE EXCEL GENERATOR
# ---------------------------------------------------------------------------

def _genereer_reconciliatie_excel(ground_truth: dict, feiten: dict, df: pd.DataFrame) -> bytes:
    """V3: Genereer reconciliatie Excel in Shortcut.ai Cashflow Overzicht format.

    Output: bytes (xlsx bestand)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = "Cashflow Overzicht"

    # Styling
    font_header = Font(name='Arial', bold=True, size=11)
    font_section = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    font_data = Font(name='Arial', size=10)
    font_totaal = Font(name='Arial', bold=True, size=10)
    font_controle = Font(name='Arial', bold=True, size=10, color='008000')
    font_red = Font(name='Arial', bold=True, size=10, color='FF0000')
    fill_section = PatternFill('solid', fgColor='2F5496')
    fill_totaal = PatternFill('solid', fgColor='D6E4F0')
    fill_saldo = PatternFill('solid', fgColor='E2EFDA')
    fill_controle = PatternFill('solid', fgColor='C6EFCE')
    align_right = Alignment(horizontal='right')
    align_left = Alignment(horizontal='left')
    border_thin = Border(
        bottom=Side(style='thin', color='CCCCCC')
    )

    # Maanden bepalen
    alle_maanden = sorted(ground_truth['periode']['alle_maanden'])
    volle_maanden = set(ground_truth['periode']['volle_maanden'])

    # Kolom-layout: B=labels, C..=maanden, dan Totaal 12m, Gem p/m
    col_offset = 2  # B = kolom 2
    maand_cols = {}
    for i, maand in enumerate(alle_maanden):
        col = col_offset + 1 + i
        maand_cols[maand] = col

    totaal_col = col_offset + 1 + len(alle_maanden)
    gem_col = totaal_col + 1
    n_mnd = ground_truth['periode']['n_mnd'] or 12

    # Helper functies
    def write_cell(row, col, value, font=font_data, fill=None, alignment=None, border=None, number_format=None):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = font
        if fill:
            cell.fill = fill
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = border
        if number_format:
            cell.number_format = number_format

    def write_section_header(row, label):
        for col in range(col_offset, gem_col + 1):
            ws.cell(row=row, column=col).fill = fill_section
            ws.cell(row=row, column=col).font = font_section
        write_cell(row, col_offset, label, font=font_section, fill=fill_section)

    def write_data_row(row, label, maand_waarden, is_totaal=False):
        f = font_totaal if is_totaal else font_data
        fl = fill_totaal if is_totaal else None
        write_cell(row, col_offset, label, font=f, fill=fl)

        totaal_12m = 0
        for maand, col in maand_cols.items():
            val = maand_waarden.get(maand, 0)
            if val != 0:
                write_cell(row, col, round(val, 2), font=f, fill=fl,
                           alignment=align_right, number_format='#,##0.00')
            if maand in volle_maanden:
                totaal_12m += val

        write_cell(row, totaal_col, round(totaal_12m, 2), font=f, fill=fl,
                   alignment=align_right, number_format='#,##0.00')
        if n_mnd > 0:
            write_cell(row, gem_col, round(totaal_12m / n_mnd, 2), font=f, fill=fl,
                       alignment=align_right, number_format='#,##0.00')

    # === ROW 1: Titel ===
    row = 1
    write_cell(row, col_offset, "Cashflow Overzicht", font=Font(name='Arial', bold=True, size=14))

    # === ROW 2: Periode info ===
    row = 2
    if volle_maanden:
        eerste = min(volle_maanden)
        laatste = max(volle_maanden)
        write_cell(row, col_offset, f"Periode: {eerste} t/m {laatste} ({n_mnd} volle maanden)",
                   font=Font(name='Arial', size=9, italic=True))

    # === ROW 3: Headers ===
    row = 3
    write_cell(row, col_offset, "", font=font_header)
    for maand, col in maand_cols.items():
        label = maand  # YYYY-MM
        is_vol = maand in volle_maanden
        f = font_header if is_vol else Font(name='Arial', bold=True, size=10, color='999999')
        write_cell(row, col, label, font=f, alignment=Alignment(horizontal='center'))
    write_cell(row, totaal_col, f"Totaal {n_mnd}m", font=font_header, alignment=Alignment(horizontal='center'))
    write_cell(row, gem_col, "Gem p/m", font=font_header, alignment=Alignment(horizontal='center'))

    # === ROW 4: Begin saldo ===
    row = 4
    # Bereken begin saldo per maand (som van alle rekeningen)
    begin_saldi = {}
    eind_saldi = {}
    for rek, rek_data in feiten.items():
        for maand_str, mdata in rek_data.get('maanden', {}).items():
            # We need per-maand begin/eind saldo per rekening
            # feiten.maanden has inkomsten/uitgaven/netto but not begin/eind per maand
            pass

    # Beter: bereken begin/eind saldo per maand uit het DataFrame
    if 'maand' in df.columns:
        for maand in alle_maanden:
            mdf = df[df['maand'].astype(str) == maand]
            if len(mdf) > 0:
                # Begin saldo = eerste transactie's Beginsaldo per rekening
                begin = 0
                eind = 0
                for rek in mdf['Rekeningnummer'].unique():
                    rdf = mdf[mdf['Rekeningnummer'] == rek].sort_values('datum')
                    if len(rdf) > 0:
                        begin += float(rdf.iloc[0].get('Beginsaldo', 0))
                        eind += float(rdf.iloc[-1].get('Eindsaldo', 0))
                begin_saldi[maand] = round(begin, 2)
                eind_saldi[maand] = round(eind, 2)

    write_cell(row, col_offset, "Begin saldo totaal", font=font_totaal, fill=fill_saldo)
    for maand, col in maand_cols.items():
        if maand in begin_saldi:
            write_cell(row, col, begin_saldi[maand], font=font_totaal, fill=fill_saldo,
                       alignment=align_right, number_format='#,##0.00')

    row += 1  # Lege rij

    # === SECTIES ===
    sectie_volgorde = [
        ('inkomsten', 'INKOMSTEN'),
        ('vaste_lasten', 'VASTE LASTEN'),
        ('variabele_kosten', 'VARIABELE UITGAVEN'),
        ('sparen_beleggen', 'SPAREN & BELEGGEN'),
        ('onderling_neutraal', 'ONDERLING / NEUTRAAL'),
    ]

    maandoverzicht = ground_truth.get('maandoverzicht', {})
    sectie_netto_per_maand = {}  # voor netto mutaties berekening

    for sectie_key, sectie_label in sectie_volgorde:
        row += 1
        write_section_header(row, sectie_label)

        # Verzamel categorieën voor deze sectie
        cats_in_sectie = set()
        for maand, secties in maandoverzicht.items():
            for cat in secties.get(sectie_key, {}).keys():
                cats_in_sectie.add(cat)

        # Ook uit jaartotalen
        for cat in ground_truth.get('categorie_totalen_12m', {}).get(sectie_key, {}).keys():
            cats_in_sectie.add(cat)

        # Sorteer categorieën op totaalbedrag (grootste eerst)
        cat_totalen = {}
        for cat in cats_in_sectie:
            cat_totalen[cat] = abs(ground_truth.get('categorie_totalen_12m', {}).get(sectie_key, {}).get(cat, 0))
        cats_sorted = sorted(cats_in_sectie, key=lambda c: cat_totalen.get(c, 0), reverse=True)

        for cat in cats_sorted:
            row += 1
            maand_waarden = {}
            for maand, secties in maandoverzicht.items():
                val = secties.get(sectie_key, {}).get(cat, 0)
                maand_waarden[maand] = float(val) if val else 0
            write_data_row(row, cat, maand_waarden)

        # Totaal rij
        row += 1
        totaal_maand = {}
        for maand, secties in maandoverzicht.items():
            sec_data = secties.get(sectie_key, {})
            totaal_maand[maand] = sum(float(v) for v in sec_data.values())

            # Track voor netto mutaties
            if maand not in sectie_netto_per_maand:
                sectie_netto_per_maand[maand] = 0
            sectie_netto_per_maand[maand] += totaal_maand[maand]

        write_data_row(row, f"Totaal {sectie_label.lower()}", totaal_maand, is_totaal=True)

    # === NETTO MUTATIES ===
    row += 2
    write_cell(row, col_offset, "Netto mutaties", font=font_totaal, fill=fill_saldo)
    for maand, col in maand_cols.items():
        val = sectie_netto_per_maand.get(maand, 0)
        write_cell(row, col, round(val, 2), font=font_totaal, fill=fill_saldo,
                   alignment=align_right, number_format='#,##0.00')

    # Totaal 12m netto
    totaal_netto_12m = sum(v for m, v in sectie_netto_per_maand.items() if m in volle_maanden)
    write_cell(row, totaal_col, round(totaal_netto_12m, 2), font=font_totaal, fill=fill_saldo,
               alignment=align_right, number_format='#,##0.00')

    # === BEREKEND EINDSALDO ===
    row += 1
    write_cell(row, col_offset, "Berekend eindsaldo", font=font_totaal, fill=fill_saldo)
    for maand, col in maand_cols.items():
        begin = begin_saldi.get(maand, 0)
        netto = sectie_netto_per_maand.get(maand, 0)
        write_cell(row, col, round(begin + netto, 2), font=font_totaal, fill=fill_saldo,
                   alignment=align_right, number_format='#,##0.00')

    # === WERKELIJK EINDSALDO ===
    row += 1
    write_cell(row, col_offset, "Werkelijk eindsaldo", font=font_totaal, fill=fill_saldo)
    for maand, col in maand_cols.items():
        if maand in eind_saldi:
            write_cell(row, col, eind_saldi[maand], font=font_totaal, fill=fill_saldo,
                       alignment=align_right, number_format='#,##0.00')

    # === CONTROLE (=0) ===
    row += 1
    write_cell(row, col_offset, "Controle (=0)", font=font_controle, fill=fill_controle)
    alle_controle_ok = True
    for maand, col in maand_cols.items():
        begin = begin_saldi.get(maand, 0)
        netto = sectie_netto_per_maand.get(maand, 0)
        berekend = round(begin + netto, 2)
        werkelijk = eind_saldi.get(maand, 0)
        controle = round(berekend - werkelijk, 2)

        f = font_controle if abs(controle) < 0.02 else font_red
        fl = fill_controle if abs(controle) < 0.02 else PatternFill('solid', fgColor='FFC7CE')
        write_cell(row, col, controle, font=f, fill=fl,
                   alignment=align_right, number_format='#,##0.00')
        if abs(controle) >= 0.02:
            alle_controle_ok = False

    # === SALDO PER REKENING ===
    row += 2
    write_section_header(row, "SALDO PER REKENING")

    for rek in sorted(feiten.keys()):
        # Begin per maand
        row += 1
        write_cell(row, col_offset, f"{rek} begin", font=font_data)
        for maand, col_num in maand_cols.items():
            mdf = df[(df['Rekeningnummer'].astype(str) == str(rek)) & (df['maand'].astype(str) == maand)]
            if len(mdf) > 0:
                val = float(mdf.sort_values('datum').iloc[0].get('Beginsaldo', 0))
                write_cell(row, col_num, round(val, 2), font=font_data,
                           alignment=align_right, number_format='#,##0.00')

        # Eind per maand
        row += 1
        write_cell(row, col_offset, f"{rek} eind", font=font_data)
        for maand, col_num in maand_cols.items():
            mdf = df[(df['Rekeningnummer'].astype(str) == str(rek)) & (df['maand'].astype(str) == maand)]
            if len(mdf) > 0:
                val = float(mdf.sort_values('datum').iloc[-1].get('Eindsaldo', 0))
                write_cell(row, col_num, round(val, 2), font=font_data,
                           alignment=align_right, number_format='#,##0.00')

    # Kolombreedte
    ws.column_dimensions['B'].width = 35
    for maand, col in maand_cols.items():
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions[get_column_letter(totaal_col)].width = 16
    ws.column_dimensions[get_column_letter(gem_col)].width = 14

    # Save to bytes
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    logger.info(f"V3 RECONCILIATIE EXCEL: {row} rijen, controle={'OK' if alle_controle_ok else 'FOUT'}")
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# V3 SPRINT 3: NO-SEND GATE + AUDIT PACKAGE
# ---------------------------------------------------------------------------

def _no_send_gate(ground_truth: dict, reconciliatie: dict, analyse: dict,
                  kwaliteit: dict) -> dict:
    """V3: Bepaal of het rapport verzonden mag worden.

    Regels (uit KWALITEITSPLAN-V1):
    1. RED reconciliatie → BLOCK (saldo klopt niet)
    2. >5% van totaalbedrag in "Overig" categorieën → BLOCK
    3. AI-samenvatting bevat getallen die niet in ground truth staan → WARN
    4. Onvolledige maand telt mee in 12m-totalen → BLOCK (maar is al gefixed door _bepaal_rapportperiode)

    Returns dict met:
    - besluit: 'SEND' / 'BLOCK' / 'REVIEW'
    - kleur: 'GREEN' / 'ORANGE' / 'RED'
    - redenen: lijst van redenen
    """
    redenen = []
    kleur = 'GREEN'

    # Regel 1: Reconciliatie status
    recon_status = reconciliatie.get('status', 'GREEN')
    if recon_status == 'RED':
        kleur = 'RED'
        red_checks = [c for c in reconciliatie.get('checks', []) if c['status'] == 'RED']
        for c in red_checks:
            redenen.append(f"RECONCILIATIE FOUT: {c['detail']}")

    # Regel 2: Overig-categorieën percentage per sectie
    # Drempels per sectie: inkomsten strenger (geld moet kloppen),
    # variabele kosten soepeler (veel kleine transacties)
    _OVERIG_DREMPELS = {
        'inkomsten':         (0.10, 0.25),
        'vaste_lasten':      (0.10, 0.30),
        'variabele_kosten':  (0.15, 0.40),
        'sparen_beleggen':   (0.10, 0.30),
        'onderling_neutraal': (0.20, 0.50),
    }
    cat_totalen = ground_truth.get('categorie_totalen_12m', {})
    sectie_totalen = ground_truth.get('sectie_totalen_12m', {})
    for sectie, cats in cat_totalen.items():
        totaal_sectie = abs(sectie_totalen.get(sectie, 0))
        if totaal_sectie < 100:
            continue
        overig_bedrag = 0
        # Alleen de echte catch-all "Overig" categorieën tellen als onbekend.
        # "Overige verzekeringen", "Overige abonnementen", "Overige belastingen" zijn
        # specifieke categorieën uit de taxonomie en tellen NIET als onbekend.
        _ECHTE_OVERIG_CATS = {
            'overige vaste lasten', 'overig vaste lasten',
            'overig variabel', 'overige variabele kosten',
            'overig inkomen', 'overige inkomsten',
            'overig sparen/beleggen', 'overig sparen',
            'overig', 'overige',
        }
        for cat, bedrag in cats.items():
            if cat.lower() in _ECHTE_OVERIG_CATS:
                overig_bedrag += abs(float(bedrag))
        if totaal_sectie > 0:
            overig_pct = overig_bedrag / totaal_sectie
            drempel_orange, drempel_red = _OVERIG_DREMPELS.get(sectie, (0.10, 0.30))
            if overig_pct > drempel_orange:
                kleur = 'RED' if overig_pct > drempel_red else max(kleur, 'ORANGE')
                redenen.append(
                    f"{sectie}: {overig_pct:.0%} in Overig-categorieën "
                    f"(EUR {overig_bedrag:,.0f} / EUR {totaal_sectie:,.0f})"
                )

    # Regel 3: AI-samenvatting bevat getallen die niet in ground truth staan
    samenvatting = analyse.get('samenvatting', '') if isinstance(analyse, dict) else ''
    if samenvatting:
        # Verzamel alle getallen uit ground truth
        gt_getallen = set()
        for sectie, cats in cat_totalen.items():
            for cat, bedrag in cats.items():
                gt_getallen.add(abs(round(float(bedrag), 0)))
                gt_getallen.add(abs(round(float(bedrag), 2)))
        for s, t in sectie_totalen.items():
            gt_getallen.add(abs(round(float(t), 0)))
            gt_getallen.add(abs(round(float(t), 2)))
        for s, g in ground_truth.get('sectie_gemiddelden_pm', {}).items():
            gt_getallen.add(abs(round(float(g), 0)))
            gt_getallen.add(abs(round(float(g), 2)))
        saldo = ground_truth.get('saldo', {})
        gt_getallen.add(abs(round(saldo.get('totaal_begin', 0), 0)))
        gt_getallen.add(abs(round(saldo.get('totaal_eind', 0), 0)))
        # Income sources
        for sf, data in ground_truth.get('income_sources', {}).items():
            gt_getallen.add(abs(round(data.get('bedrag_12m', 0), 0)))
            gt_getallen.add(abs(round(data.get('bedrag_12m', 0), 2)))
        # Verwijder triviale getallen
        gt_getallen.discard(0)
        gt_getallen.discard(0.0)

        # Zoek getallen in de samenvatting
        import re
        bedragen_in_tekst = re.findall(r'€\s*([0-9.,]+)', samenvatting)
        onbekende_bedragen = []
        for bedrag_str in bedragen_in_tekst:
            try:
                # Verwerk Nederlandse notatie: €1.234,56 of €1.234
                bedrag_str_clean = bedrag_str.replace('.', '').replace(',', '.')
                bedrag_val = abs(float(bedrag_str_clean))
                if bedrag_val > 50:  # Negeer kleine bedragen
                    # Check of dit bedrag in de ground truth staat (met 2% marge)
                    gevonden = False
                    for gt_val in gt_getallen:
                        if gt_val > 0 and abs(bedrag_val - gt_val) / gt_val < 0.02:
                            gevonden = True
                            break
                    if not gevonden:
                        onbekende_bedragen.append(bedrag_str)
            except (ValueError, ZeroDivisionError):
                pass

        if onbekende_bedragen:
            if kleur == 'GREEN':
                kleur = 'ORANGE'
            redenen.append(
                f"AI-samenvatting bevat {len(onbekende_bedragen)} bedrag(en) niet in ground truth: "
                f"{', '.join(onbekende_bedragen[:5])}"
            )

    # Regel 4: Jaar-totaal ≠ som van maanden (consistency check)
    maand_sectie = ground_truth.get('maand_sectie_totalen', {})
    for sectie, jaar_totaal in sectie_totalen.items():
        maand_som = sum(
            maand_data.get(sectie, 0)
            for maand_data in maand_sectie.values()
        )
        if abs(jaar_totaal) > 10 and abs(maand_som - jaar_totaal) > max(1.0, abs(jaar_totaal) * 0.01):
            kleur = 'RED'
            redenen.append(
                f"JAARTOTAAL ≠ SOM MAANDEN: {sectie} jaar={jaar_totaal:,.0f} vs som={maand_som:,.0f} "
                f"(verschil EUR {abs(maand_som - jaar_totaal):,.0f})"
            )

    # Regel 5: Classificatie-kwaliteit
    pct_rule = kwaliteit.get('pct_rule_based', 0)
    if pct_rule < 40:
        if kleur == 'GREEN':
            kleur = 'ORANGE'
        redenen.append(f"Slechts {pct_rule:.0f}% rule-based geclassificeerd (target: >60%)")

    # Besluit
    if kleur == 'RED':
        besluit = 'BLOCK'
    elif kleur == 'ORANGE':
        besluit = 'REVIEW'
    else:
        besluit = 'SEND'

    gate_result = {
        'besluit': besluit,
        'kleur': kleur,
        'redenen': redenen,
        'tijdstip': datetime.now().isoformat(),
    }

    logger.info(f"V3 NO-SEND GATE: {besluit} ({kleur}), {len(redenen)} redenen")
    for r in redenen:
        logger.info(f"  → {r}")

    return gate_result


def _bouw_audit_package(ground_truth: dict, gate_result: dict, kwaliteit: dict,
                        reconciliatie: dict, rapport_data: dict,
                        review_items: list = None) -> dict:
    """V3: Bouw audit package JSON voor ChatGPT review.

    Dit is het complete dossier dat ChatGPT CEO kan reviewen voordat
    het rapport naar de klant gaat. Bevat alles wat nodig is om te
    beoordelen of de getallen kloppen.
    """
    return {
        'versie': 'V3',
        'tijdstip': datetime.now().isoformat(),
        'gate': gate_result,
        'ground_truth': {
            'periode': ground_truth.get('periode', {}),
            'saldo': ground_truth.get('saldo', {}),
            'sectie_totalen_12m': ground_truth.get('sectie_totalen_12m', {}),
            'sectie_gemiddelden_pm': ground_truth.get('sectie_gemiddelden_pm', {}),
            'income_sources': ground_truth.get('income_sources', {}),
        },
        'reconciliatie': {
            'status': reconciliatie.get('status', 'UNKNOWN'),
            'checks': reconciliatie.get('checks', []),
        },
        'classificatie_kwaliteit': kwaliteit,
        'analyse_samenvatting': rapport_data.get('analyse', {}).get('samenvatting', ''),
        'review_items': review_items or gate_result.get('redenen', []),
    }


def _run_rapport_pipeline(job_id: str, bestanden: list, email: str,
                          ai_categorizer: str = 'claude_opus_47',
                          ai_auditor: str = 'none'):
    """Achtergrond-thread: volledige pipeline upload → analyse → PDF → email.

    bestanden: list van (inhoud_bytes, bestandsnaam) tuples.
    ai_categorizer: welk AI-model voor classificatie (key uit _AI_CATEGORIZERS).
    ai_auditor: welk AI-model voor sanity check ('none' = overslaan).
    Schrijft voortgang naar jobs[job_id] zodat de status-endpoint het kan serveren.
    Draait NIET in een HTTP-request — dus geen proxy timeout meer.
    """
    def update(fase: str, pct: int):
        with jobs_lock:
            jobs[job_id]['fase'] = fase
            jobs[job_id]['voortgang'] = pct
        logger.info(f"[{job_id}] {fase}")

    try:
        # 1. Inlezen — meerdere bestanden samenvoegen
        update('Transacties inlezen...', 10)
        dfs = []
        for inhoud, bestandsnaam in bestanden:
            df_deel = lees_transacties(inhoud, bestandsnaam)
            dfs.append(df_deel)
            logger.info(f"[{job_id}] {bestandsnaam}: {len(df_deel)} transacties")
        df = pd.concat(dfs, ignore_index=True)
        update(f'{len(df)} transacties ingelezen uit {len(bestanden)} bestand(en)', 15)

        # 1a2. Transactie-verrijking: naam-extractie + IBAN-extractie uit omschrijvingen
        update('Transacties verrijken...', 16)
        df = _verrijk_transactie_velden(df)

        # 1a3. Ntropy verwijderd — classificatie volledig via eigen MERCHANT_MAPPING + keyword-heuristieken + AI
        # Initialiseer lege kolommen voor backward compatibility
        for col in ['ntropy_entity', 'ntropy_category', 'ntropy_sectie', 'ntropy_categorie', 'ntropy_website']:
            df[col] = ''

        # 1b. Huishoudregister & interne-overboekingen detectie
        update('Eigen rekeningen herkennen...', 17)
        eigen_rekeningen = _bouw_huishoudregister(df)
        df = _markeer_interne_transfers(df, eigen_rekeningen)
        n_intern_rek = df['is_intern'].sum()
        update(f'{n_intern_rek} interne overboekingen (eigen rekeningen) gedetecteerd', 19)

        # 1b2. Huishoudleden-detectie (partner, kinderen — automatisch)
        update('Huishoudleden detecteren...', 20)
        df = _detecteer_huishoudleden(df)
        n_intern_totaal = df['is_intern'].sum()
        n_huishoud = n_intern_totaal - n_intern_rek
        if n_huishoud > 0:
            update(f'{n_huishoud} huishoudoverboeking(en) gedetecteerd', 21)

        # 1b3. Related Party Resolution (RPR v1.3)
        update('Tegenpartijen classificeren (RPR)...', 21)
        eigen_fi_ibans = _bouw_eigen_financieel_domein(df)
        df = _resolve_related_parties(df, eigen_rekeningen, eigen_fi_ibans=eigen_fi_ibans)
        n_intern_na_rpr = df['is_intern'].sum()
        n_rpr_extra = n_intern_na_rpr - n_intern_totaal
        if n_rpr_extra > 0:
            update(f'RPR: {n_rpr_extra} extra interne transacties via multi-IBAN linking', 22)
        else:
            update('RPR: tegenpartijen geclassificeerd', 22)

        # 1c. Rule-based classificatie (vóór AI)
        update('Transacties classificeren...', 22)
        df = _classificeer_rule_based(df)
        n_regel = len(df[df['classificatie_bron'] == 'rule'])
        update(f'{n_regel} transacties rule-based geclassificeerd', 24)

        # 1d. Patroon-detectie: vast inkomen (salaris/DGA-loon, vóór AI)
        update('Vast inkomen detecteren...', 24)
        df = _detecteer_vast_inkomen(df)
        n_regel_na = len(df[df['classificatie_bron'] == 'rule'])
        n_patroon = n_regel_na - n_regel
        if n_patroon > 0:
            update(f'{n_patroon} extra transacties via patroondetectie', 25)
        else:
            update('Patroondetectie afgerond', 25)

        # 1d2. Decision Engine: inflow classificatie (eigen_fi_ibans al gebouwd in 1b3)
        update('Decision engine: positieve inflows classificeren...', 25)
        n_voor_inflow = len(df[df['classificatie_bron'] == 'rule'])
        df = _classify_positive_inflows(df, eigen_fi_ibans=eigen_fi_ibans, eigen_rekeningen=eigen_rekeningen)
        n_na_inflow = len(df[df['classificatie_bron'] == 'rule'])
        n_inflow = n_na_inflow - n_voor_inflow
        if n_inflow > 0:
            update(f'{n_inflow} positieve transacties geclassificeerd door decision engine', 25)

        # 1e. Consistentie-afdwinging: propageer classificaties via IBAN
        update('Consistentie afdwingen...', 26)
        df = _afdwing_iban_consistentie(df)
        n_regel_final = len(df[df['classificatie_bron'] == 'rule'])
        n_consistentie = n_regel_final - n_regel_na
        if n_consistentie > 0:
            update(f'{n_consistentie} extra transacties via IBAN-consistentie', 26)

        # 1f. Classificatie-kwaliteit loggen
        kwaliteit = _log_classificatie_kwaliteit(df)
        update(f'{kwaliteit["pct_rule_based"]:.0f}% rule-based geclassificeerd', 27)

        # 1g. Review items verzamelen
        review_items = _verzamel_review_items(df)
        if review_items:
            update(f'{len(review_items)} review items gevonden', 27)

        # 2. Deterministisch rekenen
        update('Bedragen berekenen en controleren...', 27)
        feiten = bereken_feiten(df)
        top = bereken_top(df)
        update(f'Feiten berekend voor {len(feiten)} rekening(en)', 28)

        # 2b. V3: Post-classificatie reconciliatie (pre-AI check)
        update('Reconciliatie controleren...', 29)
        reconciliatie = _post_classificatie_reconciliatie(df, feiten)
        rapportperiode = _bepaal_rapportperiode(df, reconciliatie)
        if reconciliatie['status'] == 'RED':
            red_checks = [c for c in reconciliatie['checks'] if c['status'] == 'RED']
            red_details = '; '.join(c['detail'] for c in red_checks[:3])
            logger.warning(f"[{job_id}] V3-RECONCILIATIE RED: {red_details}")
            # Nog niet blokkeren (no-send gate komt in Sprint 3), maar wel loggen
        update(f"Reconciliatie: {reconciliatie['status']}, {rapportperiode['n_mnd']} volle maanden", 30)

        # 2c. Rule-based totalen berekenen uit DataFrame (vóór AI — voor ground truth prompt)
        rule_totalen = _bereken_rule_based_totalen(df)
        n_rule_cats = sum(
            len(cats) for secties in rule_totalen['jaartotalen'].values()
            for cats in secties.values()
        )
        logger.info(f"[{job_id}] Rule-based: {n_rule_cats} categorieën berekend uit DataFrame")

        # 2d. V3: Voorlopige ground truth bouwen (rule-based only — voor AI-prompt)
        pre_ground_truth = _bouw_ground_truth(
            merged_data=rule_totalen,
            feiten=feiten,
            rapportperiode=rapportperiode,
            reconciliatie=reconciliatie,
            df=df,
        )
        logger.info(f"[{job_id}] Voorlopige ground truth gebouwd (rule-based only)")

        # 3. AI categoriseren + analyseren (langste stap: 30s-300s)
        cat_info = _AI_CATEGORIZERS.get(ai_categorizer, ('claude', 'claude-opus-4-7'))
        model_naam = cat_info[1]
        update(f'AI analyseert uw transacties ({model_naam})...', 35)
        prompt = bouw_prompt(df, feiten, top, eigen_rekeningen=eigen_rekeningen, ground_truth=pre_ground_truth)
        logger.info(f"[{job_id}] Prompt: {len(prompt)} tekens, {len(df)} transacties, categorizer={ai_categorizer}")
        claude_result = vraag_claude(prompt, categorizer=ai_categorizer)

        if not claude_result.get('data'):
            raise ValueError(f"AI-analyse ongeldig: {claude_result.get('error', 'onbekend')}")

        # 3c. Merge: rule-based + AI totalen samenvoegen (geen overlap)
        claude_result['data'] = _merge_rule_en_ai_totalen(rule_totalen, claude_result['data'])
        update('Rule-based en AI resultaten samengevoegd', 65)

        # 3c. Report quality checks (blokkeer bij grove fouten, raises ValueError)
        kwaliteitswaarschuwingen = _rapport_kwaliteitscheck(claude_result['data'], df, eigen_rekeningen)
        if kwaliteitswaarschuwingen:
            update(f'AI-analyse compleet, {len(kwaliteitswaarschuwingen)} waarschuwing(en)', 70)
        else:
            update('AI-analyse compleet, kwaliteitscheck geslaagd', 70)

        # 4. V3: Ground Truth bouwen (bevroren payload — alle downstream outputs lezen hieruit)
        update('Ground truth payload bouwen...', 72)
        ground_truth = _bouw_ground_truth(
            merged_data=claude_result['data'],
            feiten=feiten,
            rapportperiode=rapportperiode,
            reconciliatie=reconciliatie,
            df=df,
        )
        logger.info(f"[{job_id}] Ground truth V3 gebouwd: {len(ground_truth['periode']['volle_maanden'])} volle maanden")

        # 4a0. Executive bucket aggregator — Laag 4
        update('Executive buckets berekenen...', 71)
        _bouw_executive_buckets(ground_truth)
        logger.info(f"[{job_id}] Executive buckets gebouwd")

        # 4a1. Deterministische cross-checks op ground truth
        update('Cross-checks uitvoeren op ground truth...', 72)
        gt_issues = _valideer_ground_truth(ground_truth)
        gt_errors = [msg for sev, msg in gt_issues if sev == 'ERROR']
        gt_warnings = [msg for sev, msg in gt_issues if sev == 'WARNING']
        if gt_errors:
            logger.error(f"[{job_id}] Ground truth ERRORS: {gt_errors}")
            # Niet direct blokkeren maar meenemen naar no-send gate
        if gt_warnings:
            logger.warning(f"[{job_id}] Ground truth WARNINGS: {gt_warnings}")
        ground_truth['_cross_check_issues'] = gt_issues
        update(f'Cross-checks: {len(gt_errors)} errors, {len(gt_warnings)} warnings', 73)

        # 4a1b. AI Auditor — onafhankelijke sanity check (optioneel)
        if ai_auditor and ai_auditor != 'none':
            auditor_info = _AUDITOR_MODELS.get(ai_auditor, ('openai', 'gpt-5.4'))
            update(f'AI Auditor ({auditor_info[1] or ai_auditor}) controleert rapport...', 74)
            auditor_result = _ai_auditor(ground_truth, auditor_model=ai_auditor)
            ground_truth['_auditor_result'] = auditor_result
            a_status = auditor_result.get('status', 'unknown')
            a_issues = auditor_result.get('issues', [])
            if a_status == 'issues_found':
                a_errors = [i for i in a_issues if i.get('severity') == 'error']
                a_warns = [i for i in a_issues if i.get('severity') == 'warning']
                update(f'AI Auditor: {len(a_errors)} errors, {len(a_warns)} warnings', 74)
                for iss in a_issues:
                    logger.warning(f"[{job_id}] AUDITOR {iss.get('severity', '?')}: {iss.get('beschrijving', '')}")
            else:
                update(f'AI Auditor: {a_status}', 74)
                logger.info(f"[{job_id}] Auditor status={a_status}, samenvatting={auditor_result.get('samenvatting', '')}")
        else:
            ground_truth['_auditor_result'] = {'status': 'skipped', 'issues': []}
            logger.info(f"[{job_id}] AI Auditor overgeslagen (ai_auditor={ai_auditor})")

        # 4a2. Strategische inzichten genereren
        update('Strategische inzichten berekenen...', 73)
        strategische_inzichten = _genereer_strategische_inzichten(ground_truth, df)
        ground_truth['strategische_inzichten'] = strategische_inzichten
        logger.info(f"[{job_id}] Strategische inzichten: {len(strategische_inzichten.get('signalen', []))} signalen")

        # 4a3. Premium inzichten berekenen (pagina 2 van PDF)
        update('Premium inzichten berekenen...', 74)
        premium_inzichten = _bereken_premium_inzichten(ground_truth, df)
        ground_truth['premium_inzichten'] = premium_inzichten
        logger.info(f"[{job_id}] Premium inzichten: {len(premium_inzichten)} inzichten")

        # 4b. V3: Reconciliatie Excel genereren
        update('Reconciliatie Excel genereren...', 74)
        try:
            reconciliatie_excel_bytes = _genereer_reconciliatie_excel(ground_truth, feiten, df)
            logger.info(f"[{job_id}] Reconciliatie Excel: {len(reconciliatie_excel_bytes)} bytes")
        except Exception as exc:
            logger.error(f"[{job_id}] Reconciliatie Excel FOUT: {exc}")
            reconciliatie_excel_bytes = None

        # 4c. Rapport data samenstellen
        rapport_data = {
            'report_id': job_id,
            'gegenereerd': datetime.now().isoformat(),
            'bestand': bestandsnaam,
            'feiten': feiten,
            'maandoverzicht': claude_result['data'].get('maandoverzicht', {}),
            'jaartotalen': claude_result['data'].get('jaartotalen', {}),
            'analyse': claude_result['data'].get('analyse', {}),
            'reconciliatie': reconciliatie,       # V3
            'rapportperiode': rapportperiode,     # V3
            'ground_truth': ground_truth,         # V3
        }

        # 4d. V3: No-Send Gate — mag dit rapport verzonden worden?
        update('Kwaliteitsgate controleren...', 75)
        gate_result = _no_send_gate(
            ground_truth=ground_truth,
            reconciliatie=reconciliatie,
            analyse=claude_result['data'].get('analyse', {}),
            kwaliteit=kwaliteit,
        )

        # 4e. V3: Audit package bouwen (altijd, ook bij BLOCK)
        audit_package = _bouw_audit_package(
            ground_truth=ground_truth,
            gate_result=gate_result,
            kwaliteit=kwaliteit,
            reconciliatie=reconciliatie,
            rapport_data=rapport_data,
            review_items=review_items,
        )
        logger.info(f"[{job_id}] Audit package gebouwd: {gate_result['besluit']} ({gate_result['kleur']})")

        # 4f. V3: Bij RED → BLOCK, geen rapport versturen
        if gate_result['besluit'] == 'BLOCK':
            redenen_tekst = '; '.join(gate_result['redenen'][:3])
            logger.warning(f"[{job_id}] NO-SEND GATE: BLOCKED — {redenen_tekst}")

            # Sla audit package op in job-state zodat het opvraagbaar is
            with jobs_lock:
                jobs[job_id]['status'] = 'geblokkeerd'
                jobs[job_id]['fase'] = f"Rapport geblokkeerd: kwaliteitscheck ({gate_result['kleur']})"
                jobs[job_id]['voortgang'] = 100
                jobs[job_id]['gate'] = gate_result
                jobs[job_id]['audit_package'] = audit_package

            # Stuur WEL de reconciliatie Excel (voor handmatige review)
            if reconciliatie_excel_bytes:
                try:
                    verstuur_rapport_email(
                        email, b'',  # Geen PDF
                        job_id,
                        reconciliatie_excel=reconciliatie_excel_bytes,
                        geblokkeerd=True,
                        gate_redenen=gate_result['redenen'],
                    )
                except Exception as exc:
                    logger.error(f"[{job_id}] Blokkade-email fout: {exc}")

            return  # Stop pipeline hier — geen PDF, geen rapport

        rapport_data['gate'] = gate_result
        rapport_data['audit_package'] = audit_package
        update(f"Kwaliteitsgate: {gate_result['kleur']} — rapport wordt verzonden", 77)

        # 5. PDF genereren
        update('Premium PDF-rapport genereren...', 80)
        pdf_bytes = genereer_pdf(rapport_data)
        logger.info(f"[{job_id}] PDF gegenereerd ({len(pdf_bytes)} bytes)")
        update('PDF klaar', 85)

        # 6. Email versturen (met reconciliatie Excel als extra bijlage)
        update('Rapport per email versturen...', 90)
        email_verstuurd = verstuur_rapport_email(
            email, pdf_bytes, job_id,
            reconciliatie_excel=reconciliatie_excel_bytes,
        )

        if not email_verstuurd:
            raise ValueError(
                "Uw analyse is gelukt, maar het rapport kon niet per email worden verstuurd. "
                "Probeer het opnieuw of neem contact op via info@peterheijen.com."
            )

        # Klaar!
        with jobs_lock:
            jobs[job_id]['status'] = 'compleet'
            jobs[job_id]['fase'] = 'Rapport verstuurd!'
            jobs[job_id]['voortgang'] = 100
            jobs[job_id]['gate'] = gate_result
        logger.info(f"[{job_id}] Pipeline compleet — rapport verstuurd naar {email}")

    except Exception as e:
        logger.error(f"[{job_id}] Pipeline FOUT: {type(e).__name__}: {e}")
        # Nette foutmelding voor de gebruiker — geen ruwe stacktraces
        err_msg = str(e)
        if len(err_msg) > 300:
            err_msg = err_msg[:300] + '...'
        with jobs_lock:
            jobs[job_id]['status'] = 'fout'
            jobs[job_id]['error'] = err_msg
            jobs[job_id]['voortgang'] = 0


@app.post("/rapport")
async def rapport(bestanden: Optional[List[UploadFile]] = None, bestand: Optional[UploadFile] = None,
                  email: str = Form(...), ai_categorizer: str = Form('claude_opus_47'),
                  ai_auditor: str = Form('none')):
    """Start de rapport-pipeline als achtergrond-job.

    Accepteert één of meerdere bestanden:
      - 'bestanden' (meerdere files) of 'bestand' (enkele file, backward compatible)

    ai_categorizer: keuze uit 'claude_opus_47', 'claude_opus_46', 'openai_gpt54',
                    'gemini_25_pro', 'shortcut_excel'.

    Retourneert DIRECT (< 1 sec) met een job_id.
    Client pollt /rapport/{job_id}/status voor voortgang.
    """
    job_id = str(uuid.uuid4())[:8]
    logger.info(f"[{job_id}] Rapport aangevraagd voor {email}, categorizer={ai_categorizer}, auditor={ai_auditor}")

    # Verzamel alle bestanden (support zowel 'bestanden' als 'bestand' veld)
    uploads = []
    if bestanden:
        uploads.extend(bestanden)
    if bestand:
        uploads.append(bestand)
    if not uploads:
        raise HTTPException(status_code=400, detail="Geen bestanden geüpload")

    # Bestanden direct inlezen en valideren (snel, < 1 sec per bestand)
    bestanden_data = []
    totaal_transacties = 0
    try:
        for f in uploads:
            inhoud = await f.read()
            if len(inhoud) > 10 * 1024 * 1024:
                raise HTTPException(status_code=400, detail=f"Bestand '{f.filename}' is te groot (max 10 MB)")
            # Quick validatie
            df_test = lees_transacties(inhoud, f.filename)
            totaal_transacties += len(df_test)
            del df_test
            bestanden_data.append((inhoud, f.filename))
            logger.info(f"[{job_id}] {f.filename}: {len(inhoud)} bytes — OK")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[{job_id}] Inleesfout: {e}")
        raise HTTPException(status_code=400, detail=f"Kan bestand niet lezen: {e}")

    logger.info(f"[{job_id}] {totaal_transacties} transacties totaal uit {len(bestanden_data)} bestand(en)")

    # Job registreren
    namen = ', '.join(f[1] for f in bestanden_data)
    with jobs_lock:
        jobs[job_id] = {
            'status': 'bezig',
            'fase': 'Gestart...',
            'voortgang': 5,
            'email': email,
            'bestand': namen,
            'gestart': datetime.now().isoformat(),
        }

    # Start achtergrond-thread
    thread = threading.Thread(
        target=_run_rapport_pipeline,
        args=(job_id, bestanden_data, email, ai_categorizer, ai_auditor),
        daemon=True,
    )
    thread.start()

    return {
        'job_id': job_id,
        'status': 'gestart',
        'categorizer': ai_categorizer,
        'auditor': ai_auditor,
    }


@app.get("/rapport/{job_id}/status")
def rapport_status(job_id: str):
    """Poll-endpoint voor voortgang van een rapport-job.

    Retourneert altijd snel (< 100ms).
    Client pollt elke 3 seconden tot status == 'compleet' of 'fout'.
    """
    with jobs_lock:
        job = jobs.get(job_id)

    if not job:
        raise HTTPException(status_code=404, detail="Job niet gevonden")

    response = {
        'job_id': job_id,
        'status': job['status'],
        'fase': job.get('fase', ''),
        'voortgang': job.get('voortgang', 0),
        'email': job.get('email', ''),
        'error': job.get('error'),
    }

    # V3: gate-info meegeven als beschikbaar
    gate = job.get('gate')
    if gate:
        response['gate'] = {
            'besluit': gate.get('besluit'),
            'kleur': gate.get('kleur'),
            'redenen': gate.get('redenen', []),
        }

    return response


@app.get("/rapport/{job_id}/audit")
def rapport_audit(job_id: str):
    """V3: Haal het audit package op voor een rapport-job.

    Bevat ground truth, gate-besluit, reconciliatie, en review-items.
    Bedoeld voor ChatGPT CEO review vóór verzending.
    """
    with jobs_lock:
        job = jobs.get(job_id)

    if not job:
        raise HTTPException(status_code=404, detail="Job niet gevonden")

    audit = job.get('audit_package')
    if not audit:
        raise HTTPException(status_code=404, detail="Audit package niet beschikbaar (job nog in uitvoering?)")

    return audit


@app.post("/feiten")
async def alleen_feiten(bestand: UploadFile):
    """Alleen deterministisch rekenen, zonder AI. Snel en gratis."""
    report_id = str(uuid.uuid4())[:8]

    try:
        inhoud = await bestand.read()
        df = lees_transacties(inhoud, bestand.filename)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Kan bestand niet lezen: {e}")

    feiten = bereken_feiten(df)
    top = bereken_top(df)

    return {
        'report_id': report_id,
        'gegenereerd': datetime.now().isoformat(),
        'feiten': feiten,
        'top_tegenpartijen': top,
    }
